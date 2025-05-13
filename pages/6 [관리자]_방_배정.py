import re
import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from collections import Counter
import random
import time
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment

# 세션 상태 초기화
if "data_loaded" not in st.session_state:
    st.session_state["data_loaded"] = False
if "df_room_request" not in st.session_state:
    st.session_state["df_room_request"] = pd.DataFrame(columns=["이름", "분류", "날짜정보"])
if "room_settings" not in st.session_state:
    st.session_state["room_settings"] = {
        "830_room_select": ['1', '2', '4', '7'],
        "900_room_select": ['10', '11', '12'],
        "930_room_select": ['5', '6', '8'],
        "1000_room_select": ['3', '9'],
        "1330_room_select": ['2', '3', '4', '9']
    }

# Google Sheets 클라이언트 초기화
def get_gspread_client():
    scope = ["https://www.googleapis.com/auth/spreadsheets"]
    service_account_info = dict(st.secrets["gspread"])
    service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
    credentials = Credentials.from_service_account_info(service_account_info, scopes=scope)
    return gspread.authorize(credentials)

# Google Sheets 업데이트 함수
def update_sheet_with_retry(worksheet, data, retries=5, delay=10):
    for attempt in range(retries):
        try:
            worksheet.clear()
            worksheet.update('A1', data, value_input_option='RAW')
            return
        except Exception as e:
            if "Quota exceeded" in str(e):
                st.warning(f"API 쿼터 초과, {delay}초 후 재시도 ({attempt+1}/{retries})")
                time.sleep(delay)
            else:
                st.error(f"업데이트 실패, {delay}초 후 재시도 ({attempt+1}/{retries}): {str(e)}")
                time.sleep(delay)
    st.error("Google Sheets 업데이트 실패: 재시도 횟수 초과")

# 데이터 로드 (캐싱 사용) - 캐시 문제 방지
def load_data_page6(month_str):
    # 캐시 강제 갱신
    st.cache_data.clear()
    
    # load_data_page6_no_cache 호출
    result = load_data_page6_no_cache(month_str)
    
    # 반환값 디버깅
    if len(result) != 3:
        st.error(f"Expected 3 return values, but got {len(result)}. Returned: {result}")
        st.stop()
    
    return result

# 데이터 로드 (캐싱 미사용)
def load_data_page6_no_cache(month_str):
    gc = get_gspread_client()
    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
    
    try:
        worksheet_schedule = sheet.worksheet(f"{month_str} 스케쥴")
        df_schedule = pd.DataFrame(worksheet_schedule.get_all_records())
    except Exception as e:
        st.error(f"스케줄 시트를 불러오는 데 실패: {e}")
        st.stop()
    
    try:
        worksheet_room_request = sheet.worksheet(f"{month_str} 방배정 요청")
        df_room_request = pd.DataFrame(worksheet_room_request.get_all_records())
        if "우선순위" in df_room_request.columns:
            df_room_request = df_room_request.drop(columns=["우선순위"])
    except:
        worksheet_room_request = sheet.add_worksheet(f"{month_str} 방배정 요청", rows=100, cols=3)
        worksheet_room_request.append_row(["이름", "분류", "날짜정보"])
        df_room_request = pd.DataFrame(columns=["이름", "분류", "날짜정보"])
    
    # 누적 시트 로드 - 첫 번째 열을 이름으로 처리
    try:
        worksheet_cumulative = sheet.worksheet(f"{month_str} 누적")
        df_cumulative = pd.DataFrame(worksheet_cumulative.get_all_records())
        if df_cumulative.empty:
            st.warning(f"{month_str} 누적 시트가 비어 있습니다. 빈 DataFrame으로 초기화합니다.")
            df_cumulative = pd.DataFrame(columns=[f"{month_str}", "오전누적", "오후누적", "오전당직 (온콜)", "오후당직"])
        else:
            # 첫 번째 열 이름을 "이름"으로 변경
            df_cumulative.rename(columns={f"{month_str}": "이름"}, inplace=True)
    except:
        st.warning(f"{month_str} 누적 시트가 없습니다. 빈 DataFrame으로 초기화합니다.")
        df_cumulative = pd.DataFrame(columns=["이름", "오전누적", "오후누적", "오전당직 (온콜)", "오후당직"])
    
    st.session_state["df_schedule"] = df_schedule
    st.session_state["df_room_request"] = df_room_request
    st.session_state["worksheet_room_request"] = worksheet_room_request
    st.session_state["df_cumulative"] = df_cumulative  # 누적 데이터 저장
    st.session_state["data_loaded"] = True
    
    # 정확히 3개 값만 반환
    result = (df_schedule, df_room_request, worksheet_room_request)
    return result

# 근무 가능 일자 계산
@st.cache_data
def get_user_available_dates(name, df_schedule, month_start, month_end):
    available_dates = []
    weekday_map = {0: "월", 1: "화", 2: "수", 3: "목", 4: "금", 5: "토", 6: "일"}
    
    personnel_columns = [str(i) for i in range(1, 13)] + [f'오후{i}' for i in range(1, 5)]
    all_personnel = set()
    for col in personnel_columns:
        for val in df_schedule[col].dropna():
            all_personnel.add(str(val).strip())
    if name not in all_personnel:
        st.warning(f"{name}이 df_schedule의 근무자 목록에 없습니다. 데이터 확인 필요: {sorted(all_personnel)}")
    
    for _, row in df_schedule.iterrows():
        date_str = row['날짜']
        try:
            if "월" in date_str:
                date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=2025).date()
            else:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            continue
        if month_start <= date_obj <= month_end and row['요일'] not in ['토요일', '일요일']:
            morning_personnel = [str(row[str(i)]).strip() for i in range(1, 13) if pd.notna(row[str(i)]) and row[str(i)]]
            afternoon_personnel = [str(row[f'오후{i}']).strip() for i in range(1, 5) if pd.notna(row[f'오후{i}']) and row[f'오후{i}']]
            display_date = f"{date_obj.month}월 {date_obj.day}일({weekday_map[date_obj.weekday()]})"
            save_date_am = f"{date_obj.strftime('%Y-%m-%d')} (오전)"
            save_date_pm = f"{date_obj.strftime('%Y-%m-%d')} (오후)"
            if name in morning_personnel:
                available_dates.append((date_obj, f"{display_date} 오전", save_date_am))
            if name in afternoon_personnel:
                available_dates.append((date_obj, f"{display_date} 오후", save_date_pm))
    
    available_dates.sort(key=lambda x: x[0])
    sorted_dates = [(display_str, save_str) for _, display_str, save_str in available_dates]
    if not sorted_dates:
        st.warning(f"{name}의 근무 가능 일자가 없습니다. df_schedule 데이터를 확인하세요.")
    return sorted_dates

# 요청 저장 (df_room_request용)
def save_to_gsheet(name, categories, dates, month_str, worksheet):
    gc = get_gspread_client()
    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
    worksheet = sheet.worksheet(f"{month_str} 방배정 요청")
    df = pd.DataFrame(worksheet.get_all_records())
    if "우선순위" in df.columns:
        df = df.drop(columns=["우선순위"])
    
    new_rows = []
    for date in dates:
        for cat in categories:
            new_rows.append({"이름": name, "분류": cat, "날짜정보": date})
    
    df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
    update_sheet_with_retry(worksheet, [df.columns.tolist()] + df.values.tolist())
    return df

# df_schedule_md 생성
def create_df_schedule_md(df_schedule):
    df_schedule_md = df_schedule.copy().fillna('')
    for idx, row in df_schedule_md.iterrows():
        date_str = row['날짜']
        oncall_worker = row['오전당직(온콜)']
        
        try:
            if isinstance(date_str, (float, int)):
                date_str = str(int(date_str))
            date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=2025) if "월" in date_str else datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError as e:
            st.error(f"날짜 파싱 오류: {date_str}, 오류: {str(e)}")
            continue
        
        afternoon_cols = ['오후1', '오후2', '오후3', '오후4', '오후5']
        if all(row[col] == '' for col in afternoon_cols):
            df_schedule_md.at[idx, '오전당직(온콜)'] = ''
            continue
        
        if pd.isna(oncall_worker) or oncall_worker == '':
            oncall_worker = ''
            df_schedule_md.at[idx, '오전당직(온콜)'] = ''
        
        if oncall_worker:
            morning_cols = [str(i) for i in range(1, 13)]
            for col in morning_cols + afternoon_cols:
                if row[col] == oncall_worker:
                    df_schedule_md.at[idx, col] = ''
        
        morning_cols = [str(i) for i in range(1, 13)]
        morning_workers = [row[col] for col in morning_cols if row[col]]
        if len(morning_workers) > 11:
            morning_workers = morning_workers[:11]
        morning_workers.extend([''] * (11 - len(morning_workers)))
        for i, col in enumerate([str(i) for i in range(1, 12)], 1):
            df_schedule_md.at[idx, col] = morning_workers[i-1]
        
        afternoon_workers = [row[col] for col in afternoon_cols if row[col]]
        if len(afternoon_workers) > 4:
            afternoon_workers = afternoon_workers[:4]
        afternoon_workers.extend([''] * (4 - len(afternoon_workers)))
        for i, col in enumerate(['오후1', '오후2', '오후3', '오후4'], 1):
            df_schedule_md.at[idx, col] = afternoon_workers[i-1]
    
    df_schedule_md = df_schedule_md.drop(columns=['12', '오후5'], errors='ignore')
    return df_schedule_md

# 메인
month_str = "2025년 04월"
next_month_start = date(2025, 4, 1)
next_month_end = date(2025, 4, 30)

# 로그인 체크
if "login_success" not in st.session_state or not st.session_state["login_success"]:
    st.warning("⚠️ Home 페이지에서 비밀번호와 사번을 먼저 입력해주세요.")
    st.stop()

# 관리자 권한 체크
if not st.session_state.get("is_admin_authenticated", False):
    st.warning("⚠️ 관리자 권한이 없습니다.")
    st.stop()

# 사이드바
st.sidebar.write(f"현재 사용자: {st.session_state['name']} ({str(st.session_state['employee_id']).zfill(5)})")
if st.sidebar.button("로그아웃"):
    st.session_state.clear()
    st.success("로그아웃되었습니다.")
    st.rerun()

# 데이터 로드 호출
df_schedule, df_room_request, worksheet_room_request = load_data_page6(month_str)
st.session_state["df_room_request"] = df_room_request
st.session_state["worksheet_room_request"] = worksheet_room_request

# df_schedule_md 초기화
if "df_schedule_md" not in st.session_state:
    st.session_state["df_schedule_md"] = create_df_schedule_md(df_schedule)

# 새로고침 버튼
if st.button("🔄 새로고침 (R)"):
    st.cache_data.clear()
    df_schedule, df_room_request, worksheet_room_request = load_data_page6_no_cache(month_str)
    st.session_state["df_schedule"] = df_schedule
    st.session_state["df_room_request"] = df_room_request
    st.session_state["worksheet_room_request"] = worksheet_room_request
    st.session_state["df_schedule_md"] = create_df_schedule_md(df_schedule)
    st.success("데이터가 새로고침되었습니다.")
    st.rerun()

# 근무자 명단
st.subheader("📋 근무자 명단")
st.dataframe(st.session_state["df_schedule_md"])

# 방 설정 UI
st.divider()
st.subheader("📋 방 설정")
room_options = [str(i) for i in range(1, 13)]

st.write(" ")
st.markdown("**🔷 8:30 시간대**")
col1, col2, col3 = st.columns([1, 2, 1])
with col1:
    num_830 = st.number_input("방 개수", min_value=0, value=4, key="830_rooms")
with col2:
    if len(st.session_state["room_settings"]["830_room_select"]) > num_830:
        st.session_state["room_settings"]["830_room_select"] = st.session_state["room_settings"]["830_room_select"][:num_830]
    rooms_830 = st.multiselect(
        "방 번호",
        room_options,
        default=st.session_state["room_settings"]["830_room_select"],
        max_selections=num_830,
        key="830_room_select"
    )
    if len(rooms_830) < num_830:
        st.warning(f"방 개수({num_830})에 맞게 방 번호를 {num_830}개 선택해주세요. 현재 {len(rooms_830)}개 선택됨.")
    st.session_state["room_settings"]["830_room_select"] = rooms_830
with col3:
    duty_830_options = rooms_830 if rooms_830 else room_options
    duty_830 = st.selectbox("당직방", duty_830_options, index=0, key="830_duty")
    st.session_state["room_settings"]["830_duty"] = duty_830

st.markdown("**🔷 9:00 시간대**")
col1, col2 = st.columns([1, 3])
with col1:
    num_900 = st.number_input("방 개수", min_value=0, value=3, key="900_rooms")
with col2:
    if len(st.session_state["room_settings"]["900_room_select"]) > num_900:
        st.session_state["room_settings"]["900_room_select"] = st.session_state["room_settings"]["900_room_select"][:num_900]
    rooms_900 = st.multiselect(
        "방 번호",
        room_options,
        default=st.session_state["room_settings"]["900_room_select"],
        max_selections=num_900,
        key="900_room_select"
    )
    if len(rooms_900) < num_900:
        st.warning(f"방 개수({num_900})에 맞게 방 번호를 {num_900}개 선택해주세요. 현재 {len(rooms_900)}개 선택됨.")
    st.session_state["room_settings"]["900_room_select"] = rooms_900

st.markdown("**🔷 9:30 시간대**")
col1, col2 = st.columns([1, 3])
with col1:
    num_930 = st.number_input("방 개수", min_value=0, value=3, key="930_rooms")
with col2:
    if len(st.session_state["room_settings"]["930_room_select"]) > num_930:
        st.session_state["room_settings"]["930_room_select"] = st.session_state["room_settings"]["930_room_select"][:num_930]
    rooms_930 = st.multiselect(
        "방 번호",
        room_options,
        default=st.session_state["room_settings"]["930_room_select"],
        max_selections=num_930,
        key="930_room_select"
    )
    if len(rooms_930) < num_930:
        st.warning(f"방 개수({num_930})에 맞게 방 번호를 {num_930}개 선택해주세요. 현재 {len(rooms_930)}개 선택됨.")
    st.session_state["room_settings"]["930_room_select"] = rooms_930

st.markdown("**🔷 10:00 시간대**")
col1, col2 = st.columns([1, 3])
with col1:
    num_1000 = st.number_input("방 개수", min_value=0, value=2, key="1000_rooms")
with col2:
    if len(st.session_state["room_settings"]["1000_room_select"]) > num_1000:
        st.session_state["room_settings"]["1000_room_select"] = st.session_state["room_settings"]["1000_room_select"][:num_1000]
    rooms_1000 = st.multiselect(
        "방 번호",
        room_options,
        default=st.session_state["room_settings"]["1000_room_select"],
        max_selections=num_1000,
        key="1000_room_select"
    )
    if len(rooms_1000) < num_1000:
        st.warning(f"방 개수({num_1000})에 맞게 방 번호를 {num_1000}개 선택해주세요. 현재 {len(rooms_1000)}개 선택됨.")
    st.session_state["room_settings"]["1000_room_select"] = rooms_1000

st.markdown("**🔶 13:30 시간대**")
col1, col2 = st.columns([3, 1])
with col1:
    num_1330 = 4
    if len(st.session_state["room_settings"]["1330_room_select"]) > num_1330:
        st.session_state["room_settings"]["1330_room_select"] = st.session_state["room_settings"]["1330_room_select"][:num_1330]
    rooms_1330 = st.multiselect(
        "방 번호",
        room_options,
        default=st.session_state["room_settings"]["1330_room_select"],
        max_selections=num_1330,
        key="1330_room_select"
    )
    if len(rooms_1330) < num_1330:
        st.warning(f"방 개수({num_1330})에 맞게 방 번호를 {num_1330}개 선택해주세요. 현재 {len(rooms_1330)}개 선택됨.")
    st.session_state["room_settings"]["1330_room_select"] = rooms_1330
with col2:
    duty_1330_options = rooms_1330 if rooms_1330 else room_options
    duty_1330 = st.selectbox("당직방", duty_1330_options, index=0, key="1330_duty")
    st.session_state["room_settings"]["1330_duty"] = duty_1330

# 중복 방 번호 검증
all_selected_rooms = (
    st.session_state["room_settings"]["830_room_select"] +
    st.session_state["room_settings"]["900_room_select"] +
    st.session_state["room_settings"]["930_room_select"] +
    st.session_state["room_settings"]["1000_room_select"] +
    st.session_state["room_settings"]["1330_room_select"]
)

# 배정 요청 입력 UI
st.divider()
st.subheader("📋 배정 요청 관리")
st.write("- 모든 인원의 배정 요청(고정 및 우선)을 추가 및 수정할 수 있습니다.")

요청분류 = ["1번방", "2번방", "3번방", "4번방", "5번방", "6번방", "7번방", "8번방", "9번방", "10번방", "11번방",
           "8:30", "9:00", "9:30", "10:00", "당직 아닌 이른방", "이른방 제외", "늦은방 제외", "오후 당직 제외"]

st.write(" ")
st.markdown("**🟢 방 배정 요청 추가**")
col1, col2, col3 = st.columns([2, 2, 3])
with col1:
    names = sorted([str(name).strip() for name in df_schedule.iloc[:, 2:].stack().dropna().unique() if str(name).strip()])
    name = st.selectbox("근무자", names, key="request_employee_select")
with col2:
    categories = st.multiselect("요청 분류", 요청분류, key="request_category_select")
with col3:
    st.cache_data.clear()
    available_dates = get_user_available_dates(name, df_schedule, next_month_start, next_month_end)
    date_options = [display_str for display_str, _ in available_dates]
    dates = st.multiselect("요청 일자", date_options, key="request_date_select")
    selected_save_dates = [save_str for display_str, save_str in available_dates if display_str in dates]

if st.button("📅 추가", key="request_add_button"):
    if not categories or not selected_save_dates:
        st.error("요청 분류와 날짜를 선택해주세요.")
    else:
        new_rows = []
        for date in selected_save_dates:
            for cat in categories:
                new_rows.append({"이름": name, "분류": cat, "날짜정보": date})
        df_room_request = pd.concat([df_room_request, pd.DataFrame(new_rows)], ignore_index=True)
        st.session_state["df_room_request"] = df_room_request
        try:
            update_sheet_with_retry(st.session_state["worksheet_room_request"], [df_room_request.columns.tolist()] + df_room_request.values.tolist())
            st.cache_data.clear()
            st.success("방 배정 요청 저장 완료!")
        except Exception as e:
            st.error(f"Google Sheets 업데이트 실패: {str(e)}")
            st.write("로컬 df_room_request는 업데이트되었습니다. 아래에서 확인 후 Google Sheets 동기화를 다시 시도해주세요.")
            st.dataframe(st.session_state["df_room_request"])

# 방 배정 요청 삭제 섹션
st.write(" ")
st.markdown("**🔴 방 배정 요청 삭제**")
if not df_room_request.empty:
    col0, col1 = st.columns([1, 2])
    with col0:
        selected_employee = st.selectbox("근무자 선택", df_room_request["이름"].unique(), key="delete_request_employee_select")
    with col1:
        df_request_filtered = df_room_request[df_room_request["이름"] == selected_employee]
        if not df_request_filtered.empty:
            options = [f"{row['분류']} - {row['날짜정보']}" for _, row in df_request_filtered.iterrows()]
            selected_items = st.multiselect("삭제할 항목)", options, key="delete_request_select")
        else:
            st.info("📍 선택한 근무자에 대한 방 배정 요청이 없습니다.")
            selected_items = []
    
    if st.button("📅 삭제", key="request_delete_button"):
        if selected_items:
            indices = []
            for item in selected_items:
                for idx, row in df_request_filtered.iterrows():
                    if f"{row['분류']} - {row['날짜정보']}" == item:
                        indices.append(idx)
            df_room_request = df_room_request.drop(indices).reset_index(drop=True)
            st.session_state["df_room_request"] = df_room_request
            try:
                update_sheet_with_retry(st.session_state["worksheet_room_request"], [df_room_request.columns.tolist()] + df_room_request.values.tolist())
                st.cache_data.clear()
                st.success("선택한 방 배정 요청 삭제 완료!")
            except Exception as e:
                st.error(f"Google Sheets 업데이트 실패: {str(e)}")
                st.write("로컬 df_room_request는 업데이트되었습니다. 아래에서 확인 후 Google Sheets 동기화를 다시 시도해주세요.")
                st.dataframe(st.session_state["df_room_request"])
else:
    st.info("📍 방 배정 요청이 없습니다.")

st.write(" ")
st.markdown("**🙋‍♂️ 현재 방 배정 요청 목록**")
if df_room_request.empty:
    st.info("☑️ 현재 방 배정 요청이 없습니다.")
else:
    st.dataframe(df_room_request, use_container_width=True)

# 날짜정보 파싱 함수
def parse_date_info(date_info):
    try:
        date_part = date_info.split('(')[0].strip()
        date_obj = datetime.strptime(date_part, '%Y-%m-%d')
        is_morning = '오전' in date_info
        parsed_date = date_obj.strftime('%Y-%m-%d')
        return parsed_date, is_morning
    except ValueError as e:
        st.warning(f"Failed to parse date_info: {date_info}, error: {str(e)}")
        return None, False

# random_assign 함수 - 오전/오후 당직 분리
def random_assign(personnel, slots, request_assignments, time_groups, total_stats, morning_personnel, afternoon_personnel, afternoon_duty_counts):
    assignment = [None] * len(slots)
    assigned_personnel_morning = set()  # 오전 시간대 배정된 인원 추적
    assigned_personnel_afternoon = set()  # 오후 시간대 배정된 인원 추적
    daily_stats = {
        'early': Counter(),
        'late': Counter(),
        'morning_duty': Counter(),  # 오전 당직 (8:30)
        'afternoon_duty': Counter(),  # 오후 당직 (13:30)
        'rooms': {str(i): Counter() for i in range(1, 13)}
    }

    # 슬롯 분류
    morning_slots = [s for s in slots if s.startswith(('8:30', '9:00', '9:30', '10:00')) and '_당직' not in s]
    afternoon_slots = [s for s in slots if s.startswith('13:30')]
    afternoon_duty_slot = '13:30(2)_당직'  # 오후당직 슬롯

    # 1. 배정 요청 먼저 처리 (중복 배정 방지, 균등 배정 고려)
    for slot, person in request_assignments.items():
        if person in personnel and slot in slots:
            slot_idx = slots.index(slot)
            if assignment[slot_idx] is None:
                # 시간대 제약 확인
                if (slot in morning_slots and person in morning_personnel) or \
                   (slot in afternoon_slots and person in afternoon_personnel):
                    # 오전/오후 중복 체크
                    if slot in morning_slots and person in assigned_personnel_morning:
                        st.warning(f"중복 배정 방지: {person}은 이미 오전 시간대({slot})에 배정됨")
                        continue
                    if slot in afternoon_slots and person in assigned_personnel_afternoon:
                        st.warning(f"중복 배정 방지: {person}은 이미 오후 시간대({slot})에 배정됨")
                        continue

                    assignment[slot_idx] = person
                    if slot in morning_slots:
                        assigned_personnel_morning.add(person)
                    else:
                        assigned_personnel_afternoon.add(person)
                    room_num = slot.split('(')[1].split(')')[0]
                    daily_stats['rooms'][room_num][person] += 1
                    if slot.startswith('8:30') and '_당직' not in slot:
                        daily_stats['early'][person] += 1
                    elif slot.startswith('10:00'):
                        daily_stats['late'][person] += 1
                    if slot.startswith('8:30') and slot.endswith('_당직'):
                        daily_stats['morning_duty'][person] += 1
                    elif slot.startswith('13:30') and slot.endswith('_당직'):
                        daily_stats['afternoon_duty'][person] += 1
                else:
                    st.warning(f"배정 요청 무시: {person}은 {slot} 시간대({'오전' if slot in morning_slots else '오후'})에 근무 불가")
            else:
                st.warning(f"배정 요청 충돌: {person}을 {slot}에 배정할 수 없음. 이미 배정됨: {assignment[slot_idx]}")

    # 2. 오후당직 우선 배정 (누적 시트 기반, 당직 균등 배정)
    afternoon_duty_slot_idx = slots.index(afternoon_duty_slot) if afternoon_duty_slot in slots else None
    if afternoon_duty_slot_idx is not None and assignment[afternoon_duty_slot_idx] is None:
        # 오후당직 배정 가능한 인원: afternoon_personnel 중 아직 오후에 배정되지 않은 인원
        available_personnel = [p for p in afternoon_personnel if p not in assigned_personnel_afternoon]
        # 오후당직 횟수가 있는 인원만 대상으로
        candidates = [p for p in available_personnel if p in afternoon_duty_counts and afternoon_duty_counts[p] > 0]
        
        if candidates:
            # 오후 당직 횟수 기준 균등 배정
            best_person = None
            min_duty_count = float('inf')
            for person in candidates:
                duty_count = total_stats['afternoon_duty'][person] + daily_stats['afternoon_duty'][person]
                if duty_count < min_duty_count:
                    min_duty_count = duty_count
                    best_person = person
            if best_person:
                assignment[afternoon_duty_slot_idx] = best_person
                assigned_personnel_afternoon.add(best_person)
                room_num = afternoon_duty_slot.split('(')[1].split(')')[0]
                daily_stats['rooms'][room_num][best_person] += 1
                daily_stats['afternoon_duty'][best_person] += 1
                # 오후당직 횟수 감소
                afternoon_duty_counts[best_person] -= 1
                if afternoon_duty_counts[best_person] <= 0:
                    del afternoon_duty_counts[best_person]

    # 3. 남은 인원 배정 (오전/오후 구분, 공란 방지, 독립적 균등 배정)
    morning_remaining = [p for p in morning_personnel if p not in assigned_personnel_morning]
    afternoon_remaining = [p for p in afternoon_personnel if p not in assigned_personnel_afternoon]
    remaining_slots = [i for i, a in enumerate(assignment) if a is None]
    
    # 오전 슬롯 배정
    morning_slot_indices = [i for i in remaining_slots if slots[i] in morning_slots]
    while morning_remaining and morning_slot_indices:
        best_person = None
        best_slot_idx = None
        min_score = float('inf')
        
        for slot_idx in morning_slot_indices:
            if assignment[slot_idx] is not None:
                continue
            slot = slots[slot_idx]
            room_num = slot.split('(')[1].split(')')[0]
            
            for person in morning_remaining:
                # 방별 배정 균등성
                room_count = total_stats['rooms'][room_num][person] + daily_stats['rooms'][room_num][person]
                # 이른방 균등성 (8:30)
                early_count = total_stats['early'][person] + daily_stats['early'][person] if slot.startswith('8:30') and '_당직' not in slot else float('inf')
                # 늦은방 균등성 (10:00)
                late_count = total_stats['late'][person] + daily_stats['late'][person] if slot.startswith('10:00') else float('inf')
                
                # 해당 슬롯 유형에 따라 스코어 선택
                if slot.startswith('8:30') and '_당직' not in slot:
                    score = early_count  # 이른방 배정 시 이른방 횟수만 고려
                elif slot.startswith('10:00'):
                    score = late_count  # 늦은방 배정 시 늦은방 횟수만 고려
                else:
                    score = room_count  # 나머지 슬롯은 방별 횟수만 고려
                
                if score < min_score:
                    min_score = score
                    best_person = person
                    best_slot_idx = slot_idx
        
        if best_slot_idx is None or best_person is None:
            st.warning(f"오전 슬롯 배정 불가: 더 이상 배정 가능한 인원 없음")
            break
        
        slot = slots[best_slot_idx]
        assignment[best_slot_idx] = best_person
        assigned_personnel_morning.add(best_person)
        morning_remaining.remove(best_person)
        morning_slot_indices.remove(best_slot_idx)
        remaining_slots.remove(best_slot_idx)
        room_num = slot.split('(')[1].split(')')[0]
        daily_stats['rooms'][room_num][best_person] += 1
        if slot.startswith('8:30') and '_당직' not in slot:
            daily_stats['early'][best_person] += 1
        elif slot.startswith('10:00'):
            daily_stats['late'][best_person] += 1
        if slot.startswith('8:30') and slot.endswith('_당직'):
            daily_stats['morning_duty'][best_person] += 1

    # 오후 슬롯 배정
    afternoon_slot_indices = [i for i in remaining_slots if slots[i] in afternoon_slots]
    while afternoon_remaining and afternoon_slot_indices:
        best_person = None
        best_slot_idx = None
        min_score = float('inf')
        
        for slot_idx in afternoon_slot_indices:
            if assignment[slot_idx] is not None:
                continue
            slot = slots[slot_idx]
            room_num = slot.split('(')[1].split(')')[0]
            
            for person in afternoon_remaining:
                # 방별 배정 균등성
                room_count = total_stats['rooms'][room_num][person] + daily_stats['rooms'][room_num][person]
                # 당직 배정 균등성
                duty_count = total_stats['afternoon_duty'][person] + daily_stats['afternoon_duty'][person] if slot.endswith('_당직') else float('inf')
                
                # 해당 슬롯 유형에 따라 스코어 선택
                if slot.endswith('_당직'):
                    score = duty_count  # 당직 슬롯은 오후 당직 횟수만 고려
                else:
                    score = room_count  # 나머지 슬롯은 방별 횟수만 고려
                
                if score < min_score:
                    min_score = score
                    best_person = person
                    best_slot_idx = slot_idx
        
        if best_slot_idx is None or best_person is None:
            st.warning(f"오후 슬롯 배정 불가: 더 이상 배정 가능한 인원 없음")
            break
        
        slot = slots[best_slot_idx]
        assignment[best_slot_idx] = best_person
        assigned_personnel_afternoon.add(best_person)
        afternoon_remaining.remove(best_person)
        afternoon_slot_indices.remove(best_slot_idx)
        room_num = slot.split('(')[1].split(')')[0]
        daily_stats['rooms'][room_num][best_person] += 1
        if slot.endswith('_당직'):
            daily_stats['afternoon_duty'][best_person] += 1

    # 모든 슬롯 채우기 (공란 방지, 독립적 균등 배정 고려)
    for slot_idx in range(len(slots)):
        if assignment[slot_idx] is None:
            slot = slots[slot_idx]
            # 오전/오후 인원 중 가능한 인원 선택
            available_personnel = morning_personnel if slot in morning_slots else afternoon_personnel
            assigned_set = assigned_personnel_morning if slot in morning_slots else assigned_personnel_afternoon
            candidates = [p for p in available_personnel if p not in assigned_set]
            
            if candidates:
                room_num = slot.split('(')[1].split(')')[0]
                best_person = None
                min_score = float('inf')
                for person in candidates:
                    early_count = total_stats['early'][person] + daily_stats['early'][person] if slot.startswith('8:30') and '_당직' not in slot else float('inf')
                    late_count = total_stats['late'][person] + daily_stats['late'][person] if slot.startswith('10:00') else float('inf')
                    morning_duty_count = total_stats['morning_duty'][person] + daily_stats['morning_duty'][person] if slot.startswith('8:30') and slot.endswith('_당직') else float('inf')
                    afternoon_duty_count = total_stats['afternoon_duty'][person] + daily_stats['afternoon_duty'][person] if slot.startswith('13:30') and slot.endswith('_당직') else float('inf')
                    room_count = total_stats['rooms'][room_num][person] + daily_stats['rooms'][room_num][person]
                    
                    if slot.startswith('8:30') and '_당직' not in slot:
                        score = early_count
                    elif slot.startswith('10:00'):
                        score = late_count
                    elif slot.startswith('8:30') and slot.endswith('_당직'):
                        score = morning_duty_count
                    elif slot.startswith('13:30') and slot.endswith('_당직'):
                        score = afternoon_duty_count
                    else:
                        score = room_count
                    
                    if score < min_score:
                        min_score = score
                        best_person = person
                
                person = best_person
                if slot in morning_slots:
                    assigned_personnel_morning.add(person)
                else:
                    assigned_personnel_afternoon.add(person)
                st.warning(f"슬롯 {slot} 공란 방지: {person} 배정 (스코어: {min_score})")
            else:
                # 이미 배정된 인원 중에서 스코어 최소인 인원 선택
                available_personnel = morning_personnel if slot in morning_slots else afternoon_personnel
                if available_personnel:
                    room_num = slot.split('(')[1].split(')')[0]
                    best_person = None
                    min_score = float('inf')
                    for person in available_personnel:
                        early_count = total_stats['early'][person] + daily_stats['early'][person] if slot.startswith('8:30') and '_당직' not in slot else float('inf')
                        late_count = total_stats['late'][person] + daily_stats['late'][person] if slot.startswith('10:00') else float('inf')
                        morning_duty_count = total_stats['morning_duty'][person] + daily_stats['morning_duty'][person] if slot.startswith('8:30') and slot.endswith('_당직') else float('inf')
                        afternoon_duty_count = total_stats['afternoon_duty'][person] + daily_stats['afternoon_duty'][person] if slot.startswith('13:30') and slot.endswith('_당직') else float('inf')
                        room_count = total_stats['rooms'][room_num][person] + daily_stats['rooms'][room_num][person]
                        
                        if slot.startswith('8:30') and '_당직' not in slot:
                            score = early_count
                        elif slot.startswith('10:00'):
                            score = late_count
                        elif slot.startswith('8:30') and slot.endswith('_당직'):
                            score = morning_duty_count
                        elif slot.startswith('13:30') and slot.endswith('_당직'):
                            score = afternoon_duty_count
                        else:
                            score = room_count
                        
                        if score < min_score:
                            min_score = score
                            best_person = person
                    
                    person = best_person
                    st.warning(f"슬롯 {slot} 공란 방지: 이미 배정된 {person} 재배정 (스코어: {min_score})")
                else:
                    st.warning(f"슬롯 {slot} 공란 방지 불가: 배정 가능한 인원 없음")
                    continue
            
            assignment[slot_idx] = person
            daily_stats['rooms'][room_num][person] += 1
            if slot.startswith('8:30') and '_당직' not in slot:
                daily_stats['early'][person] += 1
            elif slot.startswith('10:00'):
                daily_stats['late'][person] += 1
            if slot.startswith('8:30') and slot.endswith('_당직'):
                daily_stats['morning_duty'][person] += 1
            elif slot.startswith('13:30') and slot.endswith('_당직'):
                daily_stats['afternoon_duty'][person] += 1

    # 통계 업데이트
    for key in ['early', 'late', 'morning_duty', 'afternoon_duty']:
        total_stats[key].update(daily_stats[key])
    for room in daily_stats['rooms']:
        total_stats['rooms'][room].update(daily_stats['rooms'][room])

    return assignment, daily_stats

# df_room 생성 로직 - 8:30 당직 통계 반영 추가
if st.button("🚀 방배정 수행"):
    st.write(" ")
    st.subheader(f"✨ {month_str} 방배정 결과", divider='rainbow')
    # 방 설정 입력값 검증 및 처리
    time_slots = {}
    time_groups = {}
    memo_rules = {}
    if num_830 + num_900 + num_930 + num_1000 != 12:
        st.error("오전 방 개수 합계는 12여야 합니다.")
        st.stop()
    elif len(rooms_830) != num_830 or len(rooms_900) != num_900 or len(rooms_930) != num_930 or len(rooms_1000) != num_1000 or len(rooms_1330) != num_1330:
        st.error("각 시간대의 방 번호 선택을 완료해주세요.")
        st.stop()
    else:
        for room in rooms_830:
            slot = f"8:30({room})_당직" if room == duty_830 else f"8:30({room})"
            time_slots[slot] = len(time_slots)
            time_groups.setdefault('8:30', []).append(slot)
        for room in rooms_900:
            slot = f"9:00({room})"
            time_slots[slot] = len(time_slots)
            time_groups.setdefault('9:00', []).append(slot)
        for room in rooms_930:
            slot = f"9:30({room})"
            time_slots[slot] = len(time_slots)
            time_groups.setdefault('9:30', []).append(slot)
        for room in rooms_1000:
            slot = f"10:00({room})"
            time_slots[slot] = len(time_slots)
            time_groups.setdefault('10:00', []).append(slot)
        for room in rooms_1330:
            slot = f"13:30({room})_당직" if room == duty_1330 else f"13:30({room})"
            time_slots[slot] = len(time_slots)
            time_groups.setdefault('13:30', []).append(slot)
        
        memo_rules = {
            **{f'{i}번방': [s for s in time_slots if f'({i})' in s and '_당직' not in s] for i in range(1, 13)},
            '당직 아닌 이른방': [s for s in time_slots if s.startswith('8:30') and '_당직' not in s],
            '이른방 제외': [s for s in time_slots if s.startswith(('9:00', '9:30', '10:00'))],
            '늦은방 제외': [s for s in time_slots if s.startswith(('8:30', '9:00', '9:30'))],
            '8:30': [s for s in time_slots if s.startswith('8:30') and '_당직' not in s],
            '9:00': [s for s in time_slots if s.startswith('9:00')],
            '9:30': [s for s in time_slots if s.startswith('9:30')],
            '10:00': [s for s in time_slots if s.startswith('10:00')],
            '오후 당직 제외': [s for s in time_slots if s.startswith('13:30') and '_당직' not in s]
        }
        
        st.session_state["time_slots"] = time_slots
        st.session_state["time_groups"] = time_groups
        st.session_state["memo_rules"] = memo_rules
        st.session_state["morning_slots_830"] = [s for s in time_slots if s.startswith('8:30') and '_당직' not in s]
        st.session_state["morning_slots_900"] = [s for s in time_slots if s.startswith('9:00')]
        st.session_state["morning_slots_930"] = [s for s in time_slots if s.startswith('9:30')]
        st.session_state["morning_slots_1000"] = [s for s in time_slots if s.startswith('10:00')]
        st.session_state["afternoon_slots"] = [s for s in time_slots if s.startswith('13:30') and '_당직' not in s]
        st.session_state["duty_slots"] = [s for s in time_slots if s.startswith('13:30') and '_당직' in s]
    
    # all_slots 동적 생성 - 8:30 당직 방을 사용자가 설정한 값으로 반영
    morning_duty_slot = f"8:30({duty_830})_당직"
    all_slots = [morning_duty_slot] + \
                sorted([s for s in time_slots if s.startswith('8:30') and not s.endswith('_당직')]) + \
                sorted([s for s in time_slots if s.startswith('9:00')]) + \
                sorted([s for s in time_slots if s.startswith('9:30')]) + \
                sorted([s for s in time_slots if s.startswith('10:00')]) + \
                ['온콜'] + \
                sorted([s for s in time_slots if s.startswith('13:30') and s.endswith('_당직')]) + \
                sorted([s for s in time_slots if s.startswith('13:30') and not s.endswith('_당직')])
    
    # columns 정의
    columns = ['날짜', '요일'] + all_slots
    
    # 배정 로직 시작
    random.seed(time.time())
    total_stats = {
        'early': Counter(),
        'late': Counter(),
        'morning_duty': Counter(),  # 오전 당직 (8:30)
        'afternoon_duty': Counter(),  # 오후 당직 (13:30)
        'rooms': {str(i): Counter() for i in range(1, 13)}
    }
    
    # 누적 시트 데이터에서 오후당직 횟수 추출
    df_cumulative = st.session_state["df_cumulative"]
    afternoon_duty_counts = {}
    if not df_cumulative.empty:
        for _, row in df_cumulative.iterrows():
            name = row['이름']  # 첫 번째 열이 이름으로 변경됨
            try:
                duty_count = int(row['오후당직'])
                if duty_count > 0:
                    afternoon_duty_counts[name] = duty_count
            except (ValueError, KeyError):
                st.warning(f"누적 시트에서 {name}의 오후당직 횟수 파싱 실패")
                continue
    
    assignments = {}
    slots = list(st.session_state["time_slots"].keys())
    assignable_slots = [s for s in slots if not (s.startswith('8:30') and s.endswith('_당직'))]
    
    morning_slots_830 = st.session_state["morning_slots_830"]
    morning_slots_900 = st.session_state["morning_slots_900"]
    morning_slots_930 = st.session_state["morning_slots_930"]
    morning_slots_1000 = st.session_state["morning_slots_1000"]
    afternoon_slots = st.session_state["afternoon_slots"]
    duty_slots = st.session_state["duty_slots"]
    
    date_cache = {}
    request_cells = {}
    result_data = []
    weekday_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
    for _, row in st.session_state["df_schedule_md"].iterrows():
        date_str = row['날짜']
        try:
            if "월" in date_str:
                date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=2025)
            else:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            formatted_date = date_obj.strftime('%Y-%m-%d').strip()
            date_cache[date_str] = formatted_date
        except ValueError:
            st.warning(f"Invalid date format for {date_str}")
            continue
    
        try:
            day_of_week = weekday_map[date_obj.weekday()]
        except ValueError:
            st.warning(f"Invalid date format for {date_str}")
            continue
        
        result_row = [date_str, day_of_week]
        personnel = [p for p in row[2:-1] if pd.notna(p) and p]
        has_person = bool(personnel)
        
        if day_of_week == '토' and has_person:
            debug_columns = [col for col in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '오후1', '오후2', '오후3', '오후4', '오전당직(온콜)'] if col in row.index]
            saturday_personnel = [row.get(str(i), None) for i in range(1, 11)]
            slot_person_map = {slot: None for slot in all_slots}
            non_duty_slots = [s for s in all_slots if s not in [morning_duty_slot, '온콜']][:10]
            for i, slot in enumerate(non_duty_slots):
                if i < len(saturday_personnel):
                    slot_person_map[slot] = saturday_personnel[i]
            
            for slot in all_slots:
                if slot == morning_duty_slot or slot == '온콜':
                    person = row['오전당직(온콜)'] if has_person else None
                else:
                    person = slot_person_map.get(slot, None)
                result_row.append(person)
            result_data.append(result_row)
            continue
        
        morning_personnel = [row[str(i)] for i in range(1, 12) if pd.notna(row[str(i)]) and row[str(i)]]
        afternoon_personnel = [row[f'오후{i}'] for i in range(1, 5) if pd.notna(row[f'오후{i}']) and row[f'오후{i}']]
        personnel = list(set(morning_personnel + afternoon_personnel))
        
        if not personnel:
            assignments[formatted_date] = [None] * len(assignable_slots)
            result_row.extend([None] * len(all_slots))
            result_data.append(result_row)
            continue
        
        if len(morning_personnel) != 11 or len(afternoon_personnel) != 4:
            st.warning(f"{date_str}: 인원 불일치, 오전 {len(morning_personnel)}명(필요 11명), 오후 {len(afternoon_personnel)}명(필요 4명)")
        
        request_assignments = {}
        if not df_room_request.empty:
            for _, req in df_room_request.iterrows():
                req_date, is_morning = parse_date_info(req['날짜정보'])
                if req_date and req_date == formatted_date:
                    slots_for_category = st.session_state["memo_rules"].get(req['분류'], [])
                    if not slots_for_category:
                        st.warning(f"{req['분류']}으로 할당할 적합한 시간대(방)이 존재하지 않습니다.")
                        continue
                    valid_slots = [
                        s for s in slots_for_category
                        if (is_morning and s.startswith(('8:30', '9:00', '9:30', '10:00')) and '_당직' not in s) or
                           (not is_morning and s.startswith('13:30') and '_당직' not in s)
                    ]
                    if valid_slots:
                        room_totals = {s: sum(total_stats['rooms'][s.split('(')[1].split(')')[0]].values()) for s in valid_slots}
                        min_total = min(room_totals.values())
                        best_slots = [s for s, total in room_totals.items() if total == min_total]
                        selected_slot = random.choice(best_slots)
                        request_assignments[selected_slot] = req['이름']
                        request_cells[(formatted_date, selected_slot)] = {'이름': req['이름'], '분류': req['분류']}

        assignment, daily_stats = random_assign(
            personnel, assignable_slots, request_assignments, st.session_state["time_groups"],
            total_stats, morning_personnel, afternoon_personnel, afternoon_duty_counts
        )
        assignments[formatted_date] = assignment
        
        for slot in all_slots:
            if slot == morning_duty_slot or slot == '온콜':
                person = row['오전당직(온콜)'] if has_person else None
            else:
                person = assignment[assignable_slots.index(slot)] if slot in assignable_slots and assignment else None
            result_row.append(person if has_person else None)
        
        result_data.append(result_row)
    
    df_room = pd.DataFrame(result_data, columns=columns)
    st.write(" ")
    st.markdown("**✅ 통합 배치 결과**")
    st.dataframe(df_room)
    
    # 8:30 당직 통계 반영
    for idx, row in enumerate(result_data):
        formatted_date = date_cache.get(row[0], '')
        if formatted_date:
            for slot_idx, slot in enumerate(all_slots):
                if slot == morning_duty_slot:
                    person = row[slot_idx + 2]  # '날짜', '요일' 이후의 열
                    if person:
                        total_stats['morning_duty'][person] += 1

    # stats_df 생성 - 오전/오후 당직 합계 별도 포함
    stats_data = []
    all_personnel = set()
    for _, row in st.session_state["df_schedule_md"].iterrows():
        personnel = [p for p in row[2:-1].dropna() if p]
        all_personnel.update(personnel)
    
    for person in sorted(all_personnel):
        stats_data.append({
            '인원': person,
            '이른방 합계': total_stats['early'][person],
            '늦은방 합계': total_stats['late'][person],
            '오전 당직 합계': total_stats['morning_duty'][person],
            '오후 당직 합계': total_stats['afternoon_duty'][person],
            **{f'{r}번방 합계': total_stats['rooms'][r][person] for r in total_stats['rooms']}
        })
    
    stats_df = pd.DataFrame(stats_data)
    st.divider()
    st.markdown("**☑️ 인원별 통계**")
    if stats_df.empty:
        st.error("통계 데이터가 생성되지 않았습니다. total_stats 확인 필요.")
    else:
        st.dataframe(stats_df)
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Schedule"
    
    for col_idx, header in enumerate(columns, 1):
        cell = sheet.cell(1, col_idx, header)
        cell.font = Font(bold=True, name="맑은 고딕", size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        if header.startswith('8:30'):
            cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        elif header.startswith('9:00'):
            cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        elif header.startswith('9:30'):
            cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        elif header.startswith('10:00'):
            cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        elif header.startswith('13:30'):
            cell.fill = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")
        elif header == '온콜':
            cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    for row_idx, row in enumerate(result_data, 2):
        has_person = any(x for x in row[2:-1] if x is not None)
        formatted_date = date_cache.get(row[0], '')
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row_idx, col_idx, value)
            if (columns[col_idx-1].endswith('_당직') or columns[col_idx-1] == '온콜') and value:
                cell.font = Font(name="맑은 고딕", size=9, bold=True, color="FF00FF")
            else:
                cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            if col_idx == 1:
                cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            elif col_idx == 2:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                if value == '토' and has_person:
                    cell.fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
            elif not has_person and col_idx >= 3:
                cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            
            if col_idx > 2 and value and formatted_date:
                slot = columns[col_idx-1]
                if (formatted_date, slot) in request_cells and value == request_cells[(formatted_date, slot)]['이름']:
                    cell.comment = Comment(f"배정 요청: {request_cells[(formatted_date, slot)]['분류']}", "System")
    
    stats_sheet = wb.create_sheet("Stats")
    stats_columns = stats_df.columns.tolist()
    for col_idx, header in enumerate(stats_columns, 1):
        cell = stats_sheet.cell(1, col_idx, header)
        cell.font = Font(bold=True, name="맑은 고딕", size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        if header == '인원':
            cell.fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
        elif header == '이른방 합계':
            cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        elif header == '늦은방 합계':
            cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        elif header == '오전 당직 합계':
            cell.fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
        elif header == '오후 당직 합계':
            cell.fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
    
    for row_idx, row in enumerate(stats_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = stats_sheet.cell(row_idx, col_idx, value)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    st.divider()
    st.download_button(
        label="📥 최종 방배정 다운로드",
        data=output,
        file_name=f"{month_str} 방배정.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )

    # Google Sheets에 방배정 시트 저장
    gc = get_gspread_client()
    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
    try:
        worksheet_result = sheet.worksheet(f"{month_str} 방배정")
    except:
        worksheet_result = sheet.add_worksheet(f"{month_str} 방배정", rows=100, cols=len(df_room.columns))
        worksheet_result.append_row(df_room.columns.tolist())

    update_sheet_with_retry(worksheet_result, [df_room.columns.tolist()] + df_room.values.tolist())
    st.success(f"✅ {month_str} 방배정 테이블이 Google Sheets에 저장되었습니다.")