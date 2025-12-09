import streamlit as st
import pandas as pd
import numpy as np
import re
import time
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from zoneinfo import ZoneInfo
from collections import Counter
import platform
import calendar

# Google Sheets 관련 라이브러리
from google.oauth2.service_account import Credentials
import gspread
from gspread.exceptions import WorksheetNotFound, APIError

# 엑셀 생성을 위한 라이브러리
import io
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.comments import Comment

# 사용자 정의 메뉴 모듈
import menu
import os
st.session_state.current_page = os.path.basename(__file__)

# --- 페이지 설정 및 초기화 ---
st.set_page_config(page_title="스케줄 수정", page_icon="✍️", layout="wide")
menu.menu()

# --- 로그인 확인 ---
if not st.session_state.get("login_success", False):
    st.warning("⚠️ Home 페이지에서 먼저 로그인해주세요.")
    st.error("1초 후 Home 페이지로 돌아갑니다...")
    time.sleep(1)
    st.switch_page("Home.py")
    st.stop()

# --- Google Sheets API 연동 함수 ---

@st.cache_resource
def get_gspread_client():
    scope = ["https://www.googleapis.com/auth/spreadsheets"]
    try:
        service_account_info = dict(st.secrets["gspread"])
        service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
        credentials = Credentials.from_service_account_info(service_account_info, scopes=scope)
        return gspread.authorize(credentials)
    except Exception as e:
        st.error(f"⚠️ Google Sheets 클라이언트 초기화 또는 인증에 실패했습니다: {e}"); st.stop()

# ✨ [새로 추가] 스프레드시트 객체를 캐시하는 함수
@st.cache_resource
def get_spreadsheet():
    """
    스프레드시트 객체를 한 번만 열어서 캐시합니다.
    """
    try:
        gc = get_gspread_client()
        sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
        return sheet
    except Exception as e:
        st.error(f"⚠️ Google Spreadsheet를 여는 데 실패했습니다: {e}")
        st.stop()

def update_sheet_with_retry(worksheet, data, retries=3, delay=5):
    for attempt in range(retries):
        try:
            worksheet.clear(); worksheet.update(data, "A1"); return True
        except APIError as e:
            if attempt < retries - 1:
                st.warning(f"⚠️ API 요청 지연... {delay}초 후 재시도 ({attempt+1}/{retries})"); time.sleep(delay * (attempt + 1))
            else:
                st.error(f"Google Sheets API 오류: {e}"); st.stop()
    return False

# [수정] 캐시 적용(TTL 60초) 및 에러 핸들링 추가
@st.cache_data(ttl=60, show_spinner=False)
def find_schedule_versions(month_str): 
    """'ver X.X' 버전과 '최종' 버전을 모두 찾아 정렬된 딕셔너리로 반환합니다."""
    
    # 1. API 호출 시도 및 에러 처리
    try:
        # 함수 내부에서 sheet 객체를 가져옵니다.
        sheet = get_spreadsheet() 
        
        # Google Sheets API에서 모든 워크시트 제목을 한 번에 가져옵니다.
        all_titles = [ws.title for ws in sheet.worksheets()]

    except APIError as e:
        # 429 에러(Quota exceeded)일 경우 사용자 친화적 메시지 출력
        if hasattr(e, 'response') and e.response.status_code == 429:
            st.error("🚨 잦은 요청으로 구글 시트 연결이 일시 지연되고 있습니다.")
            st.warning("약 1분 뒤에 '새로고침' 버튼을 눌러주세요.")
            st.stop() # 이후 코드 실행 중단
        else:
            # 다른 에러라면 그대로 에러를 발생시킴
            raise e

    # 2. 버전 파싱 로직 (기존과 동일)
    versions = {}
    base_name = f"{month_str} 스케줄"
    
    for title in all_titles:
        # 1. "최종" 버전 확인
        if title == f"{base_name} 최종":
            versions[title] = 999.0
            continue

        # 2. "ver X.X" 버전 확인
        ver_match = re.match(f"^{re.escape(base_name)}\s*ver\s*(\d+\.\d+)$", title)
        if ver_match:
            version_num = float(ver_match.group(1))
            versions[title] = version_num
            continue

        # 3. 기본 버전 확인
        if title == base_name:
            versions[title] = 1.0

    # 버전을 기준으로 내림차순 정렬하여 반환
    return dict(sorted(versions.items(), key=lambda item: item[1], reverse=True))

# --- ▼▼▼ [신규] '베이스 누적 시트' 로드용 함수 추가 ▼▼▼ ---
def find_latest_cumulative_version(sheet, month_str):
    """
    [★복사됨★]
    주어진 월에 해당하는 누적 시트 중 가장 최신 버전을 찾습니다.
    '최종' 버전을 최우선으로 간주합니다. (공백 차이 무시)
    """
    versions = {}
    
    # 1. '최종' 시트가 있는지 먼저 확인 (공백(s+)을 허용하는 정규식 사용)
    final_pattern = re.compile(f"^{re.escape(month_str)}\s+누적\s+최종$")
    for ws in sheet.worksheets():
        if final_pattern.match(ws.title.strip()): # .strip() 추가로 앞뒤 공백 제거
            return ws.title # '최종' 버전을 찾으면 즉시 반환
    
    # 2. '최종'이 없으면 'ver X.X' 및 기본 버전('누적')을 찾음
    pattern = re.compile(f"^{re.escape(month_str)} 누적(?: ver\s*(\d+\.\d+))?$")

    for ws in sheet.worksheets():
        match = pattern.match(ws.title)
        if match:
            version_num_str = match.group(1) # ver 뒤의 숫자 부분 (예: '1.0')
            version_num = float(version_num_str) if version_num_str else 1.0
            versions[ws.title] = version_num

    if not versions:
        return None # 어떠한 버전의 시트도 찾지 못하면 None 반환

    return max(versions, key=versions.get)

# --- ▼▼▼ [교체] L108 ~ L179의 기존 load_data 함수 전체를 교체 ▼▼▼ ---
@st.cache_data(ttl=600, show_spinner="최신 데이터를 구글 시트에서 불러오는 중...")
def load_data(month_str, schedule_sheet_name):
    sheet = get_spreadsheet() 
    target_year = month_str.split('년')[0]
    
    current_month_dt = datetime.strptime(month_str, "%Y년 %m월")
    next_month_str = (current_month_dt + relativedelta(months=1)).strftime("%Y년 %-m월")

    # 1. 스케줄 시트 로드 (기존과 동일)
    try:
        ws_schedule = sheet.worksheet(schedule_sheet_name)
        df_schedule = pd.DataFrame(ws_schedule.get_all_records())
    except WorksheetNotFound:
        st.error(f"'{schedule_sheet_name}' 시트를 찾을 수 없습니다."); st.stop()
    
    # 2. 익월(결과) 누적 시트 로드 (기존과 동일)
    version_suffix = ""
    if " ver" in schedule_sheet_name:
        version_suffix = " " + schedule_sheet_name.split(" 스케줄 ")[1]
    elif "최종" in schedule_sheet_name:
        version_suffix = " 최종"
        
    display_cum_sheet_name = f"{next_month_str} 누적{version_suffix}"
    try:
        ws_display_cum = sheet.worksheet(display_cum_sheet_name)
        all_values = ws_display_cum.get_all_values()
        if not all_values or len(all_values) < 2:
            df_display_cum = pd.DataFrame()
        else:
            headers = all_values[0]
            data = all_values[1:]
            df_display_cum = pd.DataFrame(data, columns=headers)
            # (데이터 클리닝)
            for col in df_display_cum.columns:
                if col != '항목':
                    df_display_cum[col] = pd.to_numeric(df_display_cum[col], errors='coerce').fillna(0).astype(int)
    except WorksheetNotFound:
        df_display_cum = pd.DataFrame()
        st.warning(f"⚠️ '{display_cum_sheet_name}' 시트를 찾을 수 없습니다. 누적 테이블이 비어있을 수 있습니다.")

    # --- ▼▼▼ [신규] 3. 당월(지난달의 누적) 베이스 누적 시트 로드 ▼▼▼ ---
    df_cumulative_base = pd.DataFrame()
    worksheet_to_load_base = None
    latest_base_cum_name = find_latest_cumulative_version(sheet, month_str) # month_str (10월)
    
    if latest_base_cum_name:
        try:
            worksheet_to_load_base = sheet.worksheet(latest_base_cum_name)
        except WorksheetNotFound:
            st.warning(f"⚠️ '{latest_base_cum_name}' 시트를 찾았지만 열 수 없습니다.")
    else:
        st.warning(f"⚠️ '{month_str} 누적' (베이스) 시트를 찾을 수 없습니다.")

    if worksheet_to_load_base:
        all_values_base = worksheet_to_load_base.get_all_values()
        if all_values_base and len(all_values_base) > 1:
            headers_base = all_values_base[0]
            data_base = [row for row in all_values_base[1:] if any(cell.strip() for cell in row)]
            df_cumulative_base = pd.DataFrame(data_base, columns=headers_base)
    
    # (데이터 클리닝)
    if df_cumulative_base.empty or '항목' not in df_cumulative_base.columns:
            # (이름 목록을 df_display_cum에서 가져오는 것으로 대체)
            master_names_list = df_display_cum.columns[1:].tolist() if not df_display_cum.empty else []
            default_cols = ["항목"] + master_names_list
            default_data = [
                ["오전누적"] + [0] * len(master_names_list), ["오후누적"] + [0] * len(master_names_list),
                ["오전당직누적"] + [0] * len(master_names_list), ["오후당직누적"] + [0] * len(master_names_list)
            ]
            df_cumulative_base = pd.DataFrame(default_data, columns=default_cols)
    
    for col in df_cumulative_base.columns:
        if col != '항목':
            df_cumulative_base[col] = pd.to_numeric(df_cumulative_base[col], errors='coerce').fillna(0).astype(int)
    # --- ▲▲▲ [신규] 3. 베이스 누적 시트 로드 끝 ▲▲▲ ---

    # --- ▼▼▼ [신규] 4. 날짜 매핑 테이블 생성 ▼▼▼ ---
    # (ISO 날짜와 '10월 1일' 표시 형식을 매핑하기 위해)
    day_map_schedule = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
    _, last_day_schedule = calendar.monthrange(current_month_dt.year, current_month_dt.month)
    all_month_dates_schedule = pd.date_range(start=current_month_dt, end=current_month_dt.replace(day=last_day_schedule))
    df_schedule_mapping = pd.DataFrame({
        '날짜': [d.strftime('%Y-%m-%d') for d in all_month_dates_schedule],
        '요일': [day_map_schedule.get(d.weekday()) for d in all_month_dates_schedule],
        '날짜_표시': [f"{d.month}월 {d.day}일" for d in all_month_dates_schedule]
    })
    # --- ▲▲▲ [신규] 4. 날짜 매핑 테이블 생성 끝 ▲▲▲ ---

    # 5. 토요/휴일, 휴관일 로드 (기존과 동일)
    try:
        ws_special = sheet.worksheet(f"{target_year}년 토요/휴일 스케줄")
        df_yearly = pd.DataFrame(ws_special.get_all_records()); df_yearly['날짜_dt'] = pd.to_datetime(df_yearly['날짜'])
        target_month_dt = datetime.strptime(month_str, "%Y년 %m월")
        df_special = df_yearly[(df_yearly['날짜_dt'].dt.year == target_month_dt.year) & (df_yearly['날짜_dt'].dt.month == target_month_dt.month)].copy()
    except WorksheetNotFound: df_special = pd.DataFrame()

    try:
        ws_closing = sheet.worksheet(f"{target_year}년 휴관일"); df_closing = pd.DataFrame(ws_closing.get_all_records())
        closing_dates = pd.to_datetime(df_closing['날짜']).dt.strftime('%Y-%m-%d').tolist() if '날짜' in df_closing.columns and not df_closing.empty else []
    except WorksheetNotFound: closing_dates = []

    is_final_version = "최종" in schedule_sheet_name
    
    # [수정] 반환 딕셔너리에 'base_cumulative'와 'schedule_mapping' 추가
    return {
        "schedule": df_schedule, 
        "cumulative_display": df_display_cum, 
        "base_cumulative": df_cumulative_base, # (신규)
        "schedule_mapping": df_schedule_mapping, # (신규)
        "swaps": pd.DataFrame(),
        "special": df_special, 
        "requests": pd.DataFrame(), 
        "closing_dates": closing_dates,
        "is_final_version": is_final_version
    }
# --- ▲▲▲ [교체] load_data 함수 교체 끝 ▲▲▲ ---

def apply_schedule_swaps(original_schedule_df, swap_requests_df):
    df_modified = original_schedule_df.copy(); change_log = []; messages = []; applied_count = 0
    for _, request_row in swap_requests_df.iterrows():
        try:
            change_request_str = str(request_row.get('변경 요청', '')).strip(); schedule_info_str = str(request_row.get('변경 요청한 스케줄', '')).strip()
            if '➡️' not in change_request_str: continue
            person_before, person_after = [p.strip() for p in change_request_str.split('➡️')]; date_match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', schedule_info_str)
            if not date_match: continue
            date_part, time_period = date_match.groups(); date_obj = datetime.strptime(date_part, '%Y-%m-%d').date(); formatted_date_in_df = f"{date_obj.month}월 {date_obj.day}일"
            target_rows = df_modified[df_modified['날짜'] == formatted_date_in_df]
            if target_rows.empty: continue
            target_row_idx = target_rows.index[0]; on_call_person = str(df_modified.at[target_row_idx, '오전당직(온콜)']).strip()
            if time_period == '오전당직(온콜)' or person_before == on_call_person:
                cols_with_person_before = [c for c in df_modified.columns if str(df_modified.at[target_row_idx, c]).strip() == person_before]
                if not cols_with_person_before: messages.append(('error', f"❌ {schedule_info_str} - '{person_before}' 당직 근무가 없습니다.")); continue
                cols_with_person_after = [c for c in df_modified.columns if str(df_modified.at[target_row_idx, c]).strip() == person_after]
                for col in cols_with_person_before: df_modified.at[target_row_idx, col] = person_after
                for col in cols_with_person_after: df_modified.at[target_row_idx, col] = person_before
                change_log.append({'날짜': f"{formatted_date_in_df} (당직 맞교환)", '변경 전': person_before, '변경 후': person_after})
            else:
                target_cols = [str(i) for i in range(1, 18)] if time_period == '오전' else [f'오후{i}' for i in range(1, 10)]; personnel_in_period = {str(df_modified.at[target_row_idx, c]).strip() for c in target_cols if c in df_modified.columns}
                if person_after in personnel_in_period: messages.append(('warning', f"🟡 {schedule_info_str} - '{person_after}'님은 이미 해당 시간 근무자입니다.")); continue
                found_and_replaced = False
                for col in target_cols:
                    if col in df_modified.columns and str(df_modified.at[target_row_idx, col]).strip() == person_before:
                        df_modified.at[target_row_idx, col] = person_after; change_log.append({'날짜': f"{schedule_info_str}", '변경 전': person_before, '변경 후': person_after}); found_and_replaced = True; break
                if not found_and_replaced: messages.append(('error', f"❌ {schedule_info_str} - '{person_before}' 근무자를 찾을 수 없습니다.")); continue
            applied_count += 1
        except Exception as e: messages.append(('error', f"요청 처리 중 오류: {e}"))
    if applied_count > 0: messages.insert(0, ('success', f"✅ 총 {applied_count}건의 스케줄 변경 요청이 반영되었습니다."))
    elif not messages: messages.append(('info', "새롭게 적용할 스케줄 변경 요청이 없습니다."))
    st.session_state["change_log"] = change_log; return df_modified, messages

def format_sheet_date_for_display(date_string):
    match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', date_string)
    if match:
        date_part, shift_part = match.groups()
        try:
            dt_obj = datetime.strptime(date_part, '%Y-%m-%d').date(); weekday_str = ['월', '화', '수', '목', '금', '토', '일'][dt_obj.weekday()]; return f"{dt_obj.month}월 {dt_obj.day}일 ({weekday_str}) - {shift_part}"
        except ValueError: pass
    return date_string

def delete_schedule_version(month_str, sheet_to_delete):
    """선택된 스케줄 버전과 해당 누적 시트를 Google Sheets에서 삭제합니다."""
    try:
        with st.spinner(f"'{sheet_to_delete}' 버전 삭제 중..."):
            # gc = get_gspread_client()
            # sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
            sheet = get_spreadsheet()

            # 1. 스케줄 시트 삭제
            try:
                worksheet_to_delete = sheet.worksheet(sheet_to_delete)
                sheet.del_worksheet(worksheet_to_delete)
                st.info(f"'{sheet_to_delete}' 시트를 삭제했습니다.")
            except WorksheetNotFound:
                st.warning(f"'{sheet_to_delete}' 시트를 찾을 수 없어 삭제를 건너뜁니다.")

            # 2. 해당 버전의 누적 시트 이름 생성 및 삭제
            version_suffix = ""
            if " ver" in sheet_to_delete:
                # "ver X.X"가 있는 경우 (예: " ver1.0")
                version_suffix = " " + sheet_to_delete.split(" 스케줄 ")[1]
            elif "최종" in sheet_to_delete:
                # "최종"이 있는 경우 (예: " 최종")
                version_suffix = " 최종"
            
            current_month_dt = datetime.strptime(month_str, "%Y년 %m월")
            next_month_str = (current_month_dt + relativedelta(months=1)).strftime("%Y년 %-m월")
            
            # [수정] version_suffix를 사용하여 정확한 누적 시트 이름 생성
            cum_sheet_name = f"{next_month_str} 누적{version_suffix}"
            
            try:
                worksheet_cum_to_delete = sheet.worksheet(cum_sheet_name)
                sheet.del_worksheet(worksheet_cum_to_delete)
                st.info(f"'{cum_sheet_name}' 시트를 삭제했습니다.")
            except WorksheetNotFound:
                st.warning(f"'{cum_sheet_name}' 시트를 찾을 수 없어 삭제를 건너뜁니다.")
        
        st.success("선택한 버전이 성공적으로 삭제되었습니다.")
        time.sleep(2)
        
        st.cache_data.clear()
        st.cache_resource.clear()

        if "selected_sheet_name" in st.session_state:
            del st.session_state["selected_sheet_name"]
        if "data_loaded" in st.session_state:
            st.session_state["data_loaded"] = False
        
        st.rerun()
        
    except Exception as e:
        st.error(f"버전 삭제 중 오류가 발생했습니다: {e}")

# --- 1. 기존 엑셀 생성 함수 전체를 이 코드로 교체하세요 ---

def create_formatted_schedule_excel(initial_df, edited_df, edited_cumulative_df, df_special, df_requests, closing_dates, month_str):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "수정된 스케줄"

    # --- 1. 스타일 및 전체 색상 맵 정의 ---
    font_name = "맑은 고딕"
    default_font = Font(name=font_name, size=9)
    bold_font = Font(name=font_name, size=9, bold=True)
    duty_font = Font(name=font_name, size=9, bold=True, color="FF69B4")
    header_font = Font(name=font_name, size=9, color='FFFFFF', bold=True)

    color_map = {
        '휴가': 'DA9694', '학회': 'DA9694',
        '꼭 근무': 'FABF8F',
        '보충': 'FFF28F',
        '대체보충': 'A9D08E',
        '휴근': 'B1A0C7',
        '대체휴근': '95B3D7',
        '특수근무': 'D0E0E3',
        '기본': 'FFFFFF'
    }
    
    header_fill = PatternFill(start_color='000000', fill_type='solid')
    date_col_fill = PatternFill(start_color='808080', fill_type='solid')
    weekday_fill = PatternFill(start_color='FFF2CC', fill_type='solid')
    special_day_fill = PatternFill(start_color='95B3D7', fill_type='solid')
    changed_fill = PatternFill(start_color='FFFF00', fill_type='solid')
    empty_day_fill = PatternFill(start_color='808080', fill_type='solid')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    requests_map = {}
    if not df_requests.empty:
        def parse_date_range(d_str):
            if pd.isna(d_str) or not isinstance(d_str, str) or d_str.strip() == '': return []
            d_str = d_str.strip()
            if '~' in d_str:
                try:
                    start, end = [datetime.strptime(d.strip(), '%Y-%m-%d').date() for d in d_str.split('~')]
                    return [(start + timedelta(days=i)).strftime('%Y-%m-%d') for i in range((end - start).days + 1)]
                except: return []
            else:
                try:
                    return [datetime.strptime(d.strip(), '%Y-%m-%d').date().strftime('%Y-%m-%d') for d in d_str.split(',')]
                except: return []
        
        for _, row in df_requests.iterrows():
            worker = row['이름']
            status = row['분류']
            if status in ['휴가', '학회'] or '꼭 근무' in status:
                clean_status = '꼭 근무' if '꼭 근무' in status else status
                for date_iso in parse_date_range(row['날짜정보']):
                    requests_map[(worker, date_iso)] = clean_status

    # --- 2. 헤더 생성 ---
    for c, col_name in enumerate(edited_df.columns, 1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border

    # --- 3. 데이터 행 생성 및 서식 적용 ---
    for r, (idx, edited_row) in enumerate(edited_df.iterrows(), 2):
        initial_row = initial_df.loc[idx]
        
        try:
            current_date = datetime.strptime(f"{month_str.split('년')[0]}-{edited_row['날짜']}", "%Y-%m월 %d일").date()
            current_date_iso = current_date.strftime('%Y-%m-%d')
        except (ValueError, TypeError):
            current_date = None; current_date_iso = None

        is_row_empty = all(pd.isna(v) or str(v).strip() == '' for k, v in edited_row.items() if k not in ['날짜', '요일'])
        is_special_day = current_date in pd.to_datetime(df_special['날짜']).dt.date.values if current_date and not df_special.empty else False
        is_empty_day = (is_row_empty and not is_special_day) or (current_date_iso in closing_dates)
        
        weekend_oncall_worker = None
        if is_special_day:
            special_day_info = df_special[pd.to_datetime(df_special['날짜']).dt.date == current_date]
            if not special_day_info.empty and '당직' in special_day_info.columns:
                oncall_val = special_day_info['당직'].iloc[0]
                if pd.notna(oncall_val) and oncall_val != "당직 없음":
                    weekend_oncall_worker = str(oncall_val).strip()

        for c, col_name in enumerate(edited_df.columns, 1):
            cell = ws.cell(row=r, column=c, value=edited_row[col_name])
            cell.font = default_font; cell.alignment = center_align; cell.border = border

            if is_empty_day:
                cell.fill = empty_day_fill; continue

            if col_name == '날짜':
                cell.fill = date_col_fill; continue
            if col_name == '요일':
                cell.fill = special_day_fill if is_special_day else weekday_fill; continue
            
            raw_value = str(edited_row.get(col_name, '')).strip()
            
            if is_special_day:
                if str(col_name).isdigit() and raw_value:
                    cell.fill = PatternFill(start_color=color_map['특수근무'], end_color=color_map['특수근무'], fill_type='solid')
                    if raw_value == weekend_oncall_worker:
                        cell.font = duty_font
                elif '오후' in str(col_name):
                    cell.value = ""
                continue
            
            worker_name = raw_value
            status = '기본'
            
            match = re.match(r'(.+?)\((.+)\)', raw_value)
            if match:
                worker_name = match.group(1).strip(); status = match.group(2).strip()
            elif current_date_iso and worker_name:
                status = requests_map.get((worker_name, current_date_iso), '기본')

            cell.value = worker_name
            if not worker_name: continue

            fill_color_hex = color_map.get(status)
            if fill_color_hex:
                cell.fill = PatternFill(start_color=fill_color_hex, end_color=fill_color_hex, fill_type='solid')

            if col_name == '오전당직(온콜)' and worker_name:
                cell.font = duty_font
            
            initial_raw_value = str(initial_row.get(col_name, '')).strip()
            if raw_value != initial_raw_value:
                cell.fill = changed_fill
                cell.comment = Comment(f"변경 전: {initial_raw_value or '빈 값'}", "Edit Tracker")

    # --- 4. 익월 누적 현황 추가 ---
    if not edited_cumulative_df.empty:
        style_args = {'font': default_font, 'bold_font': bold_font, 'border': border}
        # 요청하신 함수에 편집된 데이터프레임을 그대로 전달
        append_summary_table_to_excel(ws, edited_cumulative_df, style_args)

    # --- 5. 열 너비 설정 ---
    ws.column_dimensions['A'].width = 11
    for col in ws.columns:
        if col[0].column_letter != 'A':
            ws.column_dimensions[col[0].column_letter].width = 9

    wb.save(output)
    return output.getvalue()

def apply_outer_border(worksheet, start_row, end_row, start_col, end_col):
    medium_side = Side(style='medium') 
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = worksheet.cell(row=r, column=c)
            top, left, bottom, right = cell.border.top, cell.border.left, cell.border.bottom, cell.border.right
            if r == start_row: top = medium_side
            if r == end_row: bottom = medium_side
            if c == start_col: left = medium_side
            if c == end_col: right = medium_side
            cell.border = Border(top=top, left=left, bottom=bottom, right=right)

def append_summary_table_to_excel(worksheet, summary_df, style_args):
    if summary_df.empty:
        return

    fills = {
        'header': PatternFill(start_color='E7E6E6', fill_type='solid'), 'yellow': PatternFill(start_color='FFF296', fill_type='solid'),
        'pink': PatternFill(start_color='FFC8CD', fill_type='solid'), 'green': PatternFill(start_color='C6E0B4', fill_type='solid'),
        'dark_green': PatternFill(start_color='82C4B5', fill_type='solid'), 'blue': PatternFill(start_color='B8CCE4', fill_type='solid'),
        'orange': PatternFill(start_color='FCE4D6', fill_type='solid'), 'lightgray': PatternFill(start_color='F2F2F2', fill_type='solid')
    }
    
    start_row = worksheet.max_row + 3
    thin_border = style_args['border'] 

    # 헤더 쓰기
    for c_idx, value in enumerate(summary_df.columns.tolist(), 1):
        cell = worksheet.cell(row=start_row, column=c_idx, value=value)
        cell.fill = fills['header']; cell.font = style_args['bold_font']; cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 데이터 행 쓰기
    for r_idx, row_data in enumerate(summary_df.itertuples(index=False), start_row + 1):
        label = row_data[0]
        for c_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            cell.font = style_args['bold_font'] if c_idx == 1 else style_args['font']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            fill_color = None
            if label in ["오전누적", "오후누적"]: fill_color = fills['pink']
            elif label in ["오전합계", "오후합계"]: fill_color = fills['blue']
            elif label == "오전당직": fill_color = fills['blue']
            elif label == "오전당직누적": fill_color = fills['pink']
            elif label == "오후당직": fill_color = fills['lightgray']
            elif label == "오후당직누적": fill_color = fills['lightgray']
            if c_idx == 1 and label in ["오전보충", "임시보충", "오후보충", "온콜검사"]: fill_color = fills['yellow']
            if fill_color: cell.fill = fill_color

    start_col, end_col = 1, len(summary_df.columns)
    labels = summary_df.iloc[:, 0].tolist()

    apply_outer_border(worksheet, start_row, start_row, start_col, end_col)
    apply_outer_border(worksheet, start_row, start_row + len(labels), start_col, start_col)
    if "오전보충" in labels and "오전누적" in labels: apply_outer_border(worksheet, start_row + 1 + labels.index("오전보충"), start_row + 1 + labels.index("오전누적"), start_col, end_col)
    if "오후보충" in labels and "오후누적" in labels: apply_outer_border(worksheet, start_row + 1 + labels.index("오후보충"), start_row + 1 + labels.index("오후누적"), start_col, end_col)
    if "오전당직" in labels and "오후당직누적" in labels: apply_outer_border(worksheet, start_row + 1 + labels.index("오전당직"), start_row + 1 + labels.index("오후당직누적"), start_col, end_col)

    legend_start_row = worksheet.max_row + 3 
    legend_data = [('A9D08E', '대체 보충'), ('FFF28F', '보충'), ('95B3D7', '대체 휴근'), ('B1A0C7', '휴근'), ('DA9694', '휴가/학회')]

    for i, (hex_color, description) in enumerate(legend_data):
        current_row = legend_start_row + i
        
        # ✨ [오류 수정] 'ws'를 'worksheet'로 변경
        color_cell = worksheet.cell(row=current_row, column=1)
        color_cell.fill = PatternFill(start_color=hex_color, fill_type='solid')
        color_cell.border = thin_border

        # ✨ [오류 수정] 'ws'를 'worksheet'로 변경
        desc_cell = worksheet.cell(row=current_row, column=2, value=description)
        desc_cell.font = style_args['font']
        desc_cell.border = thin_border
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ✨ [오류 수정] 'ws'를 'worksheet'로 변경
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 15
    for i in range(2, len(summary_df.columns) + 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 9

# --- 1. 최종본(공유용) 엑셀 생성 함수 ---
# [★ L442의 이 함수 전체를 교체하세요 ★]

# [★ L442의 이 함수 전체를 교체하세요 ★]

def create_final_schedule_excel(initial_df, edited_df, edited_cumulative_df, df_special, df_requests, closing_dates, month_str):
    """
    [공유용 최종본]
    [★ v3 수정 ★]
    - '변경된' 셀의 배경색을 'F2DCDB' (연분홍)로 변경합니다.
    - '변경된' 셀의 색상을 '상태' 색상보다 우선 적용합니다.
    - (대체보충) 상태이고 (10/6에서 대체됨) 형식의 텍스트가 괄호 안에 있을 경우,
    - 이를 엑셀 '메모'로 추가하는 로직을 이식합니다.
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "스케줄"

    # --- 스타일 정의 ---
    font_name = "맑은 고딕" if platform.system() == "Windows" else "Arial"
    default_font = Font(name=font_name, size=9)
    bold_font = Font(name=font_name, size=9, bold=True)
    duty_font = Font(name=font_name, size=9, bold=True, color="FF69B4")
    header_font = Font(name=font_name, size=9, color='FFFFFF', bold=True)
    color_map = {'휴가': 'DA9694', '학회': 'DA9694', '꼭 근무': 'FABF8F', '보충': 'FFF28F', '대체보충': 'A9D08E', '휴근': 'B1A0C7', '대체휴근': '95B3D7', '특수근무': 'D0E0E3', '기본': 'FFFFFF'}
    header_fill = PatternFill(start_color='000000', fill_type='solid')
    date_col_fill = PatternFill(start_color='808080', fill_type='solid')
    weekday_fill = PatternFill(start_color='FFF2CC', fill_type='solid')
    special_day_fill = PatternFill(start_color='95B3D7', fill_type='solid')
    
    # --- ▼▼▼ [핵심 수정] 변경된 셀 색상 F2DCDB로 수정 ▼▼▼ ---
    changed_fill = PatternFill(start_color='F2DCDB', fill_type='solid') # (연분홍)
    # --- ▲▲▲ [수정 완료] ▲▲▲ ---
    
    empty_day_fill = PatternFill(start_color='808080', fill_type='solid')
    holiday_blue_fill = PatternFill(start_color="DDEBF7", fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    # --- 고정된 열 정의 ---
    final_columns = ['날짜', '요일'] + [str(i) for i in range(1, 13)] + [''] + ['오전당직(온콜)'] + [f'오후{i}' for i in range(1, 5)]

    # --- 헤더 생성 ---
    for c, col_name in enumerate(final_columns, 1):
        cell = ws.cell(row=1, column=c, value=col_name); cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border

    # --- 데이터 행 생성 및 서식 적용 ---
    for r, (idx, edited_row) in enumerate(edited_df.iterrows(), 2):
        initial_row = initial_df.loc[idx]
        try:
            current_date = datetime.strptime(f"{month_str.split('년')[0]}-{edited_row['날짜']}", "%Y-%m월 %d일").date()
            current_date_iso = current_date.strftime('%Y-%m-%d')
        except: current_date, current_date_iso = None, None
        
        is_row_empty = all(pd.isna(v) or str(v).strip() == '' for k, v in edited_row.items() if k not in ['날짜', '요일'])
        is_special_day = current_date in pd.to_datetime(df_special['날짜']).dt.date.values if current_date and not df_special.empty else False
        is_empty_day = (is_row_empty and not is_special_day) or (current_date_iso in closing_dates)
        
        weekend_oncall_worker = None
        if is_special_day:
            special_day_info = df_special[pd.to_datetime(df_special['날짜']).dt.date == current_date]
            if not special_day_info.empty and '당직' in special_day_info.columns:
                oncall_val = special_day_info['당직'].iloc[0]
                if pd.notna(oncall_val) and oncall_val != "당직 없음": weekend_oncall_worker = str(oncall_val).strip()

        for c, col_name in enumerate(final_columns, 1):
            
            raw_value = str(edited_row.get(col_name, '')).strip()
            
            cell = ws.cell(row=r, column=c) # value는 나중에 설정
            cell.font = default_font; cell.alignment = center_align; cell.border = border
            
            if is_empty_day: cell.fill = empty_day_fill; continue
            if col_name == '날짜': cell.value = edited_row.get(col_name, ''); cell.fill = date_col_fill; continue
            if col_name == '요일': cell.value = edited_row.get(col_name, ''); cell.fill = special_day_fill if is_special_day else weekday_fill; continue
            
            worker_name = re.sub(r'\(.+\)', '', raw_value).strip()
            
            # --- ▼▼▼ [핵심 수정] 파싱 로직 교체 (v2) ▼▼▼ ---
            status_or_memo = '기본'
            match = re.match(r'.+?\((.+)\)', raw_value)
            if match: status_or_memo = match.group(1).strip()
            
            real_status = '기본'
            if status_or_memo == '기본':
                real_status = '기본'
            elif status_or_memo in color_map: 
                real_status = status_or_memo
            elif pd.notna(status_or_memo) and (re.search(r'\d{1,2}/\d{1,2}', status_or_memo) or '대체됨' in status_or_memo or '대체함' in status_or_memo):
                real_status = '대체보충' # 메모 형식은 '대체보충'으로 강제 매핑
            # --- ▲▲▲ [파싱 로직 교체 완료] ▲▲▲ ---

            cell.value = worker_name
            if not worker_name: continue
            
            if is_special_day:
                if str(col_name).isdigit():
                    cell.fill = holiday_blue_fill
                    if worker_name == weekend_oncall_worker: cell.font = duty_font
                elif '오후' in str(col_name): cell.value = ""
                continue
            
            # --- ▼▼▼ [핵심 수정] 색상 적용 로직 (F2DCDB 우선) ▼▼▼ ---
            fill_hex = color_map.get(real_status, 'FFFFFF') # status -> real_status
            
            initial_raw_value = str(initial_row.get(col_name, '')).strip()
            cell_changed = (raw_value != initial_raw_value)

            if cell_changed:
                cell.fill = changed_fill # 1순위: 변경됨 (F2DCDB)
            elif fill_hex and fill_hex != 'FFFFFF':
                cell.fill = PatternFill(start_color=fill_hex, fill_type='solid') # 2순위: 상태 색상
            # 3순위: 기본 (흰색)
            # --- ▲▲▲ [색상 적용 완료] ▲▲▲ ---

            if col_name == '오전당직(온콜)': cell.font = duty_font
            
            if cell_changed:
                cell.comment = Comment(f"변경 전: {initial_raw_value or '빈 값'}", "Edit Tracker")

            # --- ▼▼▼ [핵심 수정] 메모 추가 로직 ▼▼▼ ---
            if real_status == '대체보충' and pd.notna(status_or_memo) and re.search(r'\d{1,2}/\d{1,2}', status_or_memo):
                try:
                    # '변경 전' 코멘트가 없을 때만 '대체' 메모를 추가
                    if cell.comment is None: 
                        cell.comment = Comment(status_or_memo, "Schedule Bot")
                except Exception as e_memo:
                    pass # 코멘트 추가 실패 시 무시
            # --- ▲▲▲ [메모 추가 완료] ▲▲▲ ---

    # --- ✨ [핵심 수정] 익월 누적 현황을 올바른 형식으로 추가 ---
    if not edited_cumulative_df.empty:
        style_args = {'font': default_font, 'bold_font': bold_font, 'border': border}
        append_summary_table_to_excel(ws, edited_cumulative_df, style_args)

    # --- 열 너비 설정 ---
    ws.column_dimensions['A'].width = 11
    for i in range(2, len(final_columns) + 1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 9

    wb.save(output)
    return output.getvalue()

# [★ L531의 이 함수 전체를 교체하세요 ★]

def create_checking_schedule_excel(initial_df, edited_df, edited_cumulative_df, df_special, df_requests, closing_dates, month_str):
    """
    [관리자 확인용] - 최종 버전 다운로드용
    [★ v4 수정 ★]
    - 평일 근무일 기준: 오전 13열~, 오후 5열~ 배경색을 B2B2B2(회색)로 강제 지정
    - 목적: 휴가/휴근 등 근무 불가자를 해당 열로 빼두었을 때 시각적으로 구분하기 위함
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "스케줄 (배정 확인용)"

    # --- 스타일 정의 ---
    font_name = "맑은 고딕" if platform.system() == "Windows" else "Arial"
    default_font = Font(name=font_name, size=9)
    bold_font = Font(name=font_name, size=9, bold=True)
    duty_font = Font(name=font_name, size=9, bold=True, color="FF69B4")
    header_font = Font(name=font_name, size=9, color='FFFFFF', bold=True)
    
    color_map = {'휴가': 'DA9694', '학회': 'DA9694', '꼭 근무': 'FABF8F', '보충': 'FFF28F', '대체보충': 'A9D08E', '휴근': 'B1A0C7', '대체휴근': '95B3D7', '특수근무': 'D0E0E3', '기본': 'FFFFFF'}
    
    header_fill = PatternFill(start_color='000000', fill_type='solid')
    date_col_fill = PatternFill(start_color='808080', fill_type='solid')
    weekday_fill = PatternFill(start_color='FFF2CC', fill_type='solid')
    special_day_fill = PatternFill(start_color='95B3D7', fill_type='solid')
    
    changed_fill = PatternFill(start_color='F2DCDB', fill_type='solid') # (연분홍 - 변경됨)
    empty_day_fill = PatternFill(start_color='808080', fill_type='solid')
    holiday_blue_fill = PatternFill(start_color="DDEBF7", fill_type='solid')
    
    # [신규] 근무 불가 영역 회색 (B2B2B2)
    gray_area_fill = PatternFill(start_color='B2B2B2', fill_type='solid')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    # --- 동적 열 정의 ---
    checking_columns = edited_df.columns.tolist()

    # --- 헤더 생성 ---
    for c, col_name in enumerate(checking_columns, 1):
        cell = ws.cell(row=1, column=c, value=col_name); cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border

    # --- 데이터 행 생성 및 서식 적용 ---
    for r, (idx, edited_row) in enumerate(edited_df.iterrows(), 2):
        initial_row = initial_df.loc[idx]
        try:
            current_date = datetime.strptime(f"{month_str.split('년')[0]}-{edited_row['날짜']}", "%Y-%m월 %d일").date()
            current_date_iso = current_date.strftime('%Y-%m-%d')
        except: current_date, current_date_iso = None, None
        
        is_row_empty = all(pd.isna(v) or str(v).strip() == '' for k, v in edited_row.items() if k not in ['날짜', '요일'])
        is_special_day = current_date in pd.to_datetime(df_special['날짜']).dt.date.values if current_date and not df_special.empty else False
        is_empty_day = (is_row_empty and not is_special_day) or (current_date_iso in closing_dates)
        
        weekend_oncall_worker = None
        if is_special_day:
            special_day_info = df_special[pd.to_datetime(df_special['날짜']).dt.date == current_date]
            if not special_day_info.empty and '당직' in special_day_info.columns:
                oncall_val = special_day_info['당직'].iloc[0]
                if pd.notna(oncall_val) and oncall_val != "당직 없음": weekend_oncall_worker = str(oncall_val).strip()

        for c, col_name in enumerate(checking_columns, 1):
            
            # --- [신규] 회색 영역(B2B2B2) 판별 로직 ---
            is_gray_area = False
            col_str = str(col_name)
            
            # 조건: 평일 근무일(토/휴일X, 빈날X)이면서 -> 특정 열 번호 이상인 경우
            if not is_special_day and not is_empty_day:
                # 1. 오전 (숫자로 된 열): 13 이상
                if col_str.isdigit():
                    if int(col_str) >= 13:
                        is_gray_area = True
                # 2. 오후 ('오후'로 시작하는 열): 5 이상 (오후5, 오후6...)
                elif col_str.startswith('오후'):
                    # '오후' 뒤의 숫자를 추출
                    pm_num_match = re.search(r'오후(\d+)', col_str)
                    if pm_num_match:
                        pm_num = int(pm_num_match.group(1))
                        if pm_num >= 5:
                            is_gray_area = True
            # ---------------------------------------

            raw_value = str(edited_row.get(col_name, '')).strip()
            
            # [수정] 내용 파싱 시 괄호 내용 지우지 않음 (요청사항 반영)
            # 엑셀 셀에는 원본 텍스트 그대로 넣습니다.
            cell_value_text = raw_value 
            
            # 다만, 스타일링을 위해 이름과 상태는 분리해서 파악
            status_or_memo = '기본'
            match = re.match(r'.+?\((.+)\)', raw_value)
            if match: status_or_memo = match.group(1).strip()
            
            real_status = '기본'
            if status_or_memo == '기본': real_status = '기본'
            elif status_or_memo in color_map: real_status = status_or_memo
            elif pd.notna(status_or_memo) and (re.search(r'\d{1,2}/\d{1,2}', status_or_memo) or '대체됨' in status_or_memo or '대체함' in status_or_memo):
                real_status = '대체보충' 
            else: real_status = '기본' 
            
            # 이름만 추출 (비교 및 당직 폰트 적용용)
            worker_name = re.sub(r'\(.+\)', '', raw_value).strip()

            cell = ws.cell(row=r, column=c, value=cell_value_text)
            cell.font = default_font; cell.alignment = center_align; cell.border = border

            if is_empty_day: cell.fill = empty_day_fill; continue
            if col_name == '날짜': cell.fill = date_col_fill; continue
            if col_name == '요일': cell.fill = special_day_fill if is_special_day else weekday_fill; continue
            
            # 회색 영역이면 값 여부와 상관없이 회색 칠하기 (우선순위 최상)
            # (단, 셀에 텍스트가 있다면 그대로 보임)
            if is_gray_area:
                cell.fill = gray_area_fill
                # continue를 하지 않고 아래로 흘려보내 폰트 설정 등은 적용되게 할 수도 있으나,
                # 배경색 결정은 여기서 끝내는 게 깔끔합니다.
                # 단, '변경 전' 코멘트는 달아야 하므로 아래 로직 일부는 수행해야 합니다.
            
            # 값이 없는데 회색영역도 아니면 스킵
            if not raw_value and not is_gray_area: continue
            
            if is_special_day:
                if str(col_name).isdigit():
                    cell.fill = holiday_blue_fill
                    if worker_name == weekend_oncall_worker: cell.font = duty_font
                elif '오후' in str(col_name): cell.value = ""
                continue
            
            # --- [수정] 색상 적용 로직 (B2B2B2 최우선) ---
            
            fill_hex = color_map.get(real_status, 'FFFFFF') 
            initial_raw_value = str(initial_row.get(col_name, '')).strip()
            cell_changed = (raw_value != initial_raw_value)

            # 0순위: 회색 영역 (근무 불가 열) - 이미 위에서 is_gray_area 체크함
            if is_gray_area:
                cell.fill = gray_area_fill
            # 1순위: 변경된 셀 (F2DCDB)
            elif cell_changed:
                cell.fill = changed_fill
            # 2순위: 상태 색상
            elif fill_hex and fill_hex != 'FFFFFF':
                cell.fill = PatternFill(start_color=fill_hex, fill_type='solid') 
            # 3순위: 기본 (흰색) - 이미 위에서 처리됨 (디폴트)
            
            # ---------------------------------------------------
            
            if col_name == '오전당직(온콜)': cell.font = duty_font
            
            if cell_changed:
                cell.comment = Comment(f"변경 전: {initial_raw_value or '빈 값'}", "Edit Tracker")

            if real_status == '대체보충' and pd.notna(status_or_memo) and re.search(r'\d{1,2}/\d{1,2}', status_or_memo):
                try:
                    if cell.comment is None: 
                        cell.comment = Comment(status_or_memo, "Schedule Bot")
                except: pass
    
    # --- 익월 누적 현황 추가 ---
    if not edited_cumulative_df.empty:
        style_args = {'font': default_font, 'bold_font': bold_font, 'border': border}
        append_summary_table_to_excel(ws, edited_cumulative_df, style_args)

    # --- 열 너비 설정 ---
    ws.column_dimensions['A'].width = 11
    for i in range(2, len(checking_columns) + 1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 9

    wb.save(output)
    return output.getvalue()

def recalculate_summary_from_schedule(edited_schedule_df, df_cumulative_initial, all_names, df_schedule_mapping):
    """
    (신규 함수)
    수정된 스케줄 data_editor 내용을 실시간으로 파싱하여,
    '보충', '당직' 횟수를 재계산하고 누적 테이블 DataFrame을 반환합니다.
    
    [★ 스케줄 수정 페이지 v2 수정 ★]
    - '스케줄 배정' 페이지의 v2 파싱 로직을 이식합니다.
    - (10/6에서 대체됨)을 +1로, (대체휴근)을 -1로 인식합니다.
    """
    
    # 1. 이름별로 (보충/휴근) 횟수, (당직) 횟수를 집계할 카운터 초기화
    am_bochong_counts = Counter()
    pm_bochong_counts = Counter()
    oncall_counts = Counter()

    # 2. 스케줄 data_editor (edited_schedule_df)의 모든 셀을 순회
    for idx, row in edited_schedule_df.iterrows():
        
        # 2-1. data_editor의 날짜(예: "10월 1일")를 ISO 날짜(예: "2025-10-01")로 변환
        try:
            # '스케줄 수정' 페이지의 인덱스(idx)는 GSheet 원본 인덱스와 일치
            date_iso = df_schedule_mapping.loc[idx, '날짜']
        except Exception:
            continue 

        for col_name in edited_schedule_df.columns:
            raw_value = str(row[col_name] or '').strip()
            if not raw_value:
                continue

            # 2-2. 셀 텍스트에서 이름과 상태 파싱
            worker_name = re.sub(r'\(.+\)', '', raw_value).strip()
            status_match = re.search(r'\((.+)\)', raw_value)
            
            # --- ▼▼▼ [핵심 수정] 파싱 로직 교체 ▼▼▼ ---
            status_text = status_match.group(1).strip() if status_match else '기본'
            
            # 2-3. 열 이름(col_name)에 따라 시간대 결정
            time_slot = None
            if col_name.isdigit(): time_slot = '오전'
            elif col_name.startswith("오후"): time_slot = '오후'
            elif col_name == '오전당직(온콜)': time_slot = '오전당직'
            
            if not time_slot or not worker_name:
                continue

            # 2-4. [수정된 로직] 파싱된 텍스트(status_text)를 '실제 상태'로 변환
            real_status_effect = 0 # 0: 기본, +1: 보충, -1: 휴근
            
            if status_text in ['보충', '대체보충']:
                real_status_effect = 1
            elif status_text in ['휴근', '대체휴근']:
                real_status_effect = -1
            elif pd.notna(status_text) and (re.search(r'\d{1,2}/\d{1,2}', status_text) or '대체됨' in status_text):
                # (10/6에서 대체됨)과 같은 메모 형식은 '대체보충'(+1)으로 간주
                real_status_effect = 1

            # 2-5. 카운터 집계
            if time_slot == '오전당직':
                oncall_counts[worker_name] += 1
            
            elif time_slot == '오전':
                if real_status_effect == 1:
                    am_bochong_counts[worker_name] += 1
                elif real_status_effect == -1:
                    am_bochong_counts[worker_name] -= 1
            
            elif time_slot == '오후':
                if real_status_effect == 1:
                    pm_bochong_counts[worker_name] += 1
                elif real_status_effect == -1:
                    pm_bochong_counts[worker_name] -= 1
            # --- ▲▲▲ [핵심 수정 완료] ▲▲▲ ---

    # 3. GSheet에서 로드한 *원본* 누적 테이블을 기반으로 최종 테이블 재구성
    recalculated_summary_df = df_cumulative_initial.copy()
    if '항목' not in recalculated_summary_df.columns:
        try:
            first_col = recalculated_summary_df.columns[0]
            recalculated_summary_df = recalculated_summary_df.set_index(first_col).transpose().reset_index().rename(columns={'index':'항목'})
        except Exception:
            return df_cumulative_initial 
            
    recalculated_summary_df = recalculated_summary_df.set_index('항목')

    # 4. 모든 근무자 목록(all_names)을 순회하며 값 채우기
    for name in all_names:
        if name not in recalculated_summary_df.columns:
            recalculated_summary_df[name] = 0 
        
        # 4-1. GSheet 원본 값 가져오기 (오류 방지를 위해 .get(name, 0) 사용)
        base_am = int(recalculated_summary_df.loc['오전누적'].get(name, 0))
        base_pm = int(recalculated_summary_df.loc['오후누적'].get(name, 0))
        base_am_oncall = int(recalculated_summary_df.loc['오전당직누적'].get(name, 0))
        base_pm_oncall = int(recalculated_summary_df.loc['오후당직누적'].get(name, 0))

        # 4-2. 실시간 집계 값 가져오기
        am_bochong = am_bochong_counts.get(name, 0)
        pm_bochong = pm_bochong_counts.get(name, 0)
        am_oncall_total = oncall_counts.get(name, 0)

        # 4-3. 최종 값 계산 및 덮어쓰기
        recalculated_summary_df.at["오전보충", name] = am_bochong
        recalculated_summary_df.at["오전합계", name] = base_am 
        recalculated_summary_df.at["오전누적", name] = base_am + am_bochong

        recalculated_summary_df.at["오후보충", name] = pm_bochong
        recalculated_summary_df.at["오후합계", name] = base_pm
        recalculated_summary_df.at["오후누적", name] = base_pm + pm_bochong

        recalculated_summary_df.at["오전당직", name] = am_oncall_total
        recalculated_summary_df.at["오전당직누적", name] = base_am_oncall + am_oncall_total
        
        recalculated_summary_df.at["오후당직", name] = 0 
        recalculated_summary_df.at["오후당직누적", name] = base_pm_oncall

    # '항목' 열을 다시 복원하여 반환
    recalculated_summary_dfr = recalculated_summary_df.reset_index()

    # --- ▼▼▼ [오타 수정] dfr -> df ▼▼▼ ---
    # 원본 build_summary_table과 동일하게 모든 숫자 열을 int로 강제 변환
    for col in recalculated_summary_dfr.columns:
        if col != '항목':
            recalculated_summary_dfr[col] = pd.to_numeric(recalculated_summary_dfr[col], errors='coerce').fillna(0).astype(int)

    return recalculated_summary_dfr
    # --- ▲▲▲ [오타 수정 완료] ▲▲▲ ---

# --- ▼▼▼ [교체] L702 ~ L786의 기존 save_schedule 함수 전체를 교체 ▼▼▼ ---
def save_schedule(month_str, sheet_name, df_to_save, df_cum_to_save):
    with st.spinner(f"'{sheet_name}' 시트에 저장 중입니다..."):
        try:
            sheet = get_spreadsheet()
            
            # 1. 스케줄 시트 저장
            try: 
                worksheet = sheet.worksheet(sheet_name)
            except WorksheetNotFound: 
                worksheet = sheet.add_worksheet(title=sheet_name, rows=100, cols=50)
            
            original_cols_df = st.session_state.get("df_schedule_original")
            if original_cols_df is None:
                st.error("원본 스케줄 컬럼 정보를 찾을 수 없습니다.")
                st.stop()

            columns_to_save = original_cols_df.columns.tolist()
            df_to_save_final = pd.DataFrame(columns=columns_to_save)
            for col in columns_to_save:
                if col in df_to_save.columns:
                    df_to_save_final[col] = df_to_save[col]
                else:
                    df_to_save_final[col] = ''
            final_data = [columns_to_save] + df_to_save_final.fillna('').values.tolist()
            update_sheet_with_retry(worksheet, final_data)

            # 2. 익월 누적 시트 저장
            if not df_cum_to_save.empty:
                current_month_dt_save = datetime.strptime(month_str, "%Y년 %m월")
                next_month_str_save = (current_month_dt_save + relativedelta(months=1)).strftime("%Y년 %-m월")
                
                version_s_save = ""
                if " ver" in sheet_name:
                    version_s_save = " " + sheet_name.split(" 스케줄 ")[1]
                elif "최종" in sheet_name:
                     version_s_save = " 최종"
                
                cum_sheet_name = f"{next_month_str_save} 누적{version_s_save}"

                try: 
                    ws_cum = sheet.worksheet(cum_sheet_name)
                except WorksheetNotFound: 
                    ws_cum = sheet.add_worksheet(title=cum_sheet_name, rows=100, cols=50)
                
                df_to_save_int = df_cum_to_save.copy()
                for col in df_to_save_int.columns[1:]:
                    df_to_save_int[col] = pd.to_numeric(df_to_save_int[col], errors='coerce').fillna(0).astype(int)

                cum_data = [df_to_save_int.columns.tolist()] + df_to_save_int.astype(str).values.tolist()
                update_sheet_with_retry(ws_cum, cum_data)

            # --- ▼▼▼ [핵심 수정] 저장 성공 시 세션 상태 및 플래그 리셋 ▼▼▼ ---
            st.session_state.df_display_initial = df_to_save.copy() # (스케줄 원본)
            st.session_state.df_cumulative_next_display = df_cum_to_save.copy() # (누적 결과)
            st.session_state.df_cumulative_next_initial = df_cum_to_save.copy() # (누적 원본)
            
            st.session_state.save_successful = True # (기존)
            st.session_state.last_saved_sheet_name = sheet_name # (기존)
            
            st.success(f"🎉 스케줄과 익월 누적 데이터가 '{sheet_name}' 버전에 맞게 저장되었습니다.")
            time.sleep(1)
            st.cache_data.clear()
            st.cache_resource.clear()
            st.rerun()

        except Exception as e: 
            st.error(f"Google Sheets 저장 중 오류 발생: {e}")
# --- ▲▲▲ [교체] save_schedule 함수 교체 끝 ▲▲▲ ---

# --- 메인 UI ---
st.header("✍️ 스케줄 수정", divider='rainbow')
kst = ZoneInfo("Asia/Seoul")
month_dt_now = datetime.now(kst).replace(day=1) + relativedelta(months=1)
month_str = month_dt_now.strftime("%Y년 %-m월")

# gc = get_gspread_client()
# sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
sheet = get_spreadsheet()
versions = find_schedule_versions(month_str)

def on_version_change():
    st.session_state.data_loaded = False

# [핵심 추가] 전체 버전 목록을 다시 불러오기 위한 새로고침 버튼
col_text, col_btn = st.columns([3, 1], vertical_alignment="center")

with col_text:
    st.caption("ℹ️ 먼저 새로고침 버튼으로 최신 데이터를 불러온 뒤 진행해주세요.")

with col_btn:
    # use_container_width=True를 쓰면 버튼이 컬럼 너비에 맞춰 깔끔하게 찹니다.
    if st.button("🔄 새로고침 (R)", use_container_width=True):
        st.cache_data.clear()
        st.cache_resource.clear()
        
        # --- ▼▼▼ [수정] 기존 for 루프 대신 명시적 삭제로 변경 ▼▼▼ ---
        
        # 새로고침 시 반드시 삭제해야 하는 핵심 상태 키 목록
        keys_to_delete = [
            "data_loaded", 
            "apply_messages", 
            "change_log", 
            "is_final_version",
            "loaded_sheet_name",        # (필수) 로드된 시트 이름
            "editor_has_changes",       # (필수) 수정 플래그
            "save_successful",          # (권장) 저장 상태
            "last_saved_sheet_name"     # (권장) 저장 이름
        ]
        
        # df_로 시작하는 모든 데이터프레임 키도 삭제 목록에 추가
        df_keys = [key for key in st.session_state.keys() if key.startswith("df_")]
        keys_to_delete.extend(df_keys)

        # 세션 상태에서 해당 키들 삭제
        for key in keys_to_delete:
            if key in st.session_state:
                del st.session_state[key]
                
        # --- ▲▲▲ [수정] 코드 교체 완료 ▲▲▲ ---
        
        st.rerun()

if not versions:
    st.warning(f"'{month_str}'에 해당하는 스케줄 시트가 없습니다. 먼저 스케줄을 생성해주세요."); st.stop()

# [★ L884 ~ L886을 이 코드로 교체하세요 ★]
version_list = list(versions.keys())

# --- ▼▼▼ [핵심 버그 수정] ▼▼▼ ---
# 'index=0'을 사용하는 대신, st.session_state에 저장된 값을 찾아
# 해당 값의 인덱스를 'index'로 사용합니다.
# 이렇게 하면 '새로고침' 시 key가 일시적으로 무효화되어도
# 'index'가 올바른 값을 가리키게 됩니다.

# 1. 세션에 저장된 값이 있는지 확인
selected_value = st.session_state.get("selected_sheet_name")

# 2. 저장된 값이 version_list에 있는지 확인
try:
    # 3. 있다면, 그 값의 인덱스를 사용
    current_index = version_list.index(selected_value)
except ValueError:
    # 4. 없다면 (삭제되었거나 첫 로드 시), 0번 인덱스(최종)를 사용
    current_index = 0
# --- ▲▲▲ [수정 완료] ---

st.divider()
st.subheader("📋 스케줄 버전 선택")
selected_sheet_name = st.selectbox(
    "- 불러올 스케줄 버전을 선택하세요:", 
    options=version_list, 
    index=current_index, 
    key="selected_sheet_name", 
    on_change=on_version_change
)

# [삭제] 기존 상단 삭제 버튼 UI 코드 블록 제거됨 (col_delete 부분)

needs_load = False
if not st.session_state.get("data_loaded", False):
    needs_load = True
elif st.session_state.get("loaded_sheet_name") != selected_sheet_name:
    needs_load = True

# --- ▼▼▼ 데이터 로드 로직 (기존 유지) ▼▼▼ ---
if needs_load:
    data = load_data(month_str, selected_sheet_name)

    st.session_state["df_schedule_original"] = data["schedule"]
    st.session_state["df_cumulative_next_display"] = data["cumulative_display"]
    st.session_state["df_display_initial"] = data["schedule"].copy()
    st.session_state["df_swaps"] = data["swaps"]
    st.session_state["df_special"] = data["special"]
    st.session_state["df_requests"] = data["requests"]
    st.session_state["closing_dates"] = data["closing_dates"]
    st.session_state["is_final_version"] = data["is_final_version"]
    
    st.session_state["df_cumulative_base_initial"] = data["base_cumulative"]
    st.session_state["df_schedule_mapping"] = data["schedule_mapping"]
    
    if st.session_state.df_cumulative_next_display.empty and not data["base_cumulative"].empty:
        st.info("로드된 익월 누적 테이블이 없어, '당월(전월 누적)' 데이터를 기준으로 새로 생성합니다.")
        all_names_list = data["base_cumulative"].columns[1:].tolist()
        
        st.session_state.df_cumulative_next_display = recalculate_summary_from_schedule(
            data["schedule"], 
            data["base_cumulative"], 
            all_names_list, 
            data["schedule_mapping"]
        )

    st.session_state["df_cumulative_next_initial"] = st.session_state.df_cumulative_next_display.copy()
    
    st.session_state.data_loaded = True
    st.session_state["loaded_sheet_name"] = selected_sheet_name
# --- ▲▲▲ 데이터 로드 끝 ▲▲▲ ---

is_final_version = st.session_state.get("is_final_version", False)

if is_final_version:
    st.error("🚨 최종 버전의 초기화 및 수정은 '방배정' 페이지에서 진행 바랍니다. 이 페이지에서는 최종본 내용 확인 및 다운로드만 가능합니다.")
    st.session_state["disable_editing"] = True
else:
    st.session_state["disable_editing"] = False


# 2. 선택된 버전을 바로 다운로드하는 버튼 + 삭제 버튼 통합
st.write(" ") # 여백

version_part = ""
schedule_keyword = "스케줄 "
if schedule_keyword in selected_sheet_name:
    version_part = selected_sheet_name.split(schedule_keyword, 1)[1]

display_version = f" {version_part}" if version_part else ""

# 데이터가 로드되었는지 확인
if "df_display_initial" in st.session_state:
    is_final_version_selected = "최종" in selected_sheet_name

    # [수정] 버전 상관없이 항상 3컬럼 (메인 다운로드 / 확인용 다운로드 / 삭제) 유지
    col_down_main, col_down_sub, col_del = st.columns([1, 1, 1])

    # 1. 메인 다운로드 버튼 (공통)
    with col_down_main:
        st.download_button(
            label=f"📥 스케줄{display_version} 다운로드",
            data=create_final_schedule_excel(
                st.session_state.df_display_initial, st.session_state.df_display_initial, 
                st.session_state.df_cumulative_next_display, st.session_state.df_special, 
                st.session_state.df_requests, st.session_state.get("closing_dates", []), month_str
            ),
            file_name=f"{month_str} 스케줄{display_version}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary",
            key="download_now_final"
        )

    # 2. 서브 다운로드 버튼 (이제 최종 버전에서도 표시됨)
    with col_down_sub:
        st.download_button(
            label=f"📥 배정 확인용 다운로드",
            data=create_checking_schedule_excel(
                st.session_state.df_display_initial, st.session_state.df_display_initial,
                st.session_state.df_cumulative_next_display, st.session_state.df_special, 
                st.session_state.df_requests, st.session_state.get("closing_dates", []), month_str
            ),
            file_name=f"{month_str} 스케줄{display_version} (배정 확인용).xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="secondary",
            key="download_now_checking"
        )

    # 3. 삭제 버튼 (우측 끝 컬럼에 배치)
    with col_del:
        # 삭제는 위험한 작업이므로 Expander 안에 배치
        with st.expander("🗑️ 현재 버전 데이터 삭제"):
            st.error("이 작업은 되돌릴 수 없습니다!\nGoogle Sheets에서 해당 버전이 영구 삭제됩니다.")
            
            # 최종 삭제 확인 버튼
            if st.button("네, 삭제합니다.", type="primary", use_container_width=True, key="btn_delete_version"):
                delete_schedule_version(month_str, selected_sheet_name)

st.divider()
st.subheader("📅 배정 스케줄 수정")

# --- ▼▼▼ [신규] '스케줄 배정' 페이지의 수정/연동/로깅 로직 적용 ▼▼▼ ---

# 1. 세션에서 원본 데이터 불러오기
df_schedule_initial = st.session_state.get("df_display_initial")
df_cumulative_base = st.session_state.get("df_cumulative_base_initial")
df_schedule_mapping = st.session_state.get("df_schedule_mapping")
df_cumulative_next_initial = st.session_state.get("df_cumulative_next_initial")
all_names_list = df_cumulative_base.columns[1:].tolist() if df_cumulative_base is not None else []

# (수정 비활성화 플래그)
disable_editing = st.session_state.get("disable_editing", False)

# 2. 상단 (스케줄) data_editor
if df_schedule_initial is None or df_schedule_mapping is None or df_cumulative_next_initial is None:
    st.error("스케줄 원본, 날짜 매핑, 또는 누적 원본 데이터를 로드하지 못했습니다. 새로고침 해주세요.")
    st.stop()

# '스케줄 수정' 페이지는 GSheet에 이미 (괄호)가 포함된 텍스트를 저장하므로,
# '스케줄 배정' 페이지와 달리 별도의 (괄호) 추가 로직이 필요 없습니다.
df_to_edit_schedule = df_schedule_initial.copy()

edited_df = st.data_editor(
    df_to_edit_schedule,
    use_container_width=True,
    key="schedule_editor",
    disabled=['날짜', '요일'] if not disable_editing else df_to_edit_schedule.columns.tolist(),
)

# 3. 상단 (스케줄) 수정 로그
st.markdown("📝 **스케줄 수정사항**")
schedule_change_log = []
original_schedule_df = df_to_edit_schedule # 에디터에 렌더링된 초기값

schedule_has_changed = False  # <--- ★★★ [ 1. 이 줄을 추가 ] ★★★

if original_schedule_df is not None and not edited_df.equals(original_schedule_df):
    schedule_has_changed = True # <--- ★★★ [ 2. 이 줄을 추가 ] ★★★
    try:
        import numpy as np 
        diff_indices = np.where(edited_df.astype(str).ne(original_schedule_df.astype(str)))
        changed_cells = set(zip(diff_indices[0], diff_indices[1])) # 중복 로그 방지
        
        for row_idx, col_idx in changed_cells:
            date_str = edited_df.iloc[row_idx, 0] # '날짜' 열 (예: "10월 1일")
            slot_name = edited_df.columns[col_idx] # 변경된 열 이름 (예: "1")
            old_value = original_schedule_df.iloc[row_idx, col_idx]
            new_value = edited_df.iloc[row_idx, col_idx]
            log_msg = f"{date_str} '{slot_name}' 변경: '{old_value or '빈 값'}' → '{new_value or '빈 값'}'"
            schedule_change_log.append(log_msg)
    except Exception as e:
        schedule_change_log.append(f"[로그 오류] 스케줄 변경사항 비교 중 오류: {e}")
        
if schedule_change_log:
    st.code("\n".join(f"• {msg}" for msg in sorted(schedule_change_log)), language='text')
else:
    st.info("수정된 사항이 없습니다.")

st.divider()
st.subheader("📊 누적 테이블 수정")
st.write("- 누적 테이블은 '스케줄표 수정' 편집기에 반영된 내용을 바탕으로 자동 재계산됩니다.\n- 주의) 대체보충은 수정 시 누적 테이블을 직접 수정해주셔야 합니다.")
# 4. 하단 (누적) data_editor
if df_cumulative_base is None or not all_names_list:
    st.error("누적 테이블 베이스 데이터 또는 이름 목록을 로드하지 못했습니다. 새로고침 해주세요.")
    st.stop()

# [실시간 재계산]
if schedule_has_changed:
    # 1. 상단 스케줄이 수정됨 -> 재계산 실행
    try:
        summary_df_input = recalculate_summary_from_schedule(
            edited_df,               # (상단) 에디터의 최종 결과
            df_cumulative_base,      # (로드된) 지난달 누적 원본
            all_names_list,          # (로드된) 이름 목록
            df_schedule_mapping      # (생성된) 날짜 매핑
        )
    except Exception as e_recalc:
        st.error(f"누적 테이블 자동 재계산 중 오류 발생: {e_recalc}")
        st.exception(e_recalc)
        summary_df_input = df_cumulative_next_initial # 오류 시, 로드했던 원본으로 복구
else:
    # 2. 상단 스케줄이 수정되지 않음 (페이지 첫 로드) -> 원본 표시
    summary_df_input = df_cumulative_next_initial.copy()

column_config = {
    summary_df_input.columns[0]: st.column_config.Column(disabled=True),
    **{col: st.column_config.NumberColumn(format="%d") 
        for col in summary_df_input.columns[1:]}
}

edited_cumulative_df = st.data_editor(
    summary_df_input, # 자동 재계산된 결과를 입력
    hide_index=True,
    key="cumulative_editor",
    use_container_width=True,
    column_config=column_config,
    disabled=disable_editing, # '최종' 버전일 경우 비활성화
)

# 5. 하단 (누적) 수정 로그
st.markdown("📝 **누적 테이블 수정사항**")
summary_change_log = [] # 리스트 초기화

# --- ▼▼▼ [핵심 수정] 정렬 순서를 위한 Map 정의 ▼▼▼ ---
# 사용자가 요청한 정확한 순서
desired_order = [
    "오전보충", "임시보충", "오전합계", "오전누적", 
    "오후보충", "온콜검사", "오후합계", "오후누적", 
    "오전당직", "오전당직누적", "오후당직", "오후당직누적"
]
# 순서를 숫자로 매핑 (예: '오전보충': 0, '오전합계': 2, '오전누적': 3)
order_map = {item_name: index for index, item_name in enumerate(desired_order)}
# --- ▲▲▲ [수정 완료] ▲▲▲ ---

# (A) = 로드 시점의 원본
original_summary_df = df_cumulative_next_initial 

if original_summary_df is not None and not edited_cumulative_df.equals(original_summary_df): # (A) vs (C)
    try:
        import numpy as np 
        stats_orig_str = original_summary_df.astype(str)
        stats_edit_str = edited_cumulative_df.astype(str)
        
        diff_indices_stats = np.where(stats_edit_str.ne(stats_orig_str))
        changed_cells_stats = set(zip(diff_indices_stats[0], diff_indices_stats[1]))

        for row_idx, col_idx in changed_cells_stats:
            item_name = edited_cumulative_df.iloc[row_idx, 0] # 예: "오전당직누적"
            person_name = edited_cumulative_df.columns[col_idx] # 예: "강승주"
            
            old_value = original_summary_df.iloc[row_idx, col_idx]
            new_value = edited_cumulative_df.iloc[row_idx, col_idx]
            
            log_msg = f"'{person_name}'의 '{item_name}' 변경: {old_value} → {new_value}"
            
            # --- ▼▼▼ [수정] 로그를 튜플로 저장 (정렬 기준 포함) ▼▼▼ ---
            # 1. 항목 순서 (order_map에서 조회, 없으면 99)
            item_sort_key = order_map.get(item_name, 99) 
            # (사람이름, 항목순서, 실제로그메시지)
            summary_change_log.append((person_name, item_sort_key, log_msg))
            # --- ▲▲▲ [수정 완료] ▲▲▲ ---
            
    except Exception as e:
        # (예외 발생 시 튜플 대신 문자열로 추가)
        summary_change_log.append(("[로그 오류]", 99, f"[로그 오류] 누적 테이블 변경사항 비교 중 오류: {e}"))

if summary_change_log:
    # --- ▼▼▼ [수정] 튜플을 기준으로 정렬 (사람이름 > 항목순서) ▼▼▼ ---
    # 1. 사람 이름(x[0])으로 먼저 정렬
    # 2. 같은 사람이면 항목 순서(x[1])로 정렬
    summary_change_log.sort(key=lambda x: (x[0], x[1]))
    
    # 3. 정렬된 튜플 리스트에서 실제 로그 메시지(x[2])만 추출
    log_text_stats = "\n".join(f"• {msg_tuple[2]}" for msg_tuple in summary_change_log)
    st.code(log_text_stats, language='text')
    # --- ▲▲▲ [수정 완료] ▲▲▲ ---
else:
    st.info("수정된 사항이 없습니다.")

# --- ▲▲▲ [신규] 교체 완료 ---

st.divider()

# --- 변경사항 유무 확인 ---
has_unsaved_changes = (not edited_df.equals(st.session_state.df_display_initial)) or \
                      (not edited_cumulative_df.empty and not edited_cumulative_df.equals(st.session_state.df_cumulative_next_display))

# --- [핵심 수정] 1. 'has_unsaved_changes'를 플래그가 아닌 '실제 비교'로 정의 ---
# (이 블록을 맨 위로 이동시켰습니다)
try:
    # 스케줄 비교: (에디터 최종본) vs (로드 시점의 원본)
    original_schedule_for_compare = df_to_edit_schedule
    schedule_changed = not edited_df.equals(original_schedule_for_compare)

    # 누적 비교: (에디터 최종본) vs (로드 시점의 원본)
    original_cumulative_for_compare = df_cumulative_next_initial
    
    cumulative_changed = not edited_cumulative_df.equals(original_cumulative_for_compare)
    
    has_unsaved_changes = schedule_changed or cumulative_changed

except Exception as e:
    st.error(f"변경 사항 비교 중 오류: {e}")
    has_unsaved_changes = False # 오류 시 안전하게 비활성화

# --- [핵심 수정] 2. '저장 완료' 블록과 '수정 중' 블록을 if/else로 분리 ---

# [★ L1188 ~ L1232의 '저장 완료' 블록 전체를 이 코드로 교체하세요 ★]

# [★ A. 저장 완료 상태 ★]
# (조건: '저장 성공' 플래그가 True이고, '실제' 변경 사항이 없을 때)
if st.session_state.get("save_successful", False) and not has_unsaved_changes:
    st.subheader("✅ 저장 완료! 엑셀 파일 다운로드")
    st.write("- 수정된 스케줄을 아래 버튼으로 다운로드하세요.")

    last_saved_sheet = st.session_state.get("last_saved_sheet_name", "스케줄")
    
    version_part = ""
    schedule_keyword = "스케줄 "
    if schedule_keyword in last_saved_sheet:
        version_part = last_saved_sheet.split(schedule_keyword, 1)[1]

    display_version = f" {version_part}" if version_part else ""

    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label=f"📥 스케줄{display_version} 다운로드",
            data=create_final_schedule_excel(
                # --- ▼▼▼ [핵심 버그 수정] ▼▼▼ ---
                # 'df_display_initial' (C) 대신 'df_schedule_original' (A)을 전달
                initial_df=st.session_state.df_schedule_original, 
                # --- ▲▲▲ [수정 완료] ▲▲▲ ---
                edited_df=edited_df, 
                edited_cumulative_df=edited_cumulative_df,
                df_special=st.session_state.df_special, 
                df_requests=st.session_state.df_requests,
                closing_dates=st.session_state.get("closing_dates", []), 
                month_str=month_str
            ),
            file_name=f"{month_str} 스케줄{display_version}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary",
            key="download_saved_final"
        )
    with col2:
        st.download_button(
            label=f"📥 스케줄{display_version} 다운로드 (배정 확인용)",
            data=create_checking_schedule_excel(
                # --- ▼▼▼ [핵심 버그 수정] ▼▼▼ ---
                # 'df_display_initial' (C) 대신 'df_schedule_original' (A)을 전달
                initial_df=st.session_state.df_schedule_original, 
                # --- ▲▲▲ [수정 완료] ▲▲▲ ---
                edited_df=edited_df, 
                edited_cumulative_df=edited_cumulative_df,
                df_special=st.session_state.df_special, 
                df_requests=st.session_state.df_requests,
                closing_dates=st.session_state.get("closing_dates", []), 
                month_str=month_str
            ),
            file_name=f"{month_str} 스케줄{display_version} (배정 확인용).xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="secondary",
            key="download_saved_checking"
        )

# [★ B. 수정 중이거나, 수정할 것이 없는 기본 상태 ★]
else:
    # [★ 3. 세션 상태 키 초기화 (if/else 분리로 인해 이 위치로 이동) ★]
    if "show_save_options" not in st.session_state:
        st.session_state.show_save_options = False
    
    # '저장 완료' 상태가 True였는데, 사용자가 다시 편집을 시작한 경우
    # (has_unsaved_changes=True가 됨), '저장 완료' UI를 숨기기 위해 플래그를 리셋
    if has_unsaved_changes and st.session_state.get("save_successful", False):
        st.session_state.save_successful = False

    def toggle_save_options():
        st.session_state.show_save_options = True
        st.session_state.df_to_save_temp = edited_df.copy()
        st.session_state.df_cum_to_save_temp = edited_cumulative_df.copy()

    col1_save, col2_save = st.columns(2)

    with col1_save:
        disable_save_button = st.session_state.get("disable_editing", False) or not has_unsaved_changes
        
        st.button(
            "💾 수정사항 Google Sheet에 저장", 
            type="primary", 
            use_container_width=True, 
            disabled=disable_save_button,
            on_click=toggle_save_options
        )

        if st.session_state.get("show_save_options", False):
            df_to_save = st.session_state.get("df_to_save_temp", edited_df.copy())
            df_cum_to_save = st.session_state.get("df_cum_to_save_temp", edited_cumulative_df.copy())

            st.warning("현재 버전 덮어쓰기를 선택하시면 이전 버전으로 돌아갈 수 없습니다.")
            
            numerical_versions = [v for v in versions.values() if v < 999.0]
            if not numerical_versions: latest_version_num = 0.0
            else: latest_version_num = max(numerical_versions)
            
            new_version_num = float(int(latest_version_num) + 1)
            new_sheet_name = f"{month_str} 스케줄 ver{new_version_num:.1f}"
            
            save_option = st.radio(
                "저장 옵션 선택",
                (f"현재 버전 - '{selected_sheet_name}' 덮어쓰기", f"다음 버전 - '{new_sheet_name}'으로 새로 저장하기"),
                key="save_option",
                label_visibility="collapsed"
            )

            if st.button("저장 실행", use_container_width=True, type="secondary"):
                sheet_name_to_save = selected_sheet_name if "덮어쓰기" in save_option else new_sheet_name
                
                st.session_state.show_save_options = False
                if "df_to_save_temp" in st.session_state:
                    del st.session_state.df_to_save_temp
                if "df_cum_to_save_temp" in st.session_state:
                    del st.session_state.df_cum_to_save_temp
                    
                save_schedule(month_str, sheet_name_to_save, df_to_save, df_cum_to_save)

        elif disable_save_button and not st.session_state.get("disable_editing", False):
            pass
        elif st.session_state.get("disable_editing", False):
            st.error("🚨 스케줄 최종본은 수정할 수 없습니다.")

    with col2_save:
        if has_unsaved_changes and not st.session_state.get("disable_editing", False):
            st.error("⚠️ 수정사항이 감지되었습니다. 먼저 '수정사항 Google Sheet에 저장' 버튼을 눌러주세요.")
        else:
            st.download_button(
                label=f"📥 스케줄{display_version} 다운로드",
                data=create_final_schedule_excel(
                    st.session_state.df_display_initial, edited_df, edited_cumulative_df,
                    st.session_state.df_special, st.session_state.df_requests,
                    st.session_state.get("closing_dates", []), month_str
                ),
                file_name=f"{month_str} 스케줄{display_version}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary",
                key="download_edited_final"
            )

            st.download_button(
                label=f"📥 스케줄{display_version} 다운로드 (배정 확인용)",
                data=create_checking_schedule_excel(
                    st.session_state.df_display_initial, edited_df, edited_cumulative_df,
                    st.session_state.df_special, st.session_state.df_requests,
                    st.session_state.get("closing_dates", []), month_str
                ),
                file_name=f"{month_str} 스케줄{display_version} (배정 확인용).xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="secondary",
                key="download_edited_checking"
            )