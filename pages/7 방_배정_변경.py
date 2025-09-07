import re
import streamlit as st
import pandas as pd
import gspread
from collections import Counter
from google.oauth2.service_account import Credentials
import time
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
import menu
import os
from dateutil.relativedelta import relativedelta

# --- 페이지 기본 설정 ---
st.set_page_config(page_title="방 배정 변경", page_icon="🔄", layout="wide")
st.session_state.current_page = os.path.basename(__file__)
menu.menu()

# --- 로그인 확인 ---
if not st.session_state.get("login_success", False):
    st.warning("⚠️ Home 페이지에서 먼저 로그인해주세요.")
    st.error("1초 후 Home 페이지로 돌아갑니다...")
    time.sleep(1)
    st.switch_page("Home.py")
    st.stop()

# --- 세션 상태 초기화 ---
if "change_data_loaded" not in st.session_state:
    st.session_state["change_data_loaded"] = False
if "saved_changes_log" not in st.session_state:
    st.session_state["saved_changes_log"] = []
if "df_final_assignment" not in st.session_state:
    st.session_state["df_final_assignment"] = pd.DataFrame()
if "df_change_requests" not in st.session_state:
    st.session_state["df_change_requests"] = pd.DataFrame()
if "changed_cells_log" not in st.session_state:
    st.session_state["changed_cells_log"] = []
if "df_before_apply" not in st.session_state:
    st.session_state["df_before_apply"] = pd.DataFrame()
if "has_changes_to_revert" not in st.session_state:
    st.session_state["has_changes_to_revert"] = False
if 'download_file' not in st.session_state:
    st.session_state.download_file = None
if 'download_filename' not in st.session_state:
    st.session_state.download_filename = None

# --- Google Sheets 연동 함수 ---
def get_gspread_client():
    scope = ["https://www.googleapis.com/auth/spreadsheets"]
    try:
        service_account_info = dict(st.secrets["gspread"])
        service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
        credentials = Credentials.from_service_account_info(service_account_info, scopes=scope)
        return gspread.authorize(credentials)
    except gspread.exceptions.APIError as e:
        st.warning("⚠️ 너무 많은 요청이 접속되어 딜레이되고 있습니다. 잠시 후 재시도 해주세요.")
        st.error(f"Google Sheets API 오류 (클라이언트 초기화): {e.response.status_code} - {e.response.text}")
        st.stop()
    except NameError as e:
        st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
        st.error(f"Google Sheets 인증 정보 로드 중 오류: {type(e).__name__} - {e}")
        st.stop()
    except Exception as e:
        st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
        st.error(f"Google Sheets 클라이언트 초기화 또는 인증 실패: {type(e).__name__} - {e}")
        st.stop()

def update_sheet_with_retry(worksheet, data, retries=5, delay=10):
    for attempt in range(retries):
        try:
            worksheet.clear()
            worksheet.update('A1', data, value_input_option='RAW')
            return
        except Exception as e:
            if "Quota exceeded" in str(e):
                st.warning(f"API 쿼터 초과, {delay}초 후 재시도 ({attempt+1}/{retries})")
                time.sleep(delay)
            else:
                st.error(f"업데이트 실패, {delay}초 후 재시도 ({attempt+1}/{retries}): {str(e)}")
                time.sleep(delay)
    st.error("Google Sheets 업데이트 실패: 재시도 횟수 초과")

# --- 데이터 로드 함수 ---
@st.cache_data(ttl=600, show_spinner=False)
def load_data_for_change_page(month_str):
    try:
        gc = get_gspread_client()
        sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
    except Exception as e:
        st.error(f"스프레드시트 열기 실패: {type(e).__name__} - {e}")
        return "STOP", None

    try:
        worksheet_final = sheet.worksheet(f"{month_str} 방배정")
        df_final = pd.DataFrame(worksheet_final.get_all_records())
        if df_final.empty:
            st.info("방배정이 아직 수행되지 않았습니다.")
            return "STOP", None
        df_final = df_final.fillna('')
    except gspread.exceptions.WorksheetNotFound:
        st.info("방배정이 아직 수행되지 않았습니다.")
        return "STOP", None
    except Exception as e:
        st.error(f"'{month_str} 방배정' 시트 로드 실패: {type(e).__name__} - {e}")
        return "STOP", None

    try:
        worksheet_req = sheet.worksheet(f"{month_str} 방배정 변경요청")
        df_req = pd.DataFrame(worksheet_req.get_all_records())
    except gspread.exceptions.WorksheetNotFound:
        st.warning(f"'{month_str} 방배정 변경요청' 시트가 없습니다. 새로운 시트로 생성하였습니다.")
        df_req = pd.DataFrame(columns=['RequestID', '요청일시', '요청자', '요청자 사번', '변경 요청', '변경 요청한 방배정'])
    except Exception as e:
        st.error(f"'{month_str} 방배정 변경요청' 시트 로드 실패: {type(e).__name__} - {e}")
        df_req = pd.DataFrame()

    return df_final, df_req

@st.cache_data(ttl=600, show_spinner=False)
def load_special_schedules(month_str):
    """
    'YYYY년 토요/휴일 스케줄' 시트에서 특정 월의 데이터를 로드합니다.
    연도는 month_str에서 동적으로 추출합니다.
    """
    try:
        gc = get_gspread_client()
        if not gc: return pd.DataFrame()
        
        spreadsheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
        
        # 1. month_str에서 연도를 동적으로 추출하여 시트 이름을 생성합니다.
        target_year = month_str.split('년')[0]
        sheet_name = f"{target_year}년 토요/휴일 스케줄"
        
        worksheet = spreadsheet.worksheet(sheet_name)
        records = worksheet.get_all_records()
        
        if not records:
            return pd.DataFrame()
        
        df = pd.DataFrame(records)

        # 2. '날짜'와 '근무' 열이 있는지 확인합니다.
        if '날짜' not in df.columns or '근무' not in df.columns:
            st.error(f"'{sheet_name}' 시트에 '날짜' 또는 '근무' 열이 없습니다.")
            return pd.DataFrame()

        df.fillna('', inplace=True)
        df['날짜_dt'] = pd.to_datetime(df['날짜'], format='%Y-%m-%d', errors='coerce')
        df.dropna(subset=['날짜_dt'], inplace=True)

        # 3. 'month_str'에 해당하는 월의 데이터만 필터링합니다.
        target_month_dt = datetime.strptime(month_str, "%Y년 %m월")
        df_filtered = df[
            (df['날짜_dt'].dt.year == target_month_dt.year) &
            (df['날짜_dt'].dt.month == target_month_dt.month)
        ].copy()

        return df_filtered
        
    except gspread.exceptions.WorksheetNotFound:
        target_year = month_str.split('년')[0]
        sheet_name = f"{target_year}년 토요/휴일 스케줄"
        st.info(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"토요/휴일 데이터 로드 중 오류 발생: {str(e)}")
        return pd.DataFrame()

# ❗️기존 apply_assignment_swaps 함수를 지우고 이 코드로 전체를 교체하세요.

def apply_assignment_swaps(df_assignment, df_requests, df_special):
    df_modified = df_assignment.copy()
    df_special_modified = df_special.copy() if df_special is not None else pd.DataFrame()
    changed_log = []
    applied_count = 0
    error_found = False

    for _, req in df_requests.iterrows():
        try:
            swap_request_str = str(req.get('변경 요청', '')).strip()
            raw_slot_info = str(req.get('변경 요청한 방배정', '')).strip()

            if '➡️' not in swap_request_str: continue
            old_person, new_person = [p.strip() for p in swap_request_str.split('➡️')]
            
            slot_match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', raw_slot_info)
            if not slot_match: continue
            
            date_str, target_slot = slot_match.groups()
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            target_date_str = f"{date_obj.month}월 {date_obj.day}일"
            
            # 날짜를 기준으로 방배정표에서 해당 행 찾기
            row_indices = df_modified.index[df_modified['날짜'] == target_date_str].tolist()
            if not row_indices:
                st.warning(f"⚠️ 요청 처리 불가: 방배정표에서 날짜 '{target_date_str}'를 찾을 수 없습니다.")
                time.sleep(1.5)
                continue
            target_row_idx = row_indices[0]

            # ✅ 수정된 로직: 평일과 휴일 구분 없이 동일한 방식으로 확인
            target_col_found = None
            # '날짜', '요일'을 제외한 모든 방(컬럼)을 순회
            for col in df_modified.columns[2:]: 
                person_in_cell = str(df_modified.at[target_row_idx, col]).strip()
                # 1. 해당 칸에 있는 사람이 바꾸려는 사람과 같고,
                # 2. 해당 칸의 이름(컬럼명)이 요청한 방(슬롯) 이름과 같으면
                if person_in_cell == old_person and col == target_slot:
                    target_col_found = col # 변경할 컬럼을 찾았으므로 저장
                    break # 반복 중단
            
            if target_col_found:
                df_modified.at[target_row_idx, target_col_found] = new_person
                applied_count += 1
                
                # 토요/휴일 당직자 변경 로직 (기존 유지)
                is_special_date = False
                if df_special is not None and not df_special.empty and '날짜_dt' in df_special.columns:
                    is_special_date = not df_special[df_special['날짜_dt'].dt.date == date_obj.date()].empty
                
                if is_special_date and not df_special_modified.empty:
                    duty_row = df_special_modified[df_special_modified['날짜_dt'].dt.date == date_obj.date()]
                    if not duty_row.empty:
                        current_duty_person = str(duty_row['당직'].iloc[0]).strip()
                        if current_duty_person == old_person:
                            df_special_modified.loc[duty_row.index, '당직'] = new_person
                            st.info(f"ℹ️ {target_date_str}의 토요/휴일 당직자가 '{new_person}' (으)로 함께 변경됩니다.")

                changed_log.append({
                    '날짜': f"{target_date_str} ({'월화수목금토일'[date_obj.weekday()]})",
                    '방배정': target_slot,
                    '변경 전 인원': old_person,
                    '변경 후 인원': new_person,
                    '변경 일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
            else:
                st.error(f"❌ 적용 실패: {target_date_str}의 '{target_slot}'에 '{old_person}'이(가) 배정되어 있지 않습니다.")
                time.sleep(1.5)
                error_found = True
                
        except Exception as e:
            st.error(f"⚠️ 요청 처리 중 시스템 오류 발생: {e}")
            time.sleep(1.5)
            error_found = True

    if applied_count > 0:
        st.success(f"🎉 총 {applied_count}건의 변경 요청이 반영되었습니다.")
        time.sleep(1.5)
    elif not df_requests.empty and not error_found:
        st.info("ℹ️ 새롭게 반영할 유효한 변경 요청이 없습니다.")
        time.sleep(1.5)

    return df_modified, changed_log, df_special_modified
    
# --- 시간대 순서 정의 ---
time_order = ['8:30', '9:00', '9:30', '10:00', '13:30']

# --- 시간대 순서 정의 ---
time_order = ['8:30', '9:00', '9:30', '10:00', '13:30']

# --- 통계 계산 함수 (수정됨) ---
def calculate_statistics(result_df: pd.DataFrame, df_special: pd.DataFrame) -> pd.DataFrame:
    total_stats = {
        'early': Counter(),
        'late': Counter(),
        'morning_duty': Counter(),
        'afternoon_duty': Counter(),
        'time_room_slots': {}  # 시간대-방 쌍 통계
    }
    
    # special_schedules 날짜를 제외하기 위해 날짜 목록 생성
    special_dates = []
    if df_special is not None and not df_special.empty and '날짜_dt' in df_special.columns:
        special_dates = df_special['날짜_dt'].dt.strftime('%#m월 %#d일').tolist() if os.name != 'nt' else df_special['날짜_dt'].dt.strftime('%m월 %d일').apply(lambda x: x.lstrip("0").replace(" 0", " "))
    
    # 모든 인원 목록 생성
    all_personnel_raw = pd.unique(result_df.iloc[:, 2:].values.ravel('K'))
    all_personnel_clean = {re.sub(r'\[\d+\]', '', str(p)).strip() for p in all_personnel_raw if pd.notna(p) and str(p).strip()}
    all_personnel = sorted(list(all_personnel_clean))
    
    SMALL_TEAM_THRESHOLD = 13
    
    # 슬롯별 통계 초기화
    for slot_name in result_df.columns[2:]:
        if slot_name != '온콜':  # '온콜' 제외
            total_stats['time_room_slots'].setdefault(slot_name, Counter())
    
    for _, row in result_df.iterrows():
        date_str = str(row.get('날짜', '')).strip()
        
        # 토요/휴일 날짜는 통계에서 제외
        if date_str in special_dates:
            continue
            
        personnel_in_row = [p for p in row.iloc[2:].dropna() if p]
        if 0 < len(personnel_in_row) < SMALL_TEAM_THRESHOLD:
            continue
        
        for slot_name in result_df.columns[2:]:
            person = row.get(slot_name)
            if not person or pd.isna(person):
                continue
            
            person_clean = re.sub(r'\[\d+\]', '', str(person)).strip()
            
            # 시간대-방 쌍 통계 ('온콜' 제외)
            if slot_name != '온콜':
                total_stats['time_room_slots'][slot_name][person_clean] += 1
            
            # 기존 통계
            if slot_name.startswith('8:30') and not slot_name.endswith('_당직'):
                total_stats['early'][person_clean] += 1
            elif slot_name.startswith('10:00'):
                total_stats['late'][person_clean] += 1
            if slot_name == '온콜' or (slot_name.startswith('8:30') and slot_name.endswith('_당직')):
                total_stats['morning_duty'][person_clean] += 1
            elif slot_name.startswith('13:30') and slot_name.endswith('_당직'):
                total_stats['afternoon_duty'][person_clean] += 1
    
    # 통계 DataFrame 생성
    stats_data = []
    for p in all_personnel:
        stats_entry = {
            '인원': p,
            '이른방 합계': total_stats['early'][p],
            '늦은방 합계': total_stats['late'][p],
            '오전 당직 합계': total_stats['morning_duty'][p],
            '오후 당직 합계': total_stats['afternoon_duty'][p],
        }
        # 시간대(방) 합계 추가 (당직 제외)
        for slot in total_stats['time_room_slots']:
            if not slot.endswith('_당직'):
                stats_entry[f'{slot} 합계'] = total_stats['time_room_slots'][slot][p]
        stats_data.append(stats_entry)
    
    # 컬럼 정렬
    sorted_columns = ['인원', '이른방 합계', '늦은방 합계', '오전 당직 합계', '오후 당직 합계']
    time_slots = sorted(
        [slot for slot in total_stats['time_room_slots'].keys() if not slot.endswith('_당직')],
        key=lambda x: (
            time_order.index(x.split('(')[0]),  # 시간대 순서
            int(x.split('(')[1].split(')')[0])  # 방 번호 순서
        )
    )
    sorted_columns.extend([f'{slot} 합계' for slot in time_slots])
    
    return pd.DataFrame(stats_data)[sorted_columns]

# --- UI 및 데이터 핸들링 ---
from zoneinfo import ZoneInfo
kst = ZoneInfo("Asia/Seoul")
now = datetime.now(kst)
today = now.date()
next_month_date = today.replace(day=1) + relativedelta(months=1)
month_str = next_month_date.strftime("%Y년 %-m월")
st.header(f"🔄 {month_str} 방 배정 변경", divider='rainbow')

# 데이터 로드 및 새로고침 로직 통합
def load_and_initialize_data():
    with st.spinner("데이터를 로드하고 있습니다..."):
        df_final, df_req = load_data_for_change_page(month_str)
    
    if isinstance(df_final, str) and df_final == "STOP":
        st.stop()
        
    df_special = load_special_schedules(month_str)
    
    st.session_state.df_final_assignment = df_final
    st.session_state.df_change_requests = df_req
    st.session_state.df_special_schedules = df_special
    st.session_state.changed_cells_log = []
    st.session_state.df_before_apply = df_final.copy()
    st.session_state.has_changes_to_revert = False
    st.session_state.change_data_loaded = True

# 새로고침 버튼
st.write("- 먼저 새로고침 버튼으로 최신 데이터를 불러온 뒤, 배정을 진행해주세요.")
if st.button("🔄 새로고침 (R)"):
    st.cache_data.clear()
    st.session_state.change_data_loaded = False
    st.rerun()

# 초기 데이터 로드
if not st.session_state.change_data_loaded:
    load_and_initialize_data()

st.divider()

st.subheader("📋 방배정 변경 요청 목록")
if not st.session_state.df_change_requests.empty:
    df_display = st.session_state.df_change_requests.copy()
    def convert_date_format(x):
        x = str(x).strip()
        match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', x)
        if match:
            date_str, slot = match.groups()
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                return f"{date_obj.strftime('%-m월 %-d일')} ({'월화수목금토일'[date_obj.weekday()]}) - {slot}"
            except ValueError:
                st.warning(f"⚠️ 잘못된 날짜 형식: '{date_str}'")
                return x
        return x
    df_display['변경 요청한 방배정'] = df_display['변경 요청한 방배정'].apply(convert_date_format)
    if 'RequestID' in df_display.columns:
        df_display = df_display.drop(columns=['RequestID'])
    if '요청자 사번' in df_display.columns:
        df_display = df_display.drop(columns=['요청자 사번'])
    st.dataframe(df_display, use_container_width=True, hide_index=True)
else:
    st.info("접수된 변경 요청이 없습니다.")
st.divider()

# --- UI 및 데이터 핸들링 (수정된 부분) ---
st.subheader("✍️ 방배정 최종 수정")
st.write("- 요청사항을 일괄 적용하거나, 셀을 더블클릭하여 직접 수정한 후 **최종 저장**하세요.")
col1, col2 = st.columns(2)
with col1:
    if st.button("🔄 요청사항 일괄 적용"):
        if not st.session_state.df_change_requests.empty:
            current_df = st.session_state.df_final_assignment
            requests_df = st.session_state.df_change_requests
            special_df = st.session_state.df_special_schedules
            st.session_state.df_before_apply = current_df.copy()
            modified_df, new_changes, modified_special_df = apply_assignment_swaps(current_df, requests_df, special_df)
            st.session_state.df_final_assignment = modified_df
            st.session_state.df_special_schedules = modified_special_df
            if not isinstance(st.session_state.changed_cells_log, list):
                st.session_state.changed_cells_log = []
            # 기존 로그에 새 로그 추가 (중복 제거)
            existing_keys = {(log['날짜'], log['방배정']) for log in st.session_state.changed_cells_log}
            for change in new_changes:
                if (change['날짜'], change['방배정']) not in existing_keys:
                    st.session_state.changed_cells_log.append(change)
                    existing_keys.add((change['날짜'], change['방배정']))
            st.session_state.has_changes_to_revert = True
            st.rerun()
        else:
            st.info("ℹ️ 처리할 변경 요청이 없습니다.")
with col2:
    if st.button("⏪ 적용 취소", disabled=not st.session_state.has_changes_to_revert):
        st.session_state.df_final_assignment = st.session_state.df_before_apply.copy()
        st.session_state.changed_cells_log = []
        st.session_state.has_changes_to_revert = False
        st.info("변경사항이 취소되고 원본 스케줄로 돌아갑니다.")
        time.sleep(1.5)
        st.rerun()

# DataFrame 편집
edited_df = st.data_editor(
    st.session_state.df_final_assignment,
    use_container_width=True,
    key="assignment_editor",
    disabled=['날짜', '요일'],
    hide_index=True
)

# 수동 편집 시 변경사항 감지 및 로그 업데이트
if not edited_df.equals(st.session_state.df_final_assignment):
    st.session_state.df_before_apply = st.session_state.df_final_assignment.copy()
    diff_mask = (edited_df != st.session_state.df_final_assignment) & (edited_df.notna() | st.session_state.df_final_assignment.notna())
    current_log = st.session_state.changed_cells_log if isinstance(st.session_state.changed_cells_log, list) else []
    
    # 새로운 변경사항 기록
    newly_changed_logs = []
    existing_keys = {(log['날짜'], log['방배정']) for log in current_log}
    
    for col in diff_mask.columns:
        if diff_mask[col].any():
            for idx in diff_mask.index[diff_mask[col]]:
                date_val = edited_df.at[idx, '날짜']
                day_val = edited_df.at[idx, '요일']
                formatted_date = f"{date_val} ({day_val})"
                
                new_val = str(edited_df.at[idx, col]).strip() if pd.notna(edited_df.at[idx, col]) else ""
                old_val = str(st.session_state.df_final_assignment.at[idx, col]).strip() if pd.notna(st.session_state.df_final_assignment.at[idx, col]) else ""
                
                log_key = (formatted_date, col)
                if log_key not in existing_keys and new_val != old_val:
                    newly_changed_logs.append({
                        '날짜': formatted_date,
                        '방배정': col,
                        '변경 전 인원': old_val,
                        '변경 후 인원': new_val,
                        # '변경 유형': '수동 편집',
                        # '변경 일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                    existing_keys.add(log_key)
    
    st.session_state.changed_cells_log = current_log + newly_changed_logs
    st.session_state.df_final_assignment = edited_df.copy()
    st.session_state.has_changes_to_revert = True

st.divider()
st.caption("📝 현재까지 기록된 변경사항 로그")
if st.session_state.changed_cells_log:
    valid_logs = [log for log in st.session_state.changed_cells_log if len(log) >= 4]
    if valid_logs:
        log_df = pd.DataFrame(valid_logs)
        log_df = log_df[['날짜', '방배정', '변경 전 인원', '변경 후 인원', '변경 일시']].fillna('')
        st.dataframe(log_df.sort_values(by=['변경 일시', '날짜', '방배정']).reset_index(drop=True), use_container_width=True, hide_index=True)
    else:
        st.info("기록된 변경사항이 없습니다.")
else:
    st.info("기록된 변경사항이 없습니다.")

# 변경사항 유무를 판단하는 플래그
has_unsaved_changes = (st.session_state.changed_cells_log is not None and len(st.session_state.changed_cells_log) > 0)

col_final1, col_final2 = st.columns(2)
with col_final1:
    if st.button("✍️ 변경사항 저장", type="primary", use_container_width=True, disabled=not has_unsaved_changes):
        final_df_to_save = st.session_state.df_final_assignment
        try:
            with st.spinner("Google Sheets에 저장 중..."):
                gc = get_gspread_client()
                sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])

                # --- '방배정 최종' 시트 저장 ---
                try:
                    worksheet_final = sheet.worksheet(f"{month_str} 방배정 최종")
                except gspread.exceptions.WorksheetNotFound:
                    st.info(f"'{month_str} 방배정 최종' 시트를 새로 생성합니다.")
                    worksheet_final = sheet.add_worksheet(title=f"{month_str} 방배정 최종", rows=100, cols=len(final_df_to_save.columns))

                final_data_list = [final_df_to_save.columns.tolist()] + final_df_to_save.fillna('').values.tolist()
                update_sheet_with_retry(worksheet_final, final_data_list)

            # 로그 처리 및 페이지 상태 업데이트
            st.session_state.saved_changes_log.extend(st.session_state.changed_cells_log)
            st.session_state.changed_cells_log = []
            st.session_state.has_changes_to_revert = False
            
            st.success("✅ Google Sheets에 최종 방배정표가 성공적으로 저장되었습니다.")
            time.sleep(2)
            st.rerun()

        except Exception as e:
            st.error(f"⚠️ 저장 중 오류 발생: {e}")
            
with col_final2:
    if has_unsaved_changes:
        st.warning("⚠️ 변경사항이 있습니다. 먼저 **'변경사항 저장'** 버튼을 눌러주세요.")
    
    if st.button("🚀 방배정 수행 및 결과 보기", type="primary", use_container_width=True, disabled=has_unsaved_changes):
        with st.spinner("방 배정 중..."):
            time.sleep(3)
            st.session_state['show_final_results'] = True
    else:
        # 이 버튼을 누르지 않은 경우 show_final_results 상태를 유지하거나 False로 설정
        if 'show_final_results' not in st.session_state:
            st.session_state['show_final_results'] = False

# 이 코드는 기존의 `if st.session_state.get('show_final_results', ...):` 블록 전체를 대체합니다.
if st.session_state.get('show_final_results', False) and not has_unsaved_changes:
    st.divider()
    final_df_to_save = st.session_state.df_final_assignment
    special_df_to_update = st.session_state.df_special_schedules
    st.subheader(f"💡 {month_str} 최종 방배정 결과", divider='rainbow')
    
    if special_df_to_update is not None and not special_df_to_update.empty:
        try:
            st.info("ℹ️ 토요/휴일 스케줄의 변경된 근무 정보를 동기화합니다...")
            gc = get_gspread_client()
            sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
            
            # 1. 연도와 시트 이름 설정
            target_year = month_str.split('년')[0]
            sheet_name = f"{target_year}년 토요/휴일 스케줄"
            worksheet_special_yearly = sheet.worksheet(sheet_name)
            
            # 2. 연간 시트 전체 데이터 로드
            all_yearly_data = pd.DataFrame(worksheet_special_yearly.get_all_records())
            
            # 3. 이번 달 수정 데이터와 연간 데이터 병합
            special_df_to_update['날짜'] = pd.to_datetime(special_df_to_update['날짜_dt']).dt.strftime('%Y-%m-%d')
            update_df = special_df_to_update[['날짜', '근무', '당직']]
            
            merged_df = pd.merge(all_yearly_data, update_df, on='날짜', how='left', suffixes=('', '_new'))
            merged_df['근무'] = merged_df['근무_new'].fillna(merged_df['근무'])
            merged_df['당직'] = merged_df['당직_new'].fillna(merged_df['당직'])
            
            final_yearly_df = merged_df[all_yearly_data.columns]
            
            # 4. 수정된 전체 연간 데이터로 시트 업데이트
            special_data_list = [final_yearly_df.columns.tolist()] + final_yearly_df.fillna('').values.tolist()
            update_sheet_with_retry(worksheet_special_yearly, special_data_list)
            st.success(f"✅ '{sheet_name}' 시트가 성공적으로 동기화되었습니다.")
                
        except gspread.exceptions.WorksheetNotFound:
            st.warning(f"'{sheet_name}' 시트가 없어 업데이트를 생략합니다.")
        except Exception as e:
            st.error(f"⚠️ 토요/휴일 스케줄 동기화 중 오류 발생: {e}")
    
    st.write(" ")
    st.markdown("**✅ 통합 배치 결과**")
    st.dataframe(final_df_to_save, use_container_width=True, hide_index=True)
    
    with st.spinner("통계 정보를 계산 중입니다..."):
        stats_df = calculate_statistics(final_df_to_save, st.session_state.df_special_schedules)
    
    st.markdown("**☑️ 인원별 통계**")
    st.dataframe(stats_df, use_container_width=True, hide_index=True)
    
    with st.spinner("Excel 파일을 생성 중입니다..."):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Schedule"

        import platform

        if platform.system() == "Windows":
            font_name = "맑은 고딕"
        else:
            font_name = "Arial"

        highlight_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        duty_font = Font(name=font_name, size=9, bold=True, color="FF00FF")
        default_font = Font(name=font_name, size=9)
        
        holiday_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # 파란색 계열

        columns = final_df_to_save.columns.tolist()
        for col_idx, header in enumerate(columns, 1):
            cell = sheet.cell(1, col_idx, header)
            cell.font = Font(bold=True, name=font_name, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            if header.startswith('8:30') or header == '온콜':
                cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            elif header.startswith('9:00'):
                cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            elif header.startswith('9:30'):
                cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
            elif header.startswith('10:00'):
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            elif header.startswith('13:30'):
                cell.fill = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")
        
        all_logs = st.session_state.saved_changes_log + st.session_state.changed_cells_log
        changed_cells_set = set()
        for log in all_logs:
            if len(log) < 4: continue
            
            date_str = log['날짜']
            slot_name_raw = log['방배정']
            
            try:
                date_without_week = date_str.split(' (')[0]
                
                if date_without_week in final_df_to_save['날짜'].values:
                    df_row_index = final_df_to_save.index[final_df_to_save['날짜'] == date_without_week].tolist()[0]
                    excel_row_idx = df_row_index + 2
                    
                    target_col_idx = -1
                    
                    if "번방" in slot_name_raw:
                        room_num_match = re.search(r'(\d+)', slot_name_raw)
                        if room_num_match:
                            requested_room_num = room_num_match.group(1)
                            for col_idx, col_name in enumerate(columns):
                                if not col_name.startswith('13:30'):
                                    col_room_match = re.search(r'\((\d+)\)', col_name)
                                    if col_room_match and col_room_match.group(1) == requested_room_num:
                                        target_col_idx = col_idx + 1
                                        break
                    else:
                        if slot_name_raw in columns:
                            target_col_idx = columns.index(slot_name_raw) + 1

                    if target_col_idx != -1:
                        changed_cells_set.add((excel_row_idx, target_col_idx))

            except (ValueError, IndexError) as e:
                st.warning(f"⚠️ 로그 처리 중 오류 (무시됨): {e} - 로그: {log}")
                continue

        special_dates_list = []
        if st.session_state.df_special_schedules is not None and not st.session_state.df_special_schedules.empty:
            try:
                # [수정] Windows에서도 0을 제거하는 가장 안정적인 방식
                special_dates_list = [d.strftime('%-m월 %-d일') for d in st.session_state.df_special_schedules['날짜_dt']]
            except ValueError:
                # Windows에서 '%-m'이 작동하지 않을 경우를 대비한 예외 처리
                temp_dates = st.session_state.df_special_schedules['날짜_dt'].dt.strftime('%m월 %d일').tolist()
                special_dates_list = [re.sub(r'^0|(?<=\s)0', '', d) for d in temp_dates]

        # 데이터 렌더링
        for row_idx, row_data in enumerate(final_df_to_save.itertuples(index=False), 2):            
            current_date_str = row_data[0]
            is_special_day = current_date_str in special_dates_list
            special_df = st.session_state.df_special_schedules
            
            # 1. 그날의 당직 인원 정보를 정확히 가져옵니다.
            duty_person_for_the_day = None
            if current_date_str in special_dates_list:
                try:
                    date_obj_lookup = datetime.strptime(current_date_str, '%m월 %d일').replace(year=datetime.now().year)
                    formatted_date_lookup = date_obj_lookup.strftime('%Y-%m-%d')
                    duty_person_row = special_df[special_df['날짜'] == formatted_date_lookup]
                    if not duty_person_row.empty:
                        duty_person_raw = duty_person_row['당직'].iloc[0]
                        if pd.notna(duty_person_raw) and str(duty_person_raw).strip() and str(duty_person_raw).strip() != '당직 없음':
                            duty_person_for_the_day = str(duty_person_raw).strip()
                except Exception as e:
                    st.warning(f"Excel 스타일링 중 당직 인원 조회 오류: {e}")

            personnel_in_row = [p for p in row_data[2:] if p]
            is_no_person_day = not any(personnel_in_row)
            is_small_team_day = (0 < len(personnel_in_row) < 15)

            # 2. 셀마다 스타일을 적용합니다.
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row_idx, col_idx, value if value else None)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                # --- 배경색 먼저 적용 ---
                if col_idx == 1: # 날짜
                    cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                elif col_idx == 2: # 요일
                    if is_no_person_day: cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                    elif is_small_team_day or is_special_day: cell.fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
                    else: cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif is_no_person_day and col_idx >= 3:
                    cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                
                # [추가] 토요/휴일 근무자 셀 배경색을 파란색으로 지정
                if is_special_day and value and col_idx > 2:
                    cell.fill = holiday_blue_fill

                # 변경된 셀은 다른 색으로 덮어쓰기 (이 코드는 원래 위치에 그대로 둡니다)
                if (row_idx, col_idx) in changed_cells_set:
                    cell.fill = highlight_fill
                
                # --- 폰트 나중에 적용 (덮어쓰기 방지) ---
                cell.font = default_font # 기본 폰트 먼저 적용

                if value: # 셀에 값이 있을 때만 폰트 변경 고려
                    slot_name = columns[col_idx-1]
                    
                    # 👇 is_special_day가 True일 때 (토요/휴일일 때)
                    if is_special_day:
                        # (2) 당직 인원 이름과 셀의 이름이 같을 때만 duty_font (핑크 볼드체) 적용
                        if duty_person_for_the_day and value == duty_person_for_the_day:
                            cell.font = duty_font
                        # (3) 위 조건이 아니면 그냥 기본 폰트. slot_name.endswith('_당직')은 체크하지 않음!
                            
                    # 👇 is_special_day가 False일 때 (평일일 때)
                    else:
                        # 평일: 슬롯 이름('_당직')으로 당직자를 판단하여 핑크색 볼드체
                        if slot_name.endswith('_당직') or slot_name == '온콜':
                            cell.font = duty_font
        
        # (이하 통계 시트 작성 코드는 동일)
        stats_sheet = wb.create_sheet("Stats")
        stats_columns = stats_df.columns.tolist()
        for col_idx, header in enumerate(stats_columns, 1):
            stats_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
            cell = stats_sheet.cell(1, col_idx, header)
            cell.font = Font(bold=True, name=font_name, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            if header == '인원':
                cell.fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
            elif header == '이른방 합계':
                cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            elif header == '늦은방 합계':
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            elif '당직' in header:
                cell.fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row_idx, row in enumerate(stats_df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = stats_sheet.cell(row_idx, col_idx, value)
                cell.font = Font(name=font_name, size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        st.session_state.download_file = output
        st.session_state.download_filename = f"{month_str} 방배정_최종확정.xlsx"

if 'download_file' in st.session_state and st.session_state.download_file:
    st.download_button(
        label="📥 최종 확정본 다운로드",
        data=st.session_state.download_file,
        file_name=st.session_state.download_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )
    st.session_state.download_file = None
