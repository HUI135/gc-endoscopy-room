import streamlit as st
import pandas as pd
import calendar
from io import BytesIO
from dateutil.relativedelta import relativedelta
from google.oauth2.service_account import Credentials
import gspread
from gspread.exceptions import WorksheetNotFound, APIError
import time
import io
import xlsxwriter
import platform
import openpyxl
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.comments import Comment
from datetime import datetime, timedelta
from collections import Counter
import menu
import re

st.set_page_config(page_title="스케줄 배정", page_icon="🗓️", layout="wide")

import os
st.session_state.current_page = os.path.basename(__file__)

menu.menu()

def initialize_schedule_session_state():
    """스케줄 배정 페이지에서 사용할 모든 세션 상태 키를 초기화합니다."""
    keys_to_init = {
        "assigned": False,
        "output": None,
        "df_cumulative_next": pd.DataFrame(),
        "request_logs": [],
        "swap_logs": [],
        "adjustment_logs": [],
        "oncall_logs": [],
        "assignment_results": None,
        "show_confirmation_warning": False,
        "latest_existing_version": None,
        "editor_has_changes": False,
        # ▼▼▼ [핵심 수정] data_editor를 강제 리셋하기 위한 키 버전 ▼▼▼
        "editor_key_version": 0
    }
    for key, value in keys_to_init.items():
        if key not in st.session_state:
            st.session_state[key] = value

def set_editor_changed_flag():
    """data_editor에서 수정이 발생했음을 세션 상태에 기록합니다."""
    st.session_state.editor_has_changes = True

def get_sort_key(log_string):
    # '10월 1일'과 같은 패턴을 찾습니다.
    match = re.search(r'(\d{1,2}월 \d{1,2}일)', log_string)
    if match:
        date_str = match.group(1)
        try:
            # month_dt 변수에서 연도를 가져와 완전한 날짜 객체로 만듭니다.
            return datetime.strptime(f"{month_dt.year}년 {date_str}", "%Y년 %m월 %d일")
        except ValueError:
            # 날짜 변환에 실패하면 정렬 순서에 영향을 주지 않도록 맨 뒤로 보냅니다.
            return datetime.max
    # 로그에서 날짜를 찾지 못하면 맨 뒤로 보냅니다.
    return datetime.max

# 로그인 체크 및 자동 리디렉션
if not st.session_state.get("login_success", False):
    st.warning("⚠️ Home 페이지에서 먼저 로그인해주세요.")
    st.error("1초 후 Home 페이지로 돌아갑니다...")
    time.sleep(1)
    st.switch_page("Home.py")  # Home 페이지로 이동
    st.stop()

# 초기 데이터 로드 및 세션 상태 설정
url = st.secrets["google_sheet"]["url"]

from zoneinfo import ZoneInfo
kst = ZoneInfo("Asia/Seoul")
now = datetime.now(kst)
today = now.date()
month_dt = today.replace(day=1) + relativedelta(months=1)
month_str = month_dt.strftime("%Y년 %-m월")
month_str = "2025년 10월"
_, last_day = calendar.monthrange(month_dt.year, month_dt.month)
month_start = month_dt
month_end = month_dt.replace(day=last_day)

# ▼▼▼ [추가] month_str을 기준으로 지난달 생성 ▼▼▼
# 1. month_str을 datetime 객체로 변환
current_target_dt = datetime.strptime(month_str, "%Y년 %m월")

# 2. 한 달을 빼서 '지난달' datetime 객체를 만듦
prev_month_dt = current_target_dt - relativedelta(months=1)

# 3. '지난달'을 month_str과 동일한 형식의 문자열로 만듦
prev_month_str = prev_month_dt.strftime("%Y년 %-m월")

# Google Sheets 클라이언트 초기화
@st.cache_resource
def get_gspread_client():
    scope = ["https://www.googleapis.com/auth/spreadsheets"]
    try:
        service_account_info = dict(st.secrets["gspread"])
        service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
        credentials = Credentials.from_service_account_info(service_account_info, scopes=scope)
        gc = gspread.authorize(credentials)
        return gc
    except gspread.exceptions.APIError as e:
        st.warning("⚠️ 너무 많은 요청이 접속되어 딜레이되고 있습니다. 잠시 후 재시도 해주세요.")
        st.error(f"Google Sheets API 오류 (클라이언트 초기화): {e.response.status_code} - {e.response.text}")
        st.stop()
    except NameError as e:
        st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
        st.error(f"Google Sheets 인증 정보 로드 중 오류: {type(e).__name__} - {e}")
        st.stop()
    except Exception as e:
        st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
        st.error(f"Google Sheets 클라이언트 초기화 또는 인증 실패: {type(e).__name__} - {e}")
        st.stop()

# Google Sheets 업데이트 함수
def update_sheet_with_retry(worksheet, data, retries=3, delay=5):
    for attempt in range(retries):
        try:
            worksheet.clear()  # 시트를 완전히 비우고 새 데이터로 덮어씌움
            worksheet.update(data, "A1")
            return True
        except gspread.exceptions.APIError as e:
            if attempt < retries - 1:
                st.warning(f"⚠️ API 요청이 지연되고 있습니다. {delay}초 후 재시도합니다... ({attempt+1}/{retries})")
                time.sleep(delay)
                delay *= 2  # 지수 백오프
            else:
                st.warning("⚠️ 너무 많은 요청이 접속되어 딜레이되고 있습니다. 잠시 후 재시도 해주세요.")
                st.error(f"Google Sheets API 오류 (시트 업데이트): {str(e)}")
                st.stop()
        except Exception as e:
            if attempt < retries - 1:
                st.warning(f"⚠️ 업데이트 실패, {delay}초 후 재시도 ({attempt+1}/{retries}): {str(e)}")
                time.sleep(delay)
                delay *= 2
            else:
                st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
                st.error(f"Google Sheets 업데이트 실패: {str(e)}")
                st.stop()
    return False

# 'find_latest_schedule_version' 함수 (수정 필요)

def find_latest_schedule_version(sheet, month_str):
    """주어진 월에 해당하는 스케줄 시트 중 가장 최신 버전을 찾습니다. '최종'이 최우선입니다."""
    versions = {}
    
    # 1. '최종' 시트 존재 여부 확인 (가장 높은 우선순위)
    final_version_name = f"{month_str} 스케줄 최종"
    for ws in sheet.worksheets():
        if ws.title == final_version_name:
            return final_version_name
    
    # 2. 'ver X.X' 및 기본 버전 찾기 (기존 로직 유지)
    # 'ver 1.0', 'ver1.0' 등 다양한 형식을 모두 찾도록 정규식 수정
    pattern = re.compile(f"^{re.escape(month_str)} 스케줄(?: ver\s*(\d+\.\d+))?$")

    for ws in sheet.worksheets():
        match = pattern.match(ws.title)
        if match:
            version_num_str = match.group(1) # ver 뒤의 숫자 부분 (예: '1.0')
            # 버전 넘버가 있으면 float으로 변환, 없으면 (기본 시트면) 1.0으로 처리
            version_num = float(version_num_str) if version_num_str else 1.0
            versions[ws.title] = version_num

    if not versions:
        return None

    # 가장 높은 버전 번호를 가진 시트의 이름을 반환
    return max(versions, key=versions.get)

def find_latest_cumulative_version(sheet, month_str):
    """
    [★수정됨★]
    주어진 월에 해당하는 누적 시트 중 가장 최신 버전을 찾습니다.
    '최종' 버전을 최우선으로 간주합니다. (공백 차이 무시)
    """
    versions = {}
    
    # 1. '최종' 시트가 있는지 먼저 확인 (공백(s+)을 허용하는 정규식 사용)
    # 예: "2025년 10월 누적 최종", "2025년 10월  누적  최종" 둘 다 찾음
    final_pattern = re.compile(f"^{re.escape(month_str)}\s+누적\s+최종$")
    final_version_name = None

    for ws in sheet.worksheets():
        if final_pattern.match(ws.title.strip()): # .strip() 추가로 앞뒤 공백 제거
            final_version_name = ws.title # '최종' 버전을 찾으면 즉시 반환
            return final_version_name 
    
    # 2. '최종'이 없으면 'ver X.X' 및 기본 버전('누적')을 찾음
    pattern = re.compile(f"^{re.escape(month_str)} 누적(?: ver\s*(\d+\.\d+))?$")

    for ws in sheet.worksheets():
        match = pattern.match(ws.title)
        if match:
            version_num_str = match.group(1) # ver 뒤의 숫자 부분 (예: '1.0')
            version_num = float(version_num_str) if version_num_str else 1.0
            versions[ws.title] = version_num

    if not versions:
        return None # 어떠한 버전의 시트도 찾지 못하면 None 반환

    return max(versions, key=versions.get)

@st.cache_data(ttl=600, show_spinner="최신 데이터를 구글 시트에서 불러오는 중...")
def load_data_page5():
    url = st.secrets["google_sheet"]["url"]
    try:
        gc = get_gspread_client()
        if gc is None: st.stop()
        sheet = gc.open_by_url(url)
    except Exception as e:
        st.error(f"스프레드시트 열기 실패: {e}"); st.stop()

    # --- 마스터 시트 로드 ---
    try:
        ws1 = sheet.worksheet("마스터")
        df_master = pd.DataFrame(ws1.get_all_records())
        master_names_list = df_master["이름"].unique().tolist()
    except WorksheetNotFound:
        st.error("❌ '마스터' 시트를 찾을 수 없습니다."); st.stop()
    except Exception as e:
        st.error(f"'마스터' 시트 로드 실패: {e}"); st.stop()

    # --- 요청사항 시트 로드 ---
    try:
        ws2 = sheet.worksheet(f"{month_str} 요청")
        df_request = pd.DataFrame(ws2.get_all_records())
    except WorksheetNotFound:
        st.warning(f"⚠️ '{month_str} 요청' 시트를 찾을 수 없어 새로 생성합니다.")
        ws2 = sheet.add_worksheet(title=f"{month_str} 요청", rows=100, cols=3)
        ws2.append_row(["이름", "분류", "날짜정보"])
        df_request = pd.DataFrame(columns=["이름", "분류", "날짜정보"])
    except Exception as e:
        st.error(f"'요청' 시트 로드 실패: {e}"); st.stop()

    # --- [핵심 수정] 누적 시트 로드 로직을 단순하고 명확하게 변경 ---
    df_cumulative = pd.DataFrame()
    worksheet_to_load = None

    # --- [핵심 수정] 누적 시트 로드 로직을 단순하고 명확하게 변경 ---
    df_cumulative = pd.DataFrame()
    worksheet_to_load = None

    # 1. month_str에 해당하는 가장 최신 버전('최종' 우선)의 누적 시트 이름을 찾습니다.
    latest_cum_version_name = find_latest_cumulative_version(sheet, month_str)

    if latest_cum_version_name:
        try:
            worksheet_to_load = sheet.worksheet(latest_cum_version_name)
            # ▼▼▼ [핵심 수정] 불러올 시트 이름을 세션에 저장합니다. ▼▼▼
            st.session_state["target_cumulative_sheet_name"] = latest_cum_version_name
            # ▲▲▲ [수정 완료] ▲▲▲
        except WorksheetNotFound:
            # 시트 이름은 찾았으나 gspread에서 못 여는 예외적인 경우
            st.warning(f"⚠️ '{latest_cum_version_name}' 시트를 찾았지만 열 수 없습니다. 빈 테이블로 시작합니다.")
            # ▼▼▼ [핵심 수정] 이 경우 기본 시트 이름을 저장합니다. ▼▼▼
            st.session_state["target_cumulative_sheet_name"] = f"{month_str} 누적"
            # ▲▲▲ [수정 완료] ▲▲▲
    else:
        # month_str에 해당하는 누적 시트가 아예 없는 경우
        st.warning(f"⚠️ '{month_str} 누적' 시트를 찾을 수 없어, 빈 누적 테이블로 시작합니다.")
        # ▼▼▼ [핵심 수정] 새로 생성할 기본 시트 이름을 세션에 저장합니다. ▼▼▼
        st.session_state["target_cumulative_sheet_name"] = f"{month_str} 누적"
        # ▲▲▲ [수정 완료] ▲▲▲

    # 2. 찾은 시트에서 데이터 로드
    if worksheet_to_load:
        all_values = worksheet_to_load.get_all_values()
        if all_values and len(all_values) > 1:
            headers = all_values[0]
            data = [row for row in all_values[1:] if any(cell.strip() for cell in row)]
            df_cumulative = pd.DataFrame(data, columns=headers)
        else:
            st.warning(f"'{worksheet_to_load.title}' 시트가 비어있어, 빈 테이블로 시작합니다.")

    # 누적 시트가 비었거나 '항목' 열이 없으면 기본값으로 생성
    if df_cumulative.empty or '항목' not in df_cumulative.columns:
        default_cols = ["항목"] + master_names_list
        default_data = [
            ["오전누적"] + [0] * len(master_names_list), ["오후누적"] + [0] * len(master_names_list),
            ["오전당직누적"] + [0] * len(master_names_list), ["오후당직누적"] + [0] * len(master_names_list)
        ]
        df_cumulative = pd.DataFrame(default_data, columns=default_cols)

    # 숫자 열 변환
    for col in df_cumulative.columns:
        if col != '항목':
            df_cumulative[col] = pd.to_numeric(df_cumulative[col], errors='coerce').fillna(0).astype(int)

    # --- 근무/보충 테이블 생성 ---
    df_shift = generate_shift_table(df_master)
    df_supplement = generate_supplement_table(df_shift, master_names_list)
    
    return df_master, df_request, df_cumulative, df_shift, df_supplement

def generate_shift_table(df_master):
    def split_shift(row):
        shifts = []
        if row["근무여부"] == "오전 & 오후":
            shifts.extend([(row["이름"], row["주차"], row["요일"], "오전"), (row["이름"], row["주차"], row["요일"], "오후")])
        elif row["근무여부"] in ["오전", "오후"]:
            shifts.append((row["이름"], row["주차"], row["요일"], row["근무여부"]))
        return shifts

    shift_list = [shift for _, row in df_master.iterrows() for shift in split_shift(row)]
    df_split = pd.DataFrame(shift_list, columns=["이름", "주차", "요일", "시간대"])

    weekday_order = ["월", "화", "수", "목", "금"]
    time_slots = ["오전", "오후"]
    result = {}
    for day in weekday_order:
        for time in time_slots:
            key = f"{day} {time}"
            df_filtered = df_split[(df_split["요일"] == day) & (df_split["시간대"] == time)]
            every_week = df_filtered[df_filtered["주차"] == "매주"]["이름"].unique()
            specific_weeks = df_filtered[df_filtered["주차"] != "매주"]
            specific_week_dict = {name: sorted(specific_weeks[specific_weeks["이름"] == name]["주차"].tolist(), 
                                               key=lambda x: int(x.replace("주", ""))) 
                                  for name in specific_weeks["이름"].unique() if specific_weeks[specific_weeks["이름"] == name]["주차"].tolist()}
            employees = list(every_week) + [f"{name}({','.join(weeks)})" for name, weeks in specific_week_dict.items()]
            result[key] = ", ".join(employees) if employees else ""
    
    return pd.DataFrame(list(result.items()), columns=["시간대", "근무"])

def generate_supplement_table(df_result, names_in_master):
    supplement = []
    weekday_order = ["월", "화", "수", "목", "금"]
    shift_list = ["오전", "오후"]
    names_in_master = set(names_in_master)

    for day in weekday_order:
        for shift in shift_list:
            time_slot = f"{day} {shift}"
            row = df_result[df_result["시간대"] == time_slot].iloc[0]
            employees = set(emp.split("(")[0].strip() for emp in row["근무"].split(", ") if emp)
            supplement_employees = names_in_master - employees

            if shift == "오후":
                morning_slot = f"{day} 오전"
                morning_employees = set(df_result[df_result["시간대"] == morning_slot].iloc[0]["근무"].split(", ") 
                                        if morning_slot in df_result["시간대"].values else [])
                supplement_employees = {emp if emp in morning_employees else f"{emp}🔺" for emp in supplement_employees}

            supplement.append({"시간대": time_slot, "보충": ", ".join(sorted(supplement_employees)) if supplement_employees else ""})

    return pd.DataFrame(supplement)

def split_column_to_multiple(df, column_name, prefix):
    if column_name not in df.columns:
        st.warning(f"⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
        st.stop()
        return df
    
    split_data = df[column_name].str.split(", ", expand=True)
    
    max_cols = split_data.shape[1]
    
    new_columns = [f"{prefix}{i+1}" for i in range(max_cols)]
    split_data.columns = new_columns
    
    df = df.drop(columns=[column_name])
    
    df = pd.concat([df, split_data], axis=1)

    return df

def append_summary_table_to_excel(worksheet, summary_df, style_args):
    if summary_df.empty:
        return

    fills = {
        'header': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),
        'yellow': PatternFill(start_color='FFF296', end_color='FFF296', fill_type='solid'),
        'pink': PatternFill(start_color='FFC8CD', end_color='FFC8CD', fill_type='solid'),
        'green': PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid'),
        'dark_green': PatternFill(start_color='82C4B5', end_color='82C4B5', fill_type='solid'),
        'blue': PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid'),
        'orange': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
        'lightgray': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    }
    
    start_row = worksheet.max_row + 3
    thin_border = style_args['border'] 

    # 헤더 쓰기
    for c_idx, value in enumerate(summary_df.columns.tolist(), 1):
        cell = worksheet.cell(row=start_row, column=c_idx, value=value)
        cell.fill = fills['header']
        cell.font = style_args['bold_font']
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 데이터 행 쓰기
    for r_idx, row_data in enumerate(summary_df.itertuples(index=False), start_row + 1):
        label = row_data[0]
        for c_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            cell.font = style_args['bold_font'] if c_idx == 1 else style_args['font']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            fill_color = None
            if label in ["오전누적", "오후누적"]: fill_color = fills['pink']
            elif label in ["오전합계", "오후합계"]: fill_color = fills['blue']
            elif label == "오전당직": fill_color = fills['blue']
            elif label == "오전당직누적": fill_color = fills['pink']
            elif label == "오후당직": fill_color = fills['lightgray']
            elif label == "오후당직누적": fill_color = fills['lightgray']

            if c_idx == 1 and label in ["오전보충", "임시보충", "오후보충", "온콜검사"]:
                fill_color = fills['yellow']
            
            if fill_color:
                cell.fill = fill_color

    start_col = 1
    end_col = len(summary_df.columns)
    labels = summary_df.iloc[:, 0].tolist()

    apply_outer_border(worksheet, start_row, start_row, start_col, end_col)
    
    apply_outer_border(worksheet, start_row, start_row + len(labels), start_col, start_col)

    block1_start = start_row + 1 + labels.index("오전보충")
    block1_end = start_row + 1 + labels.index("오전누적")
    apply_outer_border(worksheet, block1_start, block1_end, start_col, end_col)

    block2_start = start_row + 1 + labels.index("오후보충")
    block2_end = start_row + 1 + labels.index("오후누적")
    apply_outer_border(worksheet, block2_start, block2_end, start_col, end_col)
    
    block3_start = start_row + 1 + labels.index("오전당직")
    block3_end = start_row + 1 + labels.index("오후당직누적")
    apply_outer_border(worksheet, block3_start, block3_end, start_col, end_col)

    legend_start_row = worksheet.max_row + 3 

    legend_data = [
        ('A9D08E', '대체 보충'),
        ('FFF28F', '보충'),
        ('95B3D7', '대체 휴근'),
        ('B1A0C7', '휴근'),
        ('DA9694', '휴가/학회'),
        ('FABF8F', '꼭근무')
    ]

    for i, (hex_color, description) in enumerate(legend_data):
        current_row = legend_start_row + i
        
        color_cell = worksheet.cell(row=current_row, column=1)
        color_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        color_cell.border = thin_border

        desc_cell = worksheet.cell(row=current_row, column=2, value=description)
        desc_cell.font = style_args['font']
        desc_cell.border = thin_border
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    worksheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 11
    for i in range(2, len(summary_df.columns) + 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 9

def apply_outer_border(worksheet, start_row, end_row, start_col, end_col):
    medium_side = Side(style='medium') 

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = worksheet.cell(row=r, column=c)
            
            top = cell.border.top
            left = cell.border.left
            bottom = cell.border.bottom
            right = cell.border.right

            if r == start_row: top = medium_side
            if r == end_row: bottom = medium_side
            if c == start_col: left = medium_side
            if c == end_col: right = medium_side
            
            cell.border = Border(top=top, left=left, bottom=bottom, right=right)

def append_final_summary_to_excel(worksheet, df_final_summary, style_args):
    if df_final_summary.empty: return
    start_row = worksheet.max_row + 3
    
    worksheet.append(df_final_summary.columns.tolist())
    for cell in worksheet[start_row]:
        cell.font = style_args['bold_font']
        cell.border = style_args['border']
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for _, row in df_final_summary.iterrows():
        worksheet.append(row.tolist())
    
    for row in worksheet.iter_rows(min_row=start_row + 1, max_row=worksheet.max_row):
        for cell in row:
            cell.font = style_args['font']
            cell.border = style_args['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')

def replace_adjustments(df):
    """
    [수정됨] 동일 인물 + 동일 시간대에서 추가보충/추가제외 -> 대체보충/대체휴근로 변경합니다.
    [★] '주차' 제약을 제거하고 월 전체에서 1:1 매칭을 수행합니다.
    [★] 메모 형식을 'm/d에서 대체됨', 'm/d로 대체함'으로 변경합니다.
    """
    color_priority = {'🟠 주황색': 0, '🟢 초록색': 1, '🟡 노란색': 2, '기본': 3, '🔴 빨간색': 4, '🔵 파란색': 5, '🟣 보라색': 6, '특수근무색': -1}

    # 1. '보충' 또는 '휴근'인 행만 필터링 (주차 정보 포함 필수)
    adjustments_df = df[df['상태'].isin(['보충', '휴근'])].copy()
    
    # 2. 그룹별로 순차 매칭을 위해 날짜순으로 정렬
    adjustments_df.sort_values(by='날짜', inplace=True)

    # 3. 그룹별로 순차 매칭 수행
    # --- ▼▼▼ [수정 1] '주차'를 groupby에서 제거 ▼▼▼ ---
    for (worker, shift), group in adjustments_df.groupby(['근무자', '시간대']):
    # --- ▲▲▲ [수정 1] 완료 ---
        
        # 날짜 순으로 정렬된 추가보충 및 추가제외 레코드 리스트를 얻습니다.
        bochung_records = group[group['상태'] == '보충'].to_dict('records')
        jeoe_records = group[group['상태'] == '휴근'].to_dict('records')

        # 대체 가능 횟수 (min(추가보충 수, 추가제외 수))
        num_swaps = min(len(bochung_records), len(jeoe_records))

        # 4. 최대 가능 횟수만큼 순차적으로 짝짓기
        for i in range(num_swaps):
            bochung = bochung_records[i]
            jeoe = jeoe_records[i]
            
            # 매칭 날짜를 YYYY-MM-DD 형식으로 가져옵니다.
            bochung_date_str = bochung['날짜']
            jeoe_date_str = jeoe['날짜']
            
            # 5. 원본 df에 상태 업데이트 (매칭된 두 레코드에 대해)
            
            # 대체보충으로 변경 (추가보충이었던 레코드)
            bochung_mask = (df['날짜'] == bochung_date_str) & \
                           (df['시간대'] == shift) & \
                           (df['근무자'] == worker) & \
                           (df['상태'] == '보충')
            
            df.loc[bochung_mask, '상태'] = '대체보충'
            df.loc[bochung_mask, '색상'] = '🟢 초록색'
            # --- ▼▼▼ [수정 2] '대체보충' 메모 형식 변경 (요청사항) ▼▼▼ ---
            df.loc[bochung_mask, '메모'] = f"{pd.to_datetime(jeoe_date_str).strftime('%-m/%-d')}에서 대체됨"
            # --- ▲▲▲ [수정 2] 완료 ---

            # 대체휴근로 변경 (추가제외였던 레코드)
            jeoe_mask = (df['날짜'] == jeoe_date_str) & \
                        (df['시간대'] == shift) & \
                        (df['근무자'] == worker) & \
                        (df['상태'] == '휴근')
            
            df.loc[jeoe_mask, '상태'] = '대체휴근'
            df.loc[jeoe_mask, '색상'] = '🔵 파란색'
            # --- ▼▼▼ [수정 3] '대체휴근' 메모 형식도 일관되게 변경 ▼▼▼ ---
            df.loc[jeoe_mask, '메모'] = f"{pd.to_datetime(bochung_date_str).strftime('%-m/%-d')}로 대체함"
            # --- ▲▲▲ [수정 3] 완료 ---
            
    # 6. 최종 결과를 반환합니다. (호출한 곳에서 최종 중복 제거 필요)
    return df

# --- 1. 최종본(공유용) 엑셀 생성 함수 ---
def create_final_schedule_excel(initial_df, edited_df, edited_cumulative_df, df_special, df_requests, closing_dates, month_str, df_final_unique, df_schedule):
    """
    [공유용 최종본]
    - [★ F2DCDB 수정 v2 ★]
    - '변경된' 셀의 배경색을 'F2DCDB' (연분홍)로 변경합니다.
    - '변경된' 셀의 색상을 '상태' 색상보다 우선 적용합니다.
    - '변경된' 셀에 '변경 전:' 메모를 추가합니다.
    - (대체보충) 메모 로직을 통합합니다.
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "스케줄"

    # --- 스타일 정의 ---
    font_name = "맑은 고딕" if platform.system() == "Windows" else "Arial"
    default_font = Font(name=font_name, size=9)
    bold_font = Font(name=font_name, size=9, bold=True)
    duty_font = Font(name=font_name, size=9, bold=True, color="FF69B4") # 핑크색 굵게
    header_font = Font(name=font_name, size=9, color='FFFFFF', bold=True)
    color_map = {'휴가': 'DA9694', '학회': 'DA9694', '꼭 근무': 'FABF8F',
                 '보충': 'FFF28F', '대체보충': 'A9D08E', '휴근': 'B1A0C7',
                 '대체휴근': '95B3D7', '특수근무': 'D0E0E3', # 토/휴일 근무용 (거의 사용 안 함)
                 '근무': 'FFFFFF', '당직': 'FFFFFF', '기본': 'FFFFFF'} # 기본 흰색
    header_fill = PatternFill(start_color='000000', fill_type='solid')
    date_col_fill = PatternFill(start_color='808080', fill_type='solid') # 날짜열 회색
    weekday_fill = PatternFill(start_color='FFF2CC', fill_type='solid') # 요일열 노란색
    special_day_fill = PatternFill(start_color='95B3D7', fill_type='solid') # 토/휴일 요일 파란색
    empty_day_fill = PatternFill(start_color='808080', fill_type='solid') # 빈 날짜 회색
    holiday_blue_fill = PatternFill(start_color="DDEBF7", fill_type='solid') # 토/휴일 오전 기본 파란색
    
    # --- ▼▼▼ [★ 1. F2DCDB 스타일 추가] ▼▼▼ ---
    changed_fill = PatternFill(start_color='F2DCDB', fill_type='solid')
    # --- ▲▲▲ [추가 완료] ▲▲▲ ---
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    if df_final_unique is None or df_schedule is None:
        st.error("Excel 생성에 필요한 최종 배정 데이터(df_final_unique or df_schedule)가 함수로 전달되지 않았습니다.")
        wb.save(output)
        return output.getvalue()
        
    final_columns = ['날짜', '요일'] + [str(i) for i in range(1, 13)] + [''] + ['오전당직(온콜)'] + [f'오후{i}' for i in range(1, 5)]

    for c, col_name in enumerate(final_columns, 1):
        cell = ws.cell(row=1, column=c, value=col_name); cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border

    # --- 데이터 행 생성 및 서식 적용 ---
    for r, (idx, edited_row) in enumerate(edited_df.iterrows(), 2):
        
        # --- ▼▼▼ [★ 2. NameError 수정] ▼▼▼ ---
        # 'initial_df' (괄호 포함 원본)에서 '수정 전' 행을 가져옵니다.
        if idx not in initial_df.index: # (안전장치)
            continue 
        initial_row = initial_df.loc[idx]
        # --- ▲▲▲ [수정 완료] ▲▲▲ ---

        try:
            display_date = edited_row['날짜']
            cleaned_display_date = display_date.replace('월','-').replace('일','').replace(' ','')
            if '날짜' in df_schedule.columns and df_schedule['날짜'].dtype == 'object':
                matched_schedule = df_schedule[df_schedule['날짜'].str.contains(cleaned_display_date, na=False)]
            else:
                matched_schedule = pd.DataFrame()

            if not matched_schedule.empty:
                current_date_iso = matched_schedule['날짜'].iloc[0]
                current_date = datetime.strptime(current_date_iso, '%Y-%m-%d').date()
            else:
                try:
                    current_date = datetime.strptime(f"{month_str.split('년')[0]}년 {display_date}", "%Y년 %m월 %d일").date()
                    current_date_iso = current_date.strftime('%Y-%m-%d')
                except ValueError:
                    st.warning(f"날짜 형식 변환 실패 (Row {r}, Date: {display_date}). 해당 행 건너뜁니다.")
                    current_date, current_date_iso = None, None
        except Exception as e:
            st.warning(f"날짜 변환 중 예상치 못한 오류 (Row {r}, Date: {edited_row.get('날짜')}): {e}")
            current_date, current_date_iso = None, None

        if not current_date_iso: continue

        is_row_empty = all(pd.isna(v) or str(v).strip() == '' for k, v in edited_row.items() if k not in ['날짜', '요일'])
        is_special_day = False
        if isinstance(df_special, pd.DataFrame) and not df_special.empty and '날짜' in df_special.columns:
            try:
                if not pd.api.types.is_datetime64_any_dtype(df_special['날짜']):
                    df_special['날짜'] = pd.to_datetime(df_special['날짜'], errors='coerce')
                is_special_day = current_date in df_special.dropna(subset=['날짜'])['날짜'].dt.date.values if current_date else False
            except Exception as e_special_date:
                st.warning(f"df_special 날짜 처리 중 오류: {e_special_date}")
                is_special_day = False

        is_empty_day = (is_row_empty and not is_special_day) or (current_date_iso in closing_dates)

        weekend_oncall_worker = None
        if is_special_day and isinstance(df_special, pd.DataFrame):
            try:
                special_day_info = df_special[df_special['날짜'].dt.date == current_date]
                if not special_day_info.empty and '당직' in special_day_info.columns:
                    oncall_val = special_day_info['당직'].iloc[0]
                    if pd.notna(oncall_val) and oncall_val != "당직 없음": weekend_oncall_worker = str(oncall_val).strip()
            except Exception as e_oncall:
                st.warning(f"주말 당직자 확인 중 오류: {e_oncall}")


        for c, col_name in enumerate(final_columns, 1):
            raw_value_edited = str(edited_row.get(col_name, '')).strip()
            worker_name_display = re.sub(r'\(.+\)', '', raw_value_edited).strip()

            cell = ws.cell(row=r, column=c, value=worker_name_display)
            cell.font = default_font; cell.alignment = center_align; cell.border = border

            if is_empty_day:
                cell.fill = empty_day_fill
                continue
            elif col_name == '날짜':
                cell.fill = date_col_fill
                continue
            elif col_name == '요일':
                cell.fill = special_day_fill if is_special_day else weekday_fill
                continue
            elif is_special_day and '오후' in str(col_name): # 토/휴일 오후는 비움
                cell.value = ""
                cell.fill = PatternFill(fill_type=None) # 배경색 없음 (기본 흰색)
                continue
            
            elif is_special_day and str(col_name).isdigit(): # 토/휴일 오전
                if worker_name_display:
                    cell.fill = holiday_blue_fill
                else:
                    cell.fill = PatternFill(fill_type=None) # 값이 없으면 기본(흰색)
            
            else: # 평일 기본 배경 없음 (기본 흰색)
                cell.fill = PatternFill(fill_type=None)

            if not worker_name_display:
                continue

            time_slot = None
            if str(col_name).isdigit(): time_slot = '오전'
            elif '오후' in str(col_name): time_slot = '오후'
            elif col_name == '오전당직(온콜)': time_slot = '오전당직'

            status_or_memo = '기본'
            match = re.match(r'.+?\((.+)\)', raw_value_edited)
            if match: 
                status_or_memo = match.group(1).strip() 

            real_status = '기본'
            if status_or_memo == '기본':
                real_status = '기본'
            elif status_or_memo in color_map: 
                real_status = status_or_memo
            elif pd.notna(status_or_memo) and ('대체됨' in status_or_memo or '대체함' in status_or_memo or re.search(r'\d{1,2}/\d{1,2}', status_or_memo)):
                real_status = '대체보충' 
            else:
                real_status = '기본' 

            color_hex = color_map.get(real_status, 'FFFFFF') 
            
            if is_special_day and time_slot == '오전' and color_hex == 'FFFFFF':
                color_hex = "DDEBF7" # holiday_blue_fill

            # --- ▼▼▼ [★ 3. F2DCDB 및 메모 로직 적용] ▼▼▼ ---
            # 'initial_row'는 괄호가 포함된 원본임
            initial_raw_value = str(initial_row.get(col_name, '')).strip()
            cell_changed = (raw_value_edited != initial_raw_value)

            if cell_changed:
                cell.fill = changed_fill # 1순위: 변경됨 (F2DCDB)
                cell.comment = Comment(f"변경 전: {initial_raw_value or '빈 값'}", "Edit Tracker")
            elif color_hex != 'FFFFFF':
                # 2순위: (변경 안됐지만) 상태 색상
                cell.fill = PatternFill(start_color=color_hex, fill_type='solid')
            # 3순위: 기본 (흰색)은 cell.fill을 따로 지정 안 함 (L598에서 이미 처리됨)

            # [수정] 기존 '대체보충' 메모 로직과 통합
            if real_status == '대체보충' and pd.notna(status_or_memo) and re.search(r'\d{1,2}/\d{1,2}', status_or_memo):
                try:
                    # '변경 전' 코멘트가 없을 때만 '대체' 메모를 추가
                    if cell.comment is None: 
                        cell.comment = Comment(status_or_memo, "Schedule Bot")
                except Exception as e_memo:
                    pass # 코멘트 추가 실패 시 무시
            # --- ▲▲▲ [신규 로직 교체 완료] ▲▲▲ ---

            # --- 당직자 폰트 적용 ---
            if col_name == '오전당직(온콜)' and worker_name_display:
                cell.font = duty_font
            elif is_special_day and time_slot == '오전' and worker_name_display == weekend_oncall_worker:
                cell.font = duty_font


    # --- 익월 누적 현황 추가 ---
    if not edited_cumulative_df.empty:
        style_args_summary = {'font': default_font, 'bold_font': bold_font, 'border': border}
        append_summary_table_to_excel(ws, edited_cumulative_df, style_args_summary)


    # --- 열 너비 설정 ---
    ws.column_dimensions['A'].width = 11
    for i, col_name in enumerate(final_columns, 1):
        if col_name != '날짜':
            col_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[col_letter].width = 9


    wb.save(output)
    return output.getvalue()

def create_checking_schedule_excel(initial_df, edited_df, edited_cumulative_df, df_special, df_requests, closing_dates, month_str):
    """
    [관리자 확인용]
    - [★ F2DCDB 수정 v2 ★]
    - '변경된' 셀의 배경색을 'F2DCDB' (연분홍)로 변경합니다.
    - '변경된' 셀의 색상을 '상태' 색상보다 우선 적용합니다.
    - '변경된' 셀에 '변경 전:' 메모를 추가합니다.
    - (대체보충) 메모 로직을 통합합니다.
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "스케줄 (배정 확인용)"

    # --- 스타일 정의 ---
    font_name = "맑은 고딕" if platform.system() == "Windows" else "Arial"
    default_font = Font(name=font_name, size=9)
    bold_font = Font(name=font_name, size=9, bold=True)
    duty_font = Font(name=font_name, size=9, bold=True, color="FF69B4")
    header_font = Font(name=font_name, size=9, color='FFFFFF', bold=True)
    color_map = {'휴가': 'DA9694', '학회': 'DA9694', '꼭 근무': 'FABF8F', '보충': 'FFF28F', '대체보충': 'A9D08E', '휴근': 'B1A0C7', '대체휴근': '95B3D7', '특수근무': 'D0E0E3', '기본': 'FFFFFF'}
    header_fill = PatternFill(start_color='000000', fill_type='solid')
    date_col_fill = PatternFill(start_color='808080', fill_type='solid')
    weekday_fill = PatternFill(start_color='FFF2CC', fill_type='solid')
    special_day_fill = PatternFill(start_color='95B3D7', fill_type='solid')
    
    # --- ▼▼▼ [★ 1. F2DCDB 스타일로 변경] ▼▼▼ ---
    changed_fill = PatternFill(start_color='F2DCDB', fill_type='solid') # (연분홍)
    # --- ▲▲▲ [수정 완료] ▲▲▲ ---
    
    empty_day_fill = PatternFill(start_color='808080', fill_type='solid')
    holiday_blue_fill = PatternFill(start_color="DDEBF7", fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    # --- 동적 열 정의 ---
    checking_columns = edited_df.columns.tolist()

    # --- 헤더 생성 ---
    for c, col_name in enumerate(checking_columns, 1):
        cell = ws.cell(row=1, column=c, value=col_name); cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border

    # --- 데이터 행 생성 및 서식 적용 ---
    for r, (idx, edited_row) in enumerate(edited_df.iterrows(), 2):
        # --- ▼▼▼ [★ 2. NameError 수정] ▼▼▼ ---
        # 'initial_df' (괄호 포함 원본)에서 '수정 전' 행을 가져옵니다.
        if idx not in initial_df.index: # (안전장치)
            continue
        initial_row = initial_df.loc[idx]
        # --- ▲▲▲ [수정 완료] ▲▲▲ ---
        
        try:
            current_date = datetime.strptime(f"{month_str.split('년')[0]}-{edited_row['날짜']}", "%Y-%m월 %d일").date()
            current_date_iso = current_date.strftime('%Y-%m-%d')
        except: current_date, current_date_iso = None, None
        
        is_row_empty = all(pd.isna(v) or str(v).strip() == '' for k, v in edited_row.items() if k not in ['날짜', '요일'])
        is_special_day = current_date in pd.to_datetime(df_special['날짜']).dt.date.values if current_date and not df_special.empty else False
        is_empty_day = (is_row_empty and not is_special_day) or (current_date_iso in closing_dates)
        
        weekend_oncall_worker = None
        if is_special_day:
            special_day_info = df_special[pd.to_datetime(df_special['날짜']).dt.date == current_date]
            if not special_day_info.empty and '당직' in special_day_info.columns:
                oncall_val = special_day_info['당직'].iloc[0]
                if pd.notna(oncall_val) and oncall_val != "당직 없음": weekend_oncall_worker = str(oncall_val).strip()

        for c, col_name in enumerate(checking_columns, 1):
            raw_value = str(edited_row.get(col_name, '')).strip()
            worker_name = re.sub(r'\(.+\)', '', raw_value).strip()
            status = '기본'
            match = re.match(r'.+?\((.+)\)', raw_value)
            if match: status = match.group(1).strip()
            
            cell = ws.cell(row=r, column=c, value=worker_name)
            cell.font = default_font; cell.alignment = center_align; cell.border = border

            if is_empty_day: cell.fill = empty_day_fill; continue
            if col_name == '날짜': cell.fill = date_col_fill; continue
            if col_name == '요일': cell.fill = special_day_fill if is_special_day else weekday_fill; continue
            
            if not worker_name: continue
            
            if is_special_day:
                if str(col_name).isdigit():
                    cell.fill = holiday_blue_fill
                    if worker_name == weekend_oncall_worker: cell.font = duty_font
                elif '오후' in str(col_name): cell.value = ""
                continue
            
            # --- ▼▼▼ [★ 3. F2DCDB 및 메모 로직 적용] ▼▼▼ ---
            status_or_memo = status 

            real_status = '기본'
            if status_or_memo == '기본':
                real_status = '기본'
            elif status_or_memo in color_map: 
                real_status = status_or_memo
            elif pd.notna(status_or_memo) and ('대체됨' in status_or_memo or '대체함' in status_or_memo or re.search(r'\d{1,2}/\d{1,2}', status_or_memo)):
                real_status = '대체보충' 
            else:
                real_status = '기본'

            fill_hex = color_map.get(real_status, 'FFFFFF') 

            # 'initial_row'는 괄호가 포함된 원본임
            initial_raw_value = str(initial_row.get(col_name, '')).strip()
            cell_changed = (raw_value != initial_raw_value)
            
            if cell_changed:
                # 1순위: 변경된 셀은 무조건 F2DCDB
                cell.fill = changed_fill
                cell.comment = Comment(f"변경 전: {initial_raw_value or '빈 값'}", "Edit Tracker")
            elif fill_hex and fill_hex != 'FFFFFF':
                # 2순위: (변경 안됐지만) 상태 색상
                cell.fill = PatternFill(start_color=fill_hex, fill_type='solid')
            else:
                # 3순위: '기본' 상태이고 변경되지도 않음 (흰색)
                cell.fill = PatternFill(start_color='FFFFFF', fill_type='solid')
            
            if col_name == '오전당직(온콜)': cell.font = duty_font
            
            # [수정] 기존 '대체보충' 메모 로직과 통합
            if real_status == '대체보충' and pd.notna(status_or_memo) and re.search(r'\d{1,2}/\d{1,2}', status_or_memo):
                try:
                    # '변경 전' 코멘트가 없을 때만 '대체' 메모를 추가
                    if cell.comment is None: 
                        cell.comment = Comment(status_or_memo, "Schedule Bot")
                except Exception as e_memo:
                    pass
            # --- ▲▲▲ [신규 로직 교체 완료] ▲▲▲ ---
    
    # --- ✨ [핵심 수정] 익월 누적 현황을 올바른 형식으로 추가 ---
    if not edited_cumulative_df.empty:
        style_args = {'font': default_font, 'bold_font': bold_font, 'border': border}
        # 요청하신 함수에 편집된 데이터프레임을 그대로 전달
        append_summary_table_to_excel(ws, edited_cumulative_df, style_args)

    # --- 열 너비 설정 ---
    ws.column_dimensions['A'].width = 11
    for i in range(2, len(checking_columns) + 1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 9

    wb.save(output)
    return output.getvalue()

def create_formatted_schedule_excel(initial_df, edited_df, edited_cumulative_df, df_special, df_requests, closing_dates, month_str):
    """
    [관리자 확인용 구버전 - create_checking_schedule_excel 로 대체 가능]
    - 이 함수는 '스케줄 수정' 페이지에서는 create_checking_schedule_excel 과 동일한 역할을 합니다.
    - 여기서는 create_checking_schedule_excel 을 대신 사용하도록 유도할 수 있으나,
    - 혹시 모르니 일단 '스케줄 수정' 페이지의 정의를 그대로 가져옵니다.
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "수정된 스케줄"

    # --- 스타일 및 맵 정의 ---
    font_name = "맑은 고딕" if platform.system() == "Windows" else "Arial"
    default_font = Font(name=font_name, size=9)
    bold_font = Font(name=font_name, size=9, bold=True)
    duty_font = Font(name=font_name, size=9, bold=True, color="FF69B4")
    header_font = Font(name=font_name, size=9, color='FFFFFF', bold=True)
    color_map = {'휴가': 'DA9694', '학회': 'DA9694', '꼭 근무': 'FABF8F', '보충': 'FFF28F', '대체보충': 'A9D08E', '휴근': 'B1A0C7', '대체휴근': '95B3D7', '특수근무': 'D0E0E3', '기본': 'FFFFFF'}
    header_fill = PatternFill(start_color='000000', fill_type='solid')
    date_col_fill = PatternFill(start_color='808080', fill_type='solid')
    weekday_fill = PatternFill(start_color='FFF2CC', fill_type='solid')
    special_day_fill = PatternFill(start_color='95B3D7', fill_type='solid')
    changed_fill = PatternFill(start_color='FFFF00', fill_type='solid') # 노란색 (변경됨)
    empty_day_fill = PatternFill(start_color='808080', fill_type='solid')
    changed_fill = PatternFill(start_color='F2DCDB', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    # 요청사항 맵 생성 (휴가, 학회, 꼭 근무)
    requests_map = {}
    if not df_requests.empty:
        def parse_date_range(d_str):
            if pd.isna(d_str) or not isinstance(d_str, str) or d_str.strip() == '': return []
            d_str = d_str.strip()
            if '~' in d_str:
                try:
                    start, end = [datetime.strptime(d.strip(), '%Y-%m-%d').date() for d in d_str.split('~')]
                    return [(start + timedelta(days=i)).strftime('%Y-%m-%d') for i in range((end - start).days + 1)]
                except: return []
            else:
                try:
                    return [datetime.strptime(d.strip(), '%Y-%m-%d').date().strftime('%Y-%m-%d') for d in d_str.split(',')]
                except: return []
        
        for _, row in df_requests.iterrows():
            worker = row['이름']
            status = row['분류']
            if status in ['휴가', '학회'] or '꼭 근무' in status:
                clean_status = '꼭 근무' if '꼭 근무' in status else status
                for date_iso in parse_date_range(row['날짜정보']):
                    requests_map[(worker, date_iso)] = clean_status

    # 헤더 생성
    for c, col_name in enumerate(edited_df.columns, 1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border

    # 데이터 행 생성
    for r, (idx, edited_row) in enumerate(edited_df.iterrows(), 2):
        initial_row = initial_df.loc[idx]
        
        try:
            current_date = datetime.strptime(f"{month_str.split('년')[0]}-{edited_row['날짜']}", "%Y-%m월 %d일").date()
            current_date_iso = current_date.strftime('%Y-%m-%d')
        except (ValueError, TypeError):
            current_date = None; current_date_iso = None

        is_row_empty = all(pd.isna(v) or str(v).strip() == '' for k, v in edited_row.items() if k not in ['날짜', '요일'])
        is_special_day = current_date in pd.to_datetime(df_special['날짜']).dt.date.values if current_date and not df_special.empty else False
        is_empty_day = (is_row_empty and not is_special_day) or (current_date_iso in closing_dates)
        
        weekend_oncall_worker = None
        if is_special_day:
            special_day_info = df_special[pd.to_datetime(df_special['날짜']).dt.date == current_date]
            if not special_day_info.empty and '당직' in special_day_info.columns:
                oncall_val = special_day_info['당직'].iloc[0]
                if pd.notna(oncall_val) and oncall_val != "당직 없음":
                    weekend_oncall_worker = str(oncall_val).strip()

        for c, col_name in enumerate(edited_df.columns, 1):
            cell = ws.cell(row=r, column=c) # value는 나중에 설정
            cell.font = default_font; cell.alignment = center_align; cell.border = border

            if is_empty_day:
                cell.fill = empty_day_fill; continue
            if col_name == '날짜':
                cell.value = edited_row[col_name]; cell.fill = date_col_fill; continue
            if col_name == '요일':
                cell.value = edited_row[col_name]; cell.fill = special_day_fill if is_special_day else weekday_fill; continue
            
            raw_value = str(edited_row.get(col_name, '')).strip()
            
            if is_special_day:
                if str(col_name).isdigit() and raw_value:
                    cell.value = raw_value
                    cell.fill = PatternFill(start_color=color_map['특수근무'], end_color=color_map['특수근무'], fill_type='solid')
                    if raw_value == weekend_oncall_worker:
                        cell.font = duty_font
                elif '오후' in str(col_name):
                    cell.value = ""
                continue
            
            worker_name = raw_value
            status = '기본'
            
            match = re.match(r'(.+?)\((.+)\)', raw_value)
            if match:
                worker_name = match.group(1).strip(); status = match.group(2).strip()
            elif current_date_iso and worker_name:
                # data_editor에서 (상태)가 없는 셀을 위해 request 맵에서 다시 찾기
                status = requests_map.get((worker_name, current_date_iso), '기본')

            cell.value = worker_name
            if not worker_name: continue

            # --- ▼▼▼ [수정] 색상 적용 로직 변경 (상태 색상 우선) ▼▼▼ ---
            initial_raw_value = str(initial_row.get(col_name, '')).strip()
            cell_changed = (raw_value != initial_raw_value)

            fill_hex = color_map.get(status) # 1. 상태에 맞는 색상 가져오기
            
            if fill_hex and fill_hex != 'FFFFFF':
                # 2. 상태 색상이 '기본'(흰색)이 아니면, 해당 색상 적용
                cell.fill = PatternFill(start_color=fill_hex, fill_type='solid')
            elif cell_changed:
                # 3. 상태 색상이 '기본'인데, 셀 내용이 변경된 경우에만 노란색 적용
                cell.fill = changed_fill
            else:
                # 4. '기본' 상태이고 변경되지도 않음 (흰색)
                cell.fill = PatternFill(start_color='FFFFFF', fill_type='solid')
            # --- ▲▲▲ [수정] 완료 ▲▲▲ ---

            if col_name == '오전당직(온콜)' and worker_name:
                cell.font = duty_font
            
            # --- ▼▼▼ [수정] 메모(Comment) 생성 라인 제거 ▼▼▼ ---
            # initial_raw_value = str(initial_row.get(col_name, '')).strip()
            # if raw_value != initial_raw_value:
            #    cell.fill = changed_fill
            #    # cell.comment = Comment(f"변경 전: {initial_raw_value or '빈 값'}", "Edit Tracker")
            # --- ▲▲▲ [수정] 완료 ▲▲▲ ---

    # 익월 누적 현황 추가
    if not edited_cumulative_df.empty:
        style_args = {'font': default_font, 'bold_font': bold_font, 'border': border}
        # 이 함수는 DataFrame을 받음
        append_summary_table_to_excel(ws, edited_cumulative_df, style_args)

    # 열 너비 설정
    ws.column_dimensions['A'].width = 11
    for col in ws.columns:
        if col[0].column_letter != 'A':
            ws.column_dimensions[col[0].column_letter].width = 9

    wb.save(output)
    return output.getvalue()

# [★ L990의 이 함수 전체를 교체하세요 ★]

def recalculate_summary_from_schedule(edited_schedule_df, df_cumulative_initial, all_names, df_schedule_mapping):
    """
    (신규 함수)
    수정된 스케줄 data_editor 내용을 실시간으로 파싱하여,
    '보충', '당직' 횟수를 재계산하고 누적 테이블 DataFrame을 반환합니다.

    [★ 버그 수정 v2 ★]
    - L1010 ~ L1040의 파싱 로직을 수정합니다.
    - data_editor에 (10/6에서 대체됨) 처럼 메모로 표시된 '대체보충' 상태를
    - 올바르게 +1로 집계하도록 로직을 수정합니다.
    """
    
    # 1. 이름별로 (보충/휴근) 횟수, (당직) 횟수를 집계할 카운터 초기화
    am_bochong_counts = Counter()
    pm_bochong_counts = Counter()
    oncall_counts = Counter()

    # 2. 스케줄 data_editor (edited_schedule_df)의 모든 셀을 순회
    for idx, row in edited_schedule_df.iterrows():
        
        # 2-1. data_editor의 날짜(예: "10월 1일")를 ISO 날짜(예: "2025-10-01")로 변환
        try:
            date_iso = df_schedule_mapping.loc[idx, '날짜']
        except Exception:
            continue

        for col_name in edited_schedule_df.columns:
            raw_value = str(row[col_name] or '').strip()
            if not raw_value:
                continue

            # 2-2. 셀 텍스트에서 이름과 상태 파싱
            worker_name = re.sub(r'\(.+\)', '', raw_value).strip()
            status_match = re.search(r'\((.+)\)', raw_value)
            
            # --- ▼▼▼ [핵심 버그 수정] L1010~L1040 교체 ▼▼▼ ---
            status_text = status_match.group(1).strip() if status_match else '기본'
            
            # 2-3. 열 이름(col_name)에 따라 시간대 결정
            time_slot = None
            if col_name.isdigit(): time_slot = '오전'
            elif col_name.startswith("오후"): time_slot = '오후'
            elif col_name == '오전당직(온콜)': time_slot = '오전당직'
            
            if not time_slot or not worker_name:
                continue

            # 2-4. [수정된 로직] 파싱된 텍스트(status_text)를 '실제 상태'로 변환
            
            real_status_effect = 0 # 0: 기본, +1: 보충, -1: 휴근
            
            if status_text in ['보충', '대체보충']:
                real_status_effect = 1
            elif status_text in ['휴근', '대체휴근']:
                real_status_effect = -1
            elif pd.notna(status_text) and (re.search(r'\d{1,2}/\d{1,2}', status_text) or '대체됨' in status_text):
                # L2158 로직에 따라, 메모(날짜)가 표시되는 경우는 '대체보충'(+1) 뿐임
                real_status_effect = 1
            # (참고: '대체휴근'은 L2161에 따라 (대체휴근)으로 표시되므로 위에서 처리됨)

            # 2-5. 카운터 집계
            if time_slot == '오전당직':
                oncall_counts[worker_name] += 1
            
            elif time_slot == '오전':
                if real_status_effect == 1:
                    am_bochong_counts[worker_name] += 1
                elif real_status_effect == -1:
                    am_bochong_counts[worker_name] -= 1
            
            elif time_slot == '오후':
                if real_status_effect == 1:
                    pm_bochong_counts[worker_name] += 1
                elif real_status_effect == -1:
                    pm_bochong_counts[worker_name] -= 1
            # --- ▲▲▲ [핵심 버그 수정 완료] ▲▲▲ ---


    # 3. GSheet에서 로드한 *원본* 누적 테이블을 기반으로 최종 테이블 재구성
    recalculated_summary_df = df_cumulative_initial.copy()
    if '항목' not in recalculated_summary_df.columns:
        # (호환성) df_cumulative가 행/열 전환된 상태일 경우
        try:
            first_col = recalculated_summary_df.columns[0]
            recalculated_summary_df = recalculated_summary_df.set_index(first_col).transpose().reset_index().rename(columns={'index':'항목'})
        except Exception:
            return df_cumulative_initial # 오류 시 원본 반환
            
    recalculated_summary_df = recalculated_summary_df.set_index('항목')

    # 4. 모든 근무자 목록(all_names)을 순회하며 값 채우기
    for name in all_names:
        if name not in recalculated_summary_df.columns:
            recalculated_summary_df[name] = 0 # 새 이름이 있으면 열 추가
        
        # 4-1. GSheet 원본 값 가져오기 (오류 방지를 위해 .get(name, 0) 사용)
        base_am = int(recalculated_summary_df.loc['오전누적'].get(name, 0))
        base_pm = int(recalculated_summary_df.loc['오후누적'].get(name, 0))
        base_am_oncall = int(recalculated_summary_df.loc['오전당직누적'].get(name, 0))
        base_pm_oncall = int(recalculated_summary_df.loc['오후당직누적'].get(name, 0))

        # 4-2. 실시간 집계 값 가져오기
        am_bochong = am_bochong_counts.get(name, 0)
        pm_bochong = pm_bochong_counts.get(name, 0)
        am_oncall_total = oncall_counts.get(name, 0)

        # 4-3. 최종 값 계산 및 덮어쓰기
        recalculated_summary_df.at["오전보충", name] = am_bochong
        recalculated_summary_df.at["오전합계", name] = base_am  # '합계'는 원본 누적값을 의미
        recalculated_summary_df.at["오전누적", name] = base_am + am_bochong

        recalculated_summary_df.at["오후보충", name] = pm_bochong
        recalculated_summary_df.at["오후합계", name] = base_pm
        recalculated_summary_df.at["오후누적", name] = base_pm + pm_bochong

        recalculated_summary_df.at["오전당직", name] = am_oncall_total
        recalculated_summary_df.at["오전당직누적", name] = base_am_oncall + am_oncall_total
        
        recalculated_summary_df.at["오후당직", name] = 0 # 오후 당직은 이 시트에서 배정 안 함
        recalculated_summary_df.at["오후당직누적", name] = base_pm_oncall

    recalculated_summary_df = recalculated_summary_df.reset_index()

    # 원본 build_summary_table과 동일하게 모든 숫자 열을 int로 강제 변환
    for col in recalculated_summary_df.columns:
        if col != '항목':
            recalculated_summary_df[col] = pd.to_numeric(recalculated_summary_df[col], errors='coerce').fillna(0).astype(int)

    return recalculated_summary_df

st.header("🗓️ 스케줄 배정", divider='rainbow')
st.write("- 먼저 새로고침 버튼으로 최신 데이터를 불러온 뒤, 배정을 진행해주세요.")
if st.button("🔄 새로고침 (R)"):
    try:
        st.cache_data.clear()
        st.cache_resource.clear()

        # ▼▼▼ [핵심 수정] 페이지에 필요한 데이터만 선택적으로 삭제합니다 ▼▼▼
        keys_to_clear = [
            "assigned", "output", "df_cumulative_next", "request_logs", 
            "swap_logs", "adjustment_logs", "oncall_logs", "assignment_results",
            "show_confirmation_warning", "latest_existing_version",
            "data_loaded", "df_master", "df_request", "df_cumulative", 
            "df_shift", "df_supplement", "edited_df_cumulative"
        ]
        
        for key in keys_to_clear:
            if key in st.session_state:
                del st.session_state[key]
        # --- 수정 끝 ---
        
        st.success("데이터가 새로고침되었습니다. 페이지를 다시 로드합니다.")
        time.sleep(1)
        st.rerun()
    except Exception as e:
        st.error(f"새로고침 중 오류 발생: {type(e).__name__} - {e}")
        st.stop()

try:
    gc = get_gspread_client()
    if gc:
        sheet = gc.open_by_url(url)
        latest_schedule = find_latest_schedule_version(sheet, month_str)
        
        if latest_schedule:
            version_str = latest_schedule.split(' 스케줄 ')[-1]
            
            # ▼▼▼ [수정] 버전 이름에 따라 다른 안내 메시지를 표시합니다. ▼▼▼
            if version_str == '최종':
                message = f"이미 '**{version_str}**' 스케줄이 존재합니다. '**{version_str}**'을 수정하시려면 **방배정 페이지**로 이동해주세요."
            else:
                message = f"이미 '**{version_str}**' 스케줄이 존재합니다. '**{version_str}**'를 수정하시려면 **스케줄 수정 페이지**로 이동해주세요."
            
            st.info(message)
            # ▲▲▲ [수정] ▲▲▲
            
        st.session_state["latest_schedule_name"] = latest_schedule

except Exception as e:
    st.error(f"최종 스케줄 버전 확인 중 오류가 발생했습니다: {e}")
    st.session_state["latest_schedule_name"] = None
except Exception as e:
    st.error(f"최종 스케줄 버전 확인 중 오류가 발생했습니다: {e}")
    st.session_state["latest_schedule_name"] = None

# get_adjustment 함수 정의 (L661)
def get_adjustment(name, time_slot, df_final_unique=None):
    """
    [수정됨] 근무자의 시간대별 *총* 보충/제외 횟수 차이를 계산합니다.
    (대체보충/대체휴근 포함)
    """
    if df_final_unique is None:
        return 0
    
    # --- ▼▼▼ [핵심 수정] '대체보충', '대체휴근'을 isin 목록에 추가 ▼▼▼ ---
    adjustments = df_final_unique[
        (df_final_unique['근무자'] == name) &
        (df_final_unique['시간대'] == time_slot) &
        (df_final_unique['상태'].isin(['보충', '휴근', '대체보충', '대체휴근'])) # <-- '대체' 상태 추가
    ]
    # --- ▲▲▲ [수정 완료] ▲▲▲ ---
    
    if adjustments.empty:
        return 0
    
    # --- ▼▼▼ [핵심 수정] '보충'과 '대체보충'을 합산, '휴근'과 '대체휴근'을 합산 ▼▼▼ ---
    count = (
        len(adjustments[adjustments['상태'].isin(['보충', '대체보충'])]) -
        len(adjustments[adjustments['상태'].isin(['휴근', '대체휴근'])])
    )
    # --- ▲▲▲ [수정 완료] ▲▲▲ ---
    
    return count

def display_cumulative_table(df_cumulative):
    if df_cumulative.empty:
        st.warning("⚠️ 누적 테이블 데이터가 비어 있습니다.")
        return
    if '항목' not in df_cumulative.columns:
        st.error(f"누적 테이블에 '항목' 열이 없습니다. 열: {df_cumulative.columns.tolist()}")
        st.stop()

def display_pivoted_summary_table(df_summary):
    if df_summary.empty:
        st.warning("⚠️ 요약 테이블 데이터가 비어 있습니다.")
        return
    st.dataframe(df_summary, use_container_width=True, hide_index=True)

def build_summary_table(df_cumulative, all_names, next_month_str, df_final_unique=None):
    """
    [수정됨] 최종 요약 테이블을 생성합니다.
    - 합계 = 이번 달 배정 횟수
    - 누적 = (GSheet에서 읽은 누적 값) + (이번 달 배정 횟수)
    """
    summary_data = {name: [""] * 12 for name in all_names}
    df_summary = pd.DataFrame(summary_data)

    row_labels = [
        "오전보충", "임시보충", "오전합계", "오전누적",
        "오후보충", "온콜검사", "오후합계", "오후누적",
        "오전당직", "오전당직누적", "오후당직", "오후당직누적"
    ]
    df_summary.index = row_labels

    df_cum_indexed = df_cumulative.set_index('항목')
    
    # 실제 배정된 당직 횟수 계산
    actual_oncall_counts = Counter(df_final_unique[df_final_unique['시간대'] == '오전당직']['근무자']) if df_final_unique is not None else Counter()

    for name in all_names:
        if name not in df_cum_indexed.columns:
            df_cum_indexed[name] = 0

        # --- 오전/오후 근무 (기존과 동일) ---
        am_hapgye = int(df_cum_indexed.loc['오전누적', name])
        pm_hapgye = int(df_cum_indexed.loc['오후누적', name])
        am_bochung = get_adjustment(name, '오전', df_final_unique)
        pm_bochung = get_adjustment(name, '오후', df_final_unique)
        
        df_summary.at["오전보충", name] = am_bochung
        df_summary.at["오전합계", name] = am_hapgye
        df_summary.at["오전누적", name] = am_hapgye + am_bochung

        df_summary.at["오후보충", name] = pm_bochung
        df_summary.at["오후합계", name] = pm_hapgye
        df_summary.at["오후누적", name] = pm_hapgye + pm_bochung

        # --- ▼▼▼ [로직 수정] 당직 파트 (요청하신 로직) ▼▼▼ ---
        
        # 1. '오전당직누적' (시작 값)을 GSheet에서 가져옴
        oncall_start_total = int(df_cum_indexed.loc['오전당직누적', name]) 
        # 2. '오후당직누적' (시작 값)을 GSheet에서 가져옴
        pm_oncall_start_total = int(df_cum_indexed.loc['오후당직누적', name])
        # 3. '이번 달 배정 횟수'를 가져옴
        oncall_this_month = actual_oncall_counts.get(name, 0)
        
        # 4. '오전당직' (이번 달 횟수) 행에 '이번 달 횟수'를 넣습니다.
        df_summary.at["오전당직", name] = oncall_this_month
        
        # 5. '오전당직누적' (최종) 행에 '시작 값 + 이번 달 횟수'를 넣습니다.
        df_summary.at["오전당직누적", name] = oncall_start_total + oncall_this_month
        
        # 6. 오후 당직 (이번 달 0회)
        df_summary.at["오후당직", name] = 0
        df_summary.at["오후당직누적", name] = pm_oncall_start_total # 시작 값 = 최종 값
        
        # ▲▲▲ [수정 완료] ▲▲▲

    df_summary.reset_index(inplace=True)
    df_summary.rename(columns={'index': '항목'}, inplace=True)

    # (유지) 모든 열을 숫자로 변환
    for col in df_summary.columns:
        if col != '항목':
            df_summary[col] = pd.to_numeric(df_summary[col], errors='coerce').fillna(0).astype(int)

    return df_summary

def build_final_summary_table(df_cumulative, df_final_unique, all_names):
    summary_data = []
    
    adjustments = df_final_unique[df_final_unique['상태'].isin(['보충', '휴근'])]
    am_adjust = adjustments[adjustments['시간대'] == '오전'].groupby('근무자')['상태'].apply(lambda x: (x == '보충').sum() - (x == '휴근').sum()).to_dict()
    pm_adjust = adjustments[adjustments['시간대'] == '오후'].groupby('근무자')['상태'].apply(lambda x: (x == '보충').sum() - (x == '휴근').sum()).to_dict()
    
    oncall_counts = df_final_unique_sorted[df_final_unique_sorted['시간대'] == '오전당직']['근무자'].value_counts().to_dict() # 여기도 _sorted로 변경

    before_dict = df_cumulative.set_index('항목').T.to_dict()

    for name in all_names:
        b = before_dict.get(name, {})
        am_change = am_adjust.get(name, 0)
        pm_change = pm_adjust.get(name, 0)
        
        summary_data.append({
            '이름': name,
            '오전누적 (시작)': b.get('오전누적', 0),
            '오전누적 (변동)': am_change,
            '오전누적 (최종)': b.get('오전누적', 0) + am_change,
            '오후누적 (시작)': b.get('오후누적', 0),
            '오후누적 (변동)': pm_change,
            '오후누적 (최종)': b.get('오후누적', 0) + pm_change,
            '오전당직': b.get('오전당직', 0),
            '오전당직 (최종)': oncall_counts.get(name, 0),
            '오후당직': b.get('오후당직', 0),
        })
        
    return pd.DataFrame(summary_data)

df_master, df_request, df_cumulative, df_shift, df_supplement = load_data_page5()

# 세션 상태에 데이터 저장 (기존 코드 유지)
st.session_state["df_master"] = df_master

# 세션 상태에 데이터 저장
st.session_state["df_master"] = df_master
st.session_state["df_request"] = df_request
if "df_cumulative" not in st.session_state or st.session_state["df_cumulative"].empty:
    st.session_state["df_cumulative"] = df_cumulative
st.session_state["df_shift"] = df_shift
st.session_state["df_supplement"] = df_supplement

# 'edited_df_cumulative'가 없거나 비어있을 경우에만 초기화
if "edited_df_cumulative" not in st.session_state or st.session_state["edited_df_cumulative"].empty:
    st.session_state["edited_df_cumulative"] = df_cumulative.copy()

if '근무' not in df_shift.columns or '보충' not in df_supplement.columns:
    st.warning("⚠️ 데이터를 불러오는 데 문제가 발생했습니다. 새로고침 버튼을 눌러 다시 시도해주세요.")
    st.stop()

st.divider()
st.subheader(f"✨ {month_str} 테이블 종합")
st.write("- 당월 근무자와 보충 가능 인원을 확인하거나, 누적 테이블을 수정할 수 있습니다.\n- 보충 테이블에서 '🔺' 표시가 있는 인원은 해당일 오전 근무가 없으므로, 보충 시 오전·오후 모두 보충되어야 함을 의미합니다.")
with st.expander("📁 테이블 펼쳐보기"):

    df_shift_processed = split_column_to_multiple(df_shift, "근무", "근무")
    df_supplement_processed = split_column_to_multiple(df_supplement, "보충", "보충")

    def excel_download(name, sheet1, name1, sheet2, name2, sheet3, name3, sheet4, name4):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            sheet1.to_excel(writer, sheet_name=name1, index=False)
            sheet2.to_excel(writer, sheet_name=name2, index=False)
            sheet3.to_excel(writer, sheet_name=name3, index=False)
            sheet4.to_excel(writer, sheet_name=name4, index=False)
        
        excel_data = output.getvalue()
        return excel_data

    st.write(" ")
    st.markdown("**✅ 근무 테이블**")
    st.dataframe(df_shift, use_container_width=True, hide_index=True)

    st.markdown("**☑️ 보충 테이블**")
    st.dataframe(df_supplement, use_container_width=True, hide_index=True)

    st.markdown(f"**➕ 전월({prev_month_str}) 배정 완료된 누적 테이블**")
    st.write("- 해당 수치를 반영하여 보충/휴근/오전당직 배정이 이뤄집니다.")
    st.write("- 변동이 있는 경우, 직접 수치를 수정 가능합니다.\n- 해당 수치를 기반으로 스케줄 배정이 이뤄집니다.")
    df_cumulative_full = st.session_state["df_cumulative"]

    # --- ✅ 오류 해결 코드 시작 ---
    # 1. '항목' 열이 있는지 확인합니다.
    if '항목' not in df_cumulative_full.columns:
        try:
            # 2. '항목' 열이 없다면, 행/열이 뒤바뀐 것으로 간주하고 원래 형태로 되돌립니다.
            #    첫 번째 열(직원 이름)을 인덱스로 설정 후 행/열 전환(transpose)
            first_column_name = df_cumulative_full.columns[0]
            df_cumulative_full = df_cumulative_full.set_index(first_column_name).transpose().reset_index()
            
            # 3. 복원된 데이터의 첫 열 이름을 '항목'으로 변경합니다.
            df_cumulative_full.rename(columns={'index': '항목'}, inplace=True)
            
            # 4. 올바르게 변환된 데이터를 세션 상태에 다시 저장하여 문제를 영구적으로 해결합니다.
            st.session_state["df_cumulative"] = df_cumulative_full.copy()
            
        except Exception as e:
            st.error(f"누적 테이블 형식 자동 변환 중 오류가 발생했습니다: {e}")
            st.stop()
    # --- ✅ 오류 해결 코드 끝 ---

    # 1. 표시할 행 이름 정의 및 원본 데이터에서 필터링
    rows_to_display = ["오전누적", "오후누적", "오전당직누적", "오후당직누적"]
    df_to_edit = df_cumulative_full[df_cumulative_full['항목'].isin(rows_to_display)]

    # 2. 필터링된 데이터를 data_editor에 표시 (display_cumulative_table 호출 제거)
    edited_partial_df = st.data_editor(
        df_to_edit,
        use_container_width=True,
        hide_index=True,
        column_config={"항목": {"editable": False}},
        key="cumulative_editor" # 고유 키 부여
    )

    # 3. 저장 버튼 로직
    if st.button("💾 누적 테이블 수정사항 저장"):
        try:
            # 원본 전체 데이터의 복사본 생성
            df_updated_full = st.session_state["df_cumulative"].copy()

            # '항목'을 인덱스로 설정하여 정확한 위치에 업데이트 준비
            df_updated_full.set_index('항목', inplace=True)
            edited_partial_df.set_index('항목', inplace=True)

            # 수정된 내용으로 원본 업데이트
            df_updated_full.update(edited_partial_df)
            df_updated_full.reset_index(inplace=True) # 인덱스를 다시 열로 복원

            # 세션 상태 및 Google Sheet 업데이트 (이제 df_updated_full이 최신 전체 데이터임)
            st.session_state["df_cumulative"] = df_updated_full.copy()
            st.session_state["edited_df_cumulative"] = df_updated_full.copy()
            
            gc = get_gspread_client()
            sheet = gc.open_by_url(url)

            # ▼▼▼ [핵심 수정] 고정된 이름 대신 세션에 저장된 시트 이름을 사용합니다. ▼▼▼
            target_sheet_name = st.session_state.get("target_cumulative_sheet_name", f"{month_str} 누적")
            try:
                worksheet4 = sheet.worksheet(target_sheet_name)
            except WorksheetNotFound:
                st.info(f"'{target_sheet_name}' 시트가 없어 새로 생성합니다.")
                worksheet4 = sheet.add_worksheet(title=target_sheet_name, rows=100, cols=len(df_updated_full.columns) + 5)
            # ▲▲▲ [수정 완료] ▲▲▲
            
            update_data = [df_updated_full.columns.tolist()] + df_updated_full.values.tolist()
            
            if update_sheet_with_retry(worksheet4, update_data):
                st.success(f"'{target_sheet_name}' 테이블이 성공적으로 저장되었습니다.")
                time.sleep(1.5)
                st.rerun()
            else:
                st.error("누적 테이블 저장 실패")
                st.stop()
        except Exception as e:
            st.error(f"누적 테이블 저장 중 오류 발생: {str(e)}")

    # 4. 다운로드 버튼 로직
    with st.container():
        excel_data = excel_download(
            name=f"{month_str} 테이블 종합",
            sheet1=df_shift_processed, name1="근무 테이블",
            sheet2=df_supplement_processed, name2="보충 테이블",
            sheet3=df_request, name3="요청사항 테이블",
            # 수정된 전체 데이터를 다운로드에 사용
            sheet4=st.session_state["edited_df_cumulative"], name4="누적 테이블"
        )
        st.download_button(
            label="📥 상단 테이블 다운로드",
            data=excel_data,
            file_name=f"{month_str} 테이블 종합.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

st.divider()
st.subheader("📋 요청사항 관리")
st.write("- 명단 및 마스터에 등록되지 않은 인원 중 스케줄 배정이 필요한 경우, 관리자가 이름을 수기로 입력하여 요청사항을 추가해야 합니다.\n- '꼭 근무'로 요청된 사항은 해당 인원이 마스터가 없거나 모두 '근무없음' 상태더라도 반드시 배정됩니다.")

if df_request["분류"].nunique() == 1 and df_request["분류"].iloc[0] == '요청 없음':
    st.warning(f"⚠️ 아직까지 {month_str}에 작성된 요청사항이 없습니다.")

요청분류 = ["휴가", "학회", "보충 어려움(오전)", "보충 어려움(오후)", "보충 불가(오전)", "보충 불가(오후)", "꼭 근무(오전)", "꼭 근무(오후)", "요청 없음"]
st.dataframe(df_request.reset_index(drop=True), use_container_width=True, hide_index=True, height=300)

def add_request_callback():
    날짜정보 = ""
    분류 = st.session_state.request_category_select
    
    if 분류 != "요청 없음":
        방식 = st.session_state.method_select
        if 방식 == "일자 선택":
            날짜 = st.session_state.get("date_multiselect", [])
            if 날짜: 날짜정보 = ", ".join([d.strftime("%Y-%m-%d") for d in 날짜])
        elif 방식 == "기간 선택":
            날짜범위 = st.session_state.get("date_range", ())
            if isinstance(날짜범위, tuple) and len(날짜범위) == 2:
                시작, 종료 = 날짜범위
                날짜정보 = f"{시작.strftime('%Y-%m-%d')} ~ {종료.strftime('%Y-%m-%d')}"
        elif 방식 == "주/요일 선택":
            선택주차 = st.session_state.get("week_select", [])
            선택요일 = st.session_state.get("day_select", [])
            if 선택주차 or 선택요일:
                c = calendar.Calendar(firstweekday=6)
                month_calendar = c.monthdatescalendar(month_dt.year, month_dt.month)
                요일_map = {"월": 0, "화": 1, "수": 2, "목": 3, "금": 4}
                선택된_요일_인덱스 = [요일_map[요일] for 요일 in 선택요일] if 선택요일 else list(요일_map.values())
                날짜목록 = []
                for i, week in enumerate(month_calendar):
                    주차_이름 = ""
                    if i == 0: 주차_이름 = "첫째주"
                    elif i == 1: 주차_이름 = "둘째주"
                    elif i == 2: 주차_이름 = "셋째주"
                    elif i == 3: 주차_이름 = "넷째주"
                    elif i == 4: 주차_이름 = "다섯째주"
                    if not 선택주차 or "매주" in 선택주차 or 주차_이름 in 선택주차:
                        for date_obj in week:
                            if date_obj.month == month_dt.month and date_obj.weekday() in 선택된_요일_인덱스:
                                날짜목록.append(date_obj.strftime("%Y-%m-%d"))
                if 날짜목록:
                    날짜정보 = ", ".join(sorted(list(set(날짜목록))))
                else:
                    add_placeholder.warning(f"⚠️ {month_str}에는 해당 주차/요일의 날짜가 없습니다. 다른 조합을 선택해주세요.")
                    return

    이름 = st.session_state.get("add_employee_select", "")
    이름_수기 = st.session_state.get("new_employee_input", "")
    최종_이름 = 이름 if 이름 else 이름_수기

    if not 최종_이름 or (분류 != "요청 없음" and not 날짜정보):
        add_placeholder.warning("⚠️ 이름과 날짜를 올바르게 선택/입력해주세요.")
        return

    with add_placeholder.container():
        with st.spinner("요청사항 확인 및 저장 중..."):
            try:
                gc = get_gspread_client()
                sheet = gc.open_by_url(url)
                worksheet2 = sheet.worksheet(f"{month_str} 요청")
                all_requests = worksheet2.get_all_records()
                df_request_live = pd.DataFrame(all_requests)

                is_duplicate = not df_request_live[
                    (df_request_live["이름"] == 최종_이름) &
                    (df_request_live["분류"] == 분류) &
                    (df_request_live["날짜정보"] == 날짜정보)
                ].empty

                if is_duplicate:
                    st.error("⚠️ 이미 존재하는 요청사항입니다.")
                    time.sleep(1.5)
                    st.rerun()
                    return

                rows_to_delete = []
                for i, req in enumerate(all_requests):
                    if req.get("이름") == 최종_이름:
                        if 분류 == "요청 없음" or req.get("분류") == "요청 없음":
                            rows_to_delete.append(i + 2)
                
                if rows_to_delete:
                    for row_idx in sorted(rows_to_delete, reverse=True):
                        worksheet2.delete_rows(row_idx)

                worksheet2.append_row([최종_이름, 분류, 날짜정보 if 분류 != "요청 없음" else ""])
                
                st.success("요청사항이 저장되었습니다.")
                time.sleep(1.5)
                
                st.session_state.add_employee_select = None
                st.session_state.new_employee_input = ""
                st.session_state.request_category_select = "휴가"
                st.session_state.method_select = "일자 선택"
                st.session_state.date_multiselect = []
                st.session_state.date_range = (month_start, month_start + timedelta(days=1))
                st.session_state.week_select = []
                st.session_state.day_select = []
                
                st.rerun()

            except Exception as e:
                st.error(f"요청사항 추가 중 오류 발생: {e}")

입력_모드 = st.selectbox("입력 모드", ["이름 선택", "이름 수기 입력"], key="input_mode_select")
col1, col2, col3, col4 = st.columns([1, 1, 1, 1.5])
with col1:
    if 입력_모드 == "이름 선택":
        sorted_names = sorted(df_master["이름"].unique()) if not df_master.empty and "이름" in df_master.columns else []
        st.selectbox("이름 선택", sorted_names, key="add_employee_select")
    else:
        이름_수기 = st.text_input("이름 입력", help="명단에 없는 새로운 인원에 대한 요청을 추가하려면 입력", key="new_employee_input")
        if 이름_수기 and 이름_수기 not in st.session_state.get("df_map", pd.DataFrame()).get("이름", pd.Series()).values:
            st.warning(f"{이름_수기}은(는) 매핑 시트에 존재하지 않습니다. 먼저 명단 관리 페이지에서 추가해주세요.")
            st.stop()
with col2:
    st.selectbox("요청 분류", 요청분류, key="request_category_select")
if st.session_state.get("request_category_select") != "요청 없음":
    with col3:
        st.selectbox("날짜 선택 방식", ["일자 선택", "기간 선택", "주/요일 선택"], key="method_select")
    with col4:
        if st.session_state.method_select == "일자 선택":
            weekday_map = {0: "월", 1: "화", 2: "수", 3: "목", 4: "금", 5: "토", 6: "일"}
            def format_date(date_obj):
                return f"{date_obj.strftime('%-m월 %-d일')} ({weekday_map[date_obj.weekday()]})"
            날짜_목록 = [month_start + timedelta(days=i) for i in range((month_end - month_start).days + 1)]
            st.multiselect("요청 일자", 날짜_목록, format_func=format_date, key="date_multiselect")
        elif st.session_state.method_select == "기간 선택":
            st.date_input("요청 기간", value=(month_start, month_start + timedelta(days=1)), min_value=month_start, max_value=month_end, key="date_range")
        elif st.session_state.method_select == "주/요일 선택":
            st.multiselect("주차 선택", ["첫째주", "둘째주", "셋째주", "넷째주", "다섯째주", "매주"], key="week_select")
            st.multiselect("요일 선택", ["월", "화", "수", "목", "금"], key="day_select")

if st.session_state.get("request_category_select") == "요청 없음":
    st.markdown("<span style='color:red;'>⚠️ 요청 없음을 추가할 경우, 기존에 입력하였던 요청사항은 전부 삭제됩니다.</span>", unsafe_allow_html=True)

st.button("📅 추가", on_click=add_request_callback)

add_placeholder = st.empty()

st.write(" ")
st.markdown("**🔴 요청사항 삭제**")
if not df_request.empty:
    col0, col1 = st.columns([1, 2])
    with col0:
        sorted_names = sorted(df_request["이름"].unique()) if not df_request.empty else []
        selected_employee_id2 = st.selectbox("이름 선택", sorted_names, key="delete_request_employee_select")
    with col1:
        df_employee2 = df_request[df_request["이름"] == selected_employee_id2]
        df_employee2_filtered = df_employee2[df_employee2["분류"] != "요청 없음"]
        if not df_employee2_filtered.empty:
            selected_rows = st.multiselect(
                "요청사항 선택",
                df_employee2_filtered.index,
                format_func=lambda x: f"{df_employee2_filtered.loc[x, '분류']} - {df_employee2_filtered.loc[x, '날짜정보']}",
                key="delete_request_select"
            )
        else:
            st.info("📍 선택한 이름에 대한 요청사항이 없습니다.")
            selected_rows = []
else:
    st.info("📍 당월 요청사항 없음")
    selected_rows = []

if st.button("📅 삭제"):
    with st.spinner("요청을 삭제하는 중입니다..."):
        try:
            if selected_rows:
                gc = get_gspread_client()
                sheet = gc.open_by_url(url)
                worksheet2 = sheet.worksheet(f"{month_str} 요청")
                all_requests = worksheet2.get_all_records()
                
                items_to_delete_set = set()
                df_request_original = st.session_state["df_request"]
                for index in selected_rows:
                    row = df_request_original.loc[index]
                    items_to_delete_set.add((row['이름'], row['분류'], row['날짜정보']))

                rows_to_delete_indices = []
                for i, record in enumerate(all_requests):
                    record_tuple = (record.get('이름'), record.get('분류'), record.get('날짜정보'))
                    if record_tuple in items_to_delete_set:
                        rows_to_delete_indices.append(i + 2)
                
                if rows_to_delete_indices:
                    for row_idx in sorted(rows_to_delete_indices, reverse=True):
                        worksheet2.delete_rows(row_idx)

                remaining_requests = worksheet2.findall(selected_employee_id2)
                if not remaining_requests:
                    worksheet2.append_row([selected_employee_id2, "요청 없음", ""])
                
                st.success("요청사항이 삭제되었습니다.")
                time.sleep(1.5)
                st.rerun()
            else:
                st.warning("삭제할 요청사항을 선택해주세요.")
        except Exception as e:
            st.error(f"요청사항 삭제 중 오류 발생: {e}")

# 근무 배정 로직
current_cumulative = {'오전': {}, '오후': {}}

_, last_day = calendar.monthrange(today.year, today.month)
next_month = today.replace(day=1) + relativedelta(months=1)
dates = pd.date_range(start=next_month, end=next_month.replace(day=calendar.monthrange(next_month.year, next_month.month)[1]))
weekdays = [d for d in dates if d.weekday() < 5]
week_numbers = {d.to_pydatetime().date(): (d.day - 1) // 7 + 1 for d in dates}
day_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금'}
df_final = pd.DataFrame(columns=['날짜', '요일', '주차', '시간대', '근무자', '상태', '메모', '색상'])

st.divider()
st.subheader(f"✨ {month_str} 스케줄 배정 수행")
st.write("- 본 페이지에서 배정된 스케줄은 ver1.0로 저장됩니다.")

def parse_date_range(date_str):
    if pd.isna(date_str) or not isinstance(date_str, str) or date_str.strip() == '':
        return []
    date_str = date_str.strip()
    result = []
    if ',' in date_str:
        for single_date in date_str.split(','):
            single_date = single_date.strip()
            try:
                parsed_date = datetime.strptime(single_date, '%Y-%m-%d')
                if parsed_date.weekday() < 5:
                    result.append(single_date)
            except ValueError:
                pass
        return result
    if '~' in date_str:
        try:
            start_date, end_date = date_str.split('~')
            start_date = start_date.strip()
            end_date = end_date.strip()
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            date_list = pd.date_range(start=start, end=end)
            return [d.strftime('%Y-%m-%d') for d in date_list if d.weekday() < 5]
        except ValueError as e:
            pass
            return []
    try:
        parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
        if parsed_date.weekday() < 5:
            return [date_str]
        return []
    except ValueError:
        pass
        return []

def update_worker_status(df, date_str, time_slot, worker, status, memo, color, day_map, week_numbers):
    date_obj = pd.to_datetime(date_str)
    worker_stripped = worker.strip()
    
    existing_indices = df.index[
        (df['날짜'] == date_str) &
        (df['시간대'] == time_slot) &
        (df['근무자'] == worker_stripped)
    ].tolist()

    if existing_indices:
        df.loc[existing_indices, ['상태', '메모', '색상']] = [status, memo, color]
    else:
        new_row = pd.DataFrame([{
            '날짜': date_str,
            '요일': day_map.get(date_obj.weekday(), ''),
            '주차': week_numbers.get(date_obj.date(), 0),
            '시간대': time_slot,
            '근무자': worker_stripped,
            '상태': status,
            '메모': memo,
            '색상': color
        }])
        df = pd.concat([df, new_row], ignore_index=True)
    return df

# 아래 코드로 함수 전체를 교체하세요.
def sync_am_to_pm_exclusions(df_final, active_weekdays, day_map, week_numbers, initial_master_assignments, current_cumulative, weekly_counts):
    """
    [v14 수정]
    오전 근무에서 제외된 근무자를 오후 근무에서도 제외 처리하여 동기화합니다.
    - df_final, current_cumulative, weekly_counts 딕셔너리를 모두 업데이트합니다.
    """
    changed = False
    for date in active_weekdays:
        date_str = date.strftime('%Y-%m-%d')
        date_obj = date.date() # 날짜 객체
        current_week = week_numbers.get(date_obj) # 현재 주차
        
        excluded_am_workers = df_final[
            (df_final['날짜'] == date_str) &
            (df_final['시간대'] == '오전') &
            (df_final['상태'].isin(['대체휴근', '휴근']))
        ]['근무자'].unique()

        for worker in excluded_am_workers:
            pm_record = df_final[
                (df_final['날짜'] == date_str) &
                (df_final['시간대'] == '오후') &
                (df_final['근무자'] == worker)
            ]

            # CASE 1: 기록이 이미 있는 경우
            if not pm_record.empty:
                if pm_record.iloc[0]['상태'] in ['근무', '대체보충', '보충']:
                    df_final = update_worker_status(
                        df_final, date_str, '오후', worker,
                        '휴근', '오전 제외로 인한 오후 제외',
                        '🟣 보라색', day_map, week_numbers
                    )
                    current_cumulative['오후'][worker] = current_cumulative['오후'].get(worker, 0) - 1
                    
                    # ▼▼▼ [핵심 수정] weekly_counts 실시간 업데이트 ▼▼▼
                    if current_week:
                        weekly_counts[worker]['오후'][current_week] = weekly_counts[worker]['오후'].get(current_week, 0) - 1
                    # ▲▲▲ [수정 완료] ▲▲▲
                    
                    changed = True
            # CASE 2: 기록이 없는 경우
            else:
                pm_master_workers = initial_master_assignments.get((date_str, '오후'), set())
                if worker in pm_master_workers:
                    df_final = update_worker_status(
                        df_final, date_str, '오후', worker,
                        '휴근', '오전 제외로 인한 오후 제외',
                        '🟣 보라색', day_map, week_numbers
                    )
                    current_cumulative['오후'][worker] = current_cumulative['오후'].get(worker, 0) - 1
                    
                    # ▼▼▼ [핵심 수정] weekly_counts 실시간 업데이트 ▼▼▼
                    if current_week:
                         weekly_counts[worker]['오후'][current_week] = weekly_counts[worker]['오후'].get(current_week, 0) - 1
                    # ▲▲▲ [수정 완료] ▲▲▲
                    
                    changed = True

    # [수정] weekly_counts 반환
    return df_final, changed, current_cumulative, weekly_counts

def is_worker_already_excluded_with_memo(df_data, date_s, time_s, worker_s):
    worker_records = df_data[
        (df_data['날짜'] == date_s) &
        (df_data['시간대'] == time_s) &
        (df_data['근무자'] == worker_s)
    ]
    if worker_records.empty:
        return False 

    excluded_records = worker_records[worker_records['상태'].isin(['대체휴근', '휴근'])]
    if excluded_records.empty:
        return False 

    return excluded_records['메모'].str.contains('보충 위해 제외됨|인원 초과로 인한 제외|오전 추가제외로 인한 오후 제외', na=False).any()

@st.cache_data(ttl=600, show_spinner=False)
def load_monthly_special_schedules(month_str):
    try:
        client = get_gspread_client()
        spreadsheet = client.open_by_url(st.secrets["google_sheet"]["url"])
        
        target_year = month_str.split('년')[0]
        sheet_name = f"{target_year}년 토요/휴일 스케줄"

        yearly_schedule_sheet = spreadsheet.worksheet(sheet_name)
        yearly_schedule_data = yearly_schedule_sheet.get_all_records()
        df_yearly_schedule = pd.DataFrame(yearly_schedule_data)

        if df_yearly_schedule.empty:
            return pd.DataFrame(), pd.DataFrame()

        target_month_dt = datetime.strptime(month_str, "%Y년 %m월")
        target_month = target_month_dt.month

        df_yearly_schedule['날짜'] = pd.to_datetime(df_yearly_schedule['날짜'])

        df_monthly_schedule = df_yearly_schedule[
            (df_yearly_schedule['날짜'].dt.year == int(target_year)) &
            (df_yearly_schedule['날짜'].dt.month == target_month)
        ].copy()

        df_display = df_monthly_schedule.copy()
        weekday_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
        df_display['날짜'] = df_display['날짜'].apply(
            lambda x: f"{x.month}월 {x.day}일 ({weekday_map[x.weekday()]})"
        )

        return df_monthly_schedule, df_display  

    except gspread.exceptions.WorksheetNotFound:
        target_year = month_str.split('년')[0]
        sheet_name = f"{target_year}년 토요/휴일 스케줄"
        st.error(f"❌ '{sheet_name}' 시트를 찾을 수 없습니다.")
        return pd.DataFrame(), pd.DataFrame()
    except Exception as e:
        st.error(f"토요/휴일 스케줄을 불러오는 중 오류가 발생했습니다: {e}")
        return pd.DataFrame(), pd.DataFrame()

@st.cache_data(ttl=600, show_spinner=False)
def load_closing_days(month_str):
    try:
        client = get_gspread_client()
        spreadsheet = client.open_by_url(st.secrets["google_sheet"]["url"])
        
        target_year = month_str.split('년')[0]
        sheet_name = f"{target_year}년 휴관일"
        
        worksheet = spreadsheet.worksheet(sheet_name)
        data = worksheet.get_all_records()
        df_closing = pd.DataFrame(data)

        if df_closing.empty or "날짜" not in df_closing.columns:
            return [], pd.DataFrame(columns=["날짜"]) 

        df_closing['날짜'] = pd.to_datetime(df_closing['날짜'])
        target_month_dt = datetime.strptime(month_str, "%Y년 %m월")
        
        df_monthly_closing = df_closing[
            df_closing['날짜'].dt.month == target_month_dt.month
        ].copy()

        df_display = df_monthly_closing.copy()
        weekday_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
        df_display['날짜'] = df_display['날짜'].apply(
            lambda x: f"{x.month}월 {x.day}일 ({weekday_map[x.weekday()]})"
        )

        closing_dates_list = df_monthly_closing['날짜'].dt.strftime('%Y-%m-%d').tolist()
        
        return closing_dates_list, df_display

    except gspread.exceptions.WorksheetNotFound:
        st.info(f"ℹ️ '{sheet_name}' 시트를 찾을 수 없어 휴관일을 불러오지 않았습니다.")
        return [], pd.DataFrame(columns=["날짜"])
    except Exception as e:
        st.error(f"휴관일 정보를 불러오는 중 오류가 발생했습니다: {e}")
        return [], pd.DataFrame(columns=["날짜"])

# [★ L1600의 이 함수 전체를 교체하세요 ★]

def transform_schedule_for_checking(df_final_unique, df_excel, month_start, month_end):
    """
    [수정] 배정 확인용 스케줄 데이터를 생성합니다.
    휴가/제외 인원을 포함한 모든 인원이 출력되도록 열 개수를 동적으로 계산합니다.
    
    [★ 버그 수정 ★]
    - L1634의 로직을 L2158의 로직(메모 확인 로직)과 동일하게 수정합니다.
    - '대체보충'의 경우 (상태)가 아닌 (메모)가 저장되도록 수정합니다.
    """
    # [핵심 수정 1] 월 전체에서 일별 최대 인원수 계산
    daily_counts = df_final_unique.groupby(['날짜', '시간대'])['근무자'].nunique().unstack(fill_value=0)
    max_am_workers = int(daily_counts.get('오전', pd.Series([0])).max())
    max_pm_workers = int(daily_counts.get('오후', pd.Series([0])).max())

    # 토요/휴일 스케줄의 최대 인원수도 고려
    if not df_excel.empty:
        # '1'부터 '12'까지의 열이 df_excel에 있는지 확인
        am_cols = [str(i) for i in range(1, 13) if str(i) in df_excel.columns]
        if am_cols:
            weekend_am_counts = df_excel[am_cols].apply(lambda row: row.str.strip().ne('').sum(), axis=1)
            if not weekend_am_counts.empty:
                max_am_workers = max(max_am_workers, weekend_am_counts.max())

    # 최종 열 개수 확정 (최소 12, 4개는 유지)
    max_am_workers = max(max_am_workers, 12)
    max_pm_workers = max(max_pm_workers, 4)

    date_range = pd.date_range(start=month_start, end=month_end)
    date_list = [f"{d.month}월 {d.day}일" for d in date_range]
    weekday_map = {'Mon': '월', 'Tue': '화', 'Wed': '수', 'Thu': '목', 'Fri': '금', 'Sat': '토', 'Sun': '일'}
    weekdays = [weekday_map[d.strftime('%a')] for d in date_range]
    target_year = month_start.year

    # [핵심 수정 2] 동적으로 계산된 열 개수로 컬럼 정의
    columns = ['날짜', '요일'] + \
              [str(i) for i in range(1, max_am_workers + 1)] + \
              ['오전당직(온콜)'] + \
              [f'오후{i}' for i in range(1, max_pm_workers + 1)]
    result_df = pd.DataFrame(columns=columns)

    for date, weekday in zip(date_list, weekdays):
        date_key = datetime.strptime(date, '%m월 %d일').replace(year=target_year).strftime('%Y-%m-%d')
        
        row_data = {'날짜': date, '요일': weekday}

        # 오전/오후 근무자 정보 처리
        for time_slot, max_workers, col_prefix in [('오전', max_am_workers, ''), ('오후', max_pm_workers, '오후')]:
            # 모든 상태의 근무자 정보를 가져옴
            workers_info = df_final_unique[
                (df_final_unique['날짜'] == date_key) &
                (df_final_unique['시간대'] == time_slot)
            ].sort_values(by=['색상_우선순위', '근무자']).to_dict('records')

            for i in range(max_workers):
                col_name = f"{col_prefix}{i+1}" if col_prefix else str(i+1)
                if i < len(workers_info):
                    info = workers_info[i]
                    worker_name = info['근무자']
                    status = info['상태']
                    
                    # --- ▼▼▼ [핵심 수정] L2158 로직 이식 ▼▼▼ ---
                    memo = info.get('메모', '') # 1. 메모 가져오기
                    
                    if status == '대체보충' and pd.notna(memo) and str(memo).strip():
                         # 2. 대체보충이고 메모가 있으면 (메모) 사용
                         row_data[col_name] = f"{worker_name}({memo})"
                    elif status not in ['근무', '당직', '기본']:
                         # 3. 그 외 (휴가, 보충, 대체휴근 등)는 (상태) 사용
                         row_data[col_name] = f"{worker_name}({status})"
                    else:
                         # 4. 기본 근무는 이름만
                         row_data[col_name] = worker_name
                    # --- ▲▲▲ [수정 완료] ▲▲▲ ---
                        
                else:
                    row_data[col_name] = ''

        # 당직 및 주말 정보 처리 (이 부분은 원본 L1645 이후와 동일)
        excel_row = df_excel[df_excel['날짜'] == date]
        if not excel_row.empty:
            # '오전당직(온콜)' 열이 df_excel에 있는지 확인
            if '오전당직(온콜)' in excel_row.columns:
                row_data['오전당직(온콜)'] = excel_row['오전당직(온콜)'].iloc[0]
            
            if weekday in ['토', '일']:
                for i in range(1, max_am_workers + 1):
                    col_str = str(i)
                    if col_str in excel_row.columns and pd.notna(excel_row[col_str].iloc[0]):
                        row_data[col_str] = excel_row[col_str].iloc[0]
                    # else:
                    #     # row_data[col_str] = '' # (이미 위에서 ''로 초기화됨)
                for i in range(1, max_pm_workers + 1):
                    row_data[f'오후{i}'] = ''
        
        # '오전당직(온콜)'이 row_data에 없는 경우(평일)를 대비해 ''로 초기화
        if '오전당직(온콜)' not in row_data:
            row_data['오전당직(온콜)'] = ''

        result_df = pd.concat([result_df, pd.DataFrame([row_data])], ignore_index=True)

    # 누락된 열이 있다면 ''로 채움 (안전장치)
    for col in columns:
        if col not in result_df.columns:
            result_df[col] = ''
            
    # 최종 열 순서 맞추기
    result_df = result_df[columns]

    return result_df

def transform_schedule_data(df, df_excel, month_start, month_end):
    # 모든 상태 포함 (제외, 추가제외 포함)
    df = df[['날짜', '시간대', '근무자', '요일', '상태', '색상', '메모']].copy()
    
    date_range = pd.date_range(start=month_start, end=month_end)
    date_list = [f"{d.month}월 {d.day}일" for d in date_range]
    weekday_list = [d.strftime('%a') for d in date_range]
    weekday_map = {'Mon': '월', 'Tue': '화', 'Wed': '수', 'Thu': '목', 'Fri': '금', 'Sat': '토', 'Sun': '일'}
    weekdays = [weekday_map[w] for w in weekday_list]
    
    target_year = month_start.year

    columns = ['날짜', '요일'] + [str(i) for i in range(1, 13)] + ['오전당직(온콜)'] + [f'오후{i}' for i in range(1, 5)]
    result_df = pd.DataFrame(columns=columns)
    
    for date, weekday in zip(date_list, weekdays):
        date_key = datetime.strptime(date, '%m월 %d일').replace(year=target_year).strftime('%Y-%m-%d')
        date_df = df[df['날짜'] == date_key]
        
        # 오전 근무자 (모든 상태 포함)
        morning_workers = date_df[date_df['시간대'] == '오전'][['근무자', '상태', '색상', '메모']].to_dict('records')
        morning_data = [''] * 12
        for i, worker_info in enumerate(morning_workers[:12]):
            morning_data[i] = worker_info['근무자']
        
        # 오후 근무자 (모든 상태 포함)
        afternoon_workers = date_df[date_df['시간대'] == '오후'][['근무자', '상태', '색상', '메모']].to_dict('records')
        afternoon_data = [''] * 4
        for i, worker_info in enumerate(afternoon_workers[:4]):
            afternoon_data[i] = worker_info['근무자']
        
        if weekday in ['토', '일']: 
            excel_row = df_excel[df_excel['날짜'] == date]
            if not excel_row.empty:
                morning_data = [excel_row[str(i)].iloc[0] if str(i) in excel_row.columns and pd.notna(excel_row[str(i)].iloc[0]) else '' for i in range(1, 13)]
        
        oncall_worker = ''
        excel_row = df_excel[df_excel['날짜'] == date]
        if not excel_row.empty:
            oncall_worker = excel_row['오전당직(온콜)'].iloc[0] if '오전당직(온콜)' in excel_row.columns else ''
        
        row_data = [date, weekday] + morning_data + [oncall_worker] + afternoon_data
        result_df = pd.concat([result_df, pd.DataFrame([row_data], columns=columns)], ignore_index=True)
    
    return result_df

df_cumulative_next = df_cumulative.copy()

initialize_schedule_session_state()

st.write("")
st.markdown(f"**📅 {month_str} 토요/휴일 스케줄**")

df_monthly_schedule, df_display = load_monthly_special_schedules(month_str)

if not df_monthly_schedule.empty:
    st.dataframe(df_display[['날짜', '근무', '당직']], use_container_width=True, hide_index=True)
else:
    st.info(f"ℹ️ '{month_str}'에 해당하는 토요/휴일 스케줄이 없습니다.")

st.write(" ")
st.markdown(f"**📅 {month_str} 휴관일 정보**")

holiday_dates, df_closing_display = load_closing_days(month_str)

if holiday_dates:
    st.write("- 아래 날짜는 근무 배정에서 제외됩니다.")
    
    formatted_dates_list = df_closing_display['날짜'].tolist()
    
    display_string = ", ".join(formatted_dates_list)
    
    st.info(f"➡️ {display_string}")
else:
    st.info(f"ℹ️ {month_str}에는 휴관일이 없습니다.")

names_in_master = set(df_master["이름"].unique().tolist())
names_in_request = set(df_request["이름"].unique().tolist())
all_names = sorted(list(names_in_master.union(names_in_request)))  

def find_afternoon_swap_possibility(worker_to_check, original_date_str, df_final, active_weekdays, target_count_pm, df_supplement_processed, df_request, initial_master_assignments, day_map, week_numbers):
    shortage_dates = []
    original_date = pd.to_datetime(original_date_str).date()

    for date in active_weekdays:
        date_str = date.strftime('%Y-%m-%d')
        if date_str == original_date_str: continue
        
        if week_numbers.get(original_date) != week_numbers.get(date.date()):
            continue

        workers_on_date = df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == '오후') & (df_final['상태'].isin(['근무', '대체보충', '보충']))]['근무자'].unique()
        if len(workers_on_date) < target_count_pm:
            shortage_dates.append(date_str)

    if not shortage_dates:
        return None

    for shortage_date in shortage_dates:
        morning_workers_on_shortage_date = set(df_final[(df_final['날짜'] == shortage_date) & (df_final['시간대'] == '오전') & (df_final['상태'].isin(['근무', '대체보충', '보충']))]['근무자'])
        if worker_to_check not in morning_workers_on_shortage_date:
            continue

        shortage_day_name = day_map.get(pd.to_datetime(shortage_date).weekday())
        supplement_row = df_supplement_processed[df_supplement_processed['시간대'] == f"{shortage_day_name} 오후"]
        if supplement_row.empty: continue
        
        supplement_pool = set()
        for col in supplement_row.columns:
            if col.startswith('보충'):
                for val in supplement_row[col].dropna():
                    supplement_pool.add(val.replace('🔺','').strip())

        if worker_to_check not in supplement_pool:
            continue
        
        if worker_to_check in initial_master_assignments.get((shortage_date, '오후'), set()):
            continue

        no_supplement_req = {r['이름'] for _, r in df_request.iterrows() if shortage_date in parse_date_range(str(r.get('날짜정보'))) and r.get('분류') == '보충 불가(오후)'}
        if worker_to_check in no_supplement_req:
            continue

        return shortage_date
    return None

# 기존 execute_adjustment_pass 함수의 내용을 아래 코드로 전체 교체하세요.

def execute_adjustment_pass(df_final, active_weekdays, time_slot, target_count, initial_master_assignments, df_supplement_processed, df_request, day_map, week_numbers, current_cumulative, df_cumulative, all_names, weekly_counts):
    from collections import defaultdict

    active_weekdays = [pd.to_datetime(date) if isinstance(date, str) else date for date in active_weekdays]
    df_cum_indexed = df_cumulative.set_index('항목').T
    
    # --- scores를 루프 시작 전 '한 번만' 정확히 계산 --- (원본 로직 유지)
    scores = {w: (df_cum_indexed.loc[w, f'{time_slot}누적'] + current_cumulative[time_slot].get(w, 0)) for w in all_names if w in df_cum_indexed.index}

    # 추가 제외 / 보충 로직
    for date in active_weekdays:
        date_str = date.strftime('%Y-%m-%d')
        date_obj = date.date() # 날짜 객체
        current_week = week_numbers.get(date_obj)
        
        # --- ▼▼▼ [핵심 수정 1] '꼭 근무' 포함 ▼▼▼ ---
        current_workers_df = df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == time_slot) & (df_final['상태'].isin(['근무', '대체보충', '보충', '꼭 근무']))]
        current_workers = current_workers_df['근무자'].unique()
        count_diff = len(current_workers) - target_count
        
        # [인원 부족 시 보충]
        if count_diff < 0:
            needed = -count_diff
            day_name = day_map.get(date.weekday())
            supplement_row = df_supplement_processed[df_supplement_processed['시간대'] == f"{day_name} {time_slot}"]
            candidates = []
            if not supplement_row.empty:
                for col in supplement_row.columns:
                    if col.startswith('보충'):
                        # [원본 로직 복원]
                        candidates.extend(val.replace('🔺', '').strip() for val in supplement_row[col].dropna())
            
            unavailable = set(current_workers)
            no_supp = {r['이름'] for _, r in df_request.iterrows() if date_str in parse_date_range(str(r.get('날짜정보'))) and r.get('분류') == f'보충 불가({time_slot})'}
            difficult_supp = {r['이름'] for _, r in df_request.iterrows() if date_str in parse_date_range(str(r.get('날짜정보'))) and r.get('분류') == f'보충 어려움({time_slot})'}
            candidates = [w for w in candidates if w not in unavailable and w not in no_supp]
            
            if time_slot == '오후' and current_week:
                candidates_filtered = []
                for w in candidates:
                    # 현재 주차의 오후 근무 횟수 확인
                    pm_shifts_this_week = weekly_counts.get(w, {}).get('오후', {}).get(current_week, 0)
                    if pm_shifts_this_week < 2:
                        candidates_filtered.append(w)
                candidates = candidates_filtered
            
            if not candidates: continue

            candidates.sort(key=lambda w: (1 if w in difficult_supp else 0, scores.get(w, 0)))

            for worker_to_add in candidates[:needed]:
                df_final = update_worker_status(df_final, date_str, time_slot, worker_to_add, '보충', '인원 부족 (균형 조정)', '🟡 노란색', day_map, week_numbers)
                current_cumulative[time_slot][worker_to_add] = current_cumulative[time_slot].get(worker_to_add, 0) + 1
                
                # ▼▼▼ [수정 3] weekly_counts 실시간 업데이트 ▼▼▼
                if current_week:
                    weekly_counts[worker_to_add][time_slot][current_week] = weekly_counts[worker_to_add][time_slot].get(current_week, 0) + 1
                # ▲▲▲ [수정 3] ▲▲▲
                
                scores[worker_to_add] = scores.get(worker_to_add, 0) + 1

        # [인원 초과 시 제외]
        elif count_diff > 0:
            over_count = count_diff
            must_work = {r['이름'] for _, r in df_request.iterrows() if date_str in parse_date_range(str(r.get('날짜정보'))) and r.get('분류') == f'꼭 근무({time_slot})'}

            for _ in range(over_count):
                # --- ▼▼▼ [핵심 수정 3] '꼭 근무' 포함 ▼▼▼ ---
                current_workers_df = df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == time_slot) & (df_final['상태'].isin(['근무', '대체보충', '보충', '꼭 근무']))]
                potential_removals = [w for w in current_workers_df['근무자'].unique() if w not in must_work]

                if not potential_removals:
                    break 

                if time_slot == '오전':
                    # --- ▼▼▼ [핵심 수정 4] '꼭 근무' 포함 ▼▼▼ ---
                    pm_workers_on_date = set(
                        df_final[
                            (df_final['날짜'] == date_str) & 
                            (df_final['시간대'] == '오후') & 
                            (df_final['상태'].isin(['근무', '대체보충', '보충', '꼭 근무'])) # <-- '꼭 근무' 추가
                        ]['근무자']
                    )
                    potential_removals.sort(
                        key=lambda w: (
                            1 if w in pm_workers_on_date else 0, 
                            -scores.get(w, 0)
                        )
                    )
                
                else: 
                    potential_removals.sort(key=lambda w: scores.get(w, 0), reverse=True)

                worker_to_remove = potential_removals[0]
                df_final = update_worker_status(df_final, date_str, time_slot, worker_to_remove, '휴근', '인원 초과 (실시간 균형 조정)', '🟣 보라색', day_map, week_numbers)

                current_cumulative[time_slot][worker_to_remove] = current_cumulative[time_slot].get(worker_to_remove, 0) - 1
                
                if current_week:
                     weekly_counts[worker_to_remove][time_slot][current_week] = weekly_counts[worker_to_remove][time_slot].get(current_week, 0) - 1

                scores[worker_to_remove] = scores.get(worker_to_remove, 0) - 1

    return df_final, current_cumulative, weekly_counts

from collections import defaultdict

def calculate_weekly_counts(df_final, all_names, week_numbers):
    """지정된 주차 정보에 따라 모든 인원의 주간 오전/오후 근무 횟수를 계산합니다."""
    weekly_counts = {worker: {'오전': defaultdict(int), '오후': defaultdict(int)} for worker in all_names}
    
    for _, row in df_final.iterrows():
        if row['상태'] in ['근무', '대체보충', '보충']:
            try:
                date_obj = pd.to_datetime(row['날짜']).date()
                week = week_numbers.get(date_obj) # .get()으로 안전하게 접근
                if week and row['근무자'] in weekly_counts:
                    weekly_counts[row['근무자']][row['시간대']][week] += 1
            except (KeyError, ValueError):
                continue
    return weekly_counts

def balance_weekly_and_cumulative(
    df_final, 
    active_weekdays_am_sorted, active_weekdays_pm_sorted,
    initial_master_assignments, df_supplement_processed, 
    df_request, day_map, week_numbers, current_cumulative, all_names, df_cumulative,
    weekly_counts 
):
    df_cum_indexed = df_cumulative.set_index('항목').T
    
    for time_slot in ['오전', '오후']:
        
        # --- ▼▼▼ [핵심 수정] 시간대에 맞는 정렬된 날짜 리스트 선택 ▼▼▼ ---
        active_weekdays_to_use = active_weekdays_am_sorted if time_slot == '오전' else active_weekdays_pm_sorted
        # --- ▲▲▲ [핵심 수정] ▲▲▲ ---

        for i in range(50):
            # [수정] 함수 시작 시 weekly_counts를 계산하는 라인 '삭제'
            # (최신 weekly_counts를 인자로 받음)

            scores = {w: (df_cum_indexed.loc[w, f'{time_slot}누적'] + current_cumulative[time_slot].get(w, 0)) for w in all_names if w in df_cum_indexed.index}
            if not scores: break
            
            min_s, max_s = min(scores.values()), max(scores.values())
            worker_scores = sorted(scores.items(), key=lambda item: item[1])
            w_l, s_l = worker_scores[0]
            w_h, s_h = worker_scores[-1]
            
            swap_found_in_iteration = False
            
            for date in active_weekdays: # [수정] active_weekdays_to_use -> active_weekdays
                date_str = date.strftime('%Y-%m-%d')
                date_obj = date.date() # 날짜 객체
                current_week = week_numbers.get(date_obj) # 현재 주차
                
                must_work = {r['이름'] for _, r in df_request.iterrows() if date_str in parse_date_range(str(r.get('날짜정보'))) and r.get('분류') == f'꼭 근무({time_slot})'}
                if w_h in must_work: continue

                # --- ▼▼▼ [핵심 수정] '꼭 근무' 포함하여 확인 ▼▼▼ ---
                is_h_working = not df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == time_slot) & (df_final['근무자'] == w_h) & (df_final['상태'].isin(['근무', '대체보충', '보충', '꼭 근무']))].empty # <-- '꼭 근무' 추가
                # --- ▲▲▲ [수정 완료] ▲▲▲ ---
                if not is_h_working: continue

                s_row = df_supplement_processed[df_supplement_processed['시간대'] == f"{day_map.get(date.weekday())} {time_slot}"]
                can_supp = any(w_l in s_row[col].dropna().str.replace('🔺', '').str.strip().tolist() for col in s_row.columns if col.startswith('보충'))
                if not can_supp: continue
                
                no_supp = {r['이름'] for _, r in df_request.iterrows() if date_str in parse_date_range(str(r.get('날짜정보'))) and r.get('분류') == f'보충 불가({time_slot})'}
                if w_l in no_supp: continue

                if time_slot == '오후':
                    am_workers = set(df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == '오전') & (df_final['상태'].isin(['근무', '대체보충', '보충', '꼭 근무']))]['근무자'])
                    if w_l not in am_workers: continue

                    # ▼▼▼ [핵심 수정] w_l (받는 사람)의 주간 2회 초과 금지 ▼▼▼
                    if current_week:
                        pm_shifts_this_week_for_wl = weekly_counts.get(w_l, {}).get('오후', {}).get(current_week, 0)
                        if pm_shifts_this_week_for_wl >= 2:
                            continue
                            
                is_master = w_l in initial_master_assignments.get((date_str, time_slot), set())
                status, color, memo = ('근무', '기본', '마스터 복귀') if is_master else ('보충', '🟡 노란색', '최종 균형 조정')
                
                # [수정] w_h (주는 사람) 업데이트
                df_final = update_worker_status(df_final, date_str, time_slot, w_h, '휴근', '최종 균형 조정', '🟣 보라색', day_map, week_numbers)
                current_cumulative[time_slot][w_h] = current_cumulative[time_slot].get(w_h, 0) - 1
                if current_week:
                    weekly_counts[w_h][time_slot][current_week] = weekly_counts[w_h][time_slot].get(current_week, 0) - 1
                
                # [수정] w_l (받는 사람) 업데이트
                df_final = update_worker_status(df_final, date_str, time_slot, w_l, status, memo, color, day_map, week_numbers)
                current_cumulative[time_slot][w_l] = current_cumulative[time_slot].get(w_l, 0) + 1
                if current_week:
                    weekly_counts[w_l][time_slot][current_week] = weekly_counts[w_l][time_slot].get(current_week, 0) + 1

                swap_found_in_iteration = True
                break

            if swap_found_in_iteration:
                continue
            else:
                break
        
        else:
            st.warning(f"⚠️ {time_slot} 균형 조정이 최대 반복 횟수({i+1}회)에 도달했습니다.")
    
    # [수정] weekly_counts는 상위에서 관리하므로 반환값에서 제거
    return df_final, current_cumulative

import pandas as pd # <-- 함수 상단에 추가 (혹시 없다면)
import streamlit as st # <-- 함수 상단에 추가 (혹시 없다면)
from collections import Counter # <-- 함수 상단에 추가 (혹시 없다면)

def balance_final_cumulative_with_weekly_check(
    df_final,
    active_weekdays_am_sorted, active_weekdays_pm_sorted,
    df_supplement_processed, df_request, day_map, week_numbers,
    current_cumulative, all_names, df_cumulative, initial_master_assignments,
    df_master,
    weekly_counts 
):
    """
    [진짜 최종 수정본 v12]
    1. '0점+마스터X' 제외 규칙 유지 (균형 조정 비대상).
    2. [핵심 수정] 성공/실패 판단을 '유효 인원 편차' (제외자 제외) 기준으로 변경.
    3. 교체 대상(w_h, w_l)은 '유효 인원'(제외자 제외) 중에서 선정.
    4. 오직 '유효 최고점자 -> 유효 최저점자' 교체만 시도.
    """
    MIN_AM_PER_WEEK = 3
    MIN_PM_PER_WEEK = 1

    # 시간대별 마스터 근무자 목록 계산
    master_workers_am = set()
    master_workers_pm = set()
    if not df_master.empty:
        for _, row in df_master.iterrows():
            worker = row['이름']
            shift_type = row['근무여부'] # 컬럼명 확인
            if shift_type in ['오전', '오전 & 오후']: master_workers_am.add(worker)
            if shift_type in ['오후', '오전 & 오후']: master_workers_pm.add(worker)

    for time_slot in ['오전', '오후']:

        active_weekdays_to_use = active_weekdays_am_sorted if time_slot == '오전' else active_weekdays_pm_sorted
        master_workers_this_slot = master_workers_am if time_slot == '오전' else master_workers_pm

        for i in range(50): # 안전장치 50회
            # 1. '바로 지금' 시점의 실시간 누적 점수 계산 (전체 인원)
            df_cum_indexed = df_cumulative.set_index('항목').T
            scores = {w: (df_cum_indexed.loc[w, f'{time_slot}누적'] + current_cumulative[time_slot].get(w, 0)) for w in all_names if w in df_cum_indexed.index}
            if not scores: break
            
            # 2. '실제' 전체 편차 계산 (로그 출력용)
            all_worker_scores_sorted = sorted(scores.items(), key=lambda item: item[1])
            if not all_worker_scores_sorted: break
            true_min_w, true_min_s = all_worker_scores_sorted[0]
            true_max_w, true_max_s = all_worker_scores_sorted[-1]
            current_true_diff = true_max_s - true_min_s # 실제 전체 편차

            # 3. 균형 조정 대상 외 인원 식별 (v10과 동일)
            excluded_workers = set()
            for w, s in scores.items():
                if s == 0 and w not in master_workers_this_slot:
                    excluded_workers.add(w)

            # 4. '유효한' 점수표 생성 및 '유효 편차' 계산 (v10과 동일)
            valid_scores = {w: s for w, s in scores.items() if w not in excluded_workers}
            
            # 5. [수정] 유효 대상이 1명 이하면 조정 불가
            if not valid_scores or len(valid_scores) < 2: 
                 st.info(f"ℹ️ [{time_slot}] 균형 조정을 고려할 유효 대상 인원이 부족합니다.")
                 # 실패 메시지 출력 전에 실제 편차 확인 (유효 대상이 없어도 전체 편차가 2 이하일 수 있음)
                 if current_true_diff > 2:
                      st.error(f"⚠️ [{time_slot}] 최종 균형 조정 중단: 유효 대상 부족. (현재 전체 편차: {current_true_diff})")
                 # (유효 대상이 없지만, 전체 편차가 2 이하면? 이미 v11의 맨 위에서 걸러졌어야 함. 
                 #  하지만 v12에서는 여기서 걸러야 함. -> [수정] 성공 조건도 여기서 체크)
                 elif current_true_diff <= 2:
                      excluded_info = f" - (균형 조정 제외: {', '.join(sorted(excluded_workers))})" if excluded_workers else ""
                      st.success(f"✅ [{time_slot}] 최종 누적 편차 2 이하 달성! (전체 편차: {current_true_diff}){excluded_info}")
                 break # i 루프 중단

            valid_worker_scores_sorted = sorted(valid_scores.items(), key=lambda item: item[1])
            min_w_valid, min_s_valid = valid_worker_scores_sorted[0]     # 유효 최저점
            max_w_valid, max_s_valid = valid_worker_scores_sorted[-1] # 유효 최고점
            current_valid_diff = max_s_valid - min_s_valid # '유효 편차'

            # --- ▼▼▼ [핵심 수정] 성공 조건: '유효 편차' 기준 ▼▼▼ ---
            # 6. 목표 달성 확인: '유효 편차'가 2 이하이면 성공!
            if current_valid_diff <= 2:
                # 성공 메시지에는 '유효 편차'와 '전체 편차'를 모두 표시
                excluded_info = f" - (균형 조정 제외: {', '.join(sorted(excluded_workers))})" if excluded_workers else ""
                st.success(f"✅ [{time_slot}] 최종 누적 편차 2 이하 달성! (유효 편차: {current_valid_diff}, 전체 편차: {current_true_diff}){excluded_info}")
                break # i 루프 중단
            # --- ▲▲▲ 성공 조건 수정 완료 ▲▲▲ ---

            # 7. [타겟 1] w_l (받는 사람): '유효 최저점자'로 고정
            w_l, s_l = min_w_valid, min_s_valid

            # 8. [타겟 2] w_h (주는 사람): '유효 최고점자'로 고정
            w_h, s_h = max_w_valid, max_s_valid

            # 9. w_h 유효성 검사: '유효 최고점자'가 교체할 근무가 있는가?
            has_shifts_to_give = df_final[
                (df_final['시간대'] == time_slot) &
                (df_final['근무자'] == w_h) &
                (df_final['상태'].isin(['근무', '대체보충', '보충', '꼭 근무']))
            ].shape[0] > 0

            # 9-1. 유효 최고점자가 교체할 근무가 없으면 포기 -> 중단!
            if not has_shifts_to_give:
                # 실패 메시지에는 '실제 전체 편차' 사용
                st.error(f"⚠️ [{time_slot}] 최종 균형 조정 중단: 유효 최고점자({w_h}, {s_h}회)가 교체할 근무가 없어 조정 불가. (현재 전체 편차: {current_true_diff})")
                break # i 루프 중단

            # 10. 교체 지점 탐색 (오직 유효 w_h -> 유효 w_l 만 시도)
            swap_found_this_pair = False
            for date in active_weekdays_to_use:
                date_str = date.strftime('%Y-%m-%d')
                date_obj = date.date() # 날짜 객체
                current_week = week_numbers.get(date_obj) # 현재 주차

                # (조건 1) w_h가 이 날 근무 중인가?
                is_working_df = df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == time_slot) & (df_final['근무자'] == w_h) & (df_final['상태'].isin(['근무', '대체보충', '보충', '꼭 근무']))]
                if is_working_df.empty: continue

                # (조건 2) w_l이 이 날 보충 가능한가?
                is_already_working = not df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == time_slot) & (df_final['근무자'] == w_l)].empty
                if is_already_working: continue
                no_supp_req = {r['이름'] for _, r in df_request.iterrows() if date_str in parse_date_range(str(r.get('날짜정보'))) and r.get('분류') == f'보충 불가({time_slot})'}
                if w_l in no_supp_req: continue
                day_name = day_map.get(date.weekday())
                supplement_row = df_supplement_processed[df_supplement_processed['시간대'] == f"{day_name} {time_slot}"]
                can_supplement = False
                if not supplement_row.empty:
                     for col in supplement_row.columns:
                         if col.startswith('보충'):
                             if w_l in [w.replace('🔺','').strip() for w in supplement_row[col].dropna()]:
                                 can_supplement = True; break
                if not can_supplement: continue

                # (조건 3) [오후 전용] w_l이 오전에 근무 중인가?
                if time_slot == '오후':
                    am_workers = set(df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == '오전') & (df_final['상태'].isin(['근무', '대체보충', '보충', '꼭 근무']))]['근무자'])
                    if w_l not in am_workers: continue

                    # ▼▼▼ [핵심 수정] w_l (받는 사람)의 주간 2회 초과 금지 ▼▼▼
                    if current_week:
                        pm_shifts_this_week_for_wl = weekly_counts.get(w_l, {}).get('오후', {}).get(current_week, 0)
                        if pm_shifts_this_week_for_wl >= 2:
                            continue
                    
                # 11. 교체 실행!
                # [수정] w_h (주는 사람) 업데이트
                df_final = update_worker_status(df_final, date_str, time_slot, w_h, '휴근', '최종 누적 균형 조정', '🟣 보라색', day_map, week_numbers)
                current_cumulative[time_slot][w_h] = current_cumulative[time_slot].get(w_h, 0) - 1
                if current_week:
                    weekly_counts[w_h][time_slot][current_week] = weekly_counts[w_h][time_slot].get(current_week, 0) - 1

                # [수정] w_l (받는 사람) 업데이트
                master_workers_on_date = initial_master_assignments.get((date_str, time_slot), set())
                status_for_wl = '근무' if w_l in master_workers_on_date else '보충'
                color_for_wl = '기본' if status_for_wl == '근무' else '🟡 노란색'
                memo_for_wl = '마스터 복귀 (균형 조정)' if status_for_wl == '근무' else '최종 누적 균형 조정'
                df_final = update_worker_status(df_final, date_str, time_slot, w_l, status_for_wl, memo_for_wl, color_for_wl, day_map, week_numbers)
                current_cumulative[time_slot][w_l] = current_cumulative[time_slot].get(w_l, 0) + 1
                if current_week:
                    weekly_counts[w_l][time_slot][current_week] = weekly_counts[w_l][time_slot].get(current_week, 0) + 1

                swap_found_this_pair = True
                break

            # 12. 교체 대상을 못 찾았다면, 최종 중단
            if not swap_found_this_pair:
                # 실패 메시지에도 '실제 전체 편차' 사용
                st.error(f"⚠️ [{time_slot}] 최종 균형 조정 중단: 최고점자({w_h})와 최저점자({w_l}) 간 교체 가능한 날짜를 찾지 못했습니다. (현재 전체 편차: {current_true_diff})")
                break # 'i' 루프 중단

        else: # for문이 break 없이 50회를 모두 돌았다면
            st.warning(f"⚠️ [{time_slot}] 최종 균형 조정이 최대 반복 횟수({i+1}회)에 도달했습니다.")

    return df_final, current_cumulative

df_cumulative_next = df_cumulative.copy()

initialize_schedule_session_state()

st.divider()
# 1단계: 메인 배정 실행 버튼
if st.button("🚀 스케줄 배정 수행", type="primary", use_container_width=True, disabled=st.session_state.get("show_confirmation_warning", False)):
    gc = get_gspread_client()
    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
    latest_version = find_latest_schedule_version(sheet, month_str)

    # 이미 버전이 존재하면 확인 단계로 넘어감
    if latest_version:
        st.session_state.show_confirmation_warning = True
        st.session_state.latest_existing_version = latest_version
        st.rerun()
    # 버전이 없으면 바로 배정 실행
    else:
        st.session_state.assigned = True
        st.session_state.assignment_results = None
        st.session_state.request_logs, st.session_state.swap_logs, st.session_state.adjustment_logs, st.session_state.oncall_logs = [], [], [], []
        
        # --- ▼▼▼ [핵심 수정] ▼▼▼ ---
        st.session_state.editor_has_changes = False # 1. 수정 플래그 리셋
        st.session_state.editor_key_version += 1 # 2. 에디터 키 버전을 올려 강제 리셋
        # --- ▲▲▲ [수정 완료] ▲▲▲ ---
            
        st.rerun()

# 2단계: 확인 경고 및 최종 실행 UI
if st.session_state.get("show_confirmation_warning", False):
    latest_version = st.session_state.get("latest_existing_version", "알 수 없는 버전")
    
    # 정규식을 사용하여 'verX.X' 부분만 추출
    version_match = re.search(r'(ver\s*\d+\.\d+)', latest_version)
    version_str = version_match.group(1) if version_match else latest_version
    
    st.warning(f"⚠️ **이미 '{version_str}' 버전이 존재합니다.**\n\n새로운 'ver1.0' 스케줄을 생성하시더라도 {version_str}은 계속 남아있습니다. 계속하시겠습니까?")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("✅ 네, 새로운 ver1.0으로 배정을 실행합니다.", use_container_width=True, type="primary"):
            st.session_state.assigned = True
            st.session_state.show_confirmation_warning = False
            st.session_state.assignment_results = None
            st.session_state.request_logs, st.session_state.swap_logs, st.session_state.adjustment_logs, st.session_state.oncall_logs = [], [], [], []
            
            # --- ▼▼▼ [핵심 수정] ▼▼▼ ---
            st.session_state.editor_has_changes = False # 1. 수정 플래그 리셋
            st.session_state.editor_key_version += 1 # 2. 에디터 키 버전을 올려 강제 리셋
            # --- ▲▲▲ [수정 완료] ▲▲▲ ---
                
            st.rerun()
    with col2:
        if st.button("❌ 아니요, 취소합니다.", use_container_width=True):
            st.session_state.show_confirmation_warning = False
            st.rerun()

if st.session_state.get('assigned', False):

    if st.session_state.get('assignment_results') is None:
        with st.spinner("근무 배정 중..."):
            st.session_state.request_logs = []
            st.session_state.swap_logs = []
            st.session_state.adjustment_logs = []
            st.session_state.oncall_logs = []
                    
            time.sleep(1)
            
            df_monthly_schedule, df_display = load_monthly_special_schedules(month_str)

            special_schedules = []
            if not df_monthly_schedule.empty:
                for index, row in df_monthly_schedule.iterrows():
                    date_str = row['날짜'].strftime('%Y-%m-%d')
                    oncall_person = row['당직']
                    workers_str = row.get('근무', '')
                    
                    if workers_str and isinstance(workers_str, str):
                        workers_list = [name.strip() for name in workers_str.split(',')]
                    else:
                        workers_list = []
                    
                    special_schedules.append((date_str, workers_list, oncall_person))

            df_final = pd.DataFrame(columns=['날짜', '요일', '주차', '시간대', '근무자', '상태', '메모', '색상'])
            month_dt = datetime.strptime(month_str, "%Y년 %m월")
            _, last_day = calendar.monthrange(month_dt.year, month_dt.month) 
            all_month_dates = pd.date_range(start=month_dt, end=month_dt.replace(day=last_day))
            weekdays = [d for d in all_month_dates if d.weekday() < 5]
            active_weekdays = [d for d in weekdays if d.strftime('%Y-%m-%d') not in holiday_dates]
            day_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}

            # --- ✨ 주차 계산 로직 변경 ---
            # 1. 월 내 모든 날짜의 ISO 주차 번호(연간 기준, 월요일 시작)를 중복 없이 구합니다.
            iso_weeks_in_month = sorted(list(set(d.isocalendar()[1] for d in all_month_dates)))
            
            # 2. ISO 주차 번호를 해당 월의 1, 2, 3... 주차로 매핑하는 사전을 만듭니다.
            # 예: {35주차: 1, 36주차: 2, 37주차: 3, ...}
            iso_to_monthly_week_map = {iso_week: i + 1 for i, iso_week in enumerate(iso_weeks_in_month)}
            
            # 3. 최종적으로 모든 날짜에 대해 '월 기준 주차'를 할당합니다.
            week_numbers = {d.to_pydatetime().date(): iso_to_monthly_week_map[d.isocalendar()[1]] for d in all_month_dates}
            # --- 로직 변경 끝 ---

            initial_master_assignments = {}
            for date in active_weekdays:
                date_str, day_name, week_num = date.strftime('%Y-%m-%d'), day_map[date.weekday()], week_numbers[date.date()]
                for ts in ['오전', '오후']:
                    shift_key, base_workers = f"{day_name} {ts}", set()
                    shift_row = df_shift_processed[df_shift_processed['시간대'] == shift_key]
                    if not shift_row.empty:
                        for col in shift_row.columns[1:]:
                            worker_info = shift_row[col].values[0]
                            if pd.notna(worker_info):
                                worker_name = str(worker_info).split('(')[0].strip()
                                if '(' in str(worker_info) and f'{week_num}주' in str(worker_info):
                                    base_workers.add(worker_name)
                                elif '(' not in str(worker_info):
                                    base_workers.add(worker_name)
                    initial_master_assignments[(date_str, ts)] = base_workers
            
            # --- ▼▼▼ [핵심 수정] 오전/오후 마스터 수에 따라 별도의 날짜 리스트 2개 생성 ▼▼▼ ---
            # st.info("🔄 오전/오후 마스터 수를 기준으로 2개의 날짜 처리 순서를 생성합니다...")
            
            # 1. 오전 난이도 계산
            date_am_master_counts = {}
            for date in active_weekdays: # (원본 시간순 리스트)
                date_str = date.strftime('%Y-%m-%d')
                am_masters = initial_master_assignments.get((date_str, '오전'), set())
                date_am_master_counts[date] = len(am_masters)
            
            # 2. 오후 난이도 계산
            date_pm_master_counts = {}
            for date in active_weekdays:
                date_str = date.strftime('%Y-%m-%d')
                pm_masters = initial_master_assignments.get((date_str, '오후'), set())
                date_pm_master_counts[date] = len(pm_masters)

            # 3. '오전용' 날짜 리스트 생성 (마스터 적은 날짜 우선)
            active_weekdays_am_sorted = sorted(active_weekdays, key=lambda d: date_am_master_counts.get(d, 999))
            # 4. '오후용' 날짜 리스트 생성 (마스터 적은 날짜 우선)
            active_weekdays_pm_sorted = sorted(active_weekdays, key=lambda d: date_pm_master_counts.get(d, 999))

            # 재정렬된 순서 로그 출력 (확인용)
            am_log = [f"{d.strftime('%-m/%d')}({date_am_master_counts.get(d, 'N/A')}명)" for d in active_weekdays_am_sorted[:5]]
            pm_log = [f"{d.strftime('%-m/%d')}({date_pm_master_counts.get(d, 'N/A')}명)" for d in active_weekdays_pm_sorted[:5]]
            # st.info(f"✨ 오전 처리 순서 (상위 5개): {', '.join(am_log)} ...")
            # st.info(f"✨ 오후 처리 순서 (상위 5개): {', '.join(pm_log)} ...")
            # time.sleep(1) # 로그를 볼 수 있도록 잠시 대기
            # --- ▲▲▲ [핵심 수정 완료] ▲▲▲ ---

            current_cumulative = {'오전': {}, '오후': {}}
            weekly_counts = calculate_weekly_counts(df_final, all_names, week_numbers)

            time_slot_am = '오전'
            target_count_am = 12

            # 오전 초기 배정
            for date in active_weekdays_am_sorted: # <-- [유지] 오전 정렬 리스트 사용
                date_str = date.strftime('%Y-%m-%d')
                requests_on_date = df_request[df_request['날짜정보'].apply(lambda x: date_str in parse_date_range(str(x)))]
                vacationers = set(requests_on_date[requests_on_date['분류'].isin(['휴가', '학회'])]['이름'].tolist())
                base_workers = initial_master_assignments.get((date_str, time_slot_am), set())
                must_work = set(requests_on_date[requests_on_date['분류'] == f'꼭 근무({time_slot_am})']['이름'].tolist())
                final_workers = (base_workers - vacationers) | (must_work - vacationers)
                
                for worker in final_workers:
                    # [핵심] '꼭 근무' 요청자는 '꼭 근무' 상태로, 나머지는 '근무' 상태로 저장
                    status = '꼭 근무' if worker in must_work else '근무'
                    color = '🟠 주황색' if worker in must_work else '기본'
                    df_final = update_worker_status(df_final, date_str, time_slot_am, worker, status, '', color, day_map, week_numbers)
                
                weekday_map_korean = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}

                # [유지] 휴가자 처리 로직
                for vac in (vacationers & base_workers):
                    if vac in final_workers: continue # '꼭 근무'가 우선
                        
                    korean_day = weekday_map_korean[date.weekday()]
                    log_date = f"{date.strftime('%-m월 %-d일')} ({korean_day})"
                    reason_series = requests_on_date[(requests_on_date['이름'] == vac) & (requests_on_date['분류'].isin(['휴가', '학회']))]['분류']
                    reason = reason_series.iloc[0] if not reason_series.empty else "휴가"
                    
                    st.session_state.request_logs.append(f"• {log_date} {vac} - {reason}로 인한 제외")
                    df_final = update_worker_status(df_final, date_str, time_slot_am, vac, reason, f'{reason}로 인한 제외', '🔴 빨간색', day_map, week_numbers)

            weekly_counts = calculate_weekly_counts(df_final, all_names, week_numbers)
            # 오전 배정 후 동기화
            # [수정] weekly_counts 전달 및 반환
            df_final, changed, current_cumulative, weekly_counts = sync_am_to_pm_exclusions(df_final, active_weekdays_am_sorted, day_map, week_numbers, initial_master_assignments, current_cumulative, weekly_counts) 
            
            # 오전 균형 맞추기 (execute_adjustment_pass)
            df_before_pass = df_final.copy()
            # [수정] weekly_counts 전달 및 반환
            df_final, current_cumulative, weekly_counts = execute_adjustment_pass(
                df_final, active_weekdays_am_sorted, time_slot_am, target_count_am, initial_master_assignments,
                df_supplement_processed, df_request, day_map, week_numbers, current_cumulative, df_cumulative, all_names,
                weekly_counts 
            )
            
            # 오전 조정 후 동기화
            # [수정] weekly_counts 전달 및 반환
            df_final, changed, current_cumulative, weekly_counts = sync_am_to_pm_exclusions(df_final, active_weekdays_am_sorted, day_map, week_numbers, initial_master_assignments, current_cumulative, weekly_counts) 

            time_slot_pm = '오후'
            target_count_pm = 4

            # 오후 초기 배정
            for date in active_weekdays_pm_sorted: # <-- [유지] 오후 정렬 리스트 사용
                date_str = date.strftime('%Y-%m-%d')
                # [수정] 오전 근무자 셀 때 '꼭 근무' 포함
                morning_workers = set(df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == '오전') & (df_final['상태'].isin(['근무', '대체보충', '보충', '꼭 근무']))]['근무자'])
                requests_on_date = df_request[df_request['날짜정보'].apply(lambda x: date_str in parse_date_range(str(x)))]
                vacationers = set(requests_on_date[requests_on_date['분류'].isin(['휴가', '학회'])]['이름'].tolist())
                base_workers = initial_master_assignments.get((date_str, time_slot_pm), set())
                must_work = set(requests_on_date[requests_on_date['분류'] == f'꼭 근무({time_slot_pm})']['이름'].tolist())
                
                eligible_workers = morning_workers | must_work
                final_workers = (base_workers & eligible_workers) - vacationers | must_work
                
                for worker in final_workers:
                    # [핵심] '꼭 근무' 요청자는 '꼭 근무' 상태로, 나머지는 '근무' 상태로 저장
                    status = '꼭 근무' if worker in must_work else '근무'
                    color = '🟠 주황색' if worker in must_work else '기본'
                    df_final = update_worker_status(df_final, date_str, time_slot_pm, worker, status, '', color, day_map, week_numbers)
                
                # [유지] 오후 휴가자 처리 로직
                for vac in (vacationers & base_workers):
                    if vac in final_workers: continue # '꼭 근무'가 우선

                    existing_record = df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == time_slot_pm) & (df_final['근무자'] == vac)]
                    if not existing_record.empty and existing_record.iloc[0]['상태'] not in ['근무', '기본']:
                         continue
                    
                    reason_series = requests_on_date[(requests_on_date['이름'] == vac) & (requests_on_date['분류'].isin(['휴가', '학회']))]['분류']
                    reason = reason_series.iloc[0] if not reason_series.empty else "휴가"
                    
                    df_final = update_worker_status(df_final, date_str, time_slot_pm, vac, reason, f'{reason}로 제외', '🔴 빨간색', day_map, week_numbers)

            # ▼▼▼ [핵심 수정] 오후 초기 배정 후, 주간 횟수를 즉시 재계산 ▼▼▼
            # (이 코드가 없으면, execute_adjustment_pass가 마스터 횟수를 0으로 착각함)
            weekly_counts = calculate_weekly_counts(df_final, all_names, week_numbers)
            # ▲▲▲ [수정 완료] ▲▲▲

            # 오후 배정 후 동기화
            # [수정] weekly_counts 전달 및 반환
            df_final, changed, current_cumulative, weekly_counts = sync_am_to_pm_exclusions(df_final, active_weekdays_pm_sorted, day_map, week_numbers, initial_master_assignments, current_cumulative, weekly_counts)
            
            # 오후 조정 패스
            # [수정] weekly_counts 전달 및 반환
            df_final, current_cumulative, weekly_counts = execute_adjustment_pass(
                df_final, active_weekdays_pm_sorted, time_slot_pm, target_count_pm, initial_master_assignments,
                df_supplement_processed, df_request, day_map, week_numbers, current_cumulative, df_cumulative, all_names,
                weekly_counts 
            )

            # [수정] 최종 균형 맞추기 전, weekly_counts를 한 번 더 최신화
            weekly_counts = calculate_weekly_counts(df_final, all_names, week_numbers)

            df_final, current_cumulative = balance_weekly_and_cumulative(
                df_final, 
                active_weekdays_am_sorted, active_weekdays_pm_sorted, 
                initial_master_assignments, df_supplement_processed,
                df_request, day_map, week_numbers, current_cumulative, all_names,
                df_cumulative,
                weekly_counts # [수정] weekly_counts 전달
            )

            # [수정] 진짜 최종 균형 맞추기 전, weekly_counts를 한 번 더 최신화
            weekly_counts = calculate_weekly_counts(df_final, all_names, week_numbers)

            df_final, current_cumulative = balance_final_cumulative_with_weekly_check(
                df_final,
                active_weekdays_am_sorted, active_weekdays_pm_sorted,
                df_supplement_processed, df_request,
                day_map, week_numbers, current_cumulative, all_names, df_cumulative,
                initial_master_assignments,
                df_master,
                weekly_counts # [수정] weekly_counts 전달
            )

            df_final = replace_adjustments(df_final)

            df_final_unique_sorted = df_final.sort_values(by=['날짜', '시간대', '근무자']).drop_duplicates(
                subset=['날짜', '시간대', '근무자'], keep='last'
            ).copy()

            # 대체 로그 생성
            df_replacements = df_final_unique_sorted[
                df_final_unique_sorted['상태'].isin(['대체보충', '대체휴근'])
            ].copy()
            df_replacements['주차'] = df_replacements['날짜'].apply(
                lambda x: week_numbers.get(pd.to_datetime(x).date())
            )

            weekly_swap_dates = {}
            for (week, worker, time_slot), group in df_replacements.groupby(['주차', '근무자', '시간대']):
                dates_excluded = sorted(group[group['상태'] == '대체휴근']['날짜'].tolist())
                dates_supplemented = sorted(group[group['상태'] == '대체보충']['날짜'].tolist())

                if dates_excluded and dates_supplemented:
                    key = (week, worker, time_slot)
                    weekly_swap_dates[key] = {
                        '제외일': dates_excluded,
                        '보충일': dates_supplemented
                    }
                    
                    # 메모 업데이트
                    memo_for_exclusion = f"{', '.join([pd.to_datetime(d).strftime('%-m월 %-d일') for d in dates_supplemented])}일과 대체"
                    memo_for_supplement = f"{', '.join([pd.to_datetime(d).strftime('%-m월 %-d일') for d in dates_excluded])}일과 대체"

                    df_final_unique_sorted.loc[
                        (df_final_unique_sorted['근무자'] == worker) &
                        (df_final_unique_sorted['시간대'] == time_slot) &
                        (df_final_unique_sorted['날짜'].isin(dates_excluded)), '메모'
                    ] = memo_for_exclusion

                    df_final_unique_sorted.loc[
                        (df_final_unique_sorted['근무자'] == worker) &
                        (df_final_unique_sorted['시간대'] == time_slot) &
                        (df_final_unique_sorted['날짜'].isin(dates_supplemented)), '메모'
                    ] = memo_for_supplement

            # 로그 생성
            st.session_state.swap_logs, st.session_state.adjustment_logs = [], []
            weekday_map_korean = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}

            # 대체 로그
            for (week, worker, time_slot), swap_info in weekly_swap_dates.items():
                excluded_dates_str = [pd.to_datetime(d).strftime('%-m월 %-d일') for d in sorted(swap_info['제외일'])]
                supplemented_dates_str = [pd.to_datetime(d).strftime('%-m월 %-d일') for d in sorted(swap_info['보충일'])]
                log_message = f"• {worker} ({time_slot}): {', '.join(excluded_dates_str)}(대체 제외) ➔ {', '.join(supplemented_dates_str)}(대체 보충)"
                if log_message not in st.session_state.swap_logs:
                    st.session_state.swap_logs.append(log_message)

            # 추가 보충/제외 로그
            for _, row in df_final_unique_sorted.iterrows():
                if row['상태'] in ['보충', '휴근']:
                    date_obj = pd.to_datetime(row['날짜'])
                    log_date_info = f"{date_obj.strftime('%-m월 %-d일')} ({weekday_map_korean[date_obj.weekday()]}) {row['시간대']}"
                    if row['상태'] == '휴근':
                        st.session_state.adjustment_logs.append(f"• {log_date_info} {row['근무자']} - {row['메모'] or '인원 초과'}로 추가 제외")
                    elif row['상태'] == '보충':
                        st.session_state.adjustment_logs.append(f"• {log_date_info} {row['근무자']} - {row['메모'] or '인원 부족'}으로 추가 보충")
                        
            # 모든 로그를 날짜 기준으로 정렬합니다.
            st.session_state.request_logs.sort(key=get_sort_key)
            st.session_state.swap_logs.sort(key=get_sort_key)
            st.session_state.adjustment_logs.sort(key=get_sort_key)          
            st.session_state.request_logs.sort(key=get_sort_key)
            st.session_state.swap_logs.sort(key=get_sort_key)
            st.session_state.adjustment_logs.sort(key=get_sort_key)

            df_cumulative_next = df_cumulative.copy()  # 인덱스 설정 제거
            for worker, count in current_cumulative.get('오전', {}).items():
                if worker not in df_cumulative_next.columns:
                    df_cumulative_next[worker] = 0  # 새로운 근무자 열 추가
                if '오전누적' not in df_cumulative_next['항목'].values:
                    new_row = pd.DataFrame([[0] * len(df_cumulative_next.columns)], columns=df_cumulative_next.columns)
                    new_row['항목'] = '오전누적'
                    df_cumulative_next = pd.concat([df_cumulative_next, new_row], ignore_index=True)
                df_cumulative_next.loc[df_cumulative_next['항목'] == '오전누적', worker] += count

            for worker, count in current_cumulative.get('오후', {}).items():
                if worker not in df_cumulative_next.columns:
                    df_cumulative_next[worker] = 0  # 새로운 근무자 열 추가
                if '오후누적' not in df_cumulative_next['항목'].values:
                    new_row = pd.DataFrame([[0] * len(df_cumulative_next.columns)], columns=df_cumulative_next.columns)
                    new_row['항목'] = '오후누적'
                    df_cumulative_next = pd.concat([df_cumulative_next, new_row], ignore_index=True)
                df_cumulative_next.loc[df_cumulative_next['항목'] == '오후누적', worker] += count

            if special_schedules:
                for date_str, workers, oncall in special_schedules:
                    if not df_final.empty: df_final = df_final[df_final['날짜'] != date_str].copy()
                    for worker in workers:
                        df_final = update_worker_status(df_final, date_str, '오전', worker, '근무', '', '특수근무색', day_map, week_numbers)

            color_priority = {'🟠 주황색': 0, '🟢 초록색': 1, '🟡 노란색': 2, '기본': 3, '🔴 빨간색': 4, '🔵 파란색': 5, '🟣 보라색': 6, '특수근무색': -1}
            df_final['색상_우선순위'] = df_final['색상'].map(color_priority)
            df_final_unique = df_final.sort_values(by=['날짜', '시간대', '근무자', '색상_우선순위']).drop_duplicates(subset=['날짜', '시간대', '근무자'], keep='last')

            all_month_dates = pd.date_range(start=month_dt, end=month_dt.replace(day=last_day))
            weekdays = [d for d in all_month_dates if d.weekday() < 5]
            active_weekdays = [d for d in weekdays if d.strftime('%Y-%m-%d') not in holiday_dates]
            day_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
            week_numbers = {d.to_pydatetime().date(): (d.day - 1) // 7 + 1 for d in all_month_dates}

            df_schedule = pd.DataFrame({
                '날짜': [d.strftime('%Y-%m-%d') for d in all_month_dates], 
                '요일': [day_map.get(d.weekday()) for d in all_month_dates],
                '날짜_표시': [f"{d.month}월 {d.day}일" for d in all_month_dates] # <-- 이 줄이 추가되었습니다.
            })
            worker_counts_all = df_final_unique.groupby(['날짜', '시간대'])['근무자'].nunique().unstack(fill_value=0)
            max_morning_workers = int(worker_counts_all.get('오전', pd.Series(data=0)).max())
            max_afternoon_workers = int(worker_counts_all.get('오후', pd.Series(data=0)).max())
            columns = ['날짜', '요일'] + [str(i) for i in range(1, max_morning_workers + 1)] + [''] + ['오전당직(온콜)'] + [f'오후{i}' for i in range(1, max_afternoon_workers + 1)]
            df_excel = pd.DataFrame(index=df_schedule.index, columns=columns)

            for idx, row in df_schedule.iterrows():
                date = row['날짜']
                date_obj = datetime.strptime(date, '%Y-%m-%d')
                df_excel.at[idx, '날짜'] = f"{date_obj.month}월 {date_obj.day}일"
                df_excel.at[idx, '요일'] = row['요일']
                df_excel.fillna("", inplace=True)
                
                morning_workers_for_excel = df_final_unique[(df_final_unique['날짜'] == date) & (df_final_unique['시간대'] == '오전')]
                morning_workers_for_excel_sorted = morning_workers_for_excel.sort_values(by=['색상_우선순위', '근무자'])['근무자'].tolist()
                for i, worker_name in enumerate(morning_workers_for_excel_sorted, 1):
                    if i <= max_morning_workers: df_excel.at[idx, str(i)] = worker_name
                
                afternoon_workers_for_excel = df_final_unique[(df_final_unique['날짜'] == date) & (df_final_unique['시간대'] == '오후')]
                afternoon_workers_for_excel_sorted = afternoon_workers_for_excel.sort_values(by=['색상_우선순위', '근무자'])['근무자'].tolist()
                for i, worker_name in enumerate(afternoon_workers_for_excel_sorted, 1):
                    if i <= max_afternoon_workers: df_excel.at[idx, f'오후{i}'] = worker_name
                
                for special_date, workers, oncall in special_schedules:
                    if date == special_date:
                        workers_padded = workers[:10] + [''] * (10 - len(workers[:10]))
                        for i in range(1, 11): df_excel.at[idx, str(i)] = workers_padded[i-1]
                        df_excel.at[idx, '오전당직(온콜)'] = oncall if oncall != "당직 없음" else ''

            ### 시작: 오전당직 배정 로직 ###
            df_cum_indexed = df_cumulative.set_index('항목')
            
            # --- ▼▼▼ [핵심 수정 1] 'oncall_targets'가 0회 목표자도 포함하도록 수정 ▼▼▼ ---
            all_workers_in_cum = [col for col in df_cumulative.columns if col != '항목']
            oncall_targets = {}
            oncall_live_counts = {}
            if '오전당직누적' in df_cum_indexed.index: # "합계" -> "누적"
                for w in all_workers_in_cum:
                    target_val = df_cum_indexed.loc['오전당직누적'].get(w) # "합계" -> "누적"            else:
                # '오전당직' 행 자체가 없는 경우
                oncall_targets = {w: 0 for w in all_workers_in_cum}
            # --- ▲▲▲ [수정 완료] ▲▲▲ ---

            ### 시작: 오전당직 배정 로직 ###
            
            # 1. (유지) 배정 가능한 날짜 목록을 시간순으로 정렬
            assignable_dates = sorted([d for d in df_final_unique['날짜'].unique() if d not in {s[0] for s in special_schedules}])
            
            # 2. [신규] 날짜별 후보자 목록 및 '총 당직 가능 횟수' 집계
            daily_candidates = {}
            total_eligibility_counts = Counter() # <--- [신규] 총 가능 횟수
            
            for date in assignable_dates:
                morning_workers = set(df_final_unique[(df_final_unique['날짜'] == date) & (df_final_unique['시간대'] == '오전') & (df_final_unique['상태'].isin(['근무', '대체보충', '보충', '꼭 근무']))]['근무자'])
                afternoon_workers = set(df_final_unique[(df_final_unique['날짜'] == date) & (df_final_unique['시간대'] == '오후') & (df_final_unique['상태'].isin(['근무', '대체보충', '보충', '꼭 근무']))]['근무자'])
                
                candidates = list(morning_workers - afternoon_workers)
                daily_candidates[date] = candidates
                
                # [신규] 총 당직 가능 횟수 집계
                for worker in candidates:
                    total_eligibility_counts[worker] += 1

            # 3. [수정] 실시간 누적 횟수(oncall_live_counts)를 '전월' 누적치로 초기화
            df_cum_indexed = df_cumulative.set_index('항목')
            all_workers_in_cum = [col for col in df_cumulative.columns if col != '항목']
            
            oncall_live_counts = {}
            if '오전당직누적' in df_cum_indexed.index:
                for w in all_workers_in_cum:
                    target_val = df_cum_indexed.loc['오전당직누적'].get(w)
                    
                    # ▼▼▼ [핵심 버그 수정] 누락된 할당 코드 추가 ▼▼▼
                    oncall_live_counts[w] = int(target_val) if pd.notna(target_val) else 0
                    # ▲▲▲ [수정 완료] ▲▲▲
            else:
                oncall_live_counts = {w: 0 for w in all_workers_in_cum}

            oncall = {} # 최종 배정 결과 (날짜 -> 근무자)
            actual_oncall_counts_this_month = Counter() # 이번 달 배정 횟수 (로그용)
            assigned_workers_by_date = {} # 연속 근무 체크용

            # 4. (유지) 날짜를 순차적으로(sequentially) 반복
            for date in assignable_dates: 
                date_str = date
                candidates_on_date = daily_candidates.get(date, [])
                
                if not candidates_on_date:
                    continue 

                # 5. (유지) 연속 근무자 제외 로직
                date_index = assignable_dates.index(date)
                previous_oncall_person = None
                if date_index > 0:
                    previous_date = assignable_dates[date_index - 1]
                    previous_oncall_person = assigned_workers_by_date.get(previous_date)

                if previous_oncall_person and len(candidates_on_date) > 1:
                    eligible_candidates = [p for p in candidates_on_date if p != previous_oncall_person]
                    if not eligible_candidates: 
                        eligible_candidates = candidates_on_date
                else:
                    eligible_candidates = candidates_on_date
                
                if not eligible_candidates:
                    continue 

                # 6. [핵심 수정] 후보자 정렬: '비율'이 아닌 '절대 횟수'가 가장 낮은 사람 우선
                def sort_key(worker):
                    # 1순위: 현재 누적 횟수 (전월 + 이번 달)
                    current_count = oncall_live_counts.get(worker, 0)
                    
                    # 2순위: (동점일 경우) 당직 가능 총 횟수가 적은 사람 (기회가 적은 사람)
                    total_eligible = total_eligibility_counts.get(worker, 1) 
                    
                    # (비율 로직 'ratio = current_count / total_eligible' 삭제)
                    
                    # 1순위: 'current_count'가 낮은 사람
                    # 2순위: 'total_eligible'이 낮은 사람
                    return (current_count, total_eligible)

                eligible_candidates.sort(key=sort_key)
                
                # 7. (유지) 최고 우선순위 후보자(0번 인덱스) 배정
                best_worker = eligible_candidates[0]
                oncall[date] = best_worker
                
                # 8. (유지) 실시간 누적 횟수 업데이트
                oncall_live_counts[best_worker] = oncall_live_counts.get(best_worker, 0) + 1
                
                # 9. (유지) 로그 및 연속체크용 변수 업데이트
                actual_oncall_counts_this_month[best_worker] += 1
                assigned_workers_by_date[date] = best_worker

            # --- 배정 종료 ---

            # --- (유지) 최종 배정 결과 로그 생성 ---
            st.session_state.oncall_logs = [] 
            for worker, count in sorted(actual_oncall_counts_this_month.items()):
                if count > 0:
                    log_message = f"• {worker}: {count}회 배정"
                    st.session_state.oncall_logs.append(log_message)

            # (유지) 엑셀 시트에 배정 결과 업데이트
            for idx, row in df_schedule.iterrows():
                date = row['날짜']
                df_excel.at[idx, '오전당직(온콜)'] = oncall.get(date, '')
            
            ### 끝: 오전당직 배정 로직 ###

            # ✨ [핵심 수정 1] 배정된 oncall 결과를 df_final에 '오전당직' 시간대로 추가
            oncall_df = pd.DataFrame([
                {
                    '날짜': date, '요일': day_map.get(pd.to_datetime(date).weekday(), ''),
                    '주차': week_numbers.get(pd.to_datetime(date).date(), 0),
                    '시간대': '오전당직', '근무자': worker, '상태': '당직',
                    '메모': '', '색상': '기본'
                } for date, worker in oncall.items()
            ])
            if not oncall_df.empty:
                df_final = pd.concat([df_final, oncall_df], ignore_index=True)

            # ✨ [핵심 수정 2] 모든 배정이 끝난 후, 최종 데이터를 정리
            color_priority = {'🟠 주황색': 0, '🟢 초록색': 1, '🟡 노란색': 2, '기본': 3, '🔴 빨간색': 4, '🔵 파란색': 5, '🟣 보라색': 6, '특수근무색': -1}
            df_final['색상_우선순위'] = df_final['색상'].map(color_priority)
            df_final_unique_sorted = df_final.sort_values(by=['날짜', '시간대', '근무자', '색상_우선순위']).drop_duplicates(
                subset=['날짜', '시간대', '근무자'], keep='last'
            )
            # create_final_schedule_excel 함수에 전달할 df_final_unique 변수도 여기서 최종본으로 다시 정의
            df_final_unique = df_final_unique_sorted 

            # ✨ [핵심 수정 3] 요약 테이블 생성에 필요한 변수들을 정의
            month_dt = datetime.strptime(month_str, "%Y년 %m월")
            next_month_dt = (month_dt + relativedelta(months=1)).replace(day=1)
            next_month_str = next_month_dt.strftime("%Y년 %-m월")

            # ✨ [핵심 수정 4] 올바른 최종 데이터로 요약 테이블 생성
            summary_df = build_summary_table(
                df_cumulative, all_names, next_month_str,
                df_final_unique=df_final_unique_sorted
            )

            if platform.system() == "Windows":
                font_name = "맑은 고딕"  
            else:
                font_name = "Arial"  

            duty_font = Font(name=font_name, size=9, bold=True, color="FF69B4")  
            default_font = Font(name=font_name, size=9)  

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "스케줄"

            color_map = {
                '🔴 빨간색': 'DA9694',  
                '🟠 주황색': 'FABF8F',  
                '🟢 초록색': 'A9D08E',  
                '🟡 노란색': 'FFF28F',  
                '🔵 파란색': '95B3D7',  
                '🟣 보라색': 'B1A0C7',  
                '기본': 'FFFFFF',        
                '특수근무색': 'D0E0E3'   
            }
            special_day_fill = PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
            empty_day_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            default_day_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

            for col_idx, col_name in enumerate(df_excel.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
                cell.font = Font(name=font_name, size=9, color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin', color='000000'),
                                    right=Side(style='thin', color='000000'),
                                    top=Side(style='thin', color='000000'),
                                    bottom=Side(style='thin', color='000000'))

            border = Border(left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000'))

            for row_idx, (idx, row) in enumerate(df_excel.iterrows(), 2):
                date_str_lookup = df_schedule.at[idx, '날짜']
                special_schedule_dates_set = {s[0] for s in special_schedules}
                is_special_day = date_str_lookup in special_schedule_dates_set
                is_empty_day = df_final_unique[df_final_unique['날짜'] == date_str_lookup].empty and not is_special_day

                oncall_person_for_row = str(row['오전당직(온콜)']).strip() if pd.notna(row['오전당직(온콜)']) else ""

                weekend_oncall_worker = None
                if is_special_day:
                    for s in special_schedules:
                        if s[0] == date_str_lookup and s[2] != "당직 없음":
                            weekend_oncall_worker = s[2]
                            break

                for col_idx, col_name in enumerate(df_excel.columns, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = row[col_name]
                    cell.font = default_font  
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    if is_empty_day:
                        cell.fill = empty_day_fill
                        continue

                    if col_name == '날짜':
                        cell.fill = empty_day_fill
                    elif col_name == '요일':
                        cell.fill = special_day_fill if is_special_day else default_day_fill
                    elif str(col_name).isdigit():  
                        worker = str(row[col_name]).strip()
                        if worker and pd.notna(worker):
                            if is_special_day and worker == weekend_oncall_worker:
                                cell.font = duty_font

                            worker_data = df_final_unique[(df_final_unique['날짜'] == date_str_lookup) & (df_final_unique['시간대'] == '오전') & (df_final_unique['근무자'] == worker)]
                            if not worker_data.empty:
                                color_name = worker_data.iloc[0]['색상']
                                cell.fill = PatternFill(start_color=color_map.get(color_name, 'FFFFFF'), end_color=color_map.get(color_name, 'FFFFFF'), fill_type='solid')
                                memo_text = worker_data.iloc[0]['메모']
                                if memo_text and ('보충' in memo_text or '이동' in memo_text or '대체' in memo_text):
                                    cell.comment = Comment(memo_text, "Schedule Bot")
                                    
                    elif '오후' in str(col_name):  
                        worker = str(row[col_name]).strip()
                        if worker and pd.notna(worker):
                            worker_data = df_final_unique[(df_final_unique['날짜'] == date_str_lookup) & (df_final_unique['시간대'] == '오후') & (df_final_unique['근무자'] == worker)]
                            if not worker_data.empty:
                                color_name = worker_data.iloc[0]['색상']
                                cell.fill = PatternFill(start_color=color_map.get(color_name, 'FFFFFF'), end_color=color_map.get(color_name, 'FFFFFF'), fill_type='solid')
                                memo_text = worker_data.iloc[0]['메모']
                                if memo_text and ('보충' in memo_text or '이동' in memo_text or '대체' in memo_text):
                                    cell.comment = Comment(memo_text, "Schedule Bot")

                    elif col_name == '오전당직(온콜)':
                        if oncall_person_for_row:
                            cell.font = duty_font

            ws.column_dimensions['A'].width = 11
            for col in ws.columns:
                 if col[0].column_letter != 'A':
                     ws.column_dimensions[col[0].column_letter].width = 9

            month_dt = datetime.strptime(month_str, "%Y년 %m월")
            next_month_dt = (month_dt + relativedelta(months=1)).replace(day=1)
            next_month_str = next_month_dt.strftime("%Y년 %-m월")
            month_start = month_dt.replace(day=1)
            month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)

            summary_df = build_summary_table(
                df_cumulative,
                all_names,
                next_month_str,
                df_final_unique=df_final_unique_sorted
            )
            style_args = {
                'font': default_font,
                'bold_font': Font(name=font_name, size=9, bold=True),
                'border': border,
            }
            append_summary_table_to_excel(ws, summary_df, style_args)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            st.session_state.output = output
            
            summary_df = build_summary_table(
                df_cumulative,
                all_names,
                next_month_str,
                df_final_unique=df_final_unique_sorted
            )

            # 사용자의 함수 정의에 맞는 인수만 전달하도록 수정
            wb_final_bytes = create_final_schedule_excel(
                initial_df=df_excel.copy(), # 초기 상태 df 전달
                edited_df=df_excel,         # 현재 상태 df 전달
                edited_cumulative_df=summary_df, # build_summary_table 결과
                df_special=df_monthly_schedule, # 로드된 토요/휴일 데이터
                df_requests=df_request,         # 로드된 요청사항 데이터
                closing_dates=holiday_dates,    # 로드된 휴관일 데이터
                month_str=month_str,            # 현재 월 문자열
                # ▼▼▼ 추가된 인수 전달 ▼▼▼
                df_final_unique=df_final_unique_sorted, # 최종 배정 결과
                df_schedule=df_schedule             # 날짜 매핑용 df
                # ▲▲▲ 추가 인수 전달 완료 ▲▲▲
            )
            # 함수가 bytes를 반환하므로 바로 BytesIO로 읽음
            output_final = io.BytesIO(wb_final_bytes)
            output_final.seek(0)
            
            month_dt = datetime.strptime(month_str, "%Y년 %m월")
            next_month_dt = (month_dt + relativedelta(months=1)).replace(day=1)
            next_month_str = next_month_dt.strftime("%Y년 %-m월")
            month_start = month_dt.replace(day=1)
            month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)

            try:
                gc = get_gspread_client()
                sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
                
                # 이 함수가 이제 동적으로 열이 생성된 데이터프레임을 반환합니다.
                df_schedule_to_save = transform_schedule_for_checking(df_final_unique, df_excel, month_start, month_end)
                
                try:
                    worksheet_schedule = sheet.worksheet(f"{month_str} 스케줄 ver1.0")
                except gspread.exceptions.WorksheetNotFound:
                    worksheet_schedule = sheet.add_worksheet(title=f"{month_str} 스케줄 ver1.0", rows=1000, cols=50) # cols는 여유있게
                
                update_sheet_with_retry(worksheet_schedule, [df_schedule_to_save.columns.tolist()] + df_schedule_to_save.astype(str).values.tolist())
                
                try:
                    # 시트 이름을 "누적 요약"으로 변경하여 기존 시트와 구분하는 것을 권장합니다.
                    worksheet_summary = sheet.worksheet(f"{next_month_str} 누적 ver1.0")
                except gspread.exceptions.WorksheetNotFound:
                    worksheet_summary = sheet.add_worksheet(title=f"{next_month_str} 누적 ver1.0", rows=100, cols=50)
                
                # [핵심] df_cumulative_next 대신 summary_df 변수를 사용하여 시트를 업데이트합니다.
                summary_df_to_save = build_summary_table(
                    df_cumulative, all_names, next_month_str,
                    df_final_unique=df_final_unique_sorted
                )

                update_sheet_with_retry(worksheet_summary, [summary_df_to_save.columns.tolist()] + summary_df_to_save.values.tolist())

            except Exception as e:
                st.error(f"Google Sheets 저장 중 오류 발생: {e}")
                st.stop()
            
            try:
                # 배정 확인용 테이블 생성 (GSheet 저장용)
                df_schedule_to_save_for_gsheet = transform_schedule_for_checking(
                    df_final_unique_sorted,
                    df_excel,
                    month_start,
                    month_end
                )

                # [수정] 세션 상태 저장을 try 블록 안으로 이동
                st.session_state.assignment_results = {
                    # --- 편집 및 다운로드에 필요한 핵심 데이터 ---
                    "df_excel_initial": df_excel.copy(),
                    "summary_df_initial": summary_df.copy(),
                    "df_schedule_for_display": df_excel,
                    "summary_df_for_display": summary_df,
                    "df_schedule_to_save_for_gsheet": df_schedule_to_save_for_gsheet, # <-- 이제 안전
                    # --- Excel 생성 시 필요한 추가 데이터 ---
                    "df_final_unique_sorted": df_final_unique_sorted,
                    "df_schedule": df_schedule,
                    "df_special": df_monthly_schedule, # df_special -> df_monthly_schedule 로 변경 (변수명 일치 확인 필요)
                    "df_requests": df_request,
                    "closing_dates": holiday_dates,
                    "month_str": month_str,
                    "all_names": all_names, # <-- [★이 줄을 추가하세요★]
                    # --- 로그 데이터 ---
                    "request_logs": st.session_state.request_logs,
                    "swap_logs": st.session_state.swap_logs,
                    "adjustment_logs": st.session_state.adjustment_logs,
                    "oncall_logs": st.session_state.oncall_logs,
                }
                # --- 세션 상태 저장 끝 ---

            except Exception as e_transform:
                # 함수 실행 중 오류 발생 시 메시지 출력 및 중단
                st.error(f"⚠️ 데이터 처리 중 오류 발생 (transform_schedule_for_checking 함수 또는 세션 상태 저장 중)")
                st.exception(e_transform) # 상세 오류 traceback 출력
                st.stop() # 스크립트 실행 중단

    month_dt = datetime.strptime(month_str, "%Y년 %m월")
    next_month_dt = (month_dt + relativedelta(months=1)).replace(day=1)
    next_month_str = next_month_dt.strftime("%Y년 %-m월")
    month_start = month_dt.replace(day=1)
    month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)

    if st.session_state.get('assigned', False):
        results = st.session_state.get('assignment_results', {})
        if results:
            with st.expander("🔍 배정 과정 상세 로그 보기", expanded=True):
                st.markdown("**📋 요청사항 반영 로그**"); st.code("\n".join(results.get("request_logs", [])) if results.get("request_logs") else "반영된 요청사항(휴가/학회)이 없습니다.", language='text')
                st.markdown("---"); st.markdown("**📞 오전당직(온콜) 배정 로그**"); st.code("\n".join(results.get("oncall_logs", [])) if results.get("oncall_logs") else "모든 오전당직(온콜)이 누적 횟수에 맞게 정상 배정되었습니다.", language='text')

            # --- [핵심 수정] 1. 스케줄 테이블 data_editor *반환값*을 변수에 저장 ---
            if "df_schedule_for_display" in results:
                st.divider()
                st.markdown(f"**➕ {month_str} 배정 스케줄 (수정 가능)**")
                st.warning("⚠️ 아래에서 내용을 수정하신 후, **'수정사항 저장'** 버튼을 누르면 Google Sheets에 반영됩니다.")

                # 1. 표시용 데이터 준비 (상태 텍스트 추가)
                df_to_edit_schedule = results["df_schedule_for_display"].copy()
                df_final_unique = results.get("df_final_unique_sorted")
                df_schedule = results.get("df_schedule")

                if df_final_unique is not None and df_schedule is not None:
                    # [★수정★] 상태와 메모를 모두 저장하는 딕셔너리로 변경
                    status_lookup = {}
                    for _, row in df_final_unique.iterrows():
                        key = (row['날짜'], row['시간대'], row['근무자'])
                        # (상태, 메모) 튜플로 저장
                        status_lookup[key] = (row['상태'], row.get('메모', '')) 

                    for idx, row in df_to_edit_schedule.iterrows():
                        if idx not in df_schedule.index: continue
                        date_str = df_schedule.at[idx, '날짜'] # YYYY-MM-DD
                        
                        for col_name in df_to_edit_schedule.columns:
                            
                            # 1. 시간대 먼저 결정
                            time_slot = None
                            if col_name.isdigit(): time_slot = '오전'
                            elif col_name.startswith("오후"): time_slot = '오후'
                            elif col_name == '오전당직(온콜)': time_slot = '오전당직'

                            # 2. 근무, 보충, 당직 셀인 경우에만
                            if time_slot:
                                worker_name_cell = str(row[col_name] or '').strip()
                                if not worker_name_cell: # 셀이 비어있으면 건너뛰기
                                    continue

                                # 3. 셀에 괄호가 이미 있는지 확인
                                match = re.match(r'.+?\((.+)\)', worker_name_cell)
                                
                                if match:
                                    pass # 이미 괄호가 있으면 (수동 편집) 그대로 둠
                                else:
                                    worker_name_only = worker_name_cell # 괄호가 없으니 이게 이름
                                    
                                    key = (date_str, time_slot, worker_name_only)
                                    lookup_result = status_lookup.get(key)
                                    
                                    # 4. [★수정★] 상태와 메모를 분리하여 조건에 맞게 괄호 추가
                                    if lookup_result:
                                        status, memo = lookup_result
                                        
                                        # [요청사항] '대체보충'이고 유효한 메모가 있으면 (메모)를 표시
                                        if status == '대체보충' and pd.notna(memo) and str(memo).strip():
                                            df_to_edit_schedule.at[idx, col_name] = f"{worker_name_only}({memo})"
                                        # [유지] 그 외 (휴가, 보충, '대체휴근' 등)
                                        elif status and status not in ['근무', '당직', '기본']:
                                            df_to_edit_schedule.at[idx, col_name] = f"{worker_name_only}({status})"

                if "df_schedule_for_comparison" not in results:
                    st.session_state.assignment_results["df_schedule_for_comparison"] = df_to_edit_schedule.copy()
                # --- ▲▲▲ [저장 완료] ▲▲▲ ---

                edited_schedule_df = st.data_editor(
                    df_to_edit_schedule,
                    # ▼▼▼ [핵심 수정] key를 동적으로 변경하여 강제 리셋 ▼▼▼
                    key=f"edited_schedule_table_{st.session_state.editor_key_version}",
                    use_container_width=True,
                    hide_index=True,
                    disabled=['날짜', '요일'],
                    on_change=set_editor_changed_flag # <--- [수정] 콜백 추가
                )
            else:
                st.warning("⚠️ 배정 스케줄 테이블 데이터를 불러올 수 없습니다.")
                edited_schedule_df = pd.DataFrame() # 오류 방지용 빈 DataFrame

            # --- ▼▼▼ [신규] 스케줄 수정사항 로그 로직 ▼▼▼ ---
            st.markdown("📝 **스케줄 수정사항**")
            schedule_change_log = []
            schedule_has_changed = False

            # [수정] 'results.get("df_excel_initial")' (괄호 없는 원본) 대신,
            # 에디터에 '입력(input)'으로 사용된 'df_to_edit_schedule' (괄호가 이미 추가된)을 
            # 비교할 원본으로 사용합니다.
            original_schedule_df = df_to_edit_schedule
            if original_schedule_df is not None and not edited_schedule_df.equals(original_schedule_df):
                schedule_has_changed = True # <--- ★★★ [ 2. 이 줄을 추가 ] ★★★
                try:
                    # (파일 상단에 'import numpy as np'가 필요합니다)
                    import numpy as np 
                    diff_indices = np.where(edited_schedule_df.astype(str).ne(original_schedule_df.astype(str)))
                    changed_cells = set(zip(diff_indices[0], diff_indices[1])) # 중복 로그 방지
                    
                    for row_idx, col_idx in changed_cells:
                        date_str = edited_schedule_df.iloc[row_idx, 0] # '날짜' 열 (예: "10월 1일")
                        slot_name = edited_schedule_df.columns[col_idx] # 변경된 열 이름 (예: "1")
                        
                        # [수정] 원본 값을 'original_schedule_df' (df_to_edit_schedule)에서 가져옵니다.
                        old_value = original_schedule_df.iloc[row_idx, col_idx]
                        new_value = edited_schedule_df.iloc[row_idx, col_idx]
                        
                        log_msg = f"{date_str} '{slot_name}' 변경: '{old_value or '빈 값'}' → '{new_value or '빈 값'}'"
                        schedule_change_log.append(log_msg)
                except Exception as e:
                    schedule_change_log.append(f"[로그 오류] 스케줄 변경사항 비교 중 오류: {e}")
                    
            if schedule_change_log:
                st.code("\n".join(f"• {msg}" for msg in sorted(schedule_change_log)), language='text')
            else:
                st.info("수정된 사항이 없습니다.")
            # --- ▲▲▲ [신규] 스케줄 로그 끝 (수정본) ---

            # [기존 코드] (L1682 근처)
            # --- [핵심 수정] 2. 누적 테이블 data_editor *반환값*을 변수에 저장 ---
            if "summary_df_for_display" in results:
                st.divider()
                st.markdown(f"**➕ {next_month_str} 누적 테이블 (수정 가능)**")
                
                # [★수정★] 누적 테이블이 자동 재계산됨을 안내
                st.write("- 누적 테이블은 '배정 스케줄' 편집기에 반영된 내용을 바탕으로 자동 재계산됩니다.\n- 주의) 대체보충은 수정 시 누적 테이블을 직접 수정해주셔야 합니다.")

                # --- ▼▼▼ [ ★ L1725~L1741을 이 블록으로 교체 ★ ] ▼▼▼ ---
                if schedule_has_changed:
                    try:
                        df_cumulative_initial = st.session_state["df_cumulative"] # GSheet 원본(A)
                        all_names_list = results.get("all_names", [])
                        df_schedule_mapping = results.get("df_schedule")

                        if not all_names_list or df_schedule_mapping is None:
                            st.error("자동 재계산에 필요한 'all_names' 또는 'df_schedule' 데이터가 없습니다.")
                            summary_df_input = results["summary_df_initial"] 
                        else:
                            # '수정된' 스케줄(edited_schedule_df)로 재계산 (B)
                            summary_df_input = recalculate_summary_from_schedule(
                                edited_schedule_df,
                                df_cumulative_initial,
                                all_names_list,
                                df_schedule_mapping
                            )
                    except Exception as e_recalc:
                        st.error(f"누적 테이블 자동 재계산 중 오류 발생: {e_recalc}")
                        summary_df_input = results["summary_df_initial"] 
                else:
                    # 2. 상단 스케줄이 수정되지 않음 (페이지 첫 로드) -> 원본(A) 표시
                    summary_df_input = results.get("summary_df_initial", pd.DataFrame()).copy() # 원본(A)을 그대로 사용
                # --- ▲▲▲ [ 교체 완료 ] ▲▲▲ ---

                # [수정] st.data_editor가 'summary_df_input' (재계산된 값)을 사용
                edited_summary_df = st.data_editor(
                    summary_df_input, # <-- 재계산된 데이터를 입력
                    # ▼▼▼ [핵심 수정] key를 동적으로 변경하여 강제 리셋 ▼▼▼
                    key=f"edited_summary_table_{st.session_state.editor_key_version}",
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        summary_df_input.columns[0]: st.column_config.Column(disabled=True),
                        **{col: st.column_config.NumberColumn(format="%d") 
                            for col in summary_df_input.columns[1:]}
                    },
                    disabled=False,
                    on_change=set_editor_changed_flag 
                )
            else:
                st.warning("⚠️ 누적 테이블 데이터를 불러올 수 없습니다.")
                edited_summary_df = pd.DataFrame() # 오류 방지용 빈 DataFrame

            # --- ▼▼▼ [누적 테이블 수동 수정사항 로그 로직 (수정됨)] ▼▼▼ ---
            st.markdown("📝 **누적 테이블 수정사항**")
            summary_change_log = [] # 리스트 초기화

            # --- ▼▼▼ [신규] 항목 순서 정렬을 위한 맵(Map) 정의 ▼▼▼ ---
            desired_order = [
                "오전보충", "임시보충", "오전합계", "오전누적", 
                "오후보충", "온콜검사", "오후합계", "오후누적", 
                "오전당직", "오전당직누적", "오후당직", "오후당직누적"
            ]
            # 항목 이름을 정렬 순서(숫자)로 매핑
            order_map = {item_name: index for index, item_name in enumerate(desired_order)}

            original_summary_df = results.get("summary_df_initial") # (A)

            cumulative_has_changed = False # <-- [★ 1. 이 줄을 추가하세요]

            if original_summary_df is not None and not edited_summary_df.equals(original_summary_df): # (A) vs (C)
                cumulative_has_changed = True # <-- [★ 2. 이 줄을 추가하세요]
                try:
                    import numpy as np
                    # 1번 수정으로 A와 C의 숫자 타입이 int로 통일되었으므로 astype(str) 비교가 안전합니다.
                    stats_orig_str = original_summary_df.astype(str) # (A)
                    stats_edit_str = edited_summary_df.astype(str) # (C)
                    
                    diff_indices_stats = np.where(stats_edit_str.ne(stats_orig_str))
                    changed_cells_stats = set(zip(diff_indices_stats[0], diff_indices_stats[1])) 

                    for row_idx, col_idx in changed_cells_stats:
                        item_name = edited_summary_df.iloc[row_idx, 0] 
                        person_name = edited_summary_df.columns[col_idx]
                        
                        # [핵심] old_value를 'original_summary_df'(A)에서 가져옵니다.
                        old_value = original_summary_df.iloc[row_idx, col_idx]
                        new_value = edited_summary_df.iloc[row_idx, col_idx]
                        
                        log_msg = f"'{person_name}'의 '{item_name}' 변경: {old_value} → {new_value}"
                        summary_change_log.append(log_msg)
                except Exception as e:
                    summary_change_log.append(f"[로그 오류] 누적 테이블 변경사항 비교 중 오류: {e}")

            if summary_change_log:
                log_text_stats = "\n".join(f"• {msg}" for msg in sorted(summary_change_log))
                st.code(log_text_stats, language='text')
            else:
                st.info("수정된 사항이 없습니다.")
            # --- ▲▲▲ [누적 테이블 로그 끝 (수정 완료)] ---

            st.divider() # 구분선 추가

            # --- ▼▼▼ [핵심 수정] 3. 저장 및 다운로드 버튼 영역 수정 ▼▼▼ ---
            col1, col2 = st.columns(2)

            with col1:
                # --- 1. Google Sheets 저장 버튼 ---

                # [★ 3. .equals() 비교 결과로 실제 변경 유무를 최종 판정 ★]
                # (schedule_has_changed는 L1822에서 이미 정의됨)
                real_has_unsaved_changes = schedule_has_changed or cumulative_has_changed
                
                if st.button("💾 수정사항 Google Sheet에 저장", 
                             type="primary", 
                             use_container_width=True, 
                             disabled=not real_has_unsaved_changes # <-- [수정 완료]
                            ):
                    # [수정] st.session_state 대신 위에서 할당받은 *변수* 사용
                    if not edited_schedule_df.empty and not edited_summary_df.empty:
                        with st.spinner("수정된 데이터 저장 중..."):
                            try:
                                # edited_schedule_df 와 edited_summary_df 변수를 직접 사용
                                df_to_save_gsheet = edited_schedule_df.copy()

                                gc = get_gspread_client()
                                sheet = gc.open_by_url(url)
                                schedule_sheet_name = f"{month_str} 스케줄 ver1.0"
                                summary_sheet_name = f"{next_month_str} 누적 ver1.0"

                                # 스케줄 시트 저장
                                try: ws_sched = sheet.worksheet(schedule_sheet_name)
                                except WorksheetNotFound: ws_sched = sheet.add_worksheet(title=schedule_sheet_name, rows=1000, cols=len(df_to_save_gsheet.columns)+5)
                                # update_sheet_with_retry가 성공하면 True 반환
                                success_sched = update_sheet_with_retry(ws_sched, [df_to_save_gsheet.columns.tolist()] + df_to_save_gsheet.astype(str).fillna('').values.tolist())

                                # 누적 시트 저장
                                try: ws_summ = sheet.worksheet(summary_sheet_name)
                                except WorksheetNotFound: ws_summ = sheet.add_worksheet(title=summary_sheet_name, rows=100, cols=len(edited_summary_df.columns)+5)
                                success_summ = update_sheet_with_retry(ws_summ, [edited_summary_df.columns.tolist()] + edited_summary_df.astype(str).fillna('').values.tolist())

                                if success_sched and success_summ:
                                    st.success(f"✅ '{schedule_sheet_name}' 및 '{summary_sheet_name}' 시트에 수정된 내용이 저장되었습니다.")

                                    # 저장 성공 후 초기 상태 업데이트
                                    # st.session_state.assignment_results["df_excel_initial"] = edited_schedule_df.copy()
                                    # st.session_state.assignment_results["summary_df_initial"] = edited_summary_df.copy()
                                    st.session_state.assignment_results["df_schedule_for_display"] = edited_schedule_df.copy()
                                    st.session_state.assignment_results["summary_df_for_display"] = edited_summary_df.copy()
                                    
                                    # ▼▼▼ [핵심 수정] 플래그 리셋 및 리런 ▼▼▼
                                    st.session_state.editor_has_changes = False 
                                    time.sleep(1)
                                    st.rerun()
                                    # ▲▲▲ [핵심 수정] ▲▲▲
                                
                                else:
                                    # update_sheet_with_retry가 False를 반환했지만 에러를 raise하지 않은 경우
                                    st.error("Google Sheets 업데이트가 완료되지 않았습니다. API 오류 로그를 확인해주세요.")

                            except Exception as e:
                                st.error(f"Google Sheets 저장 중 오류 발생: {e}")
                                # 에러 발생 시 플래그를 True로 유지 (다운로드 방지 상태)
                                st.session_state.editor_has_changes = True

                    else:
                        st.error("편집된 데이터가 없습니다.")

            with col2:
                # --- 2. Excel 다운로드 버튼 (두 종류) ---
                if not edited_schedule_df.empty and not edited_summary_df.empty:
                    try:
                        # --- 데이터 로드 (기존과 동일) ---
                        results = st.session_state.get('assignment_results', {})
                        initial_schedule_df = results.get("df_schedule_for_comparison")
                        # initial_summary_df는 사용되지 않음 (콜백 플래그가 대체)
                        df_special_dl = results.get("df_special")
                        df_requests_dl = results.get("df_requests")
                        closing_dates_dl = results.get("closing_dates")
                        month_str_dl = results.get("month_str")
                        df_final_unique_dl = results.get("df_final_unique_sorted")
                        df_schedule_dl = results.get("df_schedule")

                        # --- ▼▼▼ [핵심 수정] 플래그 확인 ▼▼▼
                        # 'editor_has_changes' 플래그가 True이면 다운로드를 막습니다.
                        has_unsaved_changes = real_has_unsaved_changes
                        # --- ▲▲▲ [핵심 수정] 완료 ▲▲▲ ---

                        if has_unsaved_changes:
                            st.error("⚠️ 수정사항이 감지되었습니다. 먼저 '수정사항 Google Sheet에 저장' 버튼을 눌러주세요.")
                            # [수정] 버튼이 아예 보이지 않도록 하거나, 여기에 disabled된 버튼을 추가할 수 있습니다.
                            # 여기서는 st.error 메시지만 표시합니다.
                        else:
                            # 변경 사항이 없거나 저장된 상태일 때만 다운로드 버튼 표시
                            if initial_schedule_df is None or month_str_dl is None or df_final_unique_dl is None or df_schedule_dl is None:
                                st.error("Excel 생성에 필요한 초기 데이터가 없습니다. 페이지를 새로고침 해주세요.")
                            else:
                                # --- 스타일 변수 재정의 (로직 유지) ---
                                if platform.system() == "Windows": font_name = "맑은 고딕"
                                else: font_name = "Arial"
                                default_font = Font(name=font_name, size=9)
                                bold_font = Font(name=font_name, size=9, bold=True)
                                duty_font = Font(name=font_name, size=9, bold=True, color="FF69B4")
                                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                                color_map = {
                                        '🔴 빨간색': 'DA9694', '🟠 주황색': 'FABF8F', '🟢 초록색': 'A9D08E',
                                        '🟡 노란색': 'FFF28F', '🔵 파란색': '95B3D7', '🟣 보라색': 'B1A0C7',
                                        '기본': 'FFFFFF', '특수근무색': 'D0E0E3'
                                    }
                                special_day_fill = PatternFill(start_color='95B3D7', fill_type='solid')
                                empty_day_fill = PatternFill(start_color='808080', fill_type='solid')
                                default_day_fill = PatternFill(start_color='FFF2CC', fill_type='solid')

                                # --- 1. 최종본(공유용) Excel 생성 및 다운로드 버튼 ---
                                excel_data_final = create_final_schedule_excel(
                                    initial_df=initial_schedule_df,
                                    edited_df=edited_schedule_df,
                                    edited_cumulative_df=edited_summary_df,
                                    df_special=df_special_dl if df_special_dl is not None else pd.DataFrame(),
                                    df_requests=df_requests_dl if df_requests_dl is not None else pd.DataFrame(),
                                    closing_dates=closing_dates_dl if closing_dates_dl is not None else [],
                                    month_str=month_str_dl,
                                    df_final_unique=df_final_unique_dl,
                                    df_schedule=df_schedule_dl
                                )
                                st.download_button(
                                    label="📥 스케줄 ver1.0 다운로드",
                                    data=excel_data_final,
                                    file_name=f"{month_str_dl} 스케줄 ver1.0.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.sheet",
                                    use_container_width=True,
                                    type="primary",
                                    key="download_edited_final"
                                )

                                # --- 2. 배정 확인용 Excel 생성 및 다운로드 버튼 ---
                                excel_data_checking = create_checking_schedule_excel(
                                    initial_df=results.get("df_schedule_for_comparison"), # (C_orig)
                                    edited_df=edited_schedule_df,
                                    edited_cumulative_df=edited_summary_df,
                                    df_special=df_special_dl if df_special_dl is not None else pd.DataFrame(),
                                    df_requests=df_requests_dl if df_requests_dl is not None else pd.DataFrame(),
                                    closing_dates=closing_dates_dl if closing_dates_dl is not None else [],
                                    month_str=month_str_dl
                                )
                                st.download_button(
                                    label="📥 스케줄 ver1.0 다운로드 (배정 확인용)",
                                    data=excel_data_checking,
                                    file_name=f"{month_str_dl} 스케줄 ver1.0 (배정 확인용).xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.sheet",
                                    use_container_width=True,
                                    type="secondary",
                                    key="download_edited_checking"
                                )
                    except Exception as e:
                        st.error(f"Excel 파일 생성 또는 변경 사항 확인 중 오류가 발생했습니다: {e}")
                        st.exception(e)
                else:
                    st.info("🔄 스케줄 데이터 로딩 중...")