import re
import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from collections import Counter
import random
import time
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment

# 세션 상태 초기화
if "data_loaded" not in st.session_state:
    st.session_state["data_loaded"] = False
if "df_room_fix" not in st.session_state:
    st.session_state["df_room_fix"] = pd.DataFrame(columns=["이름", "분류", "날짜정보"])
if "df_room_request" not in st.session_state:
    st.session_state["df_room_request"] = pd.DataFrame(columns=["이름", "분류", "날짜정보"])
# 방 설정 초기화
if "room_settings" not in st.session_state:
    st.session_state["room_settings"] = {
        "830_room_select": ['1', '2', '4', '7'],
        "900_room_select": ['10', '11', '12'],
        "930_room_select": ['5', '6', '8'],
        "1000_room_select": ['3', '9'],
        "1330_room_select": ['2', '3', '4', '9']
    }

# Google Sheets 클라이언트 초기화
def get_gspread_client():
    scope = ["https://www.googleapis.com/auth/spreadsheets"]
    service_account_info = dict(st.secrets["gspread"])
    service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
    credentials = Credentials.from_service_account_info(service_account_info, scopes=scope)
    return gspread.authorize(credentials)

# Google Sheets 업데이트 함수
def update_sheet_with_retry(worksheet, data, retries=5, delay=10):
    for attempt in range(retries):
        try:
            worksheet.batch_update([
                {"range": "A1:D", "values": [[]]},
                {"range": "A1", "values": data}
            ])
            return
        except Exception as e:
            if "Quota exceeded" in str(e):
                st.warning(f"API 쿼터 초과, {delay}초 후 재시도 ({attempt+1}/{retries})")
                time.sleep(delay)
            else:
                st.warning(f"업데이트 실패, {delay}초 후 재시도 ({attempt+1}/{retries}): {str(e)}")
                time.sleep(delay)
    st.error("Google Sheets 업데이트 실패: 재시도 횟수 초과")

# 데이터 로드 (캐싱 사용)
@st.cache_data
def load_data(month_str):
    return load_data_no_cache(month_str)

# 데이터 로드 (캐싱 미사용)
def load_data_no_cache(month_str):
    gc = get_gspread_client()
    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
    
    try:
        worksheet_schedule = sheet.worksheet(f"{month_str} 스케쥴")
        df_schedule = pd.DataFrame(worksheet_schedule.get_all_records())
    except Exception as e:
        st.error(f"스케줄 시트를 불러오는 데 실패: {e}")
        st.stop()
    
    try:
        worksheet_request = sheet.worksheet(f"{month_str} 방배정 요청")
        df_room_request = pd.DataFrame(worksheet_request.get_all_records())
        if "우선순위" in df_room_request.columns:
            df_room_request = df_room_request.drop(columns=["우선순위"])
    except:
        worksheet_request = sheet.add_worksheet(f"{month_str} 방배정 요청", rows=100, cols=3)
        worksheet_request.append_row(["이름", "분류", "날짜정보"])
        df_room_request = pd.DataFrame(columns=["이름", "분류", "날짜정보"])
    
    try:
        worksheet_fix = sheet.worksheet(f"{month_str} 방배정 고정")
        df_room_fix = pd.DataFrame(worksheet_fix.get_all_records())
    except:
        worksheet_fix = sheet.add_worksheet(f"{month_str} 방배정 고정", rows=100, cols=3)
        worksheet_fix.append_row(["이름", "분류", "날짜정보"])
        df_room_fix = pd.DataFrame(columns=["이름", "분류", "날짜정보"])
    
    st.session_state["df_schedule"] = df_schedule
    st.session_state["df_room_request"] = df_room_request
    st.session_state["df_room_fix"] = df_room_fix
    st.session_state["worksheet_request"] = worksheet_request
    st.session_state["worksheet_fix"] = worksheet_fix
    st.session_state["data_loaded"] = True
    
    return df_schedule, df_room_request, df_room_fix, worksheet_request, worksheet_fix

# 근무 가능 일자 계산
@st.cache_data
def get_user_available_dates(name, df_schedule, month_start, month_end):
    available_dates = []
    for _, row in df_schedule.iterrows():
        date_str = row['날짜']
        try:
            if "월" in date_str:
                date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=2025).date()
            else:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            continue
        if month_start <= date_obj <= month_end and row['요일'] not in ['토요일', '일요일']:
            personnel = [row[str(i)] for i in range(1, 13)] + [row[f'오후{i}'] for i in range(1, 5)]
            if name in personnel:
                available_dates.append((date_obj, f"{date_str}({row['요일']}) 오전"))
                available_dates.append((date_obj, f"{date_str}({row['요일']}) 오후"))
    
    available_dates.sort(key=lambda x: x[0])
    sorted_dates = [date_str for _, date_str in available_dates]
    return sorted_dates

# 요청 저장 (df_room_request용)
def save_to_gsheet(name, categories, dates, month_str, worksheet):
    gc = get_gspread_client()
    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
    worksheet = sheet.worksheet(f"{month_str} 방배정 요청")
    df = pd.Data.feedDataFrame(worksheet.get_all_records())
    if "우선순위" in df.columns:
        df = df.drop(columns=["우선순위"])
    
    new_rows = []
    for date in dates:
        for cat in categories:
            new_rows.append({"이름": name, "분류": cat, "날짜정보": date})
    
    df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
    update_sheet_with_retry(worksheet, [df.columns.tolist()] + df.values.tolist())
    return df

# df_schedule_md 생성
def create_df_schedule_md(df_schedule):
    df_schedule_md = df_schedule.copy().fillna('')
    for idx, row in df_schedule_md.iterrows():
        date_str = row['날짜']
        oncall_worker = row['오전당직(온콜)']
        
        try:
            if isinstance(date_str, (float, int)):
                date_str = str(int(date_str))
            date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=2025) if "월" in date_str else datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError as e:
            st.error(f"날짜 파싱 오류: {date_str}, 오류: {str(e)}")
            continue
        
        afternoon_cols = ['오후1', '오후2', '오후3', '오후4', '오후5']
        if all(row[col] == '' for col in afternoon_cols):
            df_schedule_md.at[idx, '오전당직(온콜)'] = ''
            continue
        
        if pd.isna(oncall_worker) or oncall_worker == '':
            oncall_worker = ''
            df_schedule_md.at[idx, '오전당직(온콜)'] = ''
        
        if oncall_worker:
            morning_cols = [str(i) for i in range(1, 13)]
            for col in morning_cols + afternoon_cols:
                if row[col] == oncall_worker:
                    df_schedule_md.at[idx, col] = ''
        
        morning_cols = [str(i) for i in range(1, 13)]
        morning_workers = [row[col] for col in morning_cols if row[col]]
        if len(morning_workers) > 11:
            morning_workers = morning_workers[:11]
        morning_workers.extend([''] * (11 - len(morning_workers)))
        for i, col in enumerate([str(i) for i in range(1, 12)], 1):
            df_schedule_md.at[idx, col] = morning_workers[i-1]
        
        afternoon_workers = [row[col] for col in afternoon_cols if row[col]]
        if len(afternoon_workers) > 4:
            afternoon_workers = afternoon_workers[:4]
        afternoon_workers.extend([''] * (4 - len(afternoon_workers)))
        for i, col in enumerate(['오후1', '오후2', '오후3', '오후4'], 1):
            df_schedule_md.at[idx, col] = afternoon_workers[i-1]
    
    df_schedule_md = df_schedule_md.drop(columns=['12', '오후5'], errors='ignore')
    return df_schedule_md

# 날짜정보 파싱 함수
def parse_date_info(date_info):
    try:
        date_part = date_info.split('(')[0].strip()
        date_obj = datetime.strptime(date_part, '%m월 %d일').replace(year=2025)
        return date_obj.strftime('%Y-%m-%d'), date_info.endswith('오전')
    except ValueError:
        return None, False

# random_assign 함수
def random_assign(personnel, slots, fixed_assignments, priority_assignments, time_groups, max_early, max_late, max_duty, max_room, total_stats, morning_personnel, afternoon_personnel):
    best_assignment = None
    best_stats = None
    min_unassigned = float('inf')
    
    early_slots = [s for s in slots if s.startswith('8:30') and '_당직' not in s]
    late_slots = [s for s in slots if s.startswith('10:00')]
    duty_slots = [s for s in slots if s.startswith('13:30') and '_당직' in s]
    morning_slots_830 = [s for s in slots if s.startswith('8:30') and '_당직' not in s]
    morning_slots_900 = [s for s in slots if s.startswith('9:00')]
    morning_slots_930 = [s for s in slots if s.startswith('9:30')]
    morning_slots_1000 = [s for s in slots if s.startswith('10:00')]
    afternoon_slots = [s for s in slots if s.startswith('13:30') and '_당직' not in s]
    
    for attempt in range(100):
        assignment = [None] * len(slots)
        assigned_counts = Counter()
        available_slots = list(range(len(slots)))
        daily_stats = {
            'early': Counter(),
            'late': Counter(),
            'duty': Counter(),
            'rooms': {str(i): Counter() for i in range(1, 13)}
        }
        
        fixed_personnel = set()
        for slot, person in fixed_assignments.items():
            if slot in slots and person in personnel:
                slot_idx = slots.index(slot)
                assignment[slot_idx] = person
                assigned_counts[person] += 1
                if slot_idx in available_slots:
                    available_slots.remove(slot_idx)
                fixed_personnel.add(person)
                room_num = re.search(r'\((\d+)\)', slot).group(1)
                daily_stats['rooms'][room_num][person] += 1
                if slot in early_slots:
                    daily_stats['early'][person] += 1
                elif slot in late_slots:
                    daily_stats['late'][person] += 1
                elif slot in duty_slots:
                    daily_stats['duty'][person] += 1
        
        morning_personnel_list = [p for p in morning_personnel if p not in fixed_personnel]
        afternoon_personnel_list = [p for p in afternoon_personnel if p not in fixed_personnel]
        used_morning_personnel = set(fixed_personnel)
        used_afternoon_personnel = set(fixed_personnel)
        
        morning_personnel_list.sort(key=lambda p: (
            daily_stats['early'][p],
            daily_stats['late'][p],
            sum(daily_stats['rooms'][r][p] for r in daily_stats['rooms'])
        ))
        afternoon_personnel_list.sort(key=lambda p: (
            daily_stats['duty'][p],
            sum(daily_stats['rooms'][r][p] for r in daily_stats['rooms'])
        ))
        random.shuffle(morning_personnel_list)
        random.shuffle(afternoon_personnel_list)
        
        priority_pairs = [(slot, person) for (slot, person), _ in priority_assignments.items() if slot in slots and person in personnel]
        for slot, person in priority_pairs:
            slot_idx = slots.index(slot)
            if slot_idx in available_slots and person not in used_morning_personnel and person not in fixed_personnel:
                room_num = re.search(r'\((\d+)\)', slot).group(1)
                early_count = total_stats['early'][person] + daily_stats['early'][person]
                late_count = total_stats['late'][person] + daily_stats['late'][person]
                duty_count = total_stats['duty'][person] + daily_stats['duty'][person]
                room_count = total_stats['rooms'][room_num][person] + daily_stats['rooms'][room_num][person]
                
                if (early_count < max_early and late_count < max_late and
                    duty_count < max_duty and room_count < max_room):
                    assignment[slot_idx] = person
                    assigned_counts[person] += 1
                    available_slots.remove(slot_idx)
                    used_morning_personnel.add(person) if slot in (morning_slots_830 + morning_slots_900 + morning_slots_930 + morning_slots_1000) else used_afternoon_personnel.add(person)
                    daily_stats['rooms'][room_num][person] += 1
                    if slot in early_slots:
                        daily_stats['early'][person] += 1
                    elif slot in late_slots:
                        daily_stats['late'][person] += 1
                    elif slot in duty_slots:
                        daily_stats['duty'][person] += 1
        
        morning_slots_all = morning_slots_830 + morning_slots_900 + morning_slots_930 + morning_slots_1000
        random.shuffle(morning_slots_all)
        morning_indices = [slots.index(slot) for slot in morning_slots_all if slots.index(slot) in available_slots]
        for slot_idx in morning_indices:
            slot = slots[slot_idx]
            assigned = False
            for person in morning_personnel_list:
                if person not in used_morning_personnel and person not in fixed_personnel:
                    room_num = re.search(r'\((\d+)\)', slot).group(1)
                    early_count = total_stats['early'][person] + daily_stats['early'][person]
                    late_count = total_stats['late'][person] + daily_stats['late'][person]
                    room_count = total_stats['rooms'][room_num][person] + daily_stats['rooms'][room_num][person]
                    
                    if room_count < max_room or (not assigned and room_count >= max_room):
                        assignment[slot_idx] = person
                        assigned_counts[person] += 1
                        available_slots.remove(slot_idx)
                        used_morning_personnel.add(person)
                        daily_stats['rooms'][room_num][person] += 1
                        if slot in early_slots:
                            daily_stats['early'][person] += 1
                        elif slot in late_slots:
                            daily_stats['late'][person] += 1
                        assigned = True
                        break
            if not assigned:
                st.warning(f"슬롯 {slot} 배정 실패: 적합한 오전 인원 없음")
        
        afternoon_indices = [i for i in available_slots if slots[i] in afternoon_slots]
        for slot_idx in afternoon_indices:
            slot = slots[slot_idx]
            assigned = False
            for person in afternoon_personnel_list:
                if person not in used_afternoon_personnel and person not in fixed_personnel:
                    room_num = re.search(r'\((\d+)\)', slot).group(1)
                    room_count = total_stats['rooms'][room_num][person] + daily_stats['rooms'][room_num][person]
                    
                    if room_count < max_room or (not assigned and room_count >= max_room):
                        assignment[slot_idx] = person
                        assigned_counts[person] += 1
                        available_slots.remove(slot_idx)
                        used_afternoon_personnel.add(person)
                        daily_stats['rooms'][room_num][person] += 1
                        assigned = True
                        break
            if not assigned:
                st.warning(f"슬롯 {slot} 배정 실패: 적합한 오후 인원 없음")
        
        duty_indices = [i for i in available_slots if slots[i] in duty_slots]
        for slot_idx in duty_indices:
            slot = slots[slot_idx]
            assigned = False
            for person in afternoon_personnel_list:
                if person not in used_afternoon_personnel and person not in fixed_personnel:
                    room_num = re.search(r'\((\d+)\)', slot).group(1)
                    duty_count = total_stats['duty'][person] + daily_stats['duty'][person]
                    room_count = total_stats['rooms'][room_num][person] + daily_stats['rooms'][room_num][person]
                    
                    if (duty_count < max_duty and room_count < max_room) or (not assigned and (duty_count >= max_duty or room_count >= max_room)):
                        assignment[slot_idx] = person
                        assigned_counts[person] += 1
                        available_slots.remove(slot_idx)
                        used_afternoon_personnel.add(person)
                        daily_stats['duty'][person] += 1
                        daily_stats['rooms'][room_num][person] += 1
                        assigned = True
                        break
            if not assigned:
                st.warning(f"슬롯 {slot} 배정 실패: 적합한 당직 인원 없음")
        
        unassigned_count = sum(1 for i in range(len(slots)) if assignment[i] is None)
        if unassigned_count < min_unassigned:
            min_unassigned = unassigned_count
            best_assignment = assignment.copy()
            best_stats = daily_stats.copy()
        
        if unassigned_count == 0:
            break
    
    if min_unassigned > 0:
        st.error(f"배정 실패: {min_unassigned}개의 슬롯이 비어 있음. 최대 배정 한계를 조정하거나 인원을 확인하세요.")
    
    for key in ['early', 'late', 'duty']:
        total_stats[key].update(best_stats[key])
    for room in best_stats['rooms']:
        total_stats['rooms'][room].update(best_stats['rooms'][room])
    
    return best_assignment, best_stats

# 메인
month_str = "2025년 04월"
next_month_start = date(2025, 4, 1)
next_month_end = date(2025, 4, 30)

# 로그인 체크
if "login_success" not in st.session_state or not st.session_state["login_success"]:
    st.warning("⚠️ Home 페이지에서 비밀번호와 사번을 먼저 입력해주세요.")
    st.stop()

# 관리자 권한 체크
if not st.session_state.get("is_admin_authenticated", False):
    st.warning("⚠️ 관리자 권한이 없습니다.")
    st.stop()

# 사이드바
st.sidebar.write(f"현재 사용자: {st.session_state['name']} ({str(st.session_state['employee_id']).zfill(5)})")
if st.sidebar.button("로그아웃"):
    st.session_state.clear()
    st.success("로그아웃되었습니다.")
    st.rerun()

# 최대 배정 한계
st.sidebar.header("최대 배정 한계 설정")
MAX_DUTY = st.sidebar.number_input("최대 당직 합계", min_value=1, value=3, step=1)
MAX_EARLY = st.sidebar.number_input("최대 이른방 합계", min_value=1, value=6, step=1)
MAX_LATE = st.sidebar.number_input("최대 늦은방 합계", min_value=1, value=6, step=1)
MAX_ROOM = st.sidebar.number_input("최대 방별 합계", min_value=1, value=3, step=1)

# 데이터 로드 호출
df_schedule, df_room_request, df_room_fix, worksheet_request, worksheet_fix = load_data(month_str)
st.session_state["df_room_fix"] = df_room_fix
st.session_state["df_room_request"] = df_room_request
st.session_state["worksheet_fix"] = worksheet_fix
st.session_state["worksheet_request"] = worksheet_request

# df_schedule_md 초기화
if "df_schedule_md" not in st.session_state:
    st.session_state["df_schedule_md"] = create_df_schedule_md(df_schedule)

# 새로고침 버튼 (맨 위로 이동)
if st.button("🔄 새로고침 (R)"):
    st.cache_data.clear()
    df_schedule, df_room_request, df_room_fix, worksheet_request, worksheet_fix = load_data_no_cache(month_str)
    st.session_state["df_schedule"] = df_schedule
    st.session_state["df_room_request"] = df_room_request
    st.session_state["df_room_fix"] = df_room_fix
    st.session_state["worksheet_request"] = worksheet_request
    st.session_state["worksheet_fix"] = worksheet_fix
    st.session_state["df_schedule_md"] = create_df_schedule_md(df_schedule)
    st.success("데이터가 새로고침되었습니다.")
    st.rerun()

# 근무자 명단
st.subheader("📋 근무자 명단")
st.dataframe(st.session_state["df_schedule_md"])

# 방 설정 UI
st.divider()
st.subheader("📋 방 설정")
room_options = [str(i) for i in range(1, 13)]

# 방 설정 UI
st.divider()
st.subheader("📋 방 설정")
room_options = [str(i) for i in range(1, 13)]

st.markdown("**🔷 8:30 시간대**")
col1, col2, col3 = st.columns([1, 2, 1])
with col1:
    num_830 = st.number_input("방 개수", min_value=0, value=4, key="830_rooms")
with col2:
    if len(st.session_state["room_settings"]["830_room_select"]) > num_830:
        st.session_state["room_settings"]["830_room_select"] = st.session_state["room_settings"]["830_room_select"][:num_830]
    rooms_830 = st.multiselect(
        "방 번호",
        room_options,
        default=st.session_state["room_settings"]["830_room_select"],
        max_selections=num_830,
        key="830_room_select"
    )
    if len(rooms_830) < num_830:
        st.warning(f"방 개수({num_830})에 맞게 방 번호를 {num_830}개 선택해주세요. 현재 {len(rooms_830)}개 선택됨.")
    st.session_state["room_settings"]["830_room_select"] = rooms_830
with col3:
    duty_830_options = rooms_830 if rooms_830 else room_options
    duty_830 = st.selectbox("당직방", duty_830_options, index=0, key="830_duty")
    st.session_state["room_settings"]["830_duty"] = duty_830

st.markdown("**🔷 9:00 시간대**")
col1, col2 = st.columns([1, 3])
with col1:
    num_900 = st.number_input("방 개수", min_value=0, value=3, key="900_rooms")
with col2:
    if len(st.session_state["room_settings"]["900_room_select"]) > num_900:
        st.session_state["room_settings"]["900_room_select"] = st.session_state["room_settings"]["900_room_select"][:num_900]
    rooms_900 = st.multiselect(
        "방 번호",
        room_options,
        default=st.session_state["room_settings"]["900_room_select"],
        max_selections=num_900,
        key="900_room_select"
    )
    if len(rooms_900) < num_900:
        st.warning(f"방 개수({num_900})에 맞게 방 번호를 {num_900}개 선택해주세요. 현재 {len(rooms_900)}개 선택됨.")
    st.session_state["room_settings"]["900_room_select"] = rooms_900

st.markdown("**🔷 9:30 시간대**")
col1, col2 = st.columns([1, 3])
with col1:
    num_930 = st.number_input("방 개수", min_value=0, value=3, key="930_rooms")
with col2:
    if len(st.session_state["room_settings"]["930_room_select"]) > num_930:
        st.session_state["room_settings"]["930_room_select"] = st.session_state["room_settings"]["930_room_select"][:num_930]
    rooms_930 = st.multiselect(
        "방 번호",
        room_options,
        default=st.session_state["room_settings"]["930_room_select"],
        max_selections=num_930,
        key="930_room_select"
    )
    if len(rooms_930) < num_930:
        st.warning(f"방 개수({num_930})에 맞게 방 번호를 {num_930}개 선택해주세요. 현재 {len(rooms_930)}개 선택됨.")
    st.session_state["room_settings"]["930_room_select"] = rooms_930

st.markdown("**🔷 10:00 시간대**")
col1, col2 = st.columns([1, 3])
with col1:
    num_1000 = st.number_input("방 개수", min_value=0, value=2, key="1000_rooms")
with col2:
    if len(st.session_state["room_settings"]["1000_room_select"]) > num_1000:
        st.session_state["room_settings"]["1000_room_select"] = st.session_state["room_settings"]["1000_room_select"][:num_1000]
    rooms_1000 = st.multiselect(
        "방 번호",
        room_options,
        default=st.session_state["room_settings"]["1000_room_select"],
        max_selections=num_1000,
        key="1000_room_select"
    )
    if len(rooms_1000) < num_1000:
        st.warning(f"방 개수({num_1000})에 맞게 방 번호를 {num_1000}개 선택해주세요. 현재 {len(rooms_1000)}개 선택됨.")
    st.session_state["room_settings"]["1000_room_select"] = rooms_1000

st.markdown("**🔶 13:30 시간대**")
col1, col2 = st.columns([3, 1])
with col1:
    num_1330 = 4
    if len(st.session_state["room_settings"]["1330_room_select"]) > num_1330:
        st.session_state["room_settings"]["1330_room_select"] = st.session_state["room_settings"]["1330_room_select"][:num_1330]
    rooms_1330 = st.multiselect(
        "방 번호",
        room_options,
        default=st.session_state["room_settings"]["1330_room_select"],
        max_selections=num_1330,
        key="1330_room_select"
    )
    if len(rooms_1330) < num_1330:
        st.warning(f"방 개수({num_1330})에 맞게 방 번호를 {num_1330}개 선택해주세요. 현재 {len(rooms_1330)}개 선택됨.")
    st.session_state["room_settings"]["1330_room_select"] = rooms_1330
with col2:
    duty_1330_options = rooms_1330 if rooms_1330 else room_options
    duty_1330 = st.selectbox("당직방", duty_1330_options, index=0, key="1330_duty")
    st.session_state["room_settings"]["1330_duty"] = duty_1330

# 중복 방 번호 검증
all_selected_rooms = (
    st.session_state["room_settings"]["830_room_select"] +
    st.session_state["room_settings"]["900_room_select"] +
    st.session_state["room_settings"]["930_room_select"] +
    st.session_state["room_settings"]["1000_room_select"] +
    st.session_state["room_settings"]["1330_room_select"]
)

# 고정 배치 입력 UI
st.divider()
st.subheader("📋 고정 배치 관리")
st.write("- 고정 배치 기능은 관리자만 제어할 수 있습니다. (사용자가 개별 입력 불가)")
요청분류 = ["1번방", "2번방", "3번방", "4번방", "5번방", "6번방", "7번방", "8번방", "9번방", "10번방", "11번방",
           "8:30", "9:00", "9:30", "10:00", "당직 아닌 이른방", "이른방 제외", "늦은방 제외"]
with st.form("fixed_form"):
    st.markdown("**🟢 고정 배치 추가**")
    col1, col2, col3 = st.columns([2, 2, 3])
    with col1:
        names = [str(name).strip() for name in df_schedule.iloc[:, 2:].stack().dropna().unique() if str(name).strip()]
        name = st.selectbox("근무자", names)
    with col2:
        categories = st.multiselect("요청 분류", 요청분류)
    with col3:
        dates = st.multiselect("요청 일자", get_user_available_dates(name, df_schedule, next_month_start, next_month_end))
    
    if st.form_submit_button("📅 추가"):
        if not categories or not dates:
            st.error("요청 분류와 날짜를 선택해주세요.")
        else:
            new_rows = []
            for date in dates:
                for cat in categories:
                    new_rows.append({"이름": name, "분류": cat, "날짜정보": date})
            df_room_fix = pd.concat([df_room_fix, pd.DataFrame(new_rows)], ignore_index=True)
            st.session_state["df_room_fix"] = df_room_fix
            try:
                gc = get_gspread_client()
                sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
                worksheet_fix = sheet.worksheet(f"{month_str} 방배정 고정")
                update_sheet_with_retry(worksheet_fix, [df_room_fix.columns.tolist()] + df_room_fix.values.tolist())
                st.session_state["worksheet_fix"] = worksheet_fix
                st.cache_data.clear()
                st.success("고정 배치 저장 완료!")
            except Exception as e:
                st.error(f"Google Sheets 업데이트 실패: {str(e)}")
                st.write("로컬 df_room_fix는 업데이트되었습니다. 아래에서 확인 후 Google Sheets 동기화를 다시 시도해주세요.")
                st.write(f"df_room_fix columns: {df_room_fix.columns.tolist()}")
                st.dataframe(st.session_state["df_room_fix"])
    
    if not df_room_fix.empty:
        st.markdown("**🔴 고정 배치 삭제**")
        col0, col1 = st.columns([1, 2])
        with col0:
            selected_employee = st.selectbox("근무자 선택", df_room_fix["이름"].unique(), key="delete_fix_employee_select")
        with col1:
            df_fix_filtered = df_room_fix[df_room_fix["이름"] == selected_employee]
            if not df_fix_filtered.empty:
                options = [f"{row['분류']} - {row['날짜정보']}" for _, row in df_fix_filtered.iterrows()]
                selected_items = st.multiselect("삭제할 항목 (고정 배치)", options, key="delete_fix_select")
            else:
                st.info("📍 선택한 근무자에 대한 고정 배치 요청이 없습니다.")
                selected_items = []
        
        if st.form_submit_button("📅 삭제"):
            if selected_items:
                indices = []
                for item in selected_items:
                    for idx, row in df_fix_filtered.iterrows():
                        if f"{row['분류']} - {row['날짜정보']}" == item:
                            indices.append(idx)
                df_room_fix = df_room_fix.drop(indices).reset_index(drop=True)
                st.session_state["df_room_fix"] = df_room_fix
                try:
                    gc = get_gspread_client()
                    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
                    worksheet_fix = sheet.worksheet(f"{month_str} 방배정 고정")
                    update_sheet_with_retry(worksheet_fix, [df_room_fix.columns.tolist()] + df_room_fix.values.tolist())
                    st.session_state["worksheet_fix"] = worksheet_fix
                    st.cache_data.clear()
                    st.success("선택한 고정 배치 삭제 완료!")
                except Exception as e:
                    st.error(f"Google Sheets 업데이트 실패: {str(e)}")
                    st.write("로컬 df_room_fix는 업데이트되었습니다. 아래에서 확인 후 Google Sheets 동기화를 다시 시도해주세요.")
                    st.write(f"df_room_fix columns: {df_room_fix.columns.tolist()}")
                    st.dataframe(st.session_state["df_room_fix"])
    
    st.write(" ")
    st.markdown("**🙋‍♂️ 현재 고정 배치 요청 목록**")
    if df_room_fix.empty:
        st.info("☑️ 현재 고정 배치 요청이 없습니다.")
    else:
        st.dataframe(df_room_fix, use_container_width=True)

# 우선 배치 입력 UI
st.divider()
st.subheader("📋 우선 배치 관리")
st.write("- 모든 인원의 우선 배치 요청을 추가 및 수정할 수 있습니다.")

요청분류 = ["1번방", "2번방", "3번방", "4번방", "5번방", "6번방", "7번방", "8번방", "9번방", "10번방", "11번방",
            "이른방", "당직 아닌 이른방", "8:30", "9:00", "9:30", "10:00"]
with st.form("priority_form"):
    st.markdown("**🟢 우선 배치 추가**")
    col1, col2, col3 = st.columns([2, 2, 3])
    with col1:
        names = [str(name).strip() for name in df_schedule.iloc[:, 2:].stack().dropna().unique() if str(name).strip()]
        name = st.selectbox("근무자", names)
    with col2:
        categories = st.multiselect("요청 분류", 요청분류)
    with col3:
        dates = st.multiselect("요청 일자", get_user_available_dates(name, df_schedule, next_month_start, next_month_end))
    
    if st.form_submit_button("📅 추가"):
        if not categories or not dates:
            st.error("요청 분류와 날짜를 선택해주세요.")
        else:
            new_rows = []
            for date in dates:
                for cat in categories:
                    new_rows.append({
                        "이름": name,
                        "분류": cat,
                        "날짜정보": date
                    })
            df_room_request = pd.concat([df_room_request, pd.DataFrame(new_rows)], ignore_index=True)
            st.session_state["df_room_request"] = df_room_request
            try:
                gc = get_gspread_client()
                sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
                worksheet_request = sheet.worksheet(f"{month_str} 방배정 요청")
                update_sheet_with_retry(worksheet_request, [df_room_request.columns.tolist()] + df_room_request.values.tolist())
                st.session_state["worksheet_request"] = worksheet_request
                st.cache_data.clear()
                st.success("우선 배치 저장 완료!")
            except Exception as e:
                st.error(f"Google Sheets 업데이트 실패: {str(e)}")
                st.write("로컬 df_room_request는 업데이트되었습니다. 아래에서 확인 후 Google Sheets 동기화를 다시 시도해주세요.")
                st.write(f"df_room_request columns: {df_room_request.columns.tolist()}")
                st.dataframe(st.session_state["df_room_request"])
    
    if not df_room_request.empty:
        st.markdown("**🔴 우선 배치 삭제**")
        col0, col1 = st.columns([1, 2])
        with col0:
            selected_employee = st.selectbox("근무자 선택", df_room_request["이름"].unique(), key="delete_request_employee_select")
        with col1:
            df_request_filtered = df_room_request[df_room_request["이름"] == selected_employee]
            if not df_request_filtered.empty:
                options = [f"{row['분류']} - {row['날짜정보']}" for _, row in df_request_filtered.iterrows()]
                selected_items = st.multiselect("삭제할 항목 (우선 배치)", options, key="delete_request_select")
            else:
                st.info("📍 선택한 근무자에 대한 우선 배치 요청이 없습니다.")
                selected_items = []
        
        if st.form_submit_button("📅 삭제"):
            if selected_items:
                indices = []
                for item in selected_items:
                    for idx, row in df_request_filtered.iterrows():
                        if f"{row['분류']} - {row['날짜정보']}" == item:
                            indices.append(idx)
                df_room_request = df_room_request.drop(indices).reset_index(drop=True)
                st.session_state["df_room_request"] = df_room_request
                try:
                    gc = get_gspread_client()
                    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
                    worksheet_request = sheet.worksheet(f"{month_str} 방배정 요청")
                    update_sheet_with_retry(worksheet_request, [df_room_request.columns.tolist()] + df_room_request.values.tolist())
                    st.session_state["worksheet_request"] = worksheet_request
                    st.cache_data.clear()
                    st.success("선택한 우선 배치 삭제 완료!")
                except Exception as e:
                    st.error(f"Google Sheets 업데이트 실패: {str(e)}")
                    st.write("로컬 df_room_request는 업데이트되었습니다. 아래에서 확인 후 Google Sheets 동기화를 다시 시도해주세요.")
                    st.write(f"df_room_request columns: {df_room_request.columns.tolist()}")
                    st.dataframe(st.session_state["df_room_request"])
    
    st.write(" ")
    st.markdown("**🙋‍♂️ 현재 우선 배치 요청 목록**")
    if df_room_request.empty:
        st.info("☑️ 현재 우선 배치 요청이 없습니다.")
    else:
        st.dataframe(df_room_request, use_container_width=True)

# df_room 생성 로직
st.divider()
st.subheader("✨ 2025년 04월 내시경실 배정 확인")
if st.button("🚀 방배정 시작"):
    # 방 설정 입력값 검증 및 처리
    time_slots = {}
    time_groups = {}
    memo_rules = {}
    if num_830 + num_900 + num_930 + num_1000 != 12:
        st.error("오전 방 개수 합계는 12여야 합니다.")
        st.stop()
    elif len(rooms_830) != num_830 or len(rooms_900) != num_900 or len(rooms_930) != num_930 or len(rooms_1000) != num_1000 or len(rooms_1330) != num_1330:
        st.error("각 시간대의 방 번호 선택을 완료해주세요.")
        st.stop()
    else:
        for room in rooms_830:
            slot = f"8:30({room})_당직" if room == duty_830 else f"8:30({room})"
            time_slots[slot] = len(time_slots)
            time_groups.setdefault('8:30', []).append(slot)
        for room in rooms_900:
            slot = f"9:00({room})"
            time_slots[slot] = len(time_slots)
            time_groups.setdefault('9:00', []).append(slot)
        for room in rooms_930:
            slot = f"9:30({room})"
            time_slots[slot] = len(time_slots)
            time_groups.setdefault('9:30', []).append(slot)
        for room in rooms_1000:
            slot = f"10:00({room})"
            time_slots[slot] = len(time_slots)
            time_groups.setdefault('10:00', []).append(slot)
        for room in rooms_1330:
            slot = f"13:30({room})_당직" if room == duty_1330 else f"13:30({room})"
            time_slots[slot] = len(time_slots)
            time_groups.setdefault('13:30', []).append(slot)
        
        memo_rules = {
            **{f'{i}번방': [s for s in time_slots if f'({i})' in s] for i in range(1, 13)},
            '당직 아닌 이른방': [s for s in time_slots if s.startswith('8:30') and '_당직' not in s],
            '이른방 제외': [s for s in time_slots if s.startswith(('9:00', '9:30', '10:00'))],
            '늦은방 제외': [s for s in time_slots if s.startswith(('8:30', '9:00', '9:30'))],
            '8:30': [s for s in time_slots if s.startswith('8:30')],
            '9:00': [s for s in time_slots if s.startswith('9:00')],
            '9:30': [s for s in time_slots if s.startswith('9:30')],
            '10:00': [s for s in time_slots if s.startswith('10:00')]
        }
        
        # 세션 상태에 저장
        st.session_state["time_slots"] = time_slots
        st.session_state["time_groups"] = time_groups
        st.session_state["memo_rules"] = memo_rules
        st.session_state["morning_slots_830"] = [s for s in time_slots if s.startswith('8:30') and '_당직' not in s]
        st.session_state["morning_slots_900"] = [s for s in time_slots if s.startswith('9:00')]
        st.session_state["morning_slots_930"] = [s for s in time_slots if s.startswith('9:30')]
        st.session_state["morning_slots_1000"] = [s for s in time_slots if s.startswith('10:00')]
        st.session_state["afternoon_slots"] = [s for s in time_slots if s.startswith('13:30') and '_당직' not in s]
        st.session_state["duty_slots"] = [s for s in time_slots if s.startswith('13:30') and '_당직' in s]
    
    # 배정 로직 시작
    random.seed(time.time())
    total_stats = {
        'early': Counter(),
        'late': Counter(),
        'duty': Counter(),
        'rooms': {str(i): Counter() for i in range(1, 13)}
    }
    
    assignments = {}
    slots = list(st.session_state["time_slots"].keys())
    assignable_slots = [s for s in slots if not (s.startswith('8:30') and s.endswith('_당직'))]
    
    morning_slots_830 = st.session_state["morning_slots_830"]
    morning_slots_900 = st.session_state["morning_slots_900"]
    morning_slots_930 = st.session_state["morning_slots_930"]
    morning_slots_1000 = st.session_state["morning_slots_1000"]
    afternoon_slots = st.session_state["afternoon_slots"]
    duty_slots = st.session_state["duty_slots"]
    
    date_cache = {}
    for _, row in st.session_state["df_schedule_md"].iterrows():
        date_str = row['날짜']
        try:
            if "월" in date_str:
                date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=2025)
            else:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            date_cache[date_str] = date_obj.strftime('%Y-%m-%d')
        except ValueError:
            continue
    
    fixed_cells = {}
    priority_cells = {}
    for _, row in st.session_state["df_schedule_md"].iterrows():
        date_str = row['날짜']
        if date_str not in date_cache:
            continue
        
        formatted_date = date_cache[date_str]
        morning_personnel = [row[str(i)] for i in range(1, 12) if pd.notna(row[str(i)]) and row[str(i)]]
        afternoon_personnel = [row[f'오후{i}'] for i in range(1, 5) if pd.notna(row[f'오후{i}']) and row[f'오후{i}']]
        personnel = morning_personnel + afternoon_personnel
        
        if not personnel:
            assignments[formatted_date] = [None] * len(assignable_slots)
            continue
        
        if len(morning_personnel) < 11 or len(afternoon_personnel) < 4:
            st.warning(f"{date_str}: 인원 부족, 오전 {len(morning_personnel)}명(필요 11명), 오후 {len(afternoon_personnel)}명(필요 4명)")
        
        fixed_assignments = {}
        if not df_room_fix.empty:
            for _, fix in df_room_fix.iterrows():
                fix_date, is_morning = parse_date_info(fix['날짜정보'])
                if fix_date == formatted_date:
                    for slot in st.session_state["memo_rules"].get(fix['분류'], []):
                        if slot in assignable_slots:
                            if is_morning and slot in (morning_slots_830 + morning_slots_900 + morning_slots_930 + morning_slots_1000):
                                fixed_assignments[slot] = fix['이름']
                                fixed_cells[(formatted_date, slot)] = fix['분류']
                            elif not is_morning and slot in (afternoon_slots + duty_slots):
                                fixed_assignments[slot] = fix['이름']
                                fixed_cells[(formatted_date, slot)] = fix['분류']
        
        priority_assignments = {}
        if not df_room_request.empty:
            for _, req in df_room_request.iterrows():
                req_date, is_morning = parse_date_info(req['날짜정보'])
                if req_date == formatted_date:
                    for slot in st.session_state["memo_rules"].get(req['분류'], []):
                        if slot in assignable_slots:
                            if is_morning and slot in (morning_slots_830 + morning_slots_900 + morning_slots_930 + morning_slots_1000):
                                priority_assignments[(slot, req['이름'])] = 0.5  # 우선순위 제거, 기본값 사용
                                priority_cells[(formatted_date, slot)] = req['분류']
                            elif not is_morning and slot in (afternoon_slots + duty_slots):
                                priority_assignments[(slot, req['이름'])] = 0.5
                                priority_cells[(formatted_date, slot)] = req['분류']
        
        assignment, daily_stats = random_assign(
            personnel, assignable_slots, fixed_assignments, priority_assignments,
            st.session_state["time_groups"], MAX_EARLY, MAX_LATE, MAX_DUTY, MAX_ROOM, total_stats,
            morning_personnel, afternoon_personnel
        )
        assignments[formatted_date] = assignment
    
    all_slots = ['8:30(1)_당직'] + \
                sorted([s for s in slots if s.startswith('8:30') and not s.endswith('_당직')]) + \
                sorted([s for s in slots if s.startswith('9:00')]) + \
                sorted([s for s in slots if s.startswith('9:30')]) + \
                sorted([s for s in slots if s.startswith('10:00')]) + \
                ['온콜'] + \
                sorted([s for s in slots if s.startswith('13:30') and s.endswith('_당직')]) + \
                sorted([s for s in slots if s.startswith('13:30') and not s.endswith('_당직')])
    columns = ['날짜', '요일'] + all_slots
    result_data = []
    
    for _, row in st.session_state["df_schedule_md"].iterrows():
        date_str = row['날짜']
        if date_str not in date_cache:
            continue
        
        formatted_date = date_cache[date_str]
        day_of_week = row['요일'] + "요일" if not row['요일'].endswith("요일") else row['요일']
        result_row = [date_str, day_of_week]
        
        personnel = [p for p in row[2:-1] if pd.notna(p) and p]
        has_person = bool(personnel)
        
        assignment = assignments.get(formatted_date, [None] * len(assignable_slots))
        for slot in all_slots:
            if slot == '8:30(1)_당직' or slot == '온콜':
                person = row['오전당직(온콜)'] if has_person else None
            else:
                person = assignment[assignable_slots.index(slot)] if slot in assignable_slots and assignment else None
            result_row.append(person if has_person else None)
        
        result_data.append(result_row)
    
    df_room = pd.DataFrame(result_data, columns=columns)
    st.write(" ")
    st.markdown("**✅ 통합 배치 결과**")
    st.dataframe(df_room)
    
    stats_data = []
    all_personnel = set()
    for _, row in st.session_state["df_schedule_md"].iterrows():
        personnel = [p for p in row[2:-1].dropna() if p]
        all_personnel.update(personnel)
    
    for person in sorted(all_personnel):
        stats_data.append({
            '인원': person,
            '이른방 합계': total_stats['early'][person],
            '늦은방 합계': total_stats['late'][person],
            '당직 합계': total_stats['duty'][person],
            **{f'{r}번방 합계': total_stats['rooms'][r][person] for r in total_stats['rooms']}
        })
    
    stats_df = pd.DataFrame(stats_data)
    st.divider()
    st.markdown("**☑️ 인원별 통계**")
    st.dataframe(stats_df)
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Schedule"
    
    for col_idx, header in enumerate(columns, 1):
        cell = sheet.cell(1, col_idx, header)
        cell.font = Font(bold=True, name="맑은 고딕", size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        if header.startswith('8:30'):
            cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        elif header.startswith('9:00'):
            cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        elif header.startswith('9:30'):
            cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        elif header.startswith('10:00'):
            cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        elif header.startswith('13:30'):
            cell.fill = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")
        elif header == '온콜':
            cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    for row_idx, row in enumerate(result_data, 2):
        has_person = any(x for x in row[2:-1] if x is not None)
        formatted_date = date_cache[row[0]]
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row_idx, col_idx, value)
            if (columns[col_idx-1].endswith('_당직') or columns[col_idx-1] == '온콜') and value:
                cell.font = Font(name="맑은 고딕", size=9, bold=True, color="FF00FF")
            else:
                cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            if col_idx == 1:
                cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            elif not has_person and col_idx >= 2:
                cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            elif col_idx == 2:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            if col_idx > 2 and value:
                slot = columns[col_idx-1]
                if (formatted_date, slot) in fixed_cells:
                    cell.comment = Comment(f"고정 배치: {fixed_cells[(formatted_date, slot)]}", "System")
                elif (formatted_date, slot) in priority_cells:
                    cell.comment = Comment(f"우선 배치: {priority_cells[(formatted_date, slot)]}", "System")
    
    stats_sheet = wb.create_sheet("Stats")
    stats_columns = stats_df.columns.tolist()
    for col_idx, header in enumerate(stats_columns, 1):
        cell = stats_sheet.cell(1, col_idx, header)
        cell.font = Font(bold=True, name="맑은 고딕", size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        if header == '인원':
            cell.fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
        elif header == '이른방 합계':
            cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        elif header == '늦은방 합계':
            cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        elif header == '당직 합계':
            cell.fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
    
    for row_idx, row in enumerate(stats_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = stats_sheet.cell(row_idx, col_idx, value)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    st.divider()
    st.download_button(
        label="📥 최종 방배정 다운로드",
        data=output,
        file_name=f"{datetime.today().strftime('%Y-%m-%d')}_내시경실배정.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )

    # Google Sheets에 방배정 시트 저장
    gc = get_gspread_client()
    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
    try:
        worksheet_result = sheet.worksheet(f"{month_str} 방배정")
    except:
        worksheet_result = sheet.add_worksheet(f"{month_str} 방배정", rows=100, cols=len(df_room.columns))
        worksheet_result.append_row(df_room.columns.tolist())

    update_sheet_with_retry(worksheet_result, [df_room.columns.tolist()] + df_room.values.tolist())
    st.success(f"✅ {month_str} 방배정 테이블이 Google Sheets에 저장되었습니다.")