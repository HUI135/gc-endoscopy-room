import re
import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from collections import Counter
import random
import time
from datetime import datetime, date, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
import menu
import numpy as np
from dateutil.relativedelta import relativedelta

st.set_page_config(page_title="방 배정", page_icon="", layout="wide")

import os
st.session_state.current_page = os.path.basename(__file__)

menu.menu()

# 로그인 체크 및 자동 리디렉션
if not st.session_state.get("login_success", False):
    st.warning("⚠️ Home 페이지에서 먼저 로그인해주세요.")
    st.error("1초 후 Home 페이지로 돌아갑니다...")
    time.sleep(1)
    st.switch_page("Home.py")
    st.stop()

# 세션 상태 초기화
def initialize_session_state():
    if "data_loaded" not in st.session_state:
        st.session_state["data_loaded"] = False
    if "df_room_request" not in st.session_state:
        st.session_state["df_room_request"] = pd.DataFrame(columns=["이름", "분류", "날짜정보"])
    if "room_settings" not in st.session_state:
        st.session_state["room_settings"] = {
            "830_room_select": ['1', '8', '4', '7'],
            "900_room_select": ['10', '11', '12'],
            "930_room_select": ['2', '5', '6'],
            "1000_room_select": ['9', '3'],
            "1330_room_select": ['3', '4', '9', '2']
        }
    if "weekend_room_settings" not in st.session_state:
        st.session_state["weekend_room_settings"] = {}
    if "swapped_assignments" not in st.session_state:
        st.session_state["swapped_assignments"] = set()
    if "df_schedule_original" not in st.session_state:
        st.session_state["df_schedule_original"] = pd.DataFrame()
    if "manual_change_log" not in st.session_state:
        st.session_state["manual_change_log"] = []
    if "final_change_log" not in st.session_state:
        st.session_state["final_change_log"] = []
    if "saved_changes_log" not in st.session_state:
        st.session_state["saved_changes_log"] = []
    if "df_schedule_md_initial" not in st.session_state:
        st.session_state["df_schedule_md_initial"] = pd.DataFrame()
    if "swapped_assignments_log" not in st.session_state:
        st.session_state["swapped_assignments_log"] = []
    if "df_schedule" not in st.session_state:
        st.session_state["df_schedule"] = pd.DataFrame()
    if "df_swap_requests" not in st.session_state:
        st.session_state["df_swap_requests"] = pd.DataFrame(columns=[
            "RequestID", "요청일시", "요청자", "변경 요청", "변경 요청한 스케줄"
        ])
    if "worksheet_room_request" not in st.session_state:
        st.session_state["worksheet_room_request"] = None

# Google Sheets 클라이언트 초기화
def get_gspread_client():
    scope = ["https://www.googleapis.com/auth/spreadsheets"]
    try:
        service_account_info = dict(st.secrets["gspread"])
        service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
        credentials = Credentials.from_service_account_info(service_account_info, scopes=scope)
        return gspread.authorize(credentials)
    except Exception as e:
        st.error(f"Google Sheets 인증 정보 로드 중 오류: {type(e).__name__} - {e}")
        return None

# Google Sheets 업데이트 함수
def update_sheet_with_retry(worksheet, data, retries=5, delay=10):
    for attempt in range(retries):
        try:
            worksheet.clear()
            worksheet.update('A1', data, value_input_option='RAW')
            return True
        except Exception as e:
            if "Quota exceeded" in str(e):
                st.warning(f"API 쿼터 초과, {delay}초 후 재시도 ({attempt+1}/{retries})")
                time.sleep(delay)
            else:
                st.error(f"업데이트 실패, {delay}초 후 재시도 ({attempt+1}/{retries}): {str(e)}")
                time.sleep(delay)
    st.error("Google Sheets 업데이트 실패: 재시도 횟수 초과")
    return False

# 데이터 로드 함수
def load_data_page6_no_cache(month_str, retries=3, delay=5):
    try:
        gc = get_gspread_client()
        if gc is None:
            raise Exception("Failed to initialize gspread client")
        sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])

        # 스케줄 시트
        try:
            worksheet_schedule = sheet.worksheet(f"{month_str} 스케줄")
        except gspread.exceptions.WorksheetNotFound:
            st.info("스케줄이 아직 배정되지 않았습니다.")
            return "STOP", None, None, None, None  # 실행 중단을 위한 특별 반환 값

        df_schedule = pd.DataFrame(worksheet_schedule.get_all_records())
        if df_schedule.empty:  # Fixed line
            raise Exception(f"{month_str} 스케줄 시트가 비어 있습니다.")

        # 방배정 요청 시트
        try:
            worksheet_room_request = sheet.worksheet(f"{month_str} 방배정 요청")
        except gspread.exceptions.WorksheetNotFound:
            st.warning(f"{month_str} 방배정 요청 시트가 없습니다. 빈 테이블로 시작합니다.")
            worksheet_room_request = sheet.add_worksheet(title=f"{month_str} 방배정 요청", rows=100, cols=10)
            worksheet_room_request.update('A1', [["이름", "분류", "날짜정보"]])
        
        df_room_request = pd.DataFrame(worksheet_room_request.get_all_records())
        if "우선순위" in df_room_request.columns:
            df_room_request = df_room_request.drop(columns=["우선순위"])

        # 누적 시트
        worksheet_cumulative = sheet.worksheet(f"{month_str} 누적")
        df_cumulative = pd.DataFrame(worksheet_cumulative.get_all_records())
        if df_cumulative.empty:
            df_cumulative = pd.DataFrame(columns=["이름", "오전누적", "오후누적", "오전당직 (온콜)", "오후당직"])
        else:
            df_cumulative.rename(columns={f"{month_str}": "이름"}, inplace=True)

        # 스케줄 변경요청 시트
        try:
            worksheet_swap_requests = sheet.worksheet(f"{month_str} 스케줄 변경요청")
        except gspread.exceptions.WorksheetNotFound:
            st.warning(f"{month_str} 스케줄 변경요청 시트를 찾을 수 없어 새로 생성합니다.")
            worksheet_swap_requests = sheet.add_worksheet(title=f"{month_str} 스케줄 변경요청", rows=100, cols=10)
            worksheet_swap_requests.update('A1', [["RequestID", "요청일시", "요청자", "변경 요청", "변경 요청한 스케줄"]])
        
        df_swap_requests = pd.DataFrame(worksheet_swap_requests.get_all_records())
        if df_swap_requests.empty:
            st.info(f"{month_str} 스케줄 변경요청이 아직 존재하지 않습니다.")

        return df_schedule, df_room_request, worksheet_room_request, df_cumulative, df_swap_requests

    except gspread.exceptions.APIError as e:
        st.warning(f"Google Sheets API 오류: {e.response.status_code} - {e.response.text}")
    except Exception as e:
        st.error(f"데이터 로드 중 오류: {type(e).__name__} - {e}")
        
    st.error("데이터 로드 실패: 새로고침 버튼을 눌러 다시 시도해주세요.")
    return None, None, None, None, None

# 근무 가능 일자 계산
@st.cache_data
def get_user_available_dates(name, df_schedule, month_start, month_end):
    available_dates = []
    weekday_map = {0: "월", 1: "화", 2: "수", 3: "목", 4: "금", 5: "토", 6: "일"}
        
    personnel_columns = [str(i) for i in range(1, 12)] + ['오전당직(온콜)'] + [f'오후{i}' for i in range(1, 5)]
    all_personnel = set()
    for col in personnel_columns:
        if col in df_schedule.columns:
            for val in df_schedule[col].dropna():
                all_personnel.add(str(val).strip())
    if name not in all_personnel:
        st.warning(f"{name}이 df_schedule의 근무자 목록에 없습니다. 데이터 확인 필요: {sorted(all_personnel)}")
        time.sleep(1)

    for _, row in df_schedule.iterrows():
        date_str = row['날짜']
        try:
            if "월" in date_str:
                date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=2025).date()
            else:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            continue
        if month_start <= date_obj <= month_end:
            oncall_person = str(row['오전당직(온콜)']).strip() if '오전당직(온콜)' in row and pd.notna(row['오전당직(온콜)']) else ''
            morning_personnel = [str(row[str(i)]).strip() for i in range(1, 12) if str(i) in df_schedule.columns and pd.notna(row[str(i)]) and str(row[str(i)]).strip() and str(row[str(i)]).strip() != oncall_person]
            if '오전당직(온콜)' in df_schedule.columns and pd.notna(row['오전당직(온콜)']) and oncall_person == name:
                morning_personnel.append(name)
            afternoon_personnel = [
                str(row[f'오후{i}']).strip()
                for i in range(1, 5)
                if f'오후{i}' in df_schedule.columns
                and pd.notna(row[f'오후{i}'])
                and str(row[f'오후{i}']).strip()
                and str(row[f'오후{i}']).strip() != oncall_person
            ]
            
            display_date = f"{date_obj.month}월 {date_obj.day}일 ({weekday_map[date_obj.weekday()]})"
            save_date_am = f"{date_obj.strftime('%Y-%m-%d')} (오전)"
            save_date_pm = f"{date_obj.strftime('%Y-%m-%d')} (오후)"
            
            if name in morning_personnel:
                available_dates.append((date_obj, f"{display_date} 오전", save_date_am))
            if name in afternoon_personnel:
                available_dates.append((date_obj, f"{display_date} 오후", save_date_pm))
        
    available_dates.sort(key=lambda x: x[0])
    sorted_dates = [(display_str, save_str) for _, display_str, save_str in available_dates]
    if not sorted_dates:
        st.warning(f"{name}의 근무 가능 일자가 없습니다. df_schedule 데이터를 확인하세요.")
        time.sleep(1)
    return sorted_dates

# df_schedule_md 생성 함수
def create_df_schedule_md(df_schedule):
    display_cols = ['날짜', '요일', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '오전당직(온콜)', '오후1', '오후2', '오후3', '오후4']
    df_schedule_md = pd.DataFrame(columns=display_cols)
    if not df_schedule.empty:
        df_schedule_md['날짜'] = df_schedule['날짜']
        df_schedule_md['요일'] = df_schedule['요일']
        df_schedule_md['오전당직(온콜)'] = df_schedule['오전당직(온콜)']

    for idx, row in df_schedule.iterrows():
        oncall_person = str(row['오전당직(온콜)']).strip() if '오전당직(온콜)' in df_schedule.columns else ''
        am_original_cols = [str(i) for i in range(1, 13)]
        am_personnel_list = [
            str(row[col]).strip() for col in am_original_cols
            if col in df_schedule.columns and str(row[col]).strip() and str(row[col]).strip() != oncall_person
        ]
        am_personnel_unique = list(dict.fromkeys(am_personnel_list))
        am_display_cols = [str(i) for i in range(1, 12)]
        for i, col in enumerate(am_display_cols):
            df_schedule_md.at[idx, col] = am_personnel_unique[i] if i < len(am_personnel_unique) else ''
        
        pm_original_cols = [f'오후{i}' for i in range(1, 6)]
        pm_personnel_list = [
            str(row[col]).strip() for col in pm_original_cols
            if col in df_schedule.columns and str(row[col]).strip() and str(row[col]).strip() != oncall_person
        ]
        pm_personnel_unique = list(dict.fromkeys(pm_personnel_list))
        pm_display_cols = [f'오후{i}' for i in range(1, 5)]
        for i, col in enumerate(pm_display_cols):
            df_schedule_md.at[idx, col] = pm_personnel_unique[i] if i < len(pm_personnel_unique) else ''
            
    return df_schedule_md

def apply_schedule_swaps(original_schedule_df, swap_requests_df):
    df_modified = original_schedule_df.copy()
    applied_count = 0
    swapped_assignments = set()
    
    # 오전 및 오후 열 정의
    am_cols = [str(i) for i in range(1, 12)] + ['오전당직(온콜)']
    pm_cols = [f'오후{i}' for i in range(1, 6)]
    special_cols = am_cols
    
    batch_change_log = []
    
    client = get_gspread_client()
    if client is None:
        st.error("Google Sheets 클라이언트 초기화 실패")
        return create_df_schedule_md(df_modified)
    
    spreadsheet = client.open_by_url(st.secrets["google_sheet"]["url"])
    sheet_name = f"{month_str} 토요/휴일 일자"
    try:
        worksheet_special = spreadsheet.worksheet(sheet_name)
        special_data = worksheet_special.get_all_records()
        special_df = pd.DataFrame(special_data)
    except gspread.exceptions.WorksheetNotFound:
        st.warning(f"{sheet_name} 시트가 없습니다.")
        special_df = pd.DataFrame(columns=["날짜", "당직 인원"])

    # special_dates 정의 (함수 내부에서 보장)
    special_dates = [f"{datetime.strptime(row['날짜'], '%Y-%m-%d').month}월 {datetime.strptime(row['날짜'], '%Y-%m-%d').day}일" 
                     for row in special_data if row["날짜"]]

    for _, request_row in swap_requests_df.iterrows():
        try:
            change_request_str = str(request_row.get('변경 요청', '')).strip()
            if '➡️' not in change_request_str:
                st.warning(f"⚠️ 변경 요청 형식이 올바르지 않습니다: '{change_request_str}'. '이름1 ➡️ 이름2' 형식으로 입력해주세요.")
                time.sleep(1)
                continue

            requester_name, new_assignee = [p.strip() for p in change_request_str.split('➡️')]
            
            schedule_info_str = str(request_row.get('변경 요청한 스케줄', '')).strip()
            date_match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', schedule_info_str)
            
            if not date_match:
                st.warning(f"스케줄 정보 형식이 올바르지 않습니다: '{schedule_info_str}'. 'YYYY-MM-DD (오전)' 형식으로 입력해주세요.")
                time.sleep(1)
                continue

            date_part, time_period = date_match.groups()
            
            try:
                date_obj = datetime.strptime(date_part, '%Y-%m-%d').date()
                formatted_date_in_df = f"{date_obj.month}월 {date_obj.day}일"
            except ValueError:
                st.warning(f"날짜 파싱 오류로 요청을 건너뜁니다: {date_part}")
                time.sleep(1)
                continue

            target_row_indices = df_modified[df_modified['날짜'] == formatted_date_in_df].index
            if target_row_indices.empty:
                st.warning(f"스케줄에서 '{formatted_date_in_df}' 날짜를 찾을 수 없습니다. 요청을 건너뜁니다.")
                time.sleep(1)
                continue
            target_row_idx = target_row_indices[0]
            
            is_special_date = formatted_date_in_df in special_dates
            time_period_cols = special_cols if is_special_date and time_period == '오전' else (am_cols if time_period == '오전' else pm_cols)
            
            existing_assignments = []
            for col in time_period_cols:
                if col in df_modified.columns:
                    value = str(df_modified.at[target_row_idx, col]).strip()
                    if value and value != requester_name and value != 'nan':
                        existing_assignments.append(value)
            existing_assignments = list(dict.fromkeys(existing_assignments))
            
            if new_assignee in existing_assignments and (formatted_date_in_df, time_period, new_assignee) not in swapped_assignments:
                st.warning(f"⚠️ '{new_assignee}'님은 이미 {formatted_date_in_df} {time_period} 시간대에 배정되어 있습니다. 변경을 적용할 수 없습니다.")
                time.sleep(1)
                continue
            
            matched_cols = [col for col in time_period_cols if col in df_modified.columns and str(df_modified.at[target_row_idx, col]).strip() == requester_name]
            
            if not matched_cols:
                st.error(f"❌ 적용 실패: '{formatted_date_in_df}'의 '{time_period}' 스케줄에서 '{requester_name}'를 찾을 수 없습니다.")
                time.sleep(1)
                continue
            
            is_swapped = False
            for col in matched_cols:
                df_modified.at[target_row_idx, col] = new_assignee
                is_swapped = True
            
            if is_swapped:
                weekday = df_modified.at[target_row_idx, '요일'].replace('요일', '')
                formatted_date_str = f"{formatted_date_in_df} ({weekday}) - {time_period}"
                
                batch_change_log.append({
                    '날짜': formatted_date_str,
                    '변경 전 인원': requester_name,
                    '변경 후 인원': new_assignee,
                })
                applied_count += 1
                swapped_assignments.add((formatted_date_in_df, time_period, new_assignee))
                
                if is_special_date and time_period == '오전' and not special_df[special_df['날짜'] == date_part].empty:
                    if requester_name == special_df[special_df['날짜'] == date_part]['당직 인원'].iloc[0]:
                        special_row_idx = special_df[special_df['날짜'] == date_part].index
                        special_df.at[special_row_idx[0], '당직 인원'] = new_assignee if new_assignee != "당직 없음" else ""
                        try:
                            update_sheet_with_retry(worksheet_special, [special_df.columns.tolist()] + special_df.fillna('').values.tolist())
                            st.success(f"{date_part}의 당직 인원이 {new_assignee}로 업데이트되었습니다.")
                        except Exception as e:
                            st.error(f"당직 인원 업데이트 실패: {type(e).__name__} - {str(e)}")
                
        except Exception as e:
            st.error(f"요청 처리 중 오류 발생: {type(e).__name__} - {str(e)}")
            time.sleep(1)
            continue
            
    if applied_count > 0:
        st.success(f"✅ 총 {applied_count}건의 스케줄 변경 요청이 성공적으로 반영되었습니다.")
        time.sleep(1.5)
        st.session_state["swapped_assignments_log"] = batch_change_log
    else:
        st.info("새롭게 적용할 스케줄 변경 요청이 없습니다.")
        time.sleep(1)
        
    st.session_state["swapped_assignments"] = swapped_assignments
    return create_df_schedule_md(df_modified)

def format_sheet_date_for_display(date_string):
    match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', date_string)
    if match:
        date_part, shift_part = match.groups()
        try:
            dt_obj = datetime.strptime(date_part, '%Y-%m-%d').date()
            weekday_str = ['월', '화', '수', '목', '금', '토', '일'][dt_obj.weekday()]
            return f"{dt_obj.month}월 {dt_obj.day}일 ({weekday_str}) - {shift_part}"
        except ValueError:
            return date_string
    return date_string

def format_date_str_to_display(date_str, weekday, time_period):
    if '요일' in weekday:
        weekday = weekday.replace('요일', '')
    return f"{date_str} ({weekday}) - {time_period}"

def save_to_gsheet(name, categories, selected_save_dates, month_str, worksheet):
    try:
        if not name or not categories or not selected_save_dates:
            st.warning("⚠️ 근무자, 요청 분류, 날짜 정보를 올바르게 입력해주세요.")
            return None

        with st.spinner("요청사항을 추가 중입니다..."):
            df_room_request_temp = st.session_state["df_room_request"].copy()
            new_requests = []

            for category in categories:
                for date in selected_save_dates:
                    date = date.strip()
                    existing_request = df_room_request_temp[
                        (df_room_request_temp['이름'] == name) &
                        (df_room_request_temp['날짜정보'] == date) &
                        (df_room_request_temp['분류'] == category)
                    ]
                    if existing_request.empty:
                        new_requests.append({"이름": name, "분류": category, "날짜정보": date})

            if not new_requests:
                st.info("ℹ️ 이미 존재하는 요청사항입니다.")
                return df_room_request_temp

            new_request_df = pd.DataFrame(new_requests)
            df_room_request_temp = pd.concat([df_room_request_temp, new_request_df], ignore_index=True)
            df_room_request_temp = df_room_request_temp.sort_values(by=["이름", "날짜정보"]).fillna("").reset_index(drop=True)

            if not update_sheet_with_retry(worksheet, [df_room_request_temp.columns.tolist()] + df_room_request_temp.astype(str).values.tolist()):
                st.warning("⚠️ Google Sheets 업데이트에 실패했습니다. 잠시 후 재시도 해주세요.")
                return None

            st.success("요청이 성공적으로 기록되었습니다.")
            time.sleep(1.5)
            return df_room_request_temp

    except gspread.exceptions.APIError as e:
        st.warning("⚠️ 너무 많은 요청이 접속되어 딜레이되고 있습니다. 잠시 후 재시도 해주세요.")
        st.error(f"Google Sheets API 오류 (요청 추가): {str(e)}")
        return None
    except Exception as e:
        st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
        st.error(f"요청 추가 중 오류 발생: {type(e).__name__} - {str(e)}")
        return None

# 메인
today = date.today()
next_month_date = today.replace(day=1) + relativedelta(months=1)
month_str = next_month_date.strftime("%Y년 %-m월")
this_month_start = next_month_date.replace(day=1)

# 다음 달의 마지막 날 계산
if this_month_start.month == 12:
    this_month_end = date(this_month_start.year, 12, 31)
else:
    this_month_end = (date(this_month_start.year, this_month_start.month + 1, 1) - timedelta(days=1))

# 다음 달 계산 (기존 코드 유지, 필요 시 사용)
if today.month == 12:
    next_month_start = date(today.year + 1, 1, 1)
else:
    next_month_start = date(today.year, today.month + 1, 1)
next_month_end = (next_month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

# 세션 상태 초기화
initialize_session_state()

# 데이터 로드
if not st.session_state["data_loaded"]:
    with st.spinner("데이터를 로드하고 있습니다..."):
        result = load_data_page6_no_cache(month_str)
    
    if isinstance(result[0], str) and result[0] == "STOP":  # 수정된 부분
        st.stop()

    if result[0] is None:  # 데이터 로드 실패 처리
        st.error("데이터 로드에 실패했습니다. 새로고침 버튼을 눌러 다시 시도해주세요.")
        st.stop()
    
    df_schedule, df_room_request, worksheet_room_request, df_cumulative, df_swap_requests = result
    if df_schedule.empty:
        st.error(f"⚠️ 로드된 스케줄 데이터가 비어있습니다. Google Sheets의 '{month_str} 스케줄' 시트를 확인하거나, 새로고침 버튼을 눌러 다시 시도해주세요.")
        st.stop()  # 무한 루프 방지

    st.session_state["df_schedule_original"] = df_schedule.copy()
    st.session_state["df_schedule"] = df_schedule
    st.session_state["df_room_request"] = df_room_request
    st.session_state["worksheet_room_request"] = worksheet_room_request
    st.session_state["df_cumulative"] = df_cumulative
    st.session_state["df_swap_requests"] = df_swap_requests
    st.session_state["df_schedule_md"] = create_df_schedule_md(df_schedule)
    st.session_state["df_schedule_md_initial"] = st.session_state["df_schedule_md"].copy()
    st.session_state["data_loaded"] = True
        
st.header("🚪 방 배정", divider='rainbow')

# 새로고침 버튼
st.write("- 먼저 새로고침 버튼으로 최신 데이터를 불러온 뒤, 배정을 진행해주세요.")
if st.button("🔄 새로고침 (R)"):
    try:
        st.cache_data.clear()
        st.session_state["data_loaded"] = False
        with st.spinner("데이터를 새로고침하는 중입니다..."):
            result = load_data_page6_no_cache(month_str)
        
        if isinstance(result[0], str) and result[0] == "STOP":  # 수정된 부분
            st.stop()

        if result[0] is None:
            st.error("데이터 로드에 실패했습니다. 잠시 후 새로고침 버튼을 다시 눌러 시도해주세요.")
            st.stop()
        
        df_schedule, df_room_request, worksheet_room_request, df_cumulative, df_swap_requests = result
        if df_schedule.empty:
            st.error(f"⚠️ 로드된 스케줄 데이터가 비어있습니다. Google Sheets의 '{month_str} 스케줄' 시트를 확인하거나, 새로고침 버튼을 눌러 다시 시도해주세요.")
            st.stop()  # 무한 루프 방지
        
        st.session_state["df_schedule_original"] = df_schedule.copy()
        st.session_state["df_schedule"] = df_schedule
        st.session_state["df_room_request"] = df_room_request
        st.session_state["worksheet_room_request"] = worksheet_room_request
        st.session_state["df_cumulative"] = df_cumulative
        st.session_state["df_swap_requests"] = df_swap_requests
        st.session_state["df_schedule_md"] = create_df_schedule_md(df_schedule)
        st.session_state["df_schedule_md_initial"] = st.session_state["df_schedule_md"].copy()
        st.session_state["swapped_assignments_log"] = []
        st.session_state["swapped_assignments"] = set()
        st.session_state["manual_change_log"] = []
        st.session_state["final_change_log"] = []
        st.session_state["data_loaded"] = True
        st.session_state["weekend_room_settings"] = {}
        st.success("데이터가 새로고침되었습니다.")
        time.sleep(1)
        st.rerun()
    except Exception as e:
        st.error(f"새로고침 중 오류 발생: {type(e).__name__} - {e}")
        st.stop()

# 근무자 명단 수정
st.divider()
st.subheader("📋 스케줄 변경 요청 목록")
if "df_schedule" not in st.session_state or st.session_state["df_schedule"].empty:
    st.warning("⚠️ 스케줄 데이터가 로드되지 않았습니다. 새로고침 버튼을 눌러 데이터를 다시 로드해주세요.")
    st.stop()
df_swaps_raw = st.session_state.get("df_swap_requests", pd.DataFrame())
if not df_swaps_raw.empty:
    cols_to_display = {'요청일시': '요청일시', '요청자': '요청자', '변경 요청': '변경 요청', '변경 요청한 스케줄': '변경 요청한 스케줄'}
    existing_cols = [col for col in cols_to_display.keys() if col in df_swaps_raw.columns]
    df_swaps_display = df_swaps_raw[existing_cols].rename(columns=cols_to_display)
    if '변경 요청한 스케줄' in df_swaps_display.columns:
        df_swaps_display['변경 요청한 스케줄'] = df_swaps_display['변경 요청한 스케줄'].apply(format_sheet_date_for_display)
    st.dataframe(df_swaps_display, use_container_width=True, hide_index=True)
else:
    st.info("표시할 교환 요청 데이터가 없습니다.")

st.divider()
st.subheader("✍️ 스케줄 수정")
st.write("- 요청사항을 **일괄 적용/취소**하거나, 셀을 더블클릭하여 직접 수정한 후 **최종 저장 버튼**을 누르세요.")

col1, col2 = st.columns(2)
with col1:
    if st.button("🔄 요청사항 일괄 적용"):
        df_swaps = st.session_state.get("df_swap_requests", pd.DataFrame())
        if not df_swaps.empty:
            modified_schedule = apply_schedule_swaps(st.session_state["df_schedule_original"], df_swaps)
            st.session_state["df_schedule"] = modified_schedule
            st.session_state["df_schedule_md"] = create_df_schedule_md(modified_schedule)
            st.rerun()
        else:
            st.info("ℹ️ 처리할 교환 요청이 없습니다.")

with col2:
    is_batch_applied = len(st.session_state.get("swapped_assignments_log", [])) > 0
    if st.button("⏪ 적용 취소", disabled=not is_batch_applied):
        st.session_state["df_schedule"] = st.session_state["df_schedule_original"].copy()
        st.session_state["df_schedule_md"] = create_df_schedule_md(st.session_state["df_schedule_original"])
        st.session_state["df_schedule_md_initial"] = st.session_state["df_schedule_md"].copy()
        st.session_state["swapped_assignments_log"] = []
        st.info("변경사항이 취소되고 원본 스케줄로 돌아갑니다.")
        time.sleep(1.5)
        st.rerun()

edited_df_md = st.data_editor(st.session_state["df_schedule_md"], use_container_width=True, key="schedule_editor", disabled=['날짜', '요일'])
st.write(" ")

if st.button("✍️ 변경사항 저장", type="primary", use_container_width=True):
    if edited_df_md.equals(st.session_state["df_schedule_md_initial"]):
        if st.session_state.get("swapped_assignments_log", []):
            st.info("ℹ️ 일괄 적용된 변경사항을 저장합니다.")
        else:
            st.info("ℹ️ 변경사항이 없습니다. 저장할 내용이 없습니다.")

    manual_change_log = []
    diff_indices = np.where(edited_df_md.ne(st.session_state["df_schedule_md_initial"]))
    for row_idx, col_idx in zip(diff_indices[0], diff_indices[1]):
        date_str_raw = edited_df_md.iloc[row_idx, 0]
        col_name = edited_df_md.columns[col_idx]
        old_value = st.session_state["df_schedule_md_initial"].iloc[row_idx, col_idx]
        new_value = edited_df_md.iloc[row_idx, col_idx]
        original_row = st.session_state["df_schedule_original"][st.session_state["df_schedule_original"]['날짜'] == date_str_raw].iloc[0]
        weekday = original_row['요일']
        time_period = '오후' if col_name.startswith('오후') else '오전'
        formatted_date_str = f"{date_str_raw} ({weekday.replace('요일', '')}) - {time_period}"
        manual_change_log.append({
            '날짜': formatted_date_str,
            '변경 전 인원': str(old_value),
            '변경 후 인원': str(new_value),
        })
        st.session_state["swapped_assignments"].add((date_str_raw, time_period, str(new_value).strip()))
    
    st.session_state["final_change_log"] = st.session_state.get("swapped_assignments_log", []) + manual_change_log

    df_schedule_to_save = st.session_state["df_schedule_original"].copy()
    for row_idx, row in edited_df_md.iterrows():
        date_str = row['날짜']
        original_row_idx = df_schedule_to_save[df_schedule_to_save['날짜'] == date_str].index[0]
        
        oncall_person = row['오전당직(온콜)']
        df_schedule_to_save.at[original_row_idx, '오전당직(온콜)'] = oncall_person

        am_personnel = [str(row[str(i)]).strip() for i in range(1, 12) if str(row[str(i)]).strip()]
        am_personnel_with_oncall = am_personnel + ([oncall_person] if oncall_person and oncall_person not in am_personnel else [])
        for i in range(1, 13):
            col = str(i)
            if i <= len(am_personnel_with_oncall):
                df_schedule_to_save.at[original_row_idx, col] = am_personnel_with_oncall[i-1]
            else:
                df_schedule_to_save.at[original_row_idx, col] = ''

        pm_personnel = [str(row[f'오후{i}']).strip() for i in range(1, 5) if str(row[f'오후{i}']).strip()]
        pm_personnel_with_oncall = pm_personnel + ([oncall_person] if oncall_person and oncall_person not in pm_personnel else [])
        for i in range(1, 6):
            col = f'오후{i}'
            if i <= len(pm_personnel_with_oncall):
                df_schedule_to_save.at[original_row_idx, col] = pm_personnel_with_oncall[i-1]
            else:
                df_schedule_to_save.at[original_row_idx, col] = ''

    try:
        st.info("ℹ️ 최종 스케줄을 Google Sheets에 저장합니다...")
        gc = get_gspread_client()
        if gc is None:
            raise Exception("Failed to initialize gspread client")
        sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])

        sheet_name = f"{month_str} 스케줄"
        
        try:
            worksheet_schedule = sheet.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            st.warning(f"'{sheet_name}' 시트를 찾을 수 없어 새로 생성합니다.")
            worksheet_schedule = sheet.add_worksheet(title=sheet_name, rows=100, cols=20)
            
        columns_to_save = df_schedule_to_save.columns.tolist()
        schedule_data = [columns_to_save] + df_schedule_to_save.fillna('').values.tolist()
        
        if update_sheet_with_retry(worksheet_schedule, schedule_data):
            st.session_state["df_schedule"] = df_schedule_to_save.copy()
            st.session_state["df_schedule_md"] = create_df_schedule_md(df_schedule_to_save)
            st.session_state["df_schedule_md_initial"] = st.session_state["df_schedule_md"].copy()
            st.success(f"🎉 최종 스케줄이 '{sheet_name}' 시트에 성공적으로 저장되었습니다.")
            time.sleep(1.5)
            st.rerun()
    except Exception as e:
        st.error(f"Google Sheets 저장 중 오류 발생: {type(e).__name__} - {e}")
        st.stop()

st.write("---")
st.caption("📝 현재까지 기록된 변경사항 로그")
final_log_list = st.session_state.get("final_change_log", [])
if final_log_list:
    log_df = pd.DataFrame(final_log_list)
    st.dataframe(log_df, use_container_width=True, hide_index=True)
else:
    st.info("기록된 변경사항이 없습니다.")

# 방 설정 UI
st.divider()
st.subheader("⚙️ 방 설정")
tab_weekday, tab_weekend = st.tabs(["평일 방 설정", "토요/휴일 방 설정"])

with tab_weekday:
    room_options = [str(i) for i in range(1, 13)]

    tab830, tab900, tab930, tab1000, tab1330 = st.tabs([
        "🕘 08:30", "🕘 09:00", "🕤 09:30", "🕙 10:00", "🕜 13:30 (오후)"
    ])
    with tab830:
        col1, col2 = st.columns([1, 2.5])
        with col1:
            st.markdown("###### **방 개수**")
            num_830 = st.number_input("830_rooms_count", min_value=0, max_value=12, value=4, key="830_rooms", label_visibility="collapsed")
            st.markdown("###### **오전 당직방**")
            duty_830_options = st.session_state["room_settings"]["830_room_select"]
            try:
                duty_index_830 = duty_830_options.index(st.session_state["room_settings"].get("830_duty"))
            except ValueError:
                duty_index_830 = 0
            duty_830 = st.selectbox("830_duty_room", duty_830_options, index=duty_index_830, key="830_duty", label_visibility="collapsed", help="8:30 시간대의 당직 방을 선택합니다.")
            st.session_state["room_settings"]["830_duty"] = duty_830
        with col2:
            st.markdown("###### **방 번호 선택**")
            if len(st.session_state["room_settings"]["830_room_select"]) > num_830:
                st.session_state["room_settings"]["830_room_select"] = st.session_state["room_settings"]["830_room_select"][:num_830]
            rooms_830 = st.multiselect("830_room_select_numbers", room_options, default=st.session_state["room_settings"]["830_room_select"], max_selections=num_830, key="830_room_select", label_visibility="collapsed")
            if len(rooms_830) < num_830:
                st.warning(f"방 번호를 {num_830}개 선택해주세요.")
            st.session_state["room_settings"]["830_room_select"] = rooms_830
    with tab900:
        col1, col2 = st.columns([1, 2.5])
        with col1:
            st.markdown("###### **방 개수**")
            num_900 = st.number_input("900_rooms_count", min_value=0, max_value=12, value=3, key="900_rooms", label_visibility="collapsed")
        with col2:
            st.markdown("###### **방 번호 선택**")
            if len(st.session_state["room_settings"]["900_room_select"]) > num_900:
                st.session_state["room_settings"]["900_room_select"] = st.session_state["room_settings"]["900_room_select"][:num_900]
            rooms_900 = st.multiselect("900_room_select_numbers", room_options, default=st.session_state["room_settings"]["900_room_select"], max_selections=num_900, key="900_room_select", label_visibility="collapsed")
            if len(rooms_900) < num_900:
                st.warning(f"방 번호를 {num_900}개 선택해주세요.")
            st.session_state["room_settings"]["900_room_select"] = rooms_900
    with tab930:
        col1, col2 = st.columns([1, 2.5])
        with col1:
            st.markdown("###### **방 개수**")
            num_930 = st.number_input("930_rooms_count", min_value=0, max_value=12, value=3, key="930_rooms", label_visibility="collapsed")
        with col2:
            st.markdown("###### **방 번호 선택**")
            if len(st.session_state["room_settings"]["930_room_select"]) > num_930:
                st.session_state["room_settings"]["930_room_select"] = st.session_state["room_settings"]["930_room_select"][:num_930]
            rooms_930 = st.multiselect("930_room_select_numbers", room_options, default=st.session_state["room_settings"]["930_room_select"], max_selections=num_930, key="930_room_select", label_visibility="collapsed")
            if len(rooms_930) < num_930:
                st.warning(f"방 번호를 {num_930}개 선택해주세요.")
            st.session_state["room_settings"]["930_room_select"] = rooms_930
    with tab1000:
        col1, col2 = st.columns([1, 2.5])
        with col1:
            st.markdown("###### **방 개수**")
            num_1000 = st.number_input("1000_rooms_count", min_value=0, max_value=12, value=2, key="1000_rooms", label_visibility="collapsed")
        with col2:
            st.markdown("###### **방 번호 선택**")
            if len(st.session_state["room_settings"]["1000_room_select"]) > num_1000:
                st.session_state["room_settings"]["1000_room_select"] = st.session_state["room_settings"]["1000_room_select"][:num_1000]
            rooms_1000 = st.multiselect("1000_room_select_numbers", room_options, default=st.session_state["room_settings"]["1000_room_select"], max_selections=num_1000, key="1000_room_select", label_visibility="collapsed")
            if len(rooms_1000) < num_1000:
                st.warning(f"방 번호를 {num_1000}개 선택해주세요.")
            st.session_state["room_settings"]["1000_room_select"] = rooms_1000
    with tab1330:
        col1, col2 = st.columns([1, 2.5])
        with col1:
            st.markdown("###### **방 개수**")
            st.info("4개 고정")
            num_1330 = 4
            st.markdown("###### **오후 당직방**")
            duty_1330_options = st.session_state["room_settings"]["1330_room_select"]
            try:
                duty_index_1330 = duty_1330_options.index(st.session_state["room_settings"].get("1330_duty"))
            except ValueError:
                duty_index_1330 = 0
            duty_1330 = st.selectbox("1330_duty_room", duty_1330_options, index=duty_index_1330, key="1330_duty", label_visibility="collapsed", help="13:30 시간대의 당직 방을 선택합니다.")
            st.session_state["room_settings"]["1330_duty"] = duty_1330
        with col2:
            st.markdown("###### **방 번호 선택**")
            if len(st.session_state["room_settings"]["1330_room_select"]) > num_1330:
                st.session_state["room_settings"]["1330_room_select"] = st.session_state["room_settings"]["1330_room_select"][:num_1330]
            rooms_1330 = st.multiselect("1330_room_select_numbers", room_options, default=st.session_state["room_settings"]["1330_room_select"], max_selections=num_1330, key="1330_room_select", label_visibility="collapsed")
            if len(rooms_1330) < num_1330:
                st.warning(f"방 번호를 {num_1330}개 선택해주세요.")
            st.session_state["room_settings"]["1330_room_select"] = rooms_1330

with tab_weekend:
    # 토요/휴일 데이터 로드
    special_schedules = []
    special_dates = set()
    client = get_gspread_client()
    try:
        spreadsheet = client.open_by_url(st.secrets["google_sheet"]["url"])
        sheet_name = f"{month_str} 토요/휴일 일자"
        worksheet = spreadsheet.worksheet(sheet_name)
        schedule_data = worksheet.get_all_records()
        special_df = pd.DataFrame(schedule_data) if schedule_data else pd.DataFrame(columns=["날짜", "당직 인원"])

        if not schedule_data:
            st.warning("별도의 토요/휴일 스케줄이 없습니다.")
        else:
            for row in schedule_data:
                if not row.get("날짜"): continue
                date_obj = datetime.strptime(row["날짜"], '%Y-%m-%d').date()
                formatted_date_str = f"{date_obj.month}월 {date_obj.day}일"
                date_row = st.session_state.df_schedule[st.session_state.df_schedule['날짜'] == formatted_date_str]
                personnel = [str(p).strip() for p in date_row.iloc[0, 2:14] if pd.notna(p) and str(p).strip()] if not date_row.empty else []
                special_schedules.append((date_obj, formatted_date_str, personnel))
                special_dates.add(formatted_date_str)
    except Exception as e:
        st.error(f"토요/휴일 데이터 로드 실패: {e}")
        special_df = pd.DataFrame(columns=["날짜", "당직 인원"])
    
    # 토요/휴일 UI 렌더링
    if special_schedules:
        for date_obj, date_str, personnel_for_day in sorted(special_schedules):
            weekday_map = {5: "토", 6: "일"}
            weekday_str = weekday_map.get(date_obj.weekday(), '휴')
            
            duty_person_for_date = ""
            if not special_df.empty:
                duty_row = special_df[special_df['날짜'] == date_obj.strftime('%Y-%m-%d')]
                if not duty_row.empty: duty_person_for_date = str(duty_row['당직 인원'].iloc[0]).strip()

            expander_title = (f"🗓️ {date_str} ({weekday_str}) | "
                              f"근무: {len(personnel_for_day)}명 | "
                              f"당직: {duty_person_for_date or '없음'}")

            with st.expander(expander_title):
                col1, col2 = st.columns([1, 1])
                duty_room = None
                with col1:
                    st.markdown("###### **당직 방**")
                    if duty_person_for_date:
                        duty_room_options = ["선택 안 함"] + [str(i) for i in range(1, 13)]
                        default_duty_room = st.session_state.weekend_room_settings.get(date_str, {}).get("duty_room", "선택 안 함")
                        duty_room = st.selectbox("당직 방 선택", duty_room_options, key=f"duty_room_{date_str}", 
                                                 index=duty_room_options.index(default_duty_room) if default_duty_room in duty_room_options else 0, label_visibility="collapsed")
                    else: st.info("당직 인원 없음")
                
                with col2:
                    st.markdown("###### **총 방 개수**")
                    default_room_count = st.session_state.weekend_room_settings.get(date_str, {}).get("total_room_count", len(personnel_for_day))
                    total_room_count = st.number_input("총 방 개수", min_value=0, max_value=12, value=default_room_count, 
                                                       key=f"total_rooms_{date_str}", label_visibility="collapsed")
                
                st.markdown("###### **방 번호 선택**")
                room_options = [str(i) for i in range(1, 13)]
                default_rooms = st.session_state.weekend_room_settings.get(date_str, {}).get("selected_rooms", room_options[:total_room_count])
                selected_rooms = st.multiselect("방 번호 선택", room_options, default=default_rooms, max_selections=total_room_count, 
                                                key=f"rooms_{date_str}", label_visibility="collapsed")

                st.session_state.weekend_room_settings[date_str] = {
                    "duty_room": duty_room if duty_room and duty_room != "선택 안 함" else None,
                    "total_room_count": total_room_count, "selected_rooms": selected_rooms
                }
    else: st.info("이번 달은 토요/휴일 근무가 없습니다.")

all_selected_rooms = (st.session_state["room_settings"]["830_room_select"] + 
                     st.session_state["room_settings"]["900_room_select"] + 
                     st.session_state["room_settings"]["930_room_select"] + 
                     st.session_state["room_settings"]["1000_room_select"] + 
                     st.session_state["room_settings"]["1330_room_select"])

# 배정 요청 입력 UI
st.divider()
st.subheader("📋 배정 요청 관리")
st.write("- 모든 인원의 배정 요청(고정 및 우선)을 추가 및 수정할 수 있습니다.\n - 인원별 시간대, 방, 당직 배정 균형을 위해, 일부 요청사항이 무시될 수 있습니다.")
요청분류 = ["1번방", "2번방", "3번방", "4번방", "5번방", "6번방", "7번방", "8번방", "9번방", "10번방", "11번방", "12번방", 
            "8:30", "9:00", "9:30", "10:00", "당직 아닌 이른방", "이른방 제외", "늦은방 제외", "오후 당직 제외"]
st.write(" ")
st.markdown("**🟢 방 배정 요청 추가**")
col1, col2, col3, col_button_add = st.columns([2.5, 2.5, 3.5, 1])
with col1:
    names = sorted([str(name).strip() for name in st.session_state["df_schedule"].iloc[:, 2:].stack().dropna().unique() if str(name).strip()])
    name = st.selectbox("근무자", names, key="request_employee_select", index=None, placeholder="근무자 선택")
with col2:
    categories = st.multiselect("요청 분류", 요청분류, key="request_category_select")
with col3:
    selected_save_dates = []
    if name:
        st.cache_data.clear()
        available_dates = get_user_available_dates(name, st.session_state["df_schedule"], this_month_start, this_month_end)
        date_options = [display_str for display_str, _ in available_dates]
        dates = st.multiselect("요청 일자", date_options, key="request_date_select")
        selected_save_dates = [save_str for display_str, save_str in available_dates if display_str in dates]
    else:
        dates = st.multiselect("요청 일자", [], key="request_date_select", disabled=True)
with col_button_add:
    st.markdown("<div>&nbsp;</div>", unsafe_allow_html=True)
    add_button_clicked = st.button("📅 추가", key="request_add_button")
if add_button_clicked:
    if not name:
        st.error("근무자를 먼저 선택해주세요.")
    elif not categories or not selected_save_dates:
        st.error("요청 분류와 날짜를 선택해주세요.")
    else:
        df_room_request = save_to_gsheet(name, categories, selected_save_dates, month_str, st.session_state["worksheet_room_request"])
        if df_room_request is not None:
            st.session_state["df_room_request"] = df_room_request
            st.cache_data.clear()
            time.sleep(1.5)
            st.rerun()

st.write(" ")
st.markdown("**🔴 방 배정 요청 삭제**")
if not st.session_state["df_room_request"].empty:
    col0, col1, col_button_del = st.columns([2.5, 4.5, 1])
    with col0:
        unique_names = st.session_state["df_room_request"]["이름"].unique()
        selected_employee = st.selectbox("근무자 선택", unique_names, key="delete_request_employee_select", index=None, placeholder="근무자 선택")
    with col1:
        selected_items = []
        if selected_employee:
            df_request_filtered = st.session_state["df_room_request"][st.session_state["df_room_request"]["이름"] == selected_employee]
            if not df_request_filtered.empty:
                options = [f"{row['분류']} - {row['날짜정보']}" for _, row in df_request_filtered.iterrows()]
                selected_items = st.multiselect("삭제할 항목", options, key="delete_request_select")
            else:
                st.multiselect("삭제할 항목", [], disabled=True, key="delete_request_select", help="해당 근무자의 요청이 없습니다.")
        else:
            st.multiselect("삭제할 항목", [], key="delete_request_select", disabled=True)
    with col_button_del:
        st.markdown("<div>&nbsp;</div>", unsafe_allow_html=True)
        delete_button_clicked = st.button("📅 삭제", key="request_delete_button")
    if delete_button_clicked:
        if not selected_employee or not selected_items:
            st.error("삭제할 근무자와 항목을 선택해주세요.")
        else:
            indices = []
            for item in selected_items:
                for idx, row in st.session_state["df_room_request"].iterrows():
                    if row['이름'] == selected_employee and f"{row['분류']} - {row['날짜정보']}" == item:
                        indices.append(idx)
            df_room_request = st.session_state["df_room_request"].drop(indices).reset_index(drop=True)
            st.session_state["df_room_request"] = df_room_request
            if update_sheet_with_retry(st.session_state["worksheet_room_request"], [df_room_request.columns.tolist()] + df_room_request.values.tolist()):
                st.cache_data.clear()
                st.success("요청사항이 삭제되었습니다.")
                time.sleep(1.5)
                st.rerun()
else:
    st.info("📍 방 배정 요청이 없습니다.")
st.write(" ")
st.markdown("**🙋‍♂️ 현재 방 배정 요청 목록**")
if st.session_state["df_room_request"].empty:
    st.info("☑️ 현재 방 배정 요청이 없습니다.")
else:
    st.dataframe(st.session_state["df_room_request"], use_container_width=True, hide_index=True)

def parse_date_info(date_info):
    try:
        date_part = date_info.split('(')[0].strip()
        date_obj = datetime.strptime(date_part, '%Y-%m-%d')
        is_morning = '오전' in date_info
        parsed_date = date_obj.strftime('%Y-%m-%d')
        return parsed_date, is_morning
    except ValueError as e:
        st.warning(f"Failed to parse date_info: {date_info}, error: {str(e)}")
        return None, False

def assign_special_date(personnel_for_day, date_str, settings):
    assignment_dict = {}
    assigned_personnel = set()
    
    duty_room = settings.get("duty_room", None)
    selected_rooms = settings.get("selected_rooms", [])
    total_room_count = settings.get("total_room_count", 0)
    
    # 사용자가 설정한 방 번호 순서
    preferred_room_order = ['1', '8', '4', '7', '10', '2', '5', '6', '9', '3']
    # 선택된 방을 사용자 지정 순서대로 정렬
    sorted_rooms = [room for room in preferred_room_order if room in selected_rooms][:total_room_count]
    
    # Google Sheets에서 당직 인원 확인
    client = get_gspread_client()
    if client is None:
        st.error("Google Sheets 클라이언트 초기화 실패")
        return {}, sorted_rooms
    
    spreadsheet = client.open_by_url(st.secrets["google_sheet"]["url"])
    sheet_name = f"{month_str} 토요/휴일 일자"
    try:
        worksheet = spreadsheet.worksheet(sheet_name)
        special_data = worksheet.get_all_records()
        special_df = pd.DataFrame(special_data)
        date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=2025)
        formatted_date = date_obj.strftime('%Y-%m-%d')
        duty_person_row = special_df[special_df['날짜'] == formatted_date]
        duty_person = duty_person_row['당직 인원'].iloc[0] if not duty_person_row.empty else None
    except gspread.exceptions.WorksheetNotFound:
        st.warning(f"{sheet_name} 시트가 없습니다.")
        return {}, sorted_rooms
    
    # 당직 인원 배정
    if duty_person and duty_person in personnel_for_day and duty_room and duty_room != "선택 안 함":
        if duty_person not in assigned_personnel:
            assignment_dict[f"방({duty_room})"] = duty_person
            assigned_personnel.add(duty_person)
        else:
            st.warning(f"{date_str}: {duty_person} 이미 배정됨, 당직 배정 건너뜀")
    
    # 나머지 인원을 랜덤 배정
    remaining_personnel = [p for p in personnel_for_day if p not in assigned_personnel]
    random.shuffle(remaining_personnel)
    remaining_rooms = [r for r in sorted_rooms if r != duty_room]
    
    for room in remaining_rooms:
        if remaining_personnel:
            person = remaining_personnel.pop(0)
            assignment_dict[f"방({room})"] = person
            assigned_personnel.add(person)
        else:
            st.warning(f"{date_str}: 배정 가능한 인원 부족, 방 {room} 배정 안 됨")
    
    return assignment_dict, sorted_rooms

from collections import Counter
import random
import streamlit as st

def random_assign(personnel, slots, request_assignments, time_groups, total_stats, morning_personnel, afternoon_personnel, afternoon_duty_counts):
    assignment = [None] * len(slots)
    assigned_personnel_morning = set()
    assigned_personnel_afternoon = set()
    daily_stats = {
        'early': Counter(),
        'late': Counter(),
        'morning_duty': Counter(),
        'afternoon_duty': Counter(),
        'rooms': {str(i): Counter() for i in range(1, 13)},
        'time_room_slots': {}  # 시간대-방 쌍에 대한 Counter 객체를 딕셔너리로 관리
    }

    # time_room_slots 초기화
    for slot in slots:
        daily_stats['time_room_slots'][slot] = Counter()

    # total_stats['time_room_slots'] 초기화 (외부 코드 수정 없이 함수 내에서 처리)
    if 'time_room_slots' not in total_stats:
        total_stats['time_room_slots'] = {}
    for slot in slots:
        total_stats['time_room_slots'].setdefault(slot, Counter())

    morning_slots = [s for s in slots if s.startswith(('8:30', '9:00', '9:30', '10:00')) and '_당직' not in s]
    afternoon_slots = [s for s in slots if s.startswith('13:30')]
    afternoon_duty_slot = [s for s in slots if s.startswith('13:30') and s.endswith('_당직')]

    # 요청된 배정 처리
    for slot, person in request_assignments.items():
        if person in personnel and slot in slots:
            slot_idx = slots.index(slot)
            if assignment[slot_idx] is None:
                if (slot in morning_slots and person in morning_personnel) or \
                   (slot in afternoon_slots and person in afternoon_personnel):
                    if slot in morning_slots and person in assigned_personnel_morning:
                        st.warning(f"중복 배정 방지: {person}은 이미 오전 시간대({slot})에 배정됨")
                        continue
                    if slot in afternoon_slots and person in assigned_personnel_afternoon:
                        st.warning(f"중복 배정 방지: {person}은 이미 오후 시간대({slot})에 배정됨")
                        continue

                    assignment[slot_idx] = person
                    if slot in morning_slots:
                        assigned_personnel_morning.add(person)
                    else:
                        assigned_personnel_afternoon.add(person)
                    room_num = slot.split('(')[1].split(')')[0]
                    daily_stats['rooms'][room_num][person] += 1
                    daily_stats['time_room_slots'][slot][person] += 1
                    if slot.startswith('8:30') and '_당직' not in slot:
                        daily_stats['early'][person] += 1
                    elif slot.startswith('10:00'):
                        daily_stats['late'][person] += 1
                    if slot.startswith('8:30') and slot.endswith('_당직'):
                        daily_stats['morning_duty'][person] += 1
                    elif slot.startswith('13:30') and slot.endswith('_당직'):
                        daily_stats['afternoon_duty'][person] += 1
                else:
                    st.warning(f"배정 요청 무시: {person}님은 {slot} 시간대({'오전' if slot in morning_slots else '오후'})에 근무 불가")
            else:
                st.warning(f"배정 요청 충돌: {person}을 {slot}에 배정할 수 없음. 이미 배정됨: {assignment[slot_idx]}")

    # 오후 당직 배정
    afternoon_duty_slot_idx = slots.index(afternoon_duty_slot[0]) if afternoon_duty_slot else None
    if afternoon_duty_slot_idx is not None and assignment[afternoon_duty_slot_idx] is None:
        available_personnel = [p for p in afternoon_personnel if p not in assigned_personnel_afternoon]
        candidates = [p for p in available_personnel if p in afternoon_duty_counts and afternoon_duty_counts[p] > 0]
        
        if candidates:
            best_person = None
            min_duty_count = float('inf')
            for person in candidates:
                duty_count = total_stats['afternoon_duty'][person] + daily_stats['afternoon_duty'][person]
                time_room_count = total_stats['time_room_slots'][afternoon_duty_slot[0]][person] + \
                                 daily_stats['time_room_slots'][afternoon_duty_slot[0]][person]
                score = duty_count * 100 + time_room_count
                if score < min_duty_count:
                    min_duty_count = score
                    best_person = person
            if best_person:
                assignment[afternoon_duty_slot_idx] = best_person
                assigned_personnel_afternoon.add(best_person)
                room_num = afternoon_duty_slot[0].split('(')[1].split(')')[0]
                daily_stats['rooms'][room_num][best_person] += 1
                daily_stats['time_room_slots'][afternoon_duty_slot[0]][best_person] += 1
                daily_stats['afternoon_duty'][best_person] += 1
                afternoon_duty_counts[best_person] -= 1
                if afternoon_duty_counts[best_person] <= 0:
                    del afternoon_duty_counts[best_person]

    # 오전 슬롯 배정
    morning_remaining = [p for p in morning_personnel if p not in assigned_personnel_morning]
    afternoon_remaining = [p for p in afternoon_personnel if p not in assigned_personnel_afternoon]
    remaining_slots = [i for i, a in enumerate(assignment) if a is None]
    
    morning_slot_indices = [i for i in remaining_slots if slots[i] in morning_slots]
    random.shuffle(morning_remaining)
    while morning_remaining and morning_slot_indices:
        best_person = None
        best_slot_idx = None
        min_score = float('inf')
        
        for slot_idx in morning_slot_indices:
            if assignment[slot_idx] is not None:
                continue
            slot = slots[slot_idx]
            
            for person in morning_remaining:
                time_room_count = total_stats['time_room_slots'][slot][person] + \
                                  daily_stats['time_room_slots'][slot][person]
                if slot.startswith('8:30') and '_당직' not in slot:
                    early_count = total_stats['early'][person] + daily_stats['early'][person]
                    score = early_count * 100 + time_room_count
                elif slot.startswith('10:00'):
                    late_count = total_stats['late'][person] + daily_stats['late'][person]
                    score = 10000 + late_count * 100 + time_room_count
                else:
                    score = 20000 + time_room_count
                
                if score < min_score:
                    min_score = score
                    best_person = person
                    best_slot_idx = slot_idx
        
        if best_slot_idx is None or best_person is None:
            st.warning(f"오전 슬롯 배정 불가: 더 이상 배정 가능한 인원 없음")
            break
        
        slot = slots[best_slot_idx]
        assignment[best_slot_idx] = best_person
        assigned_personnel_morning.add(best_person)
        morning_remaining.remove(best_person)
        morning_slot_indices.remove(best_slot_idx)
        remaining_slots.remove(best_slot_idx)
        room_num = slot.split('(')[1].split(')')[0]
        daily_stats['rooms'][room_num][best_person] += 1
        daily_stats['time_room_slots'][slot][best_person] += 1
        if slot.startswith('8:30') and '_당직' not in slot:
            daily_stats['early'][best_person] += 1
        elif slot.startswith('10:00'):
            daily_stats['late'][best_person] += 1
        if slot.startswith('8:30') and slot.endswith('_당직'):
            daily_stats['morning_duty'][best_person] += 1

    # 오후 슬롯 배정
    afternoon_slot_indices = [i for i in remaining_slots if slots[i] in afternoon_slots]
    while afternoon_remaining and afternoon_slot_indices:
        best_person = None
        best_slot_idx = None
        min_score = float('inf')
        
        for slot_idx in afternoon_slot_indices:
            if assignment[slot_idx] is not None:
                continue
            slot = slots[slot_idx]
            
            for person in afternoon_remaining:
                time_room_count = total_stats['time_room_slots'][slot][person] + \
                                  daily_stats['time_room_slots'][slot][person]
                duty_count = total_stats['afternoon_duty'][person] + daily_stats['afternoon_duty'][person] if slot.endswith('_당직') else float('inf')
                if slot.endswith('_당직'):
                    score = duty_count * 100 + time_room_count
                else:
                    score = time_room_count
                
                if score < min_score:
                    min_score = score
                    best_person = person
                    best_slot_idx = slot_idx
        
        if best_slot_idx is None or best_person is None:
            st.warning(f"오후 슬롯 배정 불가: 더 이상 배정 가능한 인원 없음")
            break
        
        slot = slots[best_slot_idx]
        assignment[best_slot_idx] = best_person
        assigned_personnel_afternoon.add(best_person)
        afternoon_remaining.remove(best_person)
        afternoon_slot_indices.remove(best_slot_idx)
        room_num = slot.split('(')[1].split(')')[0]
        daily_stats['rooms'][room_num][best_person] += 1
        daily_stats['time_room_slots'][slot][best_person] += 1
        if slot.endswith('_당직'):
            daily_stats['afternoon_duty'][best_person] += 1

    # 남은 빈 슬롯 처리
    for slot_idx in range(len(slots)):
        if assignment[slot_idx] is None:
            slot = slots[slot_idx]
            available_personnel = morning_personnel if slot in morning_slots else afternoon_personnel
            assigned_set = assigned_personnel_morning if slot in morning_slots else assigned_personnel_afternoon
            candidates = [p for p in available_personnel if p not in assigned_set]
            
            if candidates:
                room_num = slot.split('(')[1].split(')')[0]
                best_person = None
                min_score = float('inf')
                for person in candidates:
                    early_count = total_stats['early'][person] + daily_stats['early'][person] if slot.startswith('8:30') and '_당직' not in slot else float('inf')
                    late_count = total_stats['late'][person] + daily_stats['late'][person] if slot.startswith('10:00') else float('inf')
                    morning_duty_count = total_stats['morning_duty'][person] + daily_stats['morning_duty'][person] if slot.startswith('8:30') and slot.endswith('_당직') else float('inf')
                    afternoon_duty_count = total_stats['afternoon_duty'][person] + daily_stats['afternoon_duty'][person] if slot.startswith('13:30') and slot.endswith('_당직') else float('inf')
                    time_room_count = total_stats['time_room_slots'][slot][person] + \
                                      daily_stats['time_room_slots'][slot][person]
                    
                    if slot.startswith('8:30') and '_당직' not in slot:
                        score = early_count * 100 + time_room_count
                    elif slot.startswith('10:00'):
                        score = late_count * 100 + time_room_count
                    elif slot.startswith('8:30') and slot.endswith('_당직'):
                        score = morning_duty_count * 100 + time_room_count
                    elif slot.startswith('13:30') and slot.endswith('_당직'):
                        score = afternoon_duty_count * 100 + time_room_count
                    else:
                        score = time_room_count
                    
                    if score < min_score:
                        min_score = score
                        best_person = person
                
                person = best_person
                if slot in morning_slots:
                    assigned_personnel_morning.add(person)
                else:
                    assigned_personnel_afternoon.add(person)
                st.warning(f"슬롯 {slot} 공란 방지: {person} 배정 (스코어: {min_score})")
            else:
                available_personnel = morning_personnel if slot in morning_slots else afternoon_personnel
                if available_personnel:
                    room_num = slot.split('(')[1].split(')')[0]
                    best_person = None
                    min_score = float('inf')
                    for person in available_personnel:
                        early_count = total_stats['early'][person] + daily_stats['early'][person] if slot.startswith('8:30') and '_당직' not in slot else float('inf')
                        late_count = total_stats['late'][person] + daily_stats['late'][person] if slot.startswith('10:00') else float('inf')
                        morning_duty_count = total_stats['morning_duty'][person] + daily_stats['morning_duty'][person] if slot.startswith('8:30') and slot.endswith('_당직') else float('inf')
                        afternoon_duty_count = total_stats['afternoon_duty'][person] + daily_stats['afternoon_duty'][person] if slot.startswith('13:30') and slot.endswith('_당직') else float('inf')
                        time_room_count = total_stats['time_room_slots'][slot][person] + \
                                          daily_stats['time_room_slots'][slot][person]
                        
                        if slot.startswith('8:30') and '_당직' not in slot:
                            score = early_count * 100 + time_room_count
                        elif slot.startswith('10:00'):
                            score = late_count * 100 + time_room_count
                        elif slot.startswith('8:30') and slot.endswith('_당직'):
                            score = morning_duty_count * 100 + time_room_count
                        elif slot.startswith('13:30') and slot.endswith('_당직'):
                            score = afternoon_duty_count * 100 + time_room_count
                        else:
                            score = time_room_count
                        
                        if score < min_score:
                            min_score = score
                            best_person = person
                    
                    person = best_person
                    st.warning(f"슬롯 {slot} 공란 방지: 이미 배정된 {person} 재배정 (스코어: {min_score})")
                else:
                    st.warning(f"슬롯 {slot} 공란 방지 불가: 배정 가능한 인원 없음")
                    continue
            
            assignment[slot_idx] = person
            daily_stats['rooms'][room_num][person] += 1
            daily_stats['time_room_slots'][slot][person] += 1
            if slot.startswith('8:30') and '_당직' not in slot:
                daily_stats['early'][person] += 1
            elif slot.startswith('10:00'):
                daily_stats['late'][person] += 1
            if slot.startswith('8:30') and slot.endswith('_당직'):
                daily_stats['morning_duty'][person] += 1
            elif slot.startswith('13:30') and slot.endswith('_당직'):
                daily_stats['afternoon_duty'][person] += 1

    # total_stats 업데이트
    for key in ['early', 'late', 'morning_duty', 'afternoon_duty']:
        total_stats[key].update(daily_stats[key])
    for room in daily_stats['rooms']:
        total_stats['rooms'][room].update(daily_stats['rooms'][room])
    for slot in daily_stats['time_room_slots']:
        total_stats['time_room_slots'][slot].update(daily_stats['time_room_slots'][slot])

    return assignment, daily_stats

if st.button("🚀 방배정 수행", type="primary", use_container_width=True):
    with st.spinner("방 배정 중..."):
        time.sleep(1)

        # --- [핵심 수정] 버튼 클릭 시점에 토요/휴일 정보를 다시 로드 ---
        client = get_gspread_client()
        if client is None:
            st.error("Google Sheets 클라이언트를 초기화할 수 없습니다.")
            st.stop()
        try:
            spreadsheet = client.open_by_url(st.secrets["google_sheet"]["url"])
            sheet_name = f"{month_str} 토요/휴일 일자"
            worksheet = spreadsheet.worksheet(sheet_name)
            schedule_data = worksheet.get_all_records()
            special_df = pd.DataFrame(schedule_data) if schedule_data else pd.DataFrame(columns=["날짜", "당직 인원"])
            special_dates = {f"{datetime.strptime(row['날짜'], '%Y-%m-%d').month}월 {datetime.strptime(row['날짜'], '%Y-%m-%d').day}일" for row in schedule_data if row.get("날짜")}
        except Exception as e:
            st.warning(f"토요/휴일 정보를 불러오는 데 실패했습니다: {e}")
            special_df = pd.DataFrame(columns=["날짜", "당직 인원"])
            special_dates = set()

        # --- 최종 당직 정보 입력 검증 ---
        for date_obj, date_str, _ in special_schedules:
            settings = st.session_state.get("weekend_room_settings", {}).get(date_str, {})
            total_room_count = settings.get("total_room_count", 0)
            duty_room_selected = settings.get("duty_room")
            
            duty_person_val = ""
            if not special_df.empty:
                duty_row = special_df[special_df['날짜'] == date_obj.strftime('%Y-%m-%d')]
                if not duty_row.empty: duty_person_val = str(duty_row['당직 인원'].iloc[0]).strip()

            if total_room_count > 0 and duty_person_val and not duty_room_selected:
                st.error(f"⚠️ {date_str}: 당직 인원({duty_person_val})이 지정되어 있으므로, '토요/휴일 방 설정'에서 당직 방을 선택해야 합니다.")
                st.stop()

        # --- 평일 방 설정 검증 및 슬롯 정보 생성 (기존과 동일) ---
        time_slots, time_groups, memo_rules = {}, {}, {}
        if num_830 + num_900 + num_930 + num_1000 != 12:
            st.error(f"오전 방 개수 합계는 12개여야 합니다. (온콜 제외) 현재: {num_830 + num_900 + num_930 + num_1000}개")
            st.stop()
        elif len(rooms_830) != num_830 or len(rooms_900) != num_900 or len(rooms_930) != num_930 or len(rooms_1000) != num_1000 or len(rooms_1330) != num_1330:
            st.error("각 시간대의 방 번호 선택을 완료해주세요.")
            st.stop()
        else:
            for room in rooms_830:
                slot = f"8:30({room})_당직" if room == duty_830 else f"8:30({room})"
                time_slots[slot] = len(time_slots)
                time_groups.setdefault('8:30', []).append(slot)
            for room in rooms_900:
                slot = f"9:00({room})"
                time_slots[slot] = len(time_slots)
                time_groups.setdefault('9:00', []).append(slot)
            for room in rooms_930:
                slot = f"9:30({room})"
                time_slots[slot] = len(time_slots)
                time_groups.setdefault('9:30', []).append(slot)
            for room in rooms_1000:
                slot = f"10:00({room})"
                time_slots[slot] = len(time_slots)
                time_groups.setdefault('10:00', []).append(slot)
            for room in rooms_1330:
                slot = f"13:30({room})_당직" if room == duty_1330 else f"13:30({room})"
                time_slots[slot] = len(time_slots)
                time_groups.setdefault('13:30', []).append(slot)

            memo_rules = {
                **{f'{i}번방': [s for s in time_slots if f'({i})' in s and '_당직' not in s] for i in range(1, 13)},
                '당직 아닌 이른방': [s for s in time_slots if s.startswith('8:30') and '_당직' not in s],
                '이른방 제외': [s for s in time_slots if s.startswith(('9:00', '9:30', '10:00'))],
                '늦은방 제외': [s for s in time_slots if s.startswith(('8:30', '9:00', '9:30'))],
                '8:30': [s for s in time_slots if s.startswith('8:30') and '_당직' not in s],
                '9:00': [s for s in time_slots if s.startswith('9:00')],
                '9:30': [s for s in time_slots if s.startswith('9:30')],
                '10:00': [s for s in time_slots if s.startswith('10:00')],
                '오후 당직 제외': [s for s in time_slots if s.startswith('13:30') and '_당직' not in s]
            }

            st.session_state.update({"time_slots": time_slots, "time_groups": time_groups, "memo_rules": memo_rules})

        morning_duty_slot = f"8:30({duty_830})_당직"
        all_slots = [morning_duty_slot] + sorted([s for s in time_slots if s.startswith('8:30') and not s.endswith('_당직')]) + sorted([s for s in time_slots if s.startswith('9:00')]) + sorted([s for s in time_slots if s.startswith('9:30')]) + sorted([s for s in time_slots if s.startswith('10:00')]) + ['온콜'] + sorted([s for s in time_slots if s.startswith('13:30') and s.endswith('_당직')]) + sorted([s for s in time_slots if s.startswith('13:30') and not s.endswith('_당직')])
        columns = ['날짜', '요일'] + all_slots

        # --- 배정 로직 ---
        total_stats = {'early': Counter(), 'late': Counter(), 'morning_duty': Counter(), 'afternoon_duty': Counter(), 'rooms': {str(i): Counter() for i in range(1, 13)}, 'time_room_slots': {s: Counter() for s in time_slots}}
        df_cumulative = st.session_state["df_cumulative"]
        afternoon_duty_counts = {row['이름']: int(row['오후당직']) for _, row in df_cumulative.iterrows() if pd.notna(row.get('오후당직')) and int(row['오후당직']) > 0}

        assignments, date_cache, request_cells, result_data = {}, {}, {}, []
        assignable_slots = [s for s in st.session_state["time_slots"].keys() if not (s.startswith('8:30') and s.endswith('_당직'))]
        weekday_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}

        special_dates = [date_str for _, date_str, _ in special_schedules]

        for _, row in st.session_state["df_schedule_md"].iterrows():
            date_str = row['날짜']
            try:
                date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=2025) if "월" in date_str else datetime.strptime(date_str, '%Y-%m-%d')
                formatted_date = date_obj.strftime('%Y-%m-%d').strip()
                date_cache[date_str] = formatted_date
                day_of_week = weekday_map[date_obj.weekday()]
            except (ValueError, TypeError):
                continue

            result_row = [date_str, day_of_week]

            # --- [수정된] 토요/휴일 배정 로직 ---
            if date_str in special_dates:
                personnel = [p for p in row.iloc[2:].dropna() if p]
                settings = st.session_state["weekend_room_settings"].get(date_str, {})
                selected_rooms = settings.get("selected_rooms", [])
                duty_room_selected = settings.get("duty_room")

                duty_person = None
                date_obj_for_lookup = datetime.strptime(date_str, '%m월 %d일').replace(year=2025)
                formatted_date_for_lookup = date_obj_for_lookup.strftime('%Y-%m-%d')

                duty_person_row = special_df[special_df['날짜'] == formatted_date_for_lookup]
                if not duty_person_row.empty:
                    duty_person_raw = duty_person_row['당직 인원'].iloc[0]
                    if pd.notna(duty_person_raw) and str(duty_person_raw).strip():
                        duty_person = str(duty_person_raw).strip()

                room_to_first_slot_idx = {}
                for slot_idx, slot_name in enumerate(columns[2:]):
                    room_match = re.search(r'\((\d+)\)', str(slot_name))
                    if room_match:
                        room_num = room_match.group(1)
                        if room_num not in room_to_first_slot_idx:
                            room_to_first_slot_idx[room_num] = slot_idx

                mapped_assignment = [None] * (len(columns) - 2)
                remaining_personnel = list(personnel)

                # 당직자 우선 배정
                if duty_person and duty_room_selected and duty_person in remaining_personnel:
                    if duty_room_selected in room_to_first_slot_idx:
                        slot_idx = room_to_first_slot_idx[duty_room_selected]
                        mapped_assignment[slot_idx] = duty_person
                        remaining_personnel.remove(duty_person)
                    else:
                        st.warning(f"⚠️ {date_str}: 선택된 당직 방({duty_room_selected})에 해당하는 열을 찾을 수 없습니다.")

                # 나머지 인원 랜덤 배정
                remaining_rooms = [r for r in selected_rooms if r != duty_room_selected]
                random.shuffle(remaining_personnel)

                for room in remaining_rooms:
                    if not remaining_personnel:
                        break
                    if room in room_to_first_slot_idx:
                        person_to_assign = remaining_personnel.pop(0)
                        slot_idx = room_to_first_slot_idx[room]
                        if mapped_assignment[slot_idx] is None:
                            mapped_assignment[slot_idx] = person_to_assign
                        else:
                            # 만약 슬롯이 이미 채워져 있다면(예: 당직자로), 다른 사람을 넣지 않고 경고
                            st.warning(f"{date_str}: 슬롯이 이미 배정되어 {person_to_assign}을(를) {room}번 방에 배정하지 못했습니다.")
                            remaining_personnel.insert(0, person_to_assign) # 다시 목록에 추가
                    else:
                        st.warning(f"⚠️ {date_str}: 선택된 방({room})에 해당하는 열을 찾을 수 없습니다.")

                result_data.append(result_row + mapped_assignment)
                continue # 평일 로직 건너뛰기

            has_person = any(val for val in row.iloc[2:-1] if pd.notna(val) and val)
            personnel_for_the_day = [p for p in row.iloc[2:].dropna() if p]
                    
            # 이 코드는 사용자의 기존 `if date_str in special_dates:` 블록을 대체합니다.
            if date_str in special_dates:
                found_special_schedule = False
                # 해당 날짜의 특별 근무 일정을 찾습니다.
                for date_obj, special_date_str, personnel in special_schedules:
                    if special_date_str == date_str:
                        # Streamlit 세션 상태에서 해당 날짜의 주말/공휴일 설정을 가져옵니다.
                        settings = st.session_state["weekend_room_settings"].get(date_str, {})
                        duty_person = settings.get("duty_person", None)
                        duty_room = settings.get("duty_room", None)

                        # 설정된 인원과 방 정보를 바탕으로 배정 계획을 생성합니다.
                        # assignment_dict는 {"방번호": "담당자"} 형태의 딕셔너리입니다.
                        assignment_dict, sorted_rooms = assign_special_date(personnel, date_str, duty_person, settings)
                        
                        # 배정된 인원 수가 방 수보다 적을 경우 경고 메시지를 표시합니다.
                        if len(assignment_dict) < len(sorted_rooms):
                            st.warning(f"{date_str}: 인원 수({len(personnel)}) 부족으로 {len(sorted_rooms) - len(assignment_dict)}개 방 배정 안 됨.")
                        
                        # 1. 각 방 번호와 매칭되는 첫 번째 오전 슬롯의 인덱스를 찾습니다.
                        room_to_first_slot_idx = {}
                        # DataFrame의 최종 컬럼을 기준으로 슬롯을 순회하여 길이 불일치 문제를 해결합니다.
                        for slot_idx, slot in enumerate(columns[2:]):
                            # 오후(13:30) 슬롯이나 '온콜' 등 배정 대상이 아닌 슬롯은 건너뜁니다.
                            slot_str = str(slot)
                            if '13:30' in slot_str or '온콜' in slot_str:
                                continue
                            
                            # 정규식을 사용해 슬롯 이름에서 방 번호를 추출합니다. 예: "8:30(1)_당직" -> "1"
                            room_match = re.search(r'\((\d+)\)', slot_str)
                            if room_match:
                                room_num = room_match.group(1)
                                # 아직 맵에 없는 방 번호일 경우에만 추가하여, 각 방의 '첫 번째' 슬롯만 매핑되도록 합니다.
                                if room_num not in room_to_first_slot_idx:
                                    room_to_first_slot_idx[room_num] = slot_idx
                        
                        # 2. 배정 결과를 최종 슬롯 리스트에 매핑합니다.
                        # 최종 결과(엑셀의 한 행)를 담을 리스트를 'columns' 길이에 맞춰 초기화합니다.
                        mapped_assignment = [None] * (len(columns) - 2)
                        # 중복 배정을 방지하기 위해 이미 배정된 인원을 기록하는 세트입니다.
                        assigned_personnel = set()
                        
                        # `assignment_dict`의 모든 항목(방-사람)을 순회하며 배정합니다.
                        for room_num, person_with_room in assignment_dict.items():
                            # 담당자 이름만 추출합니다. (예: "강승주[3]" -> "강승주")
                            person = person_with_room.split('[')[0].strip()

                            # 해당 방 번호가 배정 대상인 오전 슬롯에 포함되어 있는지 확인합니다.
                            if room_num in room_to_first_slot_idx:
                                # 이미 다른 방에 배정된 인원인지 확인하여 중복을 방지합니다.
                                if person in assigned_personnel:
                                    st.warning(f"{date_str}: {person}님이 중복 배정되었습니다. 확인이 필요합니다.")
                                    continue

                                # 배정할 슬롯의 인덱스를 가져옵니다.
                                slot_idx = room_to_first_slot_idx[room_num]
                                
                                # 최종 배정 리스트의 해당 위치에 담당자 이름을 할당합니다.
                                mapped_assignment[slot_idx] = person
                                # 이 담당자를 '배정 완료' 세트에 추가합니다.
                                assigned_personnel.add(person)
                        
                        # 완성된 배정 결과를 전체 결과 데이터에 추가합니다.
                        full_row = result_row + mapped_assignment
                        result_data.append(full_row)
                        found_special_schedule = True
                        break  # 해당 날짜의 처리가 끝났으므로 내부 루프를 종료합니다.

                # 특별 근무 일정이 없는 경우 (예: 공휴일이지만 근무자가 없는 날) 빈 행을 추가합니다.
                if not found_special_schedule:
                    result_data.append(result_row + [None] * (len(columns) - 2))

                # special_date 처리가 끝났으므로, 평일 배정 로직을 건너뛰고 다음 날짜로 넘어갑니다.
                continue

            # 기존 평일 처리
            # 2. '소수 인원 근무'로 판단할 기준 인원수를 설정합니다.
            SMALL_TEAM_THRESHOLD = 15

            # 3. 근무 인원수가 설정된 기준보다 적으면, 방 배정 없이 순서대로 나열합니다.
            if len(personnel_for_the_day) < SMALL_TEAM_THRESHOLD and has_person:
                result_row.append(None)
                result_row.extend(personnel_for_the_day)
                num_slots_to_fill = len(all_slots)
                slots_filled_count = len(personnel_for_the_day) + 1  # 근무자 수 + 비워둔 1칸
                padding_needed = num_slots_to_fill - slots_filled_count
                if padding_needed > 0:
                    result_row.extend([None] * padding_needed)
                result_data.append(result_row)
                continue
            
            morning_personnel = [row[str(i)] for i in range(1, 12) if pd.notna(row[str(i)]) and row[str(i)]]
            afternoon_personnel = [row[f'오후{i}'] for i in range(1, 5) if pd.notna(row[f'오후{i}']) and row[f'오후{i}']]
            
            if not (morning_personnel or afternoon_personnel):
                result_row.extend([None] * len(all_slots))
                result_data.append(result_row)
                continue
            
            request_assignments = {}
            if not st.session_state["df_room_request"].empty:
                for _, req in st.session_state["df_room_request"].iterrows():
                    req_date, is_morning = parse_date_info(req['날짜정보'])
                    if req_date and req_date == formatted_date:
                        slots_for_category = st.session_state["memo_rules"].get(req['분류'], [])
                        if slots_for_category:
                            valid_slots = [s for s in slots_for_category if (is_morning and not s.startswith('13:30')) or (not is_morning and s.startswith('13:30'))]
                            if valid_slots:
                                selected_slot = random.choice(valid_slots)
                                request_assignments[selected_slot] = req['이름']
                                request_cells[(formatted_date, selected_slot)] = {'이름': req['이름'], '분류': req['분류']}

            assignment, _ = random_assign(list(set(morning_personnel+afternoon_personnel)), assignable_slots, request_assignments, st.session_state["time_groups"], total_stats, morning_personnel, afternoon_personnel, afternoon_duty_counts)
            
            for slot in all_slots:
                person = row['오전당직(온콜)'] if slot == morning_duty_slot or slot == '온콜' else (assignment[assignable_slots.index(slot)] if slot in assignable_slots and assignment else None)
                result_row.append(person if has_person else None)
            
            # [추가] 중복 배정 검증 로직
            assignments_for_day = dict(zip(all_slots, result_row[2:]))
            morning_slots_check = [s for s in all_slots if s.startswith(('8:30', '9:00', '9:30', '10:00'))]
            afternoon_slots_check = [s for s in all_slots if s.startswith('13:30') or s == '온콜']

            morning_counts = Counter(p for s, p in assignments_for_day.items() if s in morning_slots_check and p)
            for person, count in morning_counts.items():
                if count > 1:
                    duplicated_slots = [s for s, p in assignments_for_day.items() if p == person and s in morning_slots_check]
                    st.error(f"⚠️ {date_str}: '{person}'님이 오전에 중복 배정되었습니다 (슬롯: {', '.join(duplicated_slots)}).")
            
            afternoon_counts = Counter(p for s, p in assignments_for_day.items() if s in afternoon_slots_check and p)
            for person, count in afternoon_counts.items():
                if count > 1:
                    duplicated_slots = [s for s, p in assignments_for_day.items() if p == person and s in afternoon_slots_check]
                    st.error(f"⚠️ {date_str}: '{person}'님이 오후/온콜에 중복 배정되었습니다 (슬롯: {', '.join(duplicated_slots)}).")

            result_data.append(result_row)
        
        df_room = pd.DataFrame(result_data, columns=columns)
        st.write(" ")
        st.markdown("**✅ 통합 배치 결과**")
        st.dataframe(df_room, hide_index=True)
        
        for row_data in result_data:
            current_date_str = row_data[0]
            if current_date_str in special_dates:
                continue  # 토요일/휴일은 통계에 포함 안 함
            person_on_call = row_data[columns.index('온콜')] if '온콜' in columns else None
            if person_on_call:
                total_stats['morning_duty'][person_on_call] += 1
                
        # --- 시간대 순서 정의 ---
        time_order = ['8:30', '9:00', '9:30', '10:00', '13:30']

        # --- 통계 DataFrame 생성 ---
        stats_data, all_personnel_stats = [], set(p for _, r in st.session_state["df_schedule_md"].iterrows() for p in r[2:-1].dropna() if p)
        for person in sorted(all_personnel_stats):
            stats_entry = {
                '인원': person,
                '이른방 합계': total_stats['early'][person],
                '늦은방 합계': total_stats['late'][person],
                '오전 당직 합계': total_stats['morning_duty'][person],
                '오후 당직 합계': total_stats['afternoon_duty'][person],
            }
            # 시간대(방) 합계 추가 (당직 제외)
            for slot in st.session_state["time_slots"].keys():
                if not slot.endswith('_당직'):  # 당직 슬롯 제외
                    stats_entry[f'{slot} 합계'] = total_stats['time_room_slots'].get(slot, Counter())[person]
            stats_data.append(stats_entry)

        # 컬럼 정렬: 시간대 및 방 번호 순으로
        sorted_columns = ['인원', '이른방 합계', '늦은방 합계', '오전 당직 합계', '오후 당직 합계']
        time_slots = sorted(
            [slot for slot in st.session_state["time_slots"].keys() if not slot.endswith('_당직')],
            key=lambda x: (
                time_order.index(x.split('(')[0]),  # 시간대 순서
                int(x.split('(')[1].split(')')[0])  # 방 번호 순서
            )
        )
        sorted_columns.extend([f'{slot} 합계' for slot in time_slots])
        stats_df = pd.DataFrame(stats_data)[sorted_columns]
        st.divider()
        st.markdown("**☑️ 인원별 통계**")
        st.dataframe(stats_df, hide_index=True)
                
        # --- Excel 생성 및 다운로드 로직 ---
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Schedule"

        import platform

        # 플랫폼에 따라 폰트 선택
        if platform.system() == "Windows":
            font_name = "맑은 고딕"  # Windows에서 기본 제공
        else:
            font_name = "Arial"  # Mac에서 기본 제공, Windows에서도 사용 가능

        # 색상 및 스타일 정의
        highlight_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        sky_blue_fill = PatternFill(start_color="CCEEFF", end_color="CCEEFF", fill_type="solid")
        duty_font = Font(name=font_name, size=9, bold=True, color="FF00FF")  # 폰트 크기 9로 명시
        default_font = Font(name=font_name, size=9)  # 폰트 크기 9로 명시
        special_day_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        no_person_day_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        default_yoil_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        special_person_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # special_schedules 근무자 셀 배경색

        # 세션에서 변경된 셀 위치를 가져옴
        swapped_assignments = st.session_state.get("swapped_assignments", set())

        # 헤더 렌더링
        for col_idx, header in enumerate(columns, 1):
            cell = sheet.cell(1, col_idx, header)
            cell.font = Font(bold=True, name=font_name, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            if header.startswith('8:30') or header == '온콜':
                cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            elif header.startswith('9:00'):
                cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            elif header.startswith('9:30'):
                cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
            elif header.startswith('10:00'):
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            elif header.startswith('13:30'):
                cell.fill = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")

        # 데이터 렌더링
        for row_idx, row_data in enumerate(result_data, 2):
            current_date_str = row_data[0]
            
            duty_person_for_the_day = None
            if current_date_str in special_dates:
                try:
                    date_obj_lookup = datetime.strptime(current_date_str, '%m월 %d일').replace(year=datetime.now().year)
                    formatted_date_lookup = date_obj_lookup.strftime('%Y-%m-%d')
                    duty_person_row = special_df[special_df['날짜'] == formatted_date_lookup]
                    if not duty_person_row.empty:
                        duty_person_raw = duty_person_row['당직 인원'].iloc[0]
                        if pd.notna(duty_person_raw) and str(duty_person_raw).strip():
                            duty_person_for_the_day = str(duty_person_raw).strip()
                except Exception as e:
                    st.warning(f"Excel 스타일링 중 당직 인원 조회 오류: {e}")

            assignment_cells = row_data[2:]
            personnel_in_row = [p for p in assignment_cells if p]
            is_no_person_day = not any(personnel_in_row)
            SMALL_TEAM_THRESHOLD_FORMAT = 15
            is_small_team_day = (0 < len(personnel_in_row) < SMALL_TEAM_THRESHOLD_FORMAT) or (current_date_str in special_dates)

            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row_idx, col_idx, value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                # --- 배경색 및 폰트 스타일링 로직 통합 ---
                
                # 1. 배경색 설정
                if col_idx == 1: # 날짜
                    cell.fill = no_person_day_fill
                elif col_idx == 2: # 요일
                    if is_no_person_day: cell.fill = no_person_day_fill
                    elif is_small_team_day: cell.fill = special_day_fill
                    else: cell.fill = default_yoil_fill
                elif is_no_person_day and col_idx >= 3:
                    cell.fill = no_person_day_fill
                elif current_date_str in special_dates and col_idx > 2 and value:
                    cell.fill = special_person_fill
                
                # 변경 요청된 셀 하이라이트 (배경색 덮어쓰기)
                slot_name = columns[col_idx-1]
                cell_shift_type = ''
                if any(time_str in str(slot_name) for time_str in ['8:30', '9:00', '9:30', '10:00']): cell_shift_type = '오전'
                elif any(time_str in str(slot_name) for time_str in ['13:30', '온콜']): cell_shift_type = '오후'
                
                if (current_date_str.strip(), cell_shift_type, str(value).strip()) in swapped_assignments:
                    cell.fill = highlight_fill

                # 2. 폰트 설정
                cell.font = default_font # 기본 폰트 먼저 적용
                
                if current_date_str in special_dates:
                    if duty_person_for_the_day and value == duty_person_for_the_day:
                        cell.font = duty_font
                else:
                    slot_name = columns[col_idx-1]
                    if (slot_name.endswith('_당직') or slot_name == '온콜') and value:
                        cell.font = duty_font

                # 3. 코멘트 추가
                if col_idx > 2 and value and date_cache.get(current_date_str):
                    formatted_date_for_comment = date_cache[current_date_str]
                    if (formatted_date_for_comment, slot_name) in request_cells and value == request_cells[(formatted_date_for_comment, slot_name)]['이름']:
                        cell.comment = Comment(f"{request_cells[(formatted_date_for_comment, slot_name)]['분류']}", "System")

                slot_name = columns[col_idx-1]
                cell_shift_type = ''
                if '8:30' in slot_name or '9:00' in slot_name or '9:30' in slot_name or '10:00' in slot_name:
                    cell_shift_type = '오전'
                elif '13:30' in slot_name or '온콜' in slot_name:
                    cell_shift_type = '오후'
                
                # 셀의 배경색 적용 (변경 요청 하이라이트)
                formatted_current_date = current_date_str.strip()
                if (formatted_current_date, cell_shift_type, str(value).strip()) in swapped_assignments:
                    cell.fill = highlight_fill

                # special_dates의 경우 폰트 설정
                if current_date_str in special_dates:
                    settings = st.session_state["weekend_room_settings"].get(current_date_str, {})
                    duty_room = settings.get("duty_room", None)
                    duty_person = settings.get("duty_person", None)
                    room_match = re.search(r'\((\d+)\)', slot_name)
                    if room_match:
                        room_num = room_match.group(1)
                        if room_num == duty_room and value and duty_person and duty_person != "선택 안 함" and value == duty_person:
                            cell.font = Font(name=font_name, size=9, bold=True, color="FF00FF")  # 당직 인원: 크기 9, 굵은 글씨, 보라색
                        else:
                            cell.font = Font(name=font_name, size=9)  # 일반 인원: 크기 9, 기본 스타일
                else:
                    # 평일 당직 강조 로직
                    if slot_name.startswith('8:30') and slot_name.endswith('_당직') and value:
                        cell.font = Font(name=font_name, size=9, bold=True, color="FF00FF")  # 크기 9, 굵은 글씨, 보라색
                    elif (slot_name.startswith('13:30') and slot_name.endswith('_당직') or slot_name == '온콜') and value:
                        cell.font = Font(name=font_name, size=9, bold=True, color="FF00FF")  # 크기 9, 굵은 글씨, 보라색
                    else:
                        cell.font = Font(name=font_name, size=9)  # 크기 9, 기본 스타일

                # special_dates의 경우 value를 그대로 셀에 기록
                if current_date_str in special_dates and col_idx > 2 and value:
                    cell.value = value
                elif col_idx > 2 and value and date_cache.get(current_date_str):
                    formatted_date_for_comment = date_cache[current_date_str]
                    if (formatted_date_for_comment, slot_name) in request_cells and value == request_cells[(formatted_date_for_comment, slot_name)]['이름']:
                        cell.comment = Comment(f"{request_cells[(formatted_date_for_comment, slot_name)]['분류']}", "System")

        # --- Stats 시트 생성 ---
        stats_sheet = wb.create_sheet("Stats")
        stats_columns = stats_df.columns.tolist()
        for col_idx, header in enumerate(stats_columns, 1):
            stats_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
            cell = stats_sheet.cell(1, col_idx, header)
            cell.font = Font(bold=True, name=font_name, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            if header == '인원':
                cell.fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
            elif header == '이른방 합계':
                cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            elif header == '늦은방 합계':
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            elif '당직' in header:
                cell.fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for row_idx, row in enumerate(stats_df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = stats_sheet.cell(row_idx, col_idx, value)
                cell.font = Font(name=font_name, size=9)  # 통계 시트도 크기 9로 통일
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.divider()
        st.download_button(
            label="📥 최종 방배정 다운로드",
            data=output,
            file_name=f"{month_str} 방배정.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )

        # Google Sheets에 방배정 시트 저장
        try:
            gc = get_gspread_client()
            sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
        except gspread.exceptions.APIError as e:
            st.warning("⚠️ 너무 많은 요청이 접속되어 딜레이되고 있습니다. 잠시 후 재시도 해주세요.")
            st.error(f"Google Sheets API 오류 (연결 단계): {e.response.status_code} - {e.response.text}")
            st.stop()
        except NameError as e:
            st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
            st.error(f"Google Sheets 연결 중 오류: {type(e).__name__} - {e}")
            st.stop()
        except Exception as e:
            st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
            st.error(f"Google Sheets 연결 중 오류: {type(e).__name__} - {e}")
            st.stop()
            
        try:
            worksheet_result = sheet.worksheet(f"{month_str} 방배정")
        except gspread.exceptions.WorksheetNotFound:
            try:
                worksheet_result = sheet.add_worksheet(f"{month_str} 방배정", rows=100, cols=len(df_room.columns))
            except gspread.exceptions.APIError as e:
                st.warning("⚠️ 너무 많은 요청이 접속되어 딜레이되고 있습니다. 잠시 후 재시도 해주세요.")
                st.error(f"Google Sheets API 오류 ('{month_str} 방배정' 시트 생성): {e.response.status_code} - {e.response.text}")
                st.stop()
            except NameError as e:
                st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
                st.error(f"'{month_str} 방배정' 시트 생성 중 오류: {type(e).__name__} - {e}")
                st.stop()
            except Exception as e:
                st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
                st.error(f"'{month_str} 방배정' 시트 생성 실패: {type(e).__name__} - {e}")
                st.stop()
        except gspread.exceptions.APIError as e:
            st.warning("⚠️ 너무 많은 요청이 접속되어 딜레이되고 있습니다. 잠시 후 재시도 해주세요.")
            st.error(f"Google Sheets API 오류 ('{month_str} 방배정' 시트 로드): {e.response.status_code} - {e.response.text}")
            st.stop()
        except NameError as e:
            st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
            st.error(f"'{month_str} 방배정' 시트 로드 중 오류: {type(e).__name__} - {e}")
            st.stop()
        except Exception as e:
            st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
            st.error(f"'{month_str} 방배정' 시트 로드 실패: {type(e).__name__} - {e}")
            st.stop()
            
        update_sheet_with_retry(worksheet_result, [df_room.columns.tolist()] + df_room.fillna('').values.tolist())
        st.success(f"✅ {month_str} 방배정 테이블이 Google Sheets에 저장되었습니다.")