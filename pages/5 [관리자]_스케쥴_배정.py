import streamlit as st
import pandas as pd
import datetime
import calendar
from io import BytesIO
from dateutil.relativedelta import relativedelta
from google.oauth2.service_account import Credentials
import gspread
from gspread.exceptions import WorksheetNotFound, APIError
import time
import io
import xlsxwriter
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.comments import Comment
from datetime import timedelta

random.seed(42)

# 🔒 관리자 페이지 체크
if "login_success" not in st.session_state or not st.session_state["login_success"]:
    st.warning("⚠️ Home 페이지에서 비밀번호와 사번을 먼저 입력해주세요.")
    st.stop()

# 관리자 권한 체크
if not st.session_state.get("is_admin_authenticated", False):
    st.warning("⚠️ 관리자 권한이 없습니다.")
    st.stop()

# 사이드바
st.sidebar.write(f"현재 사용자: {st.session_state['name']} ({str(st.session_state['employee_id']).zfill(5)})")
if st.sidebar.button("로그아웃"):
    st.session_state.clear()
    st.success("로그아웃되었습니다. 🏠 Home 페이지로 돌아가 주세요.")
    time.sleep(5)
    st.rerun()

# 초기 데이터 로드 및 세션 상태 설정
url = st.secrets["google_sheet"]["url"]
month_str = "2025년 04월"

# Google Sheets 클라이언트 초기화
@st.cache_resource # 이 함수 자체를 캐싱하여 불필요한 초기화 반복 방지
def get_gspread_client():
    # st.write("DEBUG: get_gspread_client() 호출 시작") # 너무 자주 나올 수 있어 주석 처리
    scope = ["https://www.googleapis.com/auth/spreadsheets"]
    try:
        service_account_info = dict(st.secrets["gspread"])
        service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
        credentials = Credentials.from_service_account_info(service_account_info, scopes=scope)
        gc = gspread.authorize(credentials)
        # st.success("✅ Google Sheets 클라이언트 인증 성공!") # 성공 메시지는 load_data_page5에서만
        # st.write("DEBUG: get_gspread_client() 호출 종료")
        return gc
    except Exception as e:
        st.error(f"❌ Google Sheets 클라이언트 초기화 또는 인증 실패: {type(e).__name__} - {e}")
        st.exception(e) # 상세 스택 트레이스 출력
        st.stop() # 치명적인 오류이므로 앱 중단


# 데이터 로드 함수 (세션 상태 활용으로 쿼터 절약)
@st.cache_data(ttl=3600) # 데이터를 1시간 동안 캐시. 개발 중에는 ttl을 0으로 설정하거나 캐시를 자주 지우세요.
def load_data_page5():
    st.write("DEBUG: load_data_page5() 호출 시작") # 디버그 메시지
    required_keys = ["df_master", "df_request", "df_cumulative", "df_shift", "df_supplement"]
    if "data_loaded" not in st.session_state or not st.session_state["data_loaded"] or not all(key in st.session_state for key in required_keys):
        st.write("DEBUG: 데이터 로드 필요. Google Sheets에서 데이터 가져오는 중...") # 디버그 메시지
        url = st.secrets["google_sheet"]["url"]
        gc = get_gspread_client() # 캐싱된 클라이언트 가져오기
        if gc is None: # get_gspread_client에서 이미 stop()을 하지만, 방어 코드
            st.stop()

        try:
            sheet = gc.open_by_url(url)
            st.write(f"DEBUG: 스프레드시트 '{url}' 열기 성공.") # 디버그 메시지
        except APIError as e:
            st.error(f"❌ 스프레드시트 열기 API 오류: {e.response.status_code} - {e.response.text}")
            st.exception(e) # 상세 스택 트레이스 출력
            st.stop()
        except Exception as e:
            st.error(f"❌ 스프레드시트 열기 실패: {type(e).__name__} - {e}")
            st.exception(e) # 상세 스택 트레이스 출력
            st.stop()

        # 마스터 시트
        try:
            worksheet1 = sheet.worksheet("마스터")
            st.session_state["df_master"] = pd.DataFrame(worksheet1.get_all_records())
            st.session_state["worksheet1"] = worksheet1
            st.write("DEBUG: '마스터' 시트 로드 성공.") # 디버그 메시지
        except WorksheetNotFound:
            st.error("❌ '마스터' 시트를 찾을 수 없습니다. 시트 이름을 확인해주세요.")
            st.stop()
        except APIError as e:
            st.error(f"❌ '마스터' 시트 로드 API 오류: {e.response.status_code} - {e.response.text}")
            st.exception(e)
            st.stop()
        except Exception as e:
            st.error(f"❌ '마스터' 시트 로드 실패: {type(e).__name__} - {e}")
            st.exception(e)
            st.session_state["df_master"] = pd.DataFrame(columns=["이름", "주차", "요일", "근무여부"])
            st.session_state["data_loaded"] = False
            st.stop()

        # 요청사항 시트
        try:
            worksheet2 = sheet.worksheet(f"{month_str} 요청")
            st.write(f"DEBUG: '{month_str} 요청' 시트 로드 성공.") # 디버그 메시지
        except WorksheetNotFound:
            st.warning(f"⚠️ '{month_str} 요청' 시트를 찾을 수 없습니다. 새로 생성합니다.")
            try:
                worksheet2 = sheet.add_worksheet(title=f"{month_str} 요청", rows="100", cols="20")
                worksheet2.append_row(["이름", "분류", "날짜정보"])
                names_in_master = st.session_state["df_master"]["이름"].unique()
                new_rows = [[name, "요청 없음", ""] for name in names_in_master]
                for row in new_rows:
                    worksheet2.append_row(row)
                st.write(f"DEBUG: '{month_str} 요청' 시트 새로 생성 및 초기 데이터 추가 성공.") # 디버그 메시지
            except APIError as e:
                st.error(f"❌ '{month_str} 요청' 시트 생성/초기화 API 오류: {e.response.status_code} - {e.response.text}")
                st.exception(e)
                st.stop()
            except Exception as e:
                st.error(f"❌ '{month_str} 요청' 시트 생성/초기화 실패: {type(e).__name__} - {e}")
                st.exception(e)
                st.stop()

        st.session_state["df_request"] = pd.DataFrame(worksheet2.get_all_records()) if worksheet2.get_all_records() else pd.DataFrame(columns=["이름", "분류", "날짜정보"])
        st.session_state["worksheet2"] = worksheet2

        # 누적 시트
        try:
            worksheet4 = sheet.worksheet(f"{month_str} 누적")
            st.write(f"DEBUG: '{month_str} 누적' 시트 로드 성공.") # 디버그 메시지
        except WorksheetNotFound:
            st.warning(f"⚠️ '{month_str} 누적' 시트를 찾을 수 없습니다. 새로 생성합니다.")
            try:
                worksheet4 = sheet.add_worksheet(title=f"{month_str} 누적", rows="100", cols="20")
                worksheet4.append_row([f"{month_str}", "오전누적", "오후누적", "오전당직 (온콜)", "오후당직"])
                names_in_master = st.session_state["df_master"]["이름"].unique()
                new_rows = [[name, "", "", "", ""] for name in names_in_master]
                for row in new_rows:
                    worksheet4.append_row(row)
                st.write(f"DEBUG: '{month_str} 누적' 시트 새로 생성 및 초기 데이터 추가 성공.") # 디버그 메시지
            except APIError as e:
                st.error(f"❌ '{month_str} 누적' 시트 생성/초기화 API 오류: {e.response.status_code} - {e.response.text}")
                st.exception(e)
                st.stop()
            except Exception as e:
                st.error(f"❌ '{month_str} 누적' 시트 생성/초기화 실패: {type(e).__name__} - {e}")
                st.exception(e)
                st.stop()
        
        # --- 수정: df_cumulative 로드 후 첫 번째 컬럼 이름을 '이름'으로 강제 변경 및 숫자 컬럼 타입 변환 ---
        df_cumulative_temp = pd.DataFrame(worksheet4.get_all_records()) if worksheet4.get_all_records() else pd.DataFrame(columns=[f"{month_str}", "오전누적", "오후누적", "오전당직 (온콜)", "오후당직"])
        if not df_cumulative_temp.empty:
            # 첫 번째 컬럼의 실제 이름이 무엇이든 '이름'으로 변경
            df_cumulative_temp.rename(columns={df_cumulative_temp.columns[0]: '이름'}, inplace=True)
            # 모든 누적 관련 컬럼을 숫자로 변환 (오류 방지)
            for col_name in ["오전누적", "오후누적", "오전당직 (온콜)", "오후당직"]:
                if col_name in df_cumulative_temp.columns:
                    # errors='coerce'를 사용하여 변환 불가능한 값은 NaN으로 만들고, fillna(0)으로 0으로 채움
                    df_cumulative_temp[col_name] = pd.to_numeric(df_cumulative_temp[col_name], errors='coerce').fillna(0).astype(int)
        st.session_state["df_cumulative"] = df_cumulative_temp
        # --- 수정 끝 ---

        st.session_state["worksheet4"] = worksheet4

        # df_shift와 df_supplement 생성 및 세션 상태에 저장
        st.session_state["df_shift"] = generate_shift_table(st.session_state["df_master"])
        st.session_state["df_supplement"] = generate_supplement_table(st.session_state["df_shift"], st.session_state["df_master"]["이름"].unique())

        st.session_state["data_loaded"] = True
        st.write("DEBUG: load_data_page5() 호출 종료 (성공)") # 디버그 메시지


# 근무 테이블 생성 함수
def generate_shift_table(df_master):
    def split_shift(row):
        shifts = []
        if row["근무여부"] == "오전 & 오후":
            shifts.extend([(row["이름"], row["주차"], row["요일"], "오전"), (row["이름"], row["주차"], row["요일"], "오후")])
        elif row["근무여부"] in ["오전", "오후"]:
            shifts.append((row["이름"], row["주차"], row["요일"], row["근무여부"]))
        return shifts

    shift_list = [shift for _, row in df_master.iterrows() for shift in split_shift(row)]
    df_split = pd.DataFrame(shift_list, columns=["이름", "주차", "요일", "시간대"])

    weekday_order = ["월", "화", "수", "목", "금"]
    time_slots = ["오전", "오후"]
    result = {}
    for day in weekday_order:
        for time in time_slots:
            key = f"{day} {time}"
            df_filtered = df_split[(df_split["요일"] == day) & (df_split["시간대"] == time)]
            every_week = df_filtered[df_filtered["주차"] == "매주"]["이름"].unique()
            specific_weeks = df_filtered[df_filtered["주차"] != "매주"]
            specific_week_dict = {name: sorted(specific_weeks[specific_weeks["이름"] == name]["주차"].tolist(), 
                                               key=lambda x: int(x.replace("주", ""))) 
                                  for name in specific_weeks["이름"].unique() if specific_weeks[specific_weeks["이름"] == name]["주차"].tolist()}
            employees = list(every_week) + [f"{name}({','.join(weeks)})" for name, weeks in specific_week_dict.items()]
            result[key] = ", ".join(employees) if employees else ""
    
    return pd.DataFrame(list(result.items()), columns=["시간대", "근무"])

# 보충 테이블 생성 함수
def generate_supplement_table(df_result, names_in_master):
    supplement = []
    weekday_order = ["월", "화", "수", "목", "금"]
    shift_list = ["오전", "오후"]
    names_in_master = set(names_in_master)

    for day in weekday_order:
        for shift in shift_list:
            time_slot = f"{day} {shift}"
            row = df_result[df_result["시간대"] == time_slot].iloc[0]
            employees = set(emp.split("(")[0].strip() for emp in row["근무"].split(", ") if emp)
            supplement_employees = names_in_master - employees

            if shift == "오후":
                morning_slot = f"{day} 오전"
                morning_employees = set(df_result[df_result["시간대"] == morning_slot].iloc[0]["근무"].split(", ") 
                                        if morning_slot in df_result["시간대"].values else [])
                supplement_employees = {emp if emp in morning_employees else f"{emp}🔺" for emp in supplement_employees}

            supplement.append({"시간대": time_slot, "보충": ", ".join(sorted(supplement_employees)) if supplement_employees else ""})

    return pd.DataFrame(supplement)

def split_column_to_multiple(df, column_name, prefix):
    """
    데이터프레임의 특정 열을 쉼표로 분리하여 여러 열로 변환하는 함수
    
    Parameters:
    - df: 입력 데이터프레임
    - column_name: 분리할 열 이름 (예: "근무", "보충")
    - prefix: 새로운 열 이름의 접두사 (예: "근무", "보충")
    
    Returns:
    - 새로운 데이터프레임
    """
    # 줄바꿈(\n)을 쉼표로 변환
    df[column_name] = df[column_name].str.replace("\n", ", ")
    
    # 쉼표로 분리하여 리스트로 변환
    split_data = df[column_name].str.split(", ", expand=True)
    
    # 최대 열 수 계산 (가장 많은 인원을 가진 행 기준)
    max_cols = split_data.shape[1]
    
    # 새로운 열 이름 생성 (예: 근무1, 근무2, ...)
    new_columns = [f"{prefix}{i+1}" for i in range(max_cols)]
    split_data.columns = new_columns
    
    # 원래 데이터프레임에서 해당 열 삭제
    df = df.drop(columns=[column_name])
    
    # 분리된 데이터를 원래 데이터프레임에 추가
    df = pd.concat([df, split_data], axis=1)

    return df

# 새로고침 버튼 (맨 상단)
if st.button("🔄 새로고침 (R)"):
    st.cache_data.clear()
    st.cache_resource.clear() # @st.cache_resource 적용 시 캐시 초기화
    st.session_state["data_loaded"] = False  # 데이터 리로드 강제
    load_data_page5()  # load_data_page5 호출로 모든 데이터 갱신
    st.success("데이터가 새로고침되었습니다.")
    st.rerun()

# 메인 로직
if st.session_state.get("is_admin_authenticated", False):
    load_data_page5()
    # Use .get() with fallback to avoid KeyError
    df_master = st.session_state.get("df_master", pd.DataFrame(columns=["이름", "주차", "요일", "근무여부"]))
    df_request = st.session_state.get("df_request", pd.DataFrame(columns=["이름", "분류", "날짜정보"]))
    # df_cumulative 컬럼 이름은 load_data_page5에서 '이름'으로 변경되었음
    df_cumulative = st.session_state.get("df_cumulative", pd.DataFrame(columns=["이름", "오전누적", "오후누적", "오전당직 (온콜)", "오후당직"])) # fallback 컬럼도 '이름'으로 통일
    df_shift = st.session_state.get("df_shift", pd.DataFrame())  # 세션 상태에서 가져오기
    df_supplement = st.session_state.get("df_supplement", pd.DataFrame())  # 세션 상태에서 가져오기

    st.subheader(f"✨ {month_str} 테이블 종합")

    # 데이터 전처리: 근무 테이블과 보충 테이블의 열 분리
    df_shift_processed = split_column_to_multiple(df_shift, "근무", "근무")
    df_supplement_processed = split_column_to_multiple(df_supplement, "보충", "보충")

    # Excel 다운로드 함수 (다중 시트)
    def excel_download(name, sheet1, name1, sheet2, name2, sheet3, name3, sheet4, name4):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            sheet1.to_excel(writer, sheet_name=name1, index=False)
            sheet2.to_excel(writer, sheet_name=name2, index=False)
            sheet3.to_excel(writer, sheet_name=name3, index=False)
            sheet4.to_excel(writer, sheet_name=name4, index=False)
        
        excel_data = output.getvalue()
        return excel_data

    # 근무 테이블
    st.write(" ")
    st.markdown("**✅ 근무 테이블**")
    st.dataframe(df_shift, use_container_width=True)

    # 보충 테이블 (중복된 df_master 표시 제거, df_supplement 표시)
    st.markdown("**☑️ 보충 테이블**")
    st.dataframe(df_supplement, use_container_width=True)

    # 요청사항 테이블
    st.markdown("**🙋‍♂️ 요청사항 테이블**")
    st.dataframe(df_request, use_container_width=True)

    # 누적 테이블
    st.markdown("**➕ 누적 테이블**")
    st.dataframe(df_cumulative, use_container_width=True)

    # 다운로드 버튼 추가
    excel_data = excel_download(
        name=f"{month_str} 테이블 종합",
        sheet1=df_shift_processed, name1="근무 테이블",
        sheet2=df_supplement_processed, name2="보충 테이블",
        sheet3=df_request, name3="요청사항 테이블",
        sheet4=df_cumulative, name4="누적 테이블"
    )
    st.download_button(
        label="📥 상단 테이블 다운로드",
        data=excel_data,
        file_name=f"{month_str} 테이블 종합.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # 근무 배정 로직
    # 누적 근무 횟수 추적용 딕셔너리 초기화
    current_cumulative = {'오전': {}, '오후': {}}

    # 2025년 4월 평일 생성
    next_month = datetime.datetime(2025, 4, 1)
    _, last_day = calendar.monthrange(next_month.year, next_month.month)
    dates = pd.date_range(start=next_month, end=next_month.replace(day=last_day))
    weekdays = [d for d in dates if d.weekday() < 5]
    week_numbers = {d: (d.day - 1) // 7 + 1 for d in dates}
    day_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금'}

    # df_final 초기화
    df_final = pd.DataFrame(columns=['날짜', '요일', '주차', '시간대', '근무자', '상태', '메모', '색상'])

    # 데이터프레임 로드 확인 (Streamlit UI로 변경)
    st.divider()
    st.subheader(f"✨ {month_str} 스케쥴 배정 수행")
    st.write("- 근무 배정 실행 시, 입력되어있는 '스케쥴 조정사항'이 초기화되므로 주의 부탁드립니다.")
    # st.write("df_shift_processed 확인:", df_shift_processed.head())
    # st.write("df_supplement_processed 확인:", df_supplement_processed.head())
    # st.write("df_request 확인:", df_request.head())
    # st.write("df_cumulative 확인:", df_cumulative.head())

    # 날짜 범위 파싱 함수
    def parse_date_range(date_str):
        if pd.isna(date_str) or not isinstance(date_str, str) or date_str.strip() == '':
            return []
        date_str = date_str.strip()
        result = []
        if ',' in date_str:
            for single_date in date_str.split(','):
                single_date = single_date.strip()
                try:
                    parsed_date = datetime.datetime.strptime(single_date, '%Y-%m-%d')
                    if parsed_date.weekday() < 5:
                        result.append(single_date)
                except ValueError:
                    # st.write(f"잘못된 날짜 형식 무시됨: {single_date}") # DEBUG 메시지로 변경
                    pass # 이 메시지는 너무 많이 나올 수 있어 주석 처리
            return result
        if '~' in date_str:
            try:
                start_date, end_date = date_str.split('~')
                start_date = start_date.strip()
                end_date = end_date.strip()
                start = datetime.datetime.strptime(start_date, '%Y-%m-%d')
                end = datetime.datetime.strptime(end_date, '%Y-%m-%d')
                date_list = pd.date_range(start=start, end=end)
                return [d.strftime('%Y-%m-%d') for d in date_list if d.weekday() < 5]
            except ValueError as e:
                # st.write(f"잘못된 날짜 범위 무시됨: {date_str}, 에러: {e}") # DEBUG 메시지로 변경
                pass # 이 메시지는 너무 많이 나올 수 있어 주석 처리
                return []
        try:
            parsed_date = datetime.datetime.strptime(date_str, '%Y-%m-%d')
            if parsed_date.weekday() < 5:
                return [date_str]
            return []
        except ValueError:
            # st.write(f"잘못된 날짜 형식 무시됨: {date_str}") # DEBUG 메시지로 변경
            pass # 이 메시지는 너무 많이 나올 수 있어 주석 처리
            return []

    # 근무자 상태 업데이트 함수
    def update_worker_status(df, date_str, time_slot, worker, status, memo, color):
        existing = df[
            (df['날짜'] == date_str) &
            (df['시간대'] == time_slot) &
            (df['근무자'] == worker.strip())
        ]
        if not existing.empty:
            df.loc[existing.index, ['상태', '메모', '색상']] = [status, memo, color]
        else:
            new_row = pd.DataFrame({
                '날짜': [date_str],
                '요일': [day_map[pd.to_datetime(date_str).weekday()]],
                '주차': [week_numbers[pd.to_datetime(date_str)]],
                '시간대': [time_slot],
                '근무자': [worker.strip()],
                '상태': [status],
                '메모': [memo],
                '색상': [color]
            })
            df = pd.concat([df, new_row], ignore_index=True)
        return df

    # df_final에서 특정 worker가 특정 날짜, 시간대에 '제외' 상태이며 특정 메모를 가지고 있는지 확인하는 헬퍼 함수
    def is_worker_already_excluded_with_memo(df_data, date_s, time_s, worker_s):
        # 해당 날짜, 시간대, 근무자의 모든 기록을 가져옴
        worker_records = df_data[
            (df_data['날짜'] == date_s) &
            (df_data['시간대'] == time_s) &
            (df_data['근무자'] == worker_s)
        ]
        if worker_records.empty:
            return False # 해당 근무자 기록 자체가 없으면 당연히 제외되지 않음

        # '제외' 또는 '추가제외' 상태인 기록만 필터링
        excluded_records = worker_records[worker_records['상태'].isin(['제외', '추가제외'])]
        if excluded_records.empty:
            return False # 제외된 기록이 없으면 False

        # 제외된 기록 중 해당 메모를 포함하는지 확인 (str.contains가 Series를 반환하므로 .any() 사용)
        return excluded_records['메모'].str.contains('보충 위해 제외됨|인원 초과로 인한 제외|오전 추가제외로 인한 오후 제외', na=False).any()


    df_cumulative_next = df_cumulative.copy()

    # 세션 상태 초기화 (기존 코드 유지)
    if "assigned" not in st.session_state:
        st.session_state.assigned = False
    if "downloaded" not in st.session_state:
        st.session_state.downloaded = False
    if "output" not in st.session_state:
        st.session_state.output = None

    # 휴관일 선택 UI 추가
    st.write(" ")
    st.markdown("**📅 센터 휴관일 추가**")

    # month_str에 해당하는 평일 날짜 생성 (이미 정의된 weekdays 사용)
    holiday_options = []
    for date in weekdays:
        date_str = date.strftime('%Y-%m-%d')
        date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
        day_name = day_map[date_obj.weekday()]
        holiday_format = f"{date_obj.month}월 {date_obj.day}일({day_name})"
        holiday_options.append((holiday_format, date_str))

    # st.multiselect로 휴관일 선택
    selected_holidays = st.multiselect(
        label=f"{month_str} 평일 중 휴관일을 선택",
        options=[option[0] for option in holiday_options],
        default=[],
        key="holiday_select",
        help="선택한 날짜는 근무 배정에서 제외됩니다."
    )

    # 선택된 휴관일을 날짜 형식(YYYY-MM-DD)으로 변환
    holiday_dates = []
    for holiday in selected_holidays:
        for option in holiday_options:
            if option[0] == holiday:
                holiday_dates.append(option[1])
                break

    # 토요 스케쥴 입력 UI 추가
    st.markdown("**📅 토요 스케쥴 입력**")

    # df_master와 df_request에서 이름 추출 및 중복 제거
    names_in_master = set(df_master["이름"].unique().tolist())
    names_in_request = set(df_request["이름"].unique().tolist())
    all_names = sorted(list(names_in_master.union(names_in_request)))  # 중복 제거 후 정렬

    # 2025년 4월의 토요일 날짜 추출
    saturdays = [d for d in dates if d.weekday() == 5]  # 토요일은 weekday() == 5
    saturday_options = []
    for date in saturdays:
        date_str = date.strftime('%Y-%m-%d')
        date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
        saturday_format = f"{date_obj.month}월 {date_obj.day}일(토)"
        saturday_options.append((saturday_format, date_str))

    # 최대 3개의 토요일 스케쥴 입력 허용
    saturday_schedules = []
    for i in range(3):
        cols = st.columns(2)
        with cols[0]:
            selected_saturday = st.selectbox(
                label=f"토요일 날짜 선택 {i+1}",
                options=["선택 안 함"] + [option[0] for option in saturday_options],
                key=f"saturday_select_{i}"
            )
        with cols[1]:
            if selected_saturday != "선택 안 함":
                selected_workers = st.multiselect(
                    label=f"근무 인원 선택 {i+1} (최대 10명)",
                    options=all_names,  # df_master와 df_request의 모든 이름 사용
                    default=[],
                    key=f"saturday_workers_{i}",
                )
                if len(selected_workers) > 10:
                    st.warning("근무 인원은 최대 10명까지 선택 가능합니다.")
                    selected_workers = selected_workers[:10]
            else:
                selected_workers = []

        # 선택된 데이터 저장
        if selected_saturday != "선택 안 함":
            for option in saturday_options:
                if option[0] == selected_saturday:
                    saturday_date = option[1]  #-MM-DD 형식
                    saturday_schedules.append((saturday_date, selected_workers))
                    break

    # 근무 배정 버튼
    st.write(" ")
    if st.button("🚀 근무 배정 실행"):
        st.write(" ")
        st.subheader(f"💡 {month_str} 스케쥴 배정 결과", divider='rainbow') # 추가된 제목
        # 버튼 클릭 시 세션 상태 초기화
        st.session_state.assigned = False
        st.session_state.output = None
        st.session_state.downloaded = False

        with st.spinner("근무 배정 중..."):
            time.sleep(1) # 스피너가 보이도록 잠시 대기

            st.write("DEBUG: 근무 배정 로직 시작.") # DEBUG
            
            # 날짜별 오전 근무 제외 인원 추적용 딕셔너리 (모든 날짜에 대해 초기화)
            excluded_morning_workers = {date.strftime('%Y-%m-%d'): set() for date in weekdays}

            # 휴관일을 제외한 평일 리스트 생성
            active_weekdays = [date for date in weekdays if date.strftime('%Y-%m-%d') not in holiday_dates]
            st.write(f"DEBUG: 활성화된 평일 수: {len(active_weekdays)}일.") # DEBUG

            # 1단계: 모든 날짜에 대해 오전 기본 배정 및 휴가자 처리 (휴관일 제외)
            st.write("DEBUG: 1단계: 오전 기본 배정 및 휴가자 처리 시작.") # DEBUG
            for date in active_weekdays:
                day_name = day_map[date.weekday()]
                week_num = week_numbers[date]
                date_str = date.strftime('%Y-%m-%d')

                # 휴가자 및 요청 사전 처리 (기존과 동일)
                vacationers = []
                must_work_morning = []
                must_work_afternoon = []
                no_supplement_morning = []
                no_supplement_afternoon = []
                hard_supplement_morning = []
                hard_supplement_afternoon = []

                for _, row in df_request.iterrows():
                    date_info = row['날짜정보']
                    name = row['이름']
                    category = row['분류']
                    if pd.isna(date_info) or not date_info:
                        continue
                    applicable_dates = parse_date_range(date_info)
                    if date_str in applicable_dates:
                        if category == '휴가':
                            vacationers.append(name)
                        elif category == '꼭 근무(오전)':
                            must_work_morning.append(name)
                        elif category == '꼭 근무(오후)':
                            must_work_afternoon.append(name)
                        elif category == '보충 불가(오전)':
                            no_supplement_morning.append(name)
                        elif category == '보충 불가(오후)':
                            no_supplement_afternoon.append(name)
                        elif category == '보충 어려움(오전)':
                            hard_supplement_morning.append(name)
                        elif category == '보충 어려움(오후)':
                            hard_supplement_afternoon.append(name)

                # 휴가자 사전 처리 (오전만)
                time_slot = '오전'
                shift_key = f'{day_name} {time_slot}'
                shift_row = df_shift_processed[df_shift_processed['시간대'] == shift_key]
                master_workers = set()
                if not shift_row.empty:
                    for col in [f'근무{i}' for i in range(1, 15)]:
                        worker = shift_row[col].values[0] if col in shift_row.columns and pd.notna(shift_row[col].values[0]) else ''
                        if worker:
                            if '(' in worker:
                                name, weeks = worker.split('(')
                                name = name.strip()
                                weeks = weeks.rstrip(')').split(',')
                                if f'{week_num}주' in weeks:
                                    master_workers.add(name)
                            else:
                                master_workers.add(worker)

                for vac in vacationers:
                    if vac in master_workers:
                        df_final = update_worker_status(df_final, date_str, time_slot, vac, '제외', '휴가로 제외됨', '🔴 빨간색')
                        excluded_morning_workers[date_str].add(vac)

                # 오전 기본 배정
                target_count = 12
                must_work = must_work_morning
                shift_row = df_shift_processed[df_shift_processed['시간대'] == shift_key]
                workers = []
                initial_workers = set()
                if not shift_row.empty:
                    for col in [f'근무{i}' for i in range(1, 15)]:
                        worker = shift_row[col].values[0] if col in shift_row.columns and pd.notna(shift_row[col].values[0]) else ''
                        if worker:
                            if '(' in worker:
                                name, weeks = worker.split('(')
                                name = name.strip()
                                weeks = weeks.rstrip(')').split(',')
                                if f'{week_num}주' in weeks:
                                    workers.append(name)
                                    initial_workers.add(name)
                            else:
                                workers.append(worker)
                                initial_workers.add(worker)

                workers = [w for w in workers if w not in vacationers]
                initial_workers = initial_workers - set(vacationers)

                for mw in must_work:
                    if mw not in workers and mw not in vacationers:
                        workers.append(mw)
                        initial_workers.add(mw)

                for worker in workers:
                    status = '근무'
                    memo = ''
                    color = '기본'
                    if worker in must_work:
                        memo = f'꼭 근무({time_slot}) 위해 배정됨'
                        color = '🟠 주황색'
                    current_cumulative[time_slot][worker] = current_cumulative[time_slot].get(worker, 0) + 1
                    df_final = update_worker_status(df_final, date_str, time_slot, worker, status, memo, color)
            st.write("DEBUG: 1단계: 오전 기본 배정 및 휴가자 처리 완료.") # DEBUG

            # 2단계: 모든 날짜에 대해 오전 보충/제외 수행 (휴관일 제외)
            st.write("DEBUG: 2단계: 오전 보충/제외 수행 시작.") # DEBUG
            time_slot = '오전'
            target_count = 12
            # moved_workers는 다른 날짜로 이동한 사람 (초과 → 부족)을 추적하기 위해 전역적으로 유지
            moved_workers = set()
            
            for date in active_weekdays:
                date_str = date.strftime('%Y-%m-%d')
                day_name = day_map[date.weekday()]
                week_num = week_numbers[date]
                st.write(f"DEBUG:   - 처리 중인 날짜 (오전 보충/제외): {date_str}") # DEBUG

                # 요청사항 재확인 (기존과 동일)
                vacationers = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == '휴가']
                must_work = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == f'꼭 근무({time_slot})']
                no_supplement = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == f'보충 불가({time_slot})']
                hard_supplement = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == f'보충 어려움({time_slot})']

                iteration = 0
                while True: # 이 내부 루프가 문제의 핵심
                    iteration += 1
                    if iteration > 500: # DEBUG: 무한 루프 방지 (적절한 값으로 조정)
                        st.warning(f"⚠️ DEBUG: 오전 보충/제외 로직 무한 루프 의심! 날짜: {date_str}. 500회 반복 후 강제 종료.")
                        break # 루프 강제 종료

                    excess_dates_inner = [] # 내부 루프용 리스트 (현재 날짜 기준)
                    shortage_dates_inner = [] # 내부 루프용 리스트 (현재 날짜 기준)

                    # 현재 날짜의 실제 근무 인원 수 계산
                    current_workers_on_date = df_final[
                        (df_final['날짜'] == date_str) &
                        (df_final['시간대'] == time_slot) &
                        (df_final['상태'].isin(['근무', '보충'])) # 2단계 처리 전까지는 '근무'만 있을 수 있음
                    ]['근무자'].tolist()
                    count_on_date = len(current_workers_on_date)

                    if count_on_date > target_count:
                        excess_dates_inner.append((date_str, count_on_date - target_count))
                    elif count_on_date < target_count:
                        shortage_dates_inner.append((date_str, target_count - count_on_date))

                    st.write(f"DEBUG:     날짜: {date_str}, 오전 - 반복 {iteration}. 초과: {excess_dates_inner}, 부족: {shortage_dates_inner}. 현재 인원: {count_on_date}/{target_count}") # DEBUG

                    any_matched_inner = False # 내부 루프용
                    
                    # 초과 인원 처리 (다른 날짜로 이동) - 현재 날짜에만 집중
                    if excess_dates_inner:
                        excess_date_curr, excess_count_curr = excess_dates_inner[0]
                        if excess_count_curr > 0:
                            # 현재 날짜의 이동 가능한 초과 근무자
                            movable_excess_workers_on_date = [
                                w for w in current_workers_on_date
                                if w not in must_work and w not in moved_workers # 이미 이동된 사람 (이 단계에서)은 제외
                                and not is_worker_already_excluded_with_memo(df_final, excess_date_curr, time_slot, w) # 수정된 조건: ValueError 방지
                            ]
                            # 누적 근무 횟수가 높은 순으로 정렬하여 제거
                            movable_excess_workers_on_date.sort(key=lambda w: current_cumulative[time_slot].get(w, 0), reverse=True)

                            for _ in range(excess_count_curr):
                                if not movable_excess_workers_on_date:
                                    st.write(f"DEBUG:         날짜: {excess_date_curr}, 오전 - 이동 가능한 초과 근무자 없음.") # DEBUG
                                    break # 더 이상 이동할 사람 없으면 중단

                                worker_to_remove = movable_excess_workers_on_date.pop(0) # 가장 많이 근무한 사람부터 제거
                                moved_workers.add(worker_to_remove) # 이동된 사람으로 기록 (파란색 제외 대상)
                                
                                # --- 수정: '제외' (파란색) 부여 로직 ---
                                df_final = update_worker_status(df_final, excess_date_curr, time_slot, worker_to_remove, '제외', f'다른 날짜 부족분 보충 위해 제외됨', '🔵 파란색')
                                excluded_morning_workers[excess_date_curr].add(worker_to_remove) # 오전 제외 인원 기록
                                current_cumulative[time_slot][worker_to_remove] = current_cumulative[time_slot].get(worker_to_remove, 0) - 1
                                any_matched_inner = True
                                st.write(f"DEBUG:         날짜: {excess_date_curr}, 오전 - '{worker_to_remove}' 제거 (🔵파란색). 현재 인원: {len(current_workers_on_date) - (_+1)}") # DEBUG
                            
                            # 재계산하여 초과 여부 확인 (이 단계의 목표: 인원 이동)
                            current_workers_on_date = df_final[
                                (df_final['날짜'] == date_str) &
                                (df_final['시간대'] == time_slot) &
                                (df_final['상태'].isin(['근무', '보충', '제외'])) # '근무', '보충', '제외' 상태 포함 (방금 처리된 파란색 제외)
                            ]['근무자'].tolist()
                            if len(current_workers_on_date) <= target_count: # 인원 초과 해결
                                break # 목표 인원 만족하면 루프 종료
                        
                    # 부족 인원 처리 (보충) - 현재 날짜에만 집중
                    if shortage_dates_inner:
                        shortage_date_curr, shortage_count_curr = shortage_dates_inner[0]
                        if shortage_count_curr > 0:
                            # 현재 날짜의 보충 가능한 인원
                            shift_key = f'{day_name} {time_slot}'
                            supplement_row = df_supplement_processed[df_supplement_processed['시간대'] == shift_key]
                            all_possible_supplement = []
                            if not supplement_row.empty:
                                for col in [f'보충{i}' for i in range(1, 13)]:
                                    worker = supplement_row[col].values[0] if col in supplement_row.columns and pd.notna(supplement_row[col].values[0]) else ''
                                    if worker:
                                        name_only = worker.replace('🔺', '').strip()
                                        if name_only not in current_workers_on_date and \
                                           name_only not in vacationers and \
                                           name_only not in no_supplement:
                                            all_possible_supplement.append(name_only)
                            
                            # 누적 근무 횟수가 낮은 순으로 정렬하여 보충
                            all_possible_supplement_with_cumulative = [
                                (w, current_cumulative[time_slot].get(w, 0)) for w in all_possible_supplement
                            ]
                            all_possible_supplement_with_cumulative.sort(key=lambda x: x[1]) # 누적 근무 횟수 낮은 순으로 정렬

                            for _ in range(shortage_count_curr):
                                if not all_possible_supplement_with_cumulative:
                                    st.write(f"DEBUG:         날짜: {shortage_date_curr}, 오전 - 보충 가능한 인원 없음.") # DEBUG
                                    break # 더 이상 보충할 사람 없으면 중단

                                worker_to_add, _ = all_possible_supplement_with_cumulative.pop(0)
                                
                                # --- 수정: '보충' (연두색) 부여 로직 ---
                                df_final = update_worker_status(df_final, shortage_date_curr, time_slot, worker_to_add, '보충', f'다른 날짜 초과분에서 보충됨', '🟢 초록색')
                                current_cumulative[time_slot][worker_to_add] = current_cumulative[time_slot].get(worker_to_add, 0) + 1
                                any_matched_inner = True
                                st.write(f"DEBUG:         날짜: {shortage_date_curr}, 오전 - '{worker_to_add}' 보충 (🟢연두색). 현재 인원: {len(current_workers_on_date) + (_+1)}") # DEBUG
                            
                            # 재계산하여 부족 여부 확인 (이 단계의 목표: 인원 이동)
                            current_workers_on_date = df_final[
                                (df_final['날짜'] == date_str) &
                                (df_final['시간대'] == time_slot) &
                                (df_final['상태'].isin(['근무', '보충', '제외'])) # '근무', '보충', '제외' 상태 포함 (방금 처리된 초록색 보충)
                            ]['근무자'].tolist()
                            if len(current_workers_on_date) >= target_count:
                                break # 목표 인원 만족하면 루프 종료


                    # 두 리스트 모두 비어있고, 어떤 작업도 수행되지 않았다면 루프 종료
                    # 현재는 각 날짜별로 excess/shortage가 계산되므로, 여기서는 해당 날짜의 상태만 확인합니다.
                    final_count_check = len(df_final[
                        (df_final['날짜'] == date_str) &
                        (df_final['시간대'] == time_slot) &
                        (df_final['상태'].isin(['근무', '보충', '제외'])) # '근무', '보충', '제외' 상태만 포함
                    ]['근무자'].tolist())
                    if final_count_check == target_count or not any_matched_inner:
                        st.write(f"DEBUG:       날짜: {date_str}, 오전 - 내부 루프 종료. 최종 인원: {final_count_check}") # DEBUG
                        break # 목표 인원 도달 또는 더 이상 매칭 불가 시 루프 종료
            st.write("DEBUG: 2단계: 오전 보충/제외 수행 완료.") # DEBUG


            # 3단계: 모든 날짜에 대해 오후 기본 배정 (휴관일 제외)
            st.write("DEBUG: 3단계: 오후 기본 배정 시작.") # DEBUG
            for date in active_weekdays:
                day_name = day_map[date.weekday()]
                week_num = week_numbers[date]
                date_str = date.strftime('%Y-%m-%d')

                # 요청사항 재확인 (오후 관련)
                vacationers = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == '휴가']
                must_work_afternoon = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == '꼭 근무(오후)']

                # 휴가자 사전 처리 (오후만)
                time_slot = '오후'
                shift_key = f'{day_name} {time_slot}'
                shift_row = df_shift_processed[df_shift_processed['시간대'] == shift_key]
                master_workers = set()
                if not shift_row.empty:
                    for col in [f'근무{i}' for i in range(1, 15)]:
                        worker = shift_row[col].values[0] if col in shift_row.columns and pd.notna(shift_row[col].values[0]) else ''
                        if worker:
                            if '(' in worker:
                                name, weeks = worker.split('(')
                                name = name.strip()
                                weeks = weeks.rstrip(')').split(',')
                                if f'{week_num}주' in weeks:
                                    master_workers.add(name)
                            else:
                                master_workers.add(worker)

                for vac in vacationers:
                    if vac in master_workers:
                        df_final = update_worker_status(df_final, date_str, time_slot, vac, '제외', '휴가로 제외됨', '🔴 빨간색')

                # 오후 기본 배정
                target_count = 5
                must_work = must_work_afternoon
                shift_row = df_shift_processed[df_shift_processed['시간대'] == shift_key]
                workers = []
                initial_workers = set()
                if not shift_row.empty:
                    for col in [f'근무{i}' for i in range(1, 15)]:
                        worker = shift_row[col].values[0] if col in shift_row.columns and pd.notna(shift_row[col].values[0]) else ''
                        if worker:
                            if '(' in worker:
                                name, weeks = worker.split('(')
                                name = name.strip()
                                weeks = weeks.rstrip(')').split(',')
                                if f'{week_num}주' in weeks:
                                    workers.append(name)
                                    initial_workers.add(name)
                            else:
                                workers.append(worker)
                                initial_workers.add(worker)

                workers = [w for w in workers if w not in vacationers]
                initial_workers = initial_workers - set(vacationers)

                for mw in must_work:
                    if mw not in workers and mw not in vacationers:
                        workers.append(mw)
                        initial_workers.add(mw)

                # 오후 근무자: 오전 근무자 중에서 선택 (보충/제외 반영된 상태)
                morning_workers = df_final[
                    (df_final['날짜'] == date_str) &
                    (df_final['시간대'] == '오전') &
                    (df_final['상태'].isin(['근무', '보충', '추가보충'])) # 이 단계까지의 오전 근무자 상태 (추가보충 포함)
                ]['근무자'].tolist()
                workers = [w for w in workers if (w in morning_workers or w in must_work) and w not in excluded_morning_workers[date_str]]

                for worker in workers:
                    status = '근무'
                    memo = ''
                    color = '기본'
                    if worker in must_work:
                        memo = f'꼭 근무({time_slot}) 위해 배정됨'
                        color = '🟠 주황색'
                    current_cumulative[time_slot][worker] = current_cumulative[time_slot].get(worker, 0) + 1
                    df_final = update_worker_status(df_final, date_str, time_slot, worker, status, memo, color)
            st.write("DEBUG: 3단계: 오후 기본 배정 완료.") # DEBUG

            # 4단계: 모든 날짜에 대해 오후 보충/제외 수행 (휴관일 제외)
            st.write("DEBUG: 4단계: 오후 보충/제외 수행 시작.") # DEBUG
            time_slot = '오후'
            target_count = 5
            for date in active_weekdays:
                date_str = date.strftime('%Y-%m-%d')
                day_name = day_map[date.weekday()]
                week_num = week_numbers[date]
                st.write(f"DEBUG:   - 처리 중인 날짜 (오후 보충/제외): {date_str}") # DEBUG

                # 요청사항 재확인 (기존과 동일)
                vacationers = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == '휴가']
                must_work = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == f'꼭 근무({time_slot})']
                no_supplement = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == f'보충 불가({time_slot})']
                hard_supplement = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == f'보충 어려움({time_slot})']

                iteration = 0
                while True: # 이 내부 루프가 문제의 핵심
                    iteration += 1
                    if iteration > 500: # DEBUG: 무한 루프 방지 (적절한 값으로 조정)
                        st.warning(f"⚠️ DEBUG: 오후 보충/제외 로직 무한 루프 의심! 날짜: {date_str}. 500회 반복 후 강제 종료.")
                        break # 루프 강제 종료

                    excess_dates_inner = [] # 내부 루프용 리스트
                    shortage_dates_inner = [] # 내부 루프용 리스트

                    current_workers_on_date = df_final[
                        (df_final['날짜'] == date_str) &
                        (df_final['시간대'] == time_slot) &
                        (df_final['상태'].isin(['근무', '보충'])) # 이 단계 처리 전까지는 '근무'만 있을 수 있음
                    ]['근무자'].tolist()
                    count_on_date = len(current_workers_on_date)

                    if count_on_date > target_count:
                        excess_dates_inner.append((date_str, count_on_date - target_count))
                    elif count_on_date < target_count:
                        shortage_dates_inner.append((date_str, target_count - count_on_date))

                    st.write(f"DEBUG:     날짜: {date_str}, 오후 - 반복 {iteration}. 초과: {excess_dates_inner}, 부족: {shortage_dates_inner}. 현재 인원: {count_on_date}/{target_count}") # DEBUG

                    any_matched_inner = False # 내부 루프용

                    # 초과 인원 처리 (다른 날짜로 이동) - 현재 날짜에만 집중
                    if excess_dates_inner:
                        excess_date_curr, excess_count_curr = excess_dates_inner[0]
                        if excess_count_curr > 0:
                            movable_excess_workers_on_date = [
                                w for w in current_workers_on_date
                                if w not in must_work and w not in moved_workers # 이미 이동된 사람은 제외
                                and not is_worker_already_excluded_with_memo(df_final, excess_date_curr, time_slot, w) # 수정된 조건: ValueError 방지
                            ]
                            movable_excess_workers_on_date.sort(key=lambda w: current_cumulative[time_slot].get(w, 0), reverse=True)

                            for _ in range(excess_count_curr):
                                if not movable_excess_workers_on_date:
                                    st.write(f"DEBUG:         날짜: {excess_date_curr}, 오후 - 이동 가능한 초과 근무자 없음.") # DEBUG
                                    break

                                worker_to_remove = movable_excess_workers_on_date.pop(0) # 가장 많이 근무한 사람부터 제거
                                moved_workers.add(worker_to_remove) # 이동된 사람으로 기록 (파란색 제외 대상)
                                
                                # --- 수정: '제외' (파란색) 부여 로직 ---
                                df_final = update_worker_status(df_final, excess_date_curr, time_slot, worker_to_remove, '제외', f'다른 날짜 부족분 보충 위해 제외됨', '🔵 파란색')
                                current_cumulative[time_slot][worker_to_remove] = current_cumulative[time_slot].get(worker_to_remove, 0) - 1
                                any_matched_inner = True
                                st.write(f"DEBUG:         날짜: {excess_date_curr}, 오후 - '{worker_to_remove}' 제거 (🔵파란색). 현재 인원: {len(current_workers_on_date) - (_+1)}") # DEBUG
                            
                            current_workers_on_date = df_final[
                                (df_final['날짜'] == date_str) &
                                (df_final['시간대'] == time_slot) &
                                (df_final['상태'].isin(['근무', '보충', '제외'])) # '근무', '보충', '제외' 상태 포함 (방금 처리된 파란색 제외)
                            ]['근무자'].tolist()
                            if len(current_workers_on_date) <= target_count:
                                break


                    # 부족 인원 처리 (보충) - 현재 날짜에만 집중
                    if shortage_dates_inner:
                        shortage_date_curr, shortage_count_curr = shortage_dates_inner[0]
                        if shortage_count_curr > 0:
                            # 현재 날짜의 보충 가능한 인원
                            shift_key = f'{day_name} {time_slot}'
                            supplement_row = df_supplement_processed[df_supplement_processed['시간대'] == shift_key]
                            all_possible_supplement = []
                            if not supplement_row.empty:
                                for col in [f'보충{i}' for i in range(1, 13)]:
                                    worker = supplement_row[col].values[0] if col in supplement_row.columns and pd.notna(supplement_row[col].values[0]) else ''
                                    if worker:
                                        name_only = worker.replace('🔺', '').strip()
                                        if name_only not in current_workers_on_date and \
                                           name_only not in vacationers and \
                                           name_only not in no_supplement:
                                            all_possible_supplement.append(name_only)
                            
                            # 오후 보충 제약 추가: 오전 근무자 또는 꼭 근무(오후)자 중 보충 가능
                            morning_workers_for_current_date = df_final[
                                (df_final['날짜'] == date_str) &
                                (df_final['시간대'] == '오전') &
                                (df_final['상태'].isin(['근무', '보충', '추가보충'])) # 이 단계까지의 오전 근무자 상태 (추가보충 포함)
                            ]['근무자'].tolist()
                            
                            eligible_for_afternoon_supplement = [
                                w for w in all_possible_supplement
                                if (w in morning_workers_for_current_date or w in must_work) # 오전 근무자이거나 꼭 근무(오후) 요청자
                                and w not in excluded_morning_workers[date_str] # 오전 제외된 사람은 제외
                            ]
                            
                            all_possible_supplement_with_cumulative = [
                                (w, current_cumulative[time_slot].get(w, 0)) for w in eligible_for_afternoon_supplement
                            ]
                            all_possible_supplement_with_cumulative.sort(key=lambda x: x[1])

                            for _ in range(shortage_count_curr):
                                if not all_possible_supplement_with_cumulative:
                                    st.write(f"DEBUG:         날짜: {shortage_date_curr}, 오후 - 보충 가능한 인원 없음.") # DEBUG
                                    break

                                worker_to_add, _ = all_possible_supplement_with_cumulative.pop(0)
                                
                                # --- 수정: '보충' (연두색) 부여 로직 ---
                                df_final = update_worker_status(df_final, shortage_date_curr, time_slot, worker_to_add, '보충', f'다른 날짜 초과분에서 보충됨', '🟢 초록색')
                                current_cumulative[time_slot][worker_to_add] = current_cumulative[time_slot].get(worker_to_add, 0) + 1
                                any_matched_inner = True
                                st.write(f"DEBUG:         날짜: {shortage_date_curr}, 오후 - '{worker_to_add}' 보충 (🟢연두색). 현재 인원: {len(current_workers_on_date) + (_+1)}") # DEBUG
                            
                            current_workers_on_date = df_final[
                                (df_final['날짜'] == date_str) &
                                (df_final['시간대'] == time_slot) &
                                (df_final['상태'].isin(['근무', '보충', '제외'])) # '근무', '보충', '제외' 상태 포함 (방금 처리된 초록색 보충)
                            ]['근무자'].tolist()
                            if len(current_workers_on_date) >= target_count:
                                break


                    final_count_check = len(df_final[
                        (df_final['날짜'] == date_str) &
                        (df_final['시간대'] == time_slot) &
                        (df_final['상태'].isin(['근무', '보충', '제외'])) # '근무', '보충', '제외' 상태만 포함
                    ]['근무자'].tolist())
                    if final_count_check == target_count or not any_matched_inner:
                        st.write(f"DEBUG:       날짜: {date_str}, 오후 - 내부 루프 종료. 최종 인원: {final_count_check}") # DEBUG
                        break
            st.write("DEBUG: 4단계: 오후 보충/제외 수행 완료.") # DEBUG

            # 5단계: 모든 날짜에 대해 추가 보충/제외 수행 (휴관일 제외)
            st.write("DEBUG: 5단계: 추가 보충/제외 수행 시작.") # DEBUG
            for date in active_weekdays:
                date_str = date.strftime('%Y-%m-%d')
                day_name = day_map[date.weekday()]
                week_num = week_numbers[date]
                supplemented_morning_workers = df_final[
                    (df_final['날짜'] == date_str) &
                    (df_final['시간대'] == '오전') &
                    (df_final['상태'].isin(['근무', '보충', '추가보충', '제외', '추가제외', '휴가'])) # 5단계까지 진행된 오전 근무했던 사람 (모든 상태 포함)
                ]['근무자'].tolist()

                for time_slot in ['오전', '오후']:
                    target_count = 12 if time_slot == '오전' else 5
                    st.write(f"DEBUG:   - 처리 중인 날짜: {date_str}, 시간대: {time_slot}") # DEBUG

                    # 기본 보충/제외 전 근무자 출력 (현재까지 배정된 근무자만)
                    current_workers = df_final[
                        (df_final['날짜'] == date_str) &
                        (df_final['시간대'] == time_slot) &
                        (df_final['상태'].isin(['근무', '보충', '추가보충'])) # 5단계까지 완료된 상태의 근무자
                    ]['근무자'].tolist()
                    st.write(f"DEBUG:     시작 인원: {len(current_workers)}/{target_count}") # DEBUG


                    # 요청사항 재확인
                    vacationers = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == '휴가']
                    must_work = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == f'꼭 근무({time_slot})']
                    no_supplement = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == f'보충 불가({time_slot})']
                    hard_supplement = [row['이름'] for _, row in df_request.iterrows() if date_str in parse_date_range(row['날짜정보']) and row['분류'] == f'보충 어려움({time_slot})']

                    # df_shift_processed 초기 근무자
                    shift_key = f'{day_name} {time_slot}'
                    shift_row = df_shift_processed[df_shift_processed['시간대'] == shift_key]
                    initial_workers = set()
                    if not shift_row.empty:
                        for col in [f'근무{i}' for i in range(1, 15)]:
                            worker = shift_row[col].values[0] if col in shift_row.columns and pd.notna(shift_row[col].values[0]) else ''
                            if worker:
                                if '(' in worker:
                                    name, weeks = worker.split('(')
                                    name = name.strip()
                                    weeks = weeks.rstrip(')').split(',')
                                    if f'{week_num}주' in weeks:
                                        initial_workers.add(name)
                                else:
                                    initial_workers.add(worker)

                    # df_supplement_processed 보충 근무자
                    supplement_row = df_supplement_processed[df_supplement_processed['시간대'] == shift_key]
                    supplement_workers = []
                    if not supplement_row.empty:
                        for col in [f'보충{i}' for i in range(1, 13)]:
                            worker = supplement_row[col].values[0] if col in supplement_row.columns and pd.notna(supplement_row[col].values[0]) else ''
                            if worker:
                                name = worker.replace('🔺', '')
                                priority = 'low' if '🔺' in worker else 'normal'
                                if name not in vacationers and name not in no_supplement:
                                    supplement_workers.append((name, priority))
                    if time_slot == '오후':
                        for worker in supplemented_morning_workers:
                            if worker not in [w for w, _ in supplement_workers] and worker not in vacationers and worker not in no_supplement:
                                supplement_workers.append((worker, 'normal'))
                    supplement_workers = [(w, p) for w, p in supplement_workers if w not in vacationers and w not in no_supplement]

                    # 오후 보충 제약
                    morning_workers = df_final[
                        (df_final['날짜'] == date_str) &
                        (df_final['시간대'] == '오전') &
                        (df_final['상태'].isin(['근무', '보충', '추가보충', '제외', '추가제외', '휴가'])) # 모든 오전 상태 고려
                    ]['근무자'].tolist() if time_slot == '오후' else None

                    # 추가 보충
                    if len(current_workers) < target_count:
                        st.write(f"DEBUG:       - 추가 보충 필요. 현재 {len(current_workers)}명, 목표 {target_count}명.") # DEBUG
                        supplement_workers_with_cumulative = [
                            # --- 수정된 라인: 괄호 짝 맞춤 및 필터링 명확화 ---
                            (w, df_cumulative_next[df_cumulative_next['이름'].str.strip() == w][f'{time_slot}누적'].iloc[0] if (w in df_cumulative_next['이름'].str.strip().values and not df_cumulative_next[df_cumulative_next['이름'].str.strip() == w].empty) else 0, p)
                            # --- 수정 끝 ---
                            for w, p in supplement_workers if w not in current_workers
                            # 이전에 '보충' (초록색) 상태를 부여받은 사람은 여기서는 '추가보충' 대상이 아님.
                            # 즉, df_final에서 '🟢 초록색'이 아닌 사람만 추가 보충 대상.
                            and df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == time_slot) & (df_final['근무자'] == w) & (df_final['색상'] == '🟢 초록색')].empty # `.any()` 대신 `.empty` 사용
                        ]
                        supplement_workers_with_cumulative.sort(key=lambda x: (x[1], x[2] == 'low'))
                        st.write(f"DEBUG:         보충 후보 (5단계): {len(supplement_workers_with_cumulative)}명.") # DEBUG

                        while len(current_workers) < target_count and supplement_workers_with_cumulative:
                            worker, _, _ = supplement_workers_with_cumulative.pop(0)
                            st.write(f"DEBUG:           '{worker}' 추가보충 시도 중...") # DEBUG
                            if time_slot == '오후' and worker not in must_work:
                                if morning_workers is not None and (worker not in morning_workers or worker in excluded_morning_workers[date_str]):
                                    st.write(f"DEBUG:           '{worker}'는 오후 추가보충 불가 (오전 근무 아님 또는 제외됨).") # DEBUG
                                    continue
                            
                            current_workers.append(worker)
                            current_cumulative[time_slot][worker] = current_cumulative[time_slot].get(worker, 0) + 1
                            
                            # df_cumulative_next 업데이트 로직
                            if (worker in df_cumulative_next['이름'].str.strip().values and not df_cumulative_next[df_cumulative_next['이름'].str.strip() == worker].empty): # 존재 및 필터링 결과 비어있지 않은지 확인
                                df_cumulative_next.loc[df_cumulative_next['이름'].str.strip() == worker, f'{time_slot}누적'] = \
                                    df_cumulative_next.loc[df_cumulative_next['이름'].str.strip() == worker, f'{time_slot}누적'].astype(int) + 1
                            else:
                                new_row_data = {
                                    '이름': worker,
                                    "오전누적": 0, "오후누적": 0, "오전당직 (온콜)": 0, "오후당직": 0
                                }
                                new_row_data[f'{time_slot}누적'] = 1
                                new_row = pd.DataFrame([new_row_data])
                                df_cumulative_next = pd.concat([df_cumulative_next, new_row], ignore_index=True)
                                st.write(f"DEBUG:             새로운 근무자 '{worker}' 누적 테이블에 추가 (5단계).") # DEBUG

                            df_final = update_worker_status(df_final, date_str, time_slot, worker, '추가보충', '인원 부족으로 인한 추가 보충', '🟡 노란색')
                            st.write(f"DEBUG:           '{worker}' 추가보충 완료 (🟡노란색). 현재 인원: {len(current_workers)}/{target_count}") # DEBUG


                    # 추가 제외
                    if len(current_workers) > target_count:
                        st.write(f"DEBUG:       - 추가 제외 필요. 현재 {len(current_workers)}명, 목표 {target_count}명.") # DEBUG
                        removable_workers = [
                            # --- 수정된 라인: 괄호 짝 맞춤 및 필터링 명확화 ---
                            (w, df_cumulative_next[df_cumulative_next['이름'].str.strip() == w][f'{time_slot}누적'].iloc[0] if (w in df_cumulative_next['이름'].str.strip().values and not df_cumulative_next[df_cumulative_next['이름'].str.strip() == w].empty) else 0)
                            # --- 수정 끝 ---
                            for w in current_workers 
                            if w not in must_work and w not in initial_workers
                            # 이전에 '제외' (파란색) 상태를 부여받은 사람은 여기서는 '추가제외' 대상이 아님.
                            and df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == time_slot) & (df_final['근무자'] == w) & (df_final['색상'] == '🔵 파란색')].empty # `.any()` 대신 `.empty` 사용
                        ]
                        if not removable_workers: # must_work 또는 initial_workers만 남았을 경우
                            removable_workers = [
                                # --- 수정된 라인: 괄호 짝 맞춤 및 필터링 명확화 ---
                                (w, df_cumulative_next[df_cumulative_next['이름'].str.strip() == w][f'{time_slot}누적'].iloc[0] if (w in df_cumulative_next['이름'].str.strip().values and not df_cumulative_next[df_cumulative_next['이름'].str.strip() == w].empty) else 0)
                                # --- 수정 끝 ---
                                for w in current_workers if w not in must_work
                                and df_final[(df_final['날짜'] == date_str) & (df_final['시간대'] == time_slot) & (df_final['근무자'] == w) & (df_final['색상'] == '🔵 파란색')].empty # `.any()` 대신 `.empty` 사용
                            ]
                        removable_workers.sort(key=lambda x: x[1], reverse=True) # 누적 근무 횟수 높은 순으로 정렬하여 제거
                        st.write(f"DEBUG:         제외 후보 (5단계): {len(removable_workers)}명.") # DEBUG

                        while len(current_workers) > target_count and removable_workers:
                            worker, _ = removable_workers.pop(0)
                            st.write(f"DEBUG:           '{worker}' 추가제외 시도 중...") # DEBUG
                            current_workers.remove(worker)
                            current_cumulative[time_slot][worker] = current_cumulative[time_slot].get(worker, 0) - 1
                            
                            # df_cumulative_next 업데이트 로직
                            if (worker in df_cumulative_next['이름'].str.strip().values and not df_cumulative_next[df_cumulative_next['이름'].str.strip() == worker].empty): # 존재 및 필터링 결과 비어있지 않은지 확인
                                df_cumulative_next.loc[df_cumulative_next['이름'].str.strip() == worker, f'{time_slot}누적'] = \
                                    df_cumulative_next.loc[df_cumulative_next['이름'].str.strip() == worker, f'{time_slot}누적'].astype(int) - 1
                            
                            df_final = update_worker_status(df_final, date_str, time_slot, worker, '추가제외', '인원 초과로 인한 추가 제외', '🟣 보라색')
                            
                            if time_slot == '오전': # 오전에서 추가 제외된 경우 오후에도 영향 줄 수 있음
                                # 해당 근무자가 오후에 이미 '근무' 또는 '보충' 상태로 있었다면 오후에서도 제외 처리
                                existing_afternoon_assignment = df_final[
                                    (df_final['날짜'] == date_str) &
                                    (df_final['시간대'] == '오후') &
                                    (df_final['근무자'] == worker) &
                                    (df_final['상태'].isin(['근무', '보충', '추가보충', '제외', '추가제외', '휴가'])) # 5단계까지 진행된 오후 근무자 상태 (모든 상태 고려)
                                ]
                                if not existing_afternoon_assignment.empty: # 이미 오후에 배정되어 있었다면
                                    df_final = update_worker_status(df_final, date_str, '오후', worker, '추가제외', '오전 추가제외로 인한 오후 제외', '🟣 보라색')
                                    current_cumulative['오후'][worker] = current_cumulative['오후'].get(worker, 0) - 1
                                    # df_cumulative_next 업데이트 로직
                                    if (worker in df_cumulative_next['이름'].str.strip().values and not df_cumulative_next[df_cumulative_next['이름'].str.strip() == worker].empty): # 존재 및 필터링 결과 비어있지 않은지 확인
                                        df_cumulative_next.loc[df_cumulative_next['이름'].str.strip() == worker, '오후누적'] = \
                                            df_cumulative_next.loc[df_cumulative_next['이름'].str.strip() == worker, '오후누적'].astype(int) - 1
                                    st.write(f"DEBUG:             오전 추가제외로 인해 '{worker}' 오후에서도 추가제외됨 (🟣보라색).") # DEBUG

                            st.write(f"DEBUG:           '{worker}' 제외 완료 (🟣보라색). 현재 인원: {len(current_workers)}/{target_count}") # DEBUG


                    # 최종 검증 (이 단계에서 인원수가 목표에 맞는지 확인)
                    final_count = len(df_final[
                        (df_final['날짜'] == date_str) &
                        (df_final['시간대'] == time_slot) &
                        (df_final['상태'].isin(['근무', '보충', '추가보충'])) # 5단계까지 완료된 최종 인원
                    ]['근무자'].tolist())
                    st.write(f"DEBUG:     최종 검증: {date_str}, {time_slot} - 최종 인원: {final_count}/{target_count}") # DEBUG
            st.write("DEBUG: 5단계: 추가 보충/제외 수행 완료.") # DEBUG


            # 2025년 4월 전체 평일 및 주말 생성
            _, last_day = calendar.monthrange(next_month.year, next_month.month)
            dates = pd.date_range(start=next_month, end=next_month.replace(day=last_day))
            week_numbers = {d: (d.day - 1) // 7 + 1 for d in dates}
            day_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}

            # df_schedule 생성 (평일 및 주말 포함)
            df_schedule = pd.DataFrame({
                '날짜': [d.strftime('%Y-%m-%d') for d in dates],
                '요일': [day_map[d.weekday()] for d in dates]
            })
            st.write("DEBUG: df_schedule 생성 완료.") # DEBUG


            # 최대 근무자 수 계산 (모든 상태 포함)
            # 여기서는 '제외', '추가제외' 상태의 인원도 포함하여 최대 열 수를 계산합니다.
            worker_counts_all = df_final.groupby(['날짜', '시간대'])['근무자'].nunique().unstack(fill_value=0)
            max_morning_workers_all = int(worker_counts_all.get('오전', pd.Series(0)).max()) if '오전' in worker_counts_all else 0
            max_afternoon_workers_all = int(worker_counts_all.get('오후', pd.Series(0)).max()) if '오후' in worker_counts_all else 0

            # 최대 근무자 수 설정 (제한 제거)
            max_morning_workers = max_morning_workers_all
            max_afternoon_workers = max_afternoon_workers_all
            st.write(f"DEBUG: 최대 오전 근무자: {max_morning_workers}, 최대 오후 근무자: {max_afternoon_workers}.") # DEBUG

            # 색상 우선순위 정의 (순서는 출력 우선순위에 따라 중요)
            color_priority = {
                '🟠 주황색': 0, # 꼭 근무 (최우선)
                '🟢 초록색': 1, # 보충 (다른 날짜 이동)
                '🟡 노란색': 2, # 추가보충 (순수 부족으로 추가)
                '기본': 3,     # 기본 근무
                '🔴 빨간색': 4, # 휴가자 제외
                '🔵 파란색': 5, # 제외 (다른 날짜 보충 위해 이동)
                '🟣 보라색': 6, # 추가제외 (순수 초과로 제외)
            }

            # df_final에 색상 우선순위 열 추가
            df_final['색상_우선순위'] = df_final['색상'].map(color_priority)

            # df_final 중복 제거 (색상 우선순위가 높은 상태 선택)
            # 엑셀에 최종적으로 보여줄 근무자만 필터링합니다. '제외'된 사람들은 제외.
            # 하지만 '제외'된 사람들도 엑셀에 이름은 표시되지만 셀 색상만 변경하는 것이라면,
            # df_final_unique에 모든 상태의 사람을 포함하고 정렬 기준만 유지해야 합니다.
            # 현재 엑셀 생성 로직(df_excel 데이터 채우기)을 볼 때, df_final_unique는
            # 모든 상태의 근무자를 포함하고 색상/메모를 통해 구분하는 것이 맞습니다.
            # 따라서 '제외'된 사람도 포함되어야 Excel에 나타납니다.
            df_final_unique = df_final.sort_values(by=['날짜', '시간대', '근무자', '색상_우선순위']).groupby(['날짜', '시간대', '근무자']).first().reset_index()
            st.write("DEBUG: df_final_unique 생성 완료.") # DEBUG

            # df_excel 열 동적 생성
            columns = ['날짜', '요일'] + [str(i) for i in range(1, max_morning_workers + 1)] + [''] + ['오전당직(온콜)'] + [f'오후{i}' for i in range(1, max_afternoon_workers + 1)]
            df_excel = pd.DataFrame(index=df_schedule.index, columns=columns)
            st.write("DEBUG: df_excel 초기화 완료.") # DEBUG

            # 데이터 채우기
            for idx, row in df_schedule.iterrows():
                date = row['날짜']
                date_obj = datetime.datetime.strptime(date, '%Y-%m-%d')
                df_excel.at[idx, '날짜'] = f"{date_obj.month}월 {date_obj.day}일"
                df_excel.at[idx, '요일'] = row['요일']

                # 오전 근무자 (df_final_unique에서 가져옴, 모든 상태 포함)
                # 이 부분의 logic은 df_final_unique에 이미 모든 상태의 근무자가 있으므로,
                # 단순히 해당 날짜/시간대의 근무자들을 가져와서 정렬하고 Excel에 채우는 것으로 충분합니다.
                morning_workers_for_excel = df_final_unique[(df_final_unique['날짜'] == date) & (df_final_unique['시간대'] == '오전')]
                # 상태 우선순위와 이름으로 정렬
                morning_workers_for_excel_sorted = morning_workers_for_excel.sort_values(by=['색상_우선순위', '근무자'])['근무자'].tolist()

                for i, worker_name in enumerate(morning_workers_for_excel_sorted, 1):
                    if i <= max_morning_workers:
                        df_excel.at[idx, str(i)] = worker_name

                # 오후 근무자 (df_final_unique에서 가져옴, 모든 상태 포함)
                afternoon_workers_for_excel = df_final_unique[(df_final_unique['날짜'] == date) & (df_final_unique['시간대'] == '오후')]
                # 상태 우선순위와 이름으로 정렬
                afternoon_workers_for_excel_sorted = afternoon_workers_for_excel.sort_values(by=['색상_우선순위', '근무자'])['근무자'].tolist()

                for i, worker_name in enumerate(afternoon_workers_for_excel_sorted, 1):
                    if i <= max_afternoon_workers:
                        df_excel.at[idx, f'오후{i}'] = worker_name

                # 토요일 근무 인원 반영 (1~10열에 딱 10명씩 배치) - df_excel에 직접 덮어쓰기
                if row['요일'] == '토':
                    for saturday_date, workers in saturday_schedules:
                        if date == saturday_date:
                            workers_padded = workers[:10] + [''] * (10 - len(workers[:10]))
                            for i in range(1, 11): # 11 대신 13으로 수정? -> 10명 제한이므로 11까지가 맞음.
                                df_excel.at[idx, str(i)] = workers_padded[i-1]
            st.write("DEBUG: Excel 출력용 DataFrame 데이터 채우기 완료.") # DEBUG

            # 오전당직(온콜) 배정
            st.write("DEBUG: 오전당직(온콜) 배정 시작.") # DEBUG
            # --- 수정: df_cumulative.set_index('이름') 사용 ---
            oncall_counts = df_cumulative.set_index('이름')['오전당직 (온콜)'].to_dict() 
            # --- 수정 끝 ---
            oncall_assignments = {worker: int(count) if count else 0 for worker, count in oncall_counts.items()}
            oncall = {}

            # 오후 근무자 중 최종적으로 배정된 (근무, 보충, 추가보충) 인원만 고려
            afternoon_counts = df_final_unique[
                (df_final_unique['시간대'] == '오후') &
                (df_final_unique['상태'].isin(['근무', '보충', '추가보충'])) 
            ]['근무자'].value_counts().to_dict()

            workers_priority = sorted(
                oncall_assignments.items(),
                key=lambda x: (-x[1], afternoon_counts.get(x[0], 0))
            )

            all_dates = df_final_unique['날짜'].unique().tolist()
            remaining_dates = set(all_dates)

            for worker, count in workers_priority:
                if count <= 0:
                    continue

                eligible_dates = df_final_unique[
                    (df_final_unique['시간대'] == '오후') &
                    (df_final_unique['근무자'] == worker) &
                    (df_final_unique['상태'].isin(['근무', '보충', '추가보충'])) # 최종 근무자만
                ]['날짜'].unique()

                eligible_dates = [d for d in eligible_dates if d in remaining_dates]
                if not eligible_dates:
                    continue

                selected_dates = random.sample(eligible_dates, min(count, len(eligible_dates)))
                for selected_date in selected_dates:
                    oncall[selected_date] = worker
                    remaining_dates.remove(selected_date)

            random_assignments = []
            if remaining_dates:
                for date in remaining_dates:
                    afternoon_workers_df = df_final_unique[
                        (df_final_unique['날짜'] == date) &
                        (df_final_unique['시간대'] == '오후') &
                        (df_final_unique['상태'].isin(['근무', '보충', '추가보충'])) # 최종 근무자만
                    ]
                    afternoon_workers = afternoon_workers_df['근무자'].tolist()

                    if afternoon_workers:
                        selected_worker = random.choice(afternoon_workers)
                        oncall[date] = selected_worker
                        random_assignments.append((date, selected_worker))
                    else:
                        date_obj = datetime.datetime.strptime(date, '%Y-%m-%d')
                        formatted_date = date_obj.strftime('%m월 %d일').lstrip('0')
                        st.warning(f"⚠️ {formatted_date}에는 오후 근무자가 없어 오전당직(온콜)을 배정할 수 없습니다.")
                        # st.write(f"{formatted_date}에 대한 df_final_unique 데이터:") # DEBUG (너무 많을 수 있음)
                        # st.dataframe(afternoon_workers_df) # DEBUG (너무 많을 수 있음)

            # df_excel에 오전당직(온콜) 배정 반영
            for idx, row in df_schedule.iterrows():
                date = row['날짜']
                date_obj = datetime.datetime.strptime(date, '%Y-%m-%d')
                formatted_date = date_obj.strftime('%Y-%m-%d')
                df_excel.at[idx, '오전당직(온콜)'] = oncall.get(formatted_date, '')

            actual_oncall_counts = {}
            for date, worker in oncall.items():
                actual_oncall_counts[worker] = actual_oncall_counts.get(worker, 0) + 1

            for worker, actual_count in actual_oncall_counts.items():
                max_count = oncall_assignments.get(worker, 0)
                if actual_count > max_count:
                    st.info(f"오전당직(온콜) 횟수 제한 한계로, {worker} 님이 최대 배치 {max_count}회가 아닌 {actual_count}회 배치되었습니다.")
            st.write("DEBUG: 오전당직(온콜) 배정 완료.") # DEBUG
            
            # Excel 파일 생성
            st.write("DEBUG: Excel 파일 생성 시작.") # DEBUG
            wb = Workbook()
            ws = wb.active
            ws.title = "스케줄"

            # 열 헤더 추가
            for col_idx, col_name in enumerate(df_excel.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
                cell.font = Font(size=9, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 검정색 테두리 스타일
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # 색상 매핑
            color_map = {
                '🔴 빨간색': 'C00000', # 휴가자 제외
                '🟠 주황색': 'FFD966', # 꼭 근무
                '🟢 초록색': '92D050', # 보충 (다른 날짜에서 이동)
                '🟡 노란색': 'FFFF00', # 추가보충 (순수 부족으로 추가)
                '🔵 파란색': '0070C0', # 제외 (다른 날짜 보충 위해 이동)
                '🟣 보라색': '7030A0', # 추가제외 (순수 초과로 제외)
                '기본': 'FFFFFF', # 기본 근무
            }

            # 데이터 추가 및 스타일 적용
            for row_idx, (idx, row) in enumerate(df_excel.iterrows(), 2):
                for col_idx, col_name in enumerate(df_excel.columns, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = row[col_name]
                    cell.font = Font(size=9)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # 날짜 열 스타일
                    if col_name == '날짜':
                        cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

                    # 요일 열 스타일
                    elif col_name == '요일':
                        date_str_formatted = row['날짜'] # 예: "4월 1일"
                        try:
                            # 2025년도로 고정하여 datetime 객체 생성
                            date_obj_for_holiday_check = datetime.datetime.strptime(date_str_formatted, '%m월 %d일').replace(year=2025)
                            formatted_date_for_holiday_check = date_obj_for_holiday_check.strftime('%Y-%m-%d')
                        except ValueError:
                            # 파싱 실패 시, date_str이 이미"%Y-%m-%d" 형식일 경우를 대비 (여기서는 아닐 가능성 높음)
                            formatted_date_for_holiday_check = date_str_formatted
                            
                        selected_saturday_dates = [schedule[0] for schedule in saturday_schedules]
                        if formatted_date_for_holiday_check in holiday_dates:
                            cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
                        elif row['요일'] == '토' and formatted_date_for_holiday_check in selected_saturday_dates:
                            cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
                        elif row['요일'] in ['토', '일']:
                            cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

                    # 오전 근무자 색상 및 메모 적용 (토요일은 UI에서 입력된 데이터로 덮어씌워짐)
                    elif col_name in [str(i) for i in range(1, max_morning_workers + 1)]:
                        date_str_for_lookup = datetime.datetime.strptime(row['날짜'], '%m월 %d일').replace(year=2025).strftime('%Y-%m-%d')
                        worker = row[col_name]
                        if worker: # worker가 존재하면
                            worker_data = df_final_unique[(df_final_unique['날짜'] == date_str_for_lookup) & (df_final_unique['시간대'] == '오전') & (df_final_unique['근무자'] == worker)]
                            if not worker_data.empty:
                                status, memo, color = worker_data.iloc[0]['상태'], worker_data.iloc[0]['메모'], worker_data.iloc[0]['색상']
                                fill = PatternFill(start_color=color_map.get(color, 'FFFFFF'), end_color=color_map.get(color, 'FFFFFF'), fill_type='solid')
                                cell.fill = fill
                                if memo:
                                    cell.comment = Comment(memo, 'Huiyeon Kim')
                        # 토요일 근무자는 별도 처리 (색상은 기본 흰색)
                        if row['요일'] == '토' and worker and col_name in [str(i) for i in range(1, 11)]:
                             cell.fill = PatternFill(start_color=color_map['기본'], end_color=color_map['기본'], fill_type='solid') # 토요일은 기본 흰색

                    # 오후 근무자 색상 및 메모 적용
                    elif col_name.startswith('오후'):
                        date_str_for_lookup = datetime.datetime.strptime(row['날짜'], '%m월 %d일').replace(year=2025).strftime('%Y-%m-%d')
                        worker = row[col_name]
                        if worker: # worker가 존재하면
                            worker_data = df_final_unique[(df_final_unique['날짜'] == date_str_for_lookup) & (df_final_unique['시간대'] == '오후') & (df_final_unique['근무자'] == worker)]
                            if not worker_data.empty:
                                status, memo, color = worker_data.iloc[0]['상태'], worker_data.iloc[0]['메모'], worker_data.iloc[0]['색상']
                                fill = PatternFill(start_color=color_map.get(color, 'FFFFFF'), end_color=color_map.get(color, 'FFFFFF'), fill_type='solid')
                                cell.fill = fill
                                if memo:
                                    cell.comment = Comment(memo, 'Huiyeon Kim')

                    # 오전당직(온콜) 색상 적용
                    elif col_name == '오전당직(온콜)':
                        if row[col_name]:
                            cell.font = Font(size=9, bold=True, color='FF69B4')
                        else:
                            cell.font = Font(size=9)
            st.write("DEBUG: Excel 파일 스타일 적용 완료.") # DEBUG

            # 열 너비 설정
            ws.column_dimensions['A'].width = 10
            for col in ws.columns:
                if col[0].column_letter != 'A':
                    ws.column_dimensions[col[0].column_letter].width = 7

            # Excel 파일을 메모리에 저장
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            st.session_state.output = output
            st.write("DEBUG: Excel 파일 메모리 저장 완료.") # DEBUG


            # df_final_unique와 df_excel을 기반으로 스케줄 데이터 변환
            def transform_schedule_data(df, df_excel, month_start, month_end):
                # '근무', '보충', '추가보충' 상태만 필터링
                df = df[df['상태'].isin(['근무', '보충', '추가보충'])][['날짜', '시간대', '근무자', '요일']].copy()
                
                # 전체 날짜 범위 생성
                date_range = pd.date_range(start=month_start, end=month_end)
                # 날짜를 "4월 1일" 형태로 포맷팅
                date_list = [f"{d.month}월 {d.day}일" for d in date_range]
                weekday_list = [d.strftime('%a') for d in date_range]
                weekday_map = {'Mon': '월', 'Tue': '화', 'Wed': '수', 'Thu': '목', 'Fri': '금', 'Sat': '토', 'Sun': '일'}
                weekdays = [weekday_map[w] for w in weekday_list]
                
                # 결과 DataFrame 초기화
                columns = ['날짜', '요일'] + [str(i) for i in range(1, 13)] + ['오전당직(온콜)'] + [f'오후{i}' for i in range(1, 6)]
                result_df = pd.DataFrame(columns=columns)
                
                # 각 날짜별로 처리
                for date, weekday in zip(date_list, weekdays):
                    date_key = datetime.datetime.strptime(date, '%m월 %d일').replace(year=2025).strftime('%Y-%m-%d')
                    date_df = df[df['날짜'] == date_key]
                    
                    # 평일 데이터 (df_final_unique에서 가져옴)
                    morning_workers = date_df[date_df['시간대'] == '오전']['근무자'].tolist()[:12]
                    morning_data = morning_workers + [''] * (12 - len(morning_workers))
                    afternoon_workers = date_df[date_df['시간대'] == '오후']['근무자'].tolist()[:5]
                    afternoon_data = afternoon_workers + [''] * (5 - len(afternoon_workers))
                    
                    # 토요일 데이터 (df_excel에서 가져옴)
                    if weekday == '토':
                        excel_row_for_saturday_check = df_excel[df_excel['날짜'] == date] # 변수명 수정
                        if not excel_row_for_saturday_check.empty:
                            # 엑셀의 1~10열 데이터만 가져옴
                            morning_data = [excel_row_for_saturday_check[str(i)].iloc[0] if str(i) in excel_row_for_saturday_check.columns and pd.notna(excel_row_for_saturday_check[str(i)].iloc[0]) else '' for i in range(1, 11)] 
                            # 10개만 가져오도록 수정합니다.
                            morning_data = morning_data[:10] + [''] * (12 - len(morning_data[:10])) # 12명에 맞춤
                    
                    # df_excel에서 해당 날짜의 온콜 데이터 가져오기
                    oncall_worker = ''
                    excel_row_for_oncall_check = df_excel[df_excel['날짜'] == date] # 변수명 수정
                    if not excel_row_for_oncall_check.empty:
                        oncall_worker = excel_row_for_oncall_check['오전당직(온콜)'].iloc[0] if '오전당직(온콜)' in excel_row_for_oncall_check.columns else ''
                        
                    row_data = [date, weekday] + morning_data + [oncall_worker] + afternoon_data
                    result_df = pd.concat([result_df, pd.DataFrame([row_data], columns=columns)], ignore_index=True)
                
                return result_df

            # Google Sheets 저장 및 다운로드 로직 수정
            # 이 블록은 Streamlit 앱의 최상위 스크립트 레벨에서 실행됩니다.
            # 따라서 'return' 문을 사용하면 SyntaxError가 발생합니다.
            # 대신 'st.stop()'를 사용하여 앱의 현재 실행을 중단합니다.

            if st.session_state.get("is_admin_authenticated", False):
                # 날짜 설정
                month_dt = datetime.datetime.strptime(month_str, "%Y년 %m월")
                next_month_dt = (month_dt + timedelta(days=32)).replace(day=1)
                next_month_str = next_month_dt.strftime("%Y년 %m월")
                next_month_start = month_dt.replace(day=1)
                _, last_day = calendar.monthrange(month_dt.year, month_dt.month)
                next_month_end = month_dt.replace(day=last_day)

                # 구글 시트 열기
                try:
                    url = st.secrets["google_sheet"]["url"]
                    gc = get_gspread_client()
                    if gc is None: # get_gspread_client에서 이미 stop()을 하지만, 방어 코드
                        st.stop()
                    sheet = gc.open_by_url(url)
                    st.write(f"DEBUG: Google Sheet '{url}' 저장용으로 다시 열기 성공.") # DEBUG
                except APIError as e: # gspread.exceptions.APIError 명시적으로 잡기
                    st.error(f"❌ Google Sheets 연결 중 API 오류 발생 (저장 단계): {e.response.status_code} - {e.response.text}")
                    st.exception(e) # 상세 스택 트레이스 출력
                    st.stop()
                except Exception as e:
                    st.error(f"❌ Google Sheets 연결 중 예기치 않은 오류 발생 (저장 단계): {type(e).__name__} - {e}")
                    st.exception(e) # 상세 스택 트레이스 출력
                    st.stop()

                # month_str 스케쥴 조정사항 시트 초기화
                try:
                    # 시트 존재 여부 확인
                    try:
                        worksheet_adjustments = sheet.worksheet(f"{month_str} 스케쥴 조정사항")
                        st.write(f"DEBUG: '{month_str} 스케쥴 조정사항' 시트 존재 확인. 초기화 중...") # DEBUG
                        # 시트 데이터 초기화 (기존 데이터 삭제)
                        worksheet_adjustments.clear()
                        # 초기 헤더 추가 (필요 시)
                        worksheet_adjustments.update('A1', [['Timestamp', '조정사항']], value_input_option='RAW')
                    except WorksheetNotFound:
                        # 시트가 없으면 새로 생성
                        st.warning(f"⚠️ '{month_str} 스케쥴 조정사항' 시트를 찾을 수 없습니다. 새로 생성합니다.") # DEBUG
                        worksheet_adjustments = sheet.add_worksheet(title=f"{month_str} 스케쥴 조정사항", rows=100, cols=10)
                        # 초기 헤더 추가
                        worksheet_adjustments.update('A1', [['Timestamp', '조정사항']], value_input_option='RAW')
                        st.write(f"DEBUG: '{month_str} 스케쥴 조정사항' 시트 새로 생성 완료.") # DEBUG
                    
                    st.success(f"✅ {month_str} 스케쥴 조정사항 시트가 초기화되었습니다.")
                except APIError as e: # APIError 명시적으로 잡기
                    st.error(f"❌ {month_str} 스케쥴 조정사항 시트 초기화 중 API 오류 발생: {e.response.status_code} - {e.response.text}")
                    st.exception(e) # 상세 스택 트레이스 출력
                    st.stop()
                except Exception as e:
                    st.error(f"❌ {month_str} 스케쥴 조정사항 시트 초기화 중 예기치 않은 오류 발생: {type(e).__name__} - {e}")
                    st.exception(e) # 상세 스택 트레이스 출력
                    st.stop()

                # df_final_unique와 df_excel을 기반으로 스케줄 데이터 변환
                df_schedule = transform_schedule_data(df_final_unique, df_excel, next_month_start, next_month_end)
                st.write("DEBUG: 최종 df_schedule 변환 완료.") # DEBUG

                # Google Sheets에 스케쥴 저장
                try:
                    # 시트 존재 여부 확인 및 생성/재사용
                    try:
                        worksheet_schedule = sheet.worksheet(f"{month_str} 스케쥴")
                        st.write(f"DEBUG: '{month_str} 스케쥴' 시트 업데이트 준비.") # DEBUG
                    except WorksheetNotFound:
                        st.warning(f"⚠️ '{month_str} 스케쥴' 시트를 찾을 수 없습니다. 새로 생성합니다.")
                        worksheet_schedule = sheet.add_worksheet(title=f"{month_str} 스케쥴", rows=1000, cols=50)
                        st.write(f"DEBUG: '{month_str} 스케쥴' 시트 새로 생성 완료.") # DEBUG

                    # 기존 데이터 삭제 및 업데이트
                    worksheet_schedule.clear()
                    data_schedule = [df_schedule.columns.tolist()] + df_schedule.astype(str).values.tolist()
                    worksheet_schedule.update('A1', data_schedule, value_input_option='RAW')
                    st.write(f"DEBUG: '{month_str} 스케쥴' 시트 Google Sheets 저장 완료.") # DEBUG
                except Exception as e: # APIError를 포함한 모든 예외를 잡도록 변경 (APIError만 잡기에는 너무 한정적)
                    st.error(f"⚠️ {month_str} 스케쥴 테이블 저장 중 오류 발생: {str(e)}")
                    st.exception(e) # 상세 스택 트레이스 출력
                    st.stop()

                # df_cumulative_next 처리
                df_cumulative_next.rename(columns={'이름': next_month_str}, inplace=True) # 누적 테이블의 첫 컬럼명을 '이름'에서 '다음달 년월'로 변경
                st.write("DEBUG: df_cumulative_next 컬럼 이름 변경 완료.") # DEBUG

                # 다음 달 누적 시트 저장
                try:
                    # 시트 존재 여부 확인 및 생성/재사용
                    try:
                        worksheet = sheet.worksheet(f"{next_month_str} 누적")
                        st.write(f"DEBUG: '{next_month_str} 누적' 시트 업데이트 준비.") # DEBUG
                    except WorksheetNotFound:
                        st.warning(f"⚠️ '{next_month_str} 누적' 시트를 찾을 수 없습니다. 새로 생성합니다.")
                        worksheet = sheet.add_worksheet(title=f"{next_month_str} 누적", rows=1000, cols=20) 
                        st.write(f"DEBUG: '{next_month_str} 누적' 시트 새로 생성 완료.") # DEBUG

                    worksheet.clear()
                    data = [df_cumulative_next.columns.tolist()] + df_cumulative_next.values.tolist()
                    worksheet.update('A1', data, value_input_option='USER_ENTERED')
                    st.write(f"DEBUG: '{next_month_str} 누적' 시트 Google Sheets 저장 완료.") # DEBUG
                except Exception as e: # APIError를 포함한 모든 예외를 잡도록 변경
                    st.error(f"⚠️ {next_month_str} 누적 테이블 저장 중 오류 발생: {str(e)}")
                    st.exception(e) # 상세 스택 트레이스 출력
                    st.stop()

                # 세션 상태 설정
                st.session_state.assigned = True
                st.session_state.output = output
                st.session_state.sheet = sheet
                st.session_state.data_schedule = data_schedule
                st.session_state.df_cumulative_next = df_cumulative_next
                st.session_state.next_month_str = next_month_str
                st.write("DEBUG: 세션 상태 업데이트 완료.") # DEBUG

                # 1. 누적 테이블 출력
                st.write(" ")
                st.markdown(f"**➕ {next_month_str} 누적 테이블**")
                st.dataframe(df_cumulative_next)

                # 2. 누적 테이블 저장 완료 메시지
                st.success(f"✅ {next_month_str} 누적 테이블이 Google Sheets에 저장되었습니다.")

                # 3. 구분선
                st.divider()

                # 4. 스케쥴 테이블 저장 완료 메시지
                st.success(f"✅ {month_str} 스케쥴 테이블이 Google Sheets에 저장되었습니다.")

                # 5. 다운로드 버튼
                st.markdown("""
                    <style>
                    .download-button > button {
                        background: linear-gradient(90deg, #e74c3c 0%, #c0392b 100%) !important;
                        color: white !important;
                        font-weight: bold;
                        font-size: 16px;
                        border-radius: 12px;
                        padding: 12px 24px;
                        border: none;
                        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                        transition: all 0.3s ease;
                    }
                    .download-button > button:hover {
                        background: linear-gradient(90deg, #c0392b 0%, #e74c3c 100%) !important;
                        box-shadow: 0 6px 12px rgba(0, 0, 0, 0.15);
                        transform: translateY(-2px);
                    }
                    .download-button > button:active {
                        transform: translateY(0);
                        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
                    }
                    </style>
                """, unsafe_allow_html=True)

                if st.session_state.assigned and not st.session_state.downloaded:
                    with st.container():
                        st.download_button(
                            label="📥 최종 스케쥴 다운로드",
                            data=st.session_state.output,
                            file_name=f"{month_str} 스케쥴.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_schedule_button",
                            type="primary",
                            on_click=lambda: st.session_state.update({"downloaded": True})
                        )
                st.write("DEBUG: 근무 배정 로직 최종 완료.") # DEBUG

            # else: # is_admin_authenticated가 False인 경우는 이미 상단에서 처리됨
            #     st.warning("⚠️ 관리자 권한이 없습니다.")
            #     st.stop() # 상단에서 처리되므로 여기서는 불필요