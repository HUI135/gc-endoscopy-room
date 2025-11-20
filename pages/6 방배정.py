import re
import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from collections import Counter
import random
import time
from datetime import datetime, date, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
import menu
import numpy as np
from dateutil.relativedelta import relativedelta
import platform

st.set_page_config(page_title="방배정", page_icon="", layout="wide")

import os
st.session_state.current_page = os.path.basename(__file__)

menu.menu()

# 로그인 체크 및 자동 리디렉션
if not st.session_state.get("login_success", False):
    st.warning("⚠️ Home 페이지에서 먼저 로그인해주세요.")
    st.error("1초 후 Home 페이지로 돌아갑니다...")
    time.sleep(1)
    st.switch_page("Home.py")
    st.stop()

# 세션 상태 초기화
def initialize_session_state():
    if "data_loaded" not in st.session_state:
        st.session_state["data_loaded"] = False
    if "df_room_request" not in st.session_state:
        st.session_state["df_room_request"] = pd.DataFrame(columns=["이름", "분류", "날짜정보"])
    if "room_settings" not in st.session_state:
        st.session_state["room_settings"] = {
            "830_room_select": ['1', '8', '4', '7'],
            "900_room_select": ['10', '11', '12'],
            "930_room_select": ['2', '5', '6'],
            "1000_room_select": ['9', '3'],
            "1330_room_select": ['3', '4', '9', '2']
        }
    if "weekend_room_settings" not in st.session_state:
        st.session_state["weekend_room_settings"] = {}
    if "swapped_assignments" not in st.session_state:
        st.session_state["swapped_assignments"] = set()
    if "df_schedule_original" not in st.session_state:
        st.session_state["df_schedule_original"] = pd.DataFrame()
    if "manual_change_log" not in st.session_state:
        st.session_state["manual_change_log"] = []
    if "final_change_log" not in st.session_state:
        st.session_state["final_change_log"] = []
    if "saved_changes_log" not in st.session_state:
        st.session_state["saved_changes_log"] = []
    if "df_schedule_md_initial" not in st.session_state:
        st.session_state["df_schedule_md_initial"] = pd.DataFrame()
    if "swapped_assignments_log" not in st.session_state:
        st.session_state["swapped_assignments_log"] = []
    if "df_schedule" not in st.session_state:
        st.session_state["df_schedule"] = pd.DataFrame()
    if "df_swap_requests" not in st.session_state:
        st.session_state["df_swap_requests"] = pd.DataFrame(columns=[
            "RequestID", "요청일시", "요청자", "변경 요청", "변경 요청한 스케줄"
        ])
    if "worksheet_room_request" not in st.session_state:
        st.session_state["worksheet_room_request"] = None
    if "batch_apply_messages" not in st.session_state:
        st.session_state["batch_apply_messages"] = []
    if "assignment_results" not in st.session_state:
        st.session_state["assignment_results"] = None
    if "show_assignment_results" not in st.session_state:
        st.session_state["show_assignment_results"] = False
    if "df_cumulative_original" not in st.session_state:
        st.session_state["df_cumulative_original"] = pd.DataFrame()
    if "latest_cumulative_name" not in st.session_state:
        st.session_state["latest_cumulative_name"] = None
    if "editor_has_changes" not in st.session_state:
        st.session_state["editor_has_changes"] = False

def set_editor_changed_flag():
    """data_editor에서 수정이 발생했음을 세션 상태에 기록합니다."""
    st.session_state.editor_has_changes = True

def clean_name(name):
    """이름 뒤에 붙는 (상태) 문자열을 제거합니다."""
    if not isinstance(name, str):
        return ""
    return re.sub(r'\s*\(.*\)', '', name).strip()

# Google Sheets 클라이언트 초기화
def get_gspread_client():
    scope = ["https://www.googleapis.com/auth/spreadsheets"]
    try:
        service_account_info = dict(st.secrets["gspread"])
        service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
        credentials = Credentials.from_service_account_info(service_account_info, scopes=scope)
        return gspread.authorize(credentials)
    except Exception as e:
        st.error(f"Google Sheets 인증 정보 로드 중 오류: {type(e).__name__} - {e}")
        return None

# Google Sheets 업데이트 함수
def update_sheet_with_retry(worksheet, data, retries=5, delay=10):
    for attempt in range(retries):
        try:
            worksheet.clear()
            
            data_forced_text = []
            for i, row in enumerate(data): # i = row index
                new_row = []
                for j, cell in enumerate(row): # j = col index
                    cell_str = str(cell).strip()
                    
                    # --- ▼▼▼ [수정] 정규 표현식(Regex)으로 숫자 감지 ---
                    is_numeric = False
                    if cell_str: # 비어있지 않은 셀만 검사
                        # 정규식: 선택적 -(마이너스), 숫자, 선택적 .소수점
                        # (예: '1', '-2', '3.14', '-0.5' 모두 감지)
                        if re.match(r'^-?\d+(\.\d+)?$', cell_str):
                            is_numeric = True
                        else:
                            is_numeric = False
                    
                    # 헤더(i=0)가 아니고, '항목' 열(j=0)이 아니며,
                    # 숫자 형태의 텍스트인 경우
                    if i > 0 and j != 0 and is_numeric:
                        new_row.append(f"'{cell_str}") # (예: '1', '-2', '0')
                    else:
                        new_row.append(cell_str) # (예: '항목', '강승주', '')
                    # --- ▲▲▲ [수정 완료] ---

                data_forced_text.append(new_row)
            
            # [수정] data_as_text -> data_forced_text
            worksheet.update('A1', data_forced_text, value_input_option='USER_ENTERED')
            
            return True
        except Exception as e:
            if "Quota exceeded" in str(e):
                st.warning(f"API 쿼터 초과, {delay}초 후 재시도 ({attempt+1}/{retries})")
                time.sleep(delay)
            else:
                st.error(f"업데이트 실패, {delay}초 후 재시도 ({attempt+1}/{retries}): {str(e)}")
                time.sleep(delay)
    st.error("Google Sheets 업데이트 실패: 재시도 횟수 초과")
    return False

# --- find_latest_version 함수를 아래 코드로 완전히 교체하세요 ---

def find_latest_version(sheet, month_str, sheet_type):
    """
    주어진 월과 시트 타입에 해당하는 시트 중 가장 최신 버전을 찾습니다.
    '최종' 버전을 최우선으로 합니다.
    """
    versions = {}
    # 1. '최종' 버전 시트 이름을 먼저 정의합니다.
    final_version_name = f"{month_str} {sheet_type} 최종"
    
    all_worksheet_titles = [ws.title for ws in sheet.worksheets()]
    
    # 2. '최종' 버전이 있는지 확인하고, 있으면 즉시 반환합니다.
    if final_version_name in all_worksheet_titles:
        return final_version_name
            
    # 3. '최종' 버전이 없으면, 다른 버전(ver X.X 또는 기본)들을 찾습니다.
    pattern = re.compile(f"^{re.escape(month_str)} {re.escape(sheet_type)}(?: ver\s*(\d+\.\d+))?$")
    for title in all_worksheet_titles:
        match = pattern.match(title)
        if match:
            version_num_str = match.group(1)
            version_num = float(version_num_str) if version_num_str else 1.0
            versions[title] = version_num
    
    if not versions:
        return None
    
    # 4. 찾은 버전들 중 가장 최신 버전을 반환합니다.
    return max(versions, key=versions.get)

def find_latest_cumulative_version(sheet, month_str):
    """
    [수정됨] 주어진 월에 해당하는 누적 시트 중 '최종'을 제외하고 
    'ver X.X' 또는 기본('누적') 시트 중에서 가장 최신 버전을 찾습니다.
    """
    versions = {}
    
    # 1. '최종' 시트(f"{month_str} 누적 최종")는 의도적으로 무시합니다.
    
    # 2. 'ver X.X' 및 기본 버전('누적')만 찾습니다.
    # 'ver 1.0', 'ver1.0' 등 다양한 형식을 모두 찾도록 정규식 수정
    pattern = re.compile(f"^{re.escape(month_str)} 누적(?: ver\s*(\d+\.\d+))?$")

    for ws in sheet.worksheets():
        match = pattern.match(ws.title)
        if match:
            version_num_str = match.group(1) # ver 뒤의 숫자 부분 (예: '1.0')
            # 버전 넘버가 있으면 float으로 변환, 없으면(기본 '누적' 시트) 1.0으로 처리
            version_num = float(version_num_str) if version_num_str else 1.0
            versions[ws.title] = version_num

    if not versions:
        return None # '최종'을 제외한 어떠한 버전의 시트도 찾지 못하면 None 반환

    # 'ver'가 붙은 시트 중 가장 높은 버전 번호를 가진 시트의 이름을 반환
    return max(versions, key=versions.get)

# [1단계: 이 함수를 스크립트 상단 L100 부근에 추가하세요]

@st.cache_data(ttl=300, show_spinner=False)
def check_final_sheets_exist(month_str, next_month_str):
    """방배정, 스케줄 최종, 누적 최종 시트가 모두 존재하는지 확인합니다."""
    sheets_to_check = [
        f"{month_str} 방배정",
        f"{month_str} 스케줄 최종",
        f"{next_month_str} 누적 최종"
    ]
    try:
        gc = get_gspread_client()
        if not gc: 
            return False
        sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
        
        # 3개 시트가 모두 존재하는지 확인
        for sheet_name in sheets_to_check:
            sheet.worksheet(sheet_name) # 이 줄이 실패하면 except로 이동
        
        return True # 3개 모두 존재
    except gspread.exceptions.WorksheetNotFound:
        return False # 1개라도 없으면 False
    except Exception:
        return False # API 오류 등 기타 문제

# 데이터 로드 함수
def load_data_page6_no_cache(month_str):
    try:
        gc = get_gspread_client()
        if gc is None:
            raise Exception("Failed to initialize gspread client")
        sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])

        # --- 스케줄 시트 로드 ---
        latest_schedule_name = find_latest_version(sheet, month_str, "스케줄") 
        if not latest_schedule_name:
            return pd.DataFrame(), pd.DataFrame(), None, pd.DataFrame(), pd.DataFrame(), None
        worksheet_schedule = sheet.worksheet(latest_schedule_name)
        df_schedule = pd.DataFrame(worksheet_schedule.get_all_records())
        if df_schedule.empty:
            return pd.DataFrame(), pd.DataFrame(), None, pd.DataFrame(), pd.DataFrame(), latest_schedule_name

        # --- 방배정 요청 시트 로드 ---
        try:
            worksheet_room_request = sheet.worksheet(f"{month_str} 방배정 요청")
        except gspread.exceptions.WorksheetNotFound:
            worksheet_room_request = sheet.add_worksheet(title=f"{month_str} 방배정 요청", rows=100, cols=10)
            worksheet_room_request.update('A1', [["이름", "분류", "날짜정보"]])
        df_room_request = pd.DataFrame(worksheet_room_request.get_all_records())

        # --- [수정] 누적 시트 로드 로직 변경 ---
        target_month_dt = datetime.strptime(month_str, "%Y년 %m월")
        next_month_dt = target_month_dt + relativedelta(months=1)
        next_month_str = next_month_dt.strftime("%Y년 %-m월")

        latest_cumulative_name = find_latest_cumulative_version(sheet, next_month_str)

        # [핵심 변경] 다음 달 누적 시트가 없으면 즉시 중단
        if not latest_cumulative_name:
            st.error(f"🚨 '{next_month_str} 누적' 시트를 찾을 수 없습니다. 방배정을 진행할 수 없습니다.")
            st.stop() # 여기서 스크립트 실행을 중단합니다.

        # [추가] 찾은 시트 이름을 세션에 저장하여 '저장' 버튼에서 사용
        st.session_state["latest_cumulative_name"] = latest_cumulative_name

        # 다음 달 누적 시트를 성공적으로 찾은 경우에만 아래 로직 실행
        worksheet_cumulative = sheet.worksheet(latest_cumulative_name)
        all_values = worksheet_cumulative.get_all_values()
        if not all_values or len(all_values) < 2 or all_values[0][0] != '항목':
            st.error(f"🚨 '{latest_cumulative_name}' 시트의 형식이 올바르지 않습니다. A1셀에 '항목'이 있는지 확인해주세요.")
            st.stop()
        else:
            headers, data = all_values[0], all_values[1:]
            
            # [추가] 원본 데이터(transpose 전)를 DataFrame으로 만들어 세션 상태에 저장
            df_cumulative_original = pd.DataFrame(data, columns=headers)
            st.session_state["df_cumulative_original"] = df_cumulative_original
            
            # 기존 로직은 다른 기능에서 사용하므로 유지
            df_transposed = df_cumulative_original.set_index('항목')
            df_cumulative = df_transposed.transpose().reset_index().rename(columns={'index': '이름'})
            for col in ['오전누적', '오후누적', '오전당직', '오후당직']:
                if col in df_cumulative.columns:
                    df_cumulative[col] = pd.to_numeric(df_cumulative[col], errors='coerce').fillna(0).astype(int)
                    
        # --- 스케줄 변경요청 시트 로드 ---
        try:
            worksheet_swap_requests = sheet.worksheet(f"{month_str} 스케줄 변경요청")
        except gspread.exceptions.WorksheetNotFound:
            worksheet_swap_requests = sheet.add_worksheet(title=f"{month_str} 스케줄 변경요청", rows=100, cols=10)
            worksheet_swap_requests.update('A1', [["RequestID", "요청일시", "요청자", "변경 요청", "변경 요청한 스케줄"]])
        df_swap_requests = pd.DataFrame(worksheet_swap_requests.get_all_records())

        return df_schedule, df_room_request, worksheet_room_request, df_cumulative, df_swap_requests, latest_schedule_name

    except Exception as e:
        st.error(f"데이터 로드 중 오류 발생: {type(e).__name__} - {e}")
        return pd.DataFrame(), pd.DataFrame(), None, pd.DataFrame(), pd.DataFrame(), None
    
def find_latest_schedule_version(sheet, month_str):
    """주어진 월에 해당하는 스케줄 시트 중 가장 최신 버전을 찾습니다. '최종'이 최우선입니다."""
    versions = {}
    
    # 1. '최종' 시트 존재 여부 확인 (가장 높은 우선순위)
    final_version_name = f"{month_str} 스케줄 최종"
    for ws in sheet.worksheets():
        if ws.title == final_version_name:
            return final_version_name
    
    # 2. 'ver X.X' 및 기본 버전 찾기 (기존 로직 유지)
    # 'ver 1.0', 'ver1.0' 등 다양한 형식을 모두 찾도록 정규식 수정
    pattern = re.compile(f"^{re.escape(month_str)} 스케줄(?: ver\s*(\d+\.\d+))?$")

    for ws in sheet.worksheets():
        match = pattern.match(ws.title)
        if match:
            version_num_str = match.group(1) # ver 뒤의 숫자 부분 (예: '1.0')
            # 버전 넘버가 있으면 float으로 변환, 없으면 (기본 시트면) 1.0으로 처리
            version_num = float(version_num_str) if version_num_str else 1.0
            versions[ws.title] = version_num

    if not versions:
        return None

    # 가장 높은 버전 번호를 가진 시트의 이름을 반환
    return max(versions, key=versions.get)

@st.cache_data(ttl=300, show_spinner=False)
def load_schedule_data(month_str):
    """가장 최신 버전의 스케줄 데이터를 불러온 후, 필요한 열만 남도록 필터링합니다."""
    try:
        gc = get_gspread_client()
        if not gc:
            return pd.DataFrame(), None

        spreadsheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
        latest_version_name = find_latest_schedule_version(spreadsheet, month_str)
        
        if not latest_version_name:
            st.info(f"{month_str} 스케줄이 아직 배정되지 않았습니다.")
            return pd.DataFrame(), None

        worksheet = spreadsheet.worksheet(latest_version_name)
        records = worksheet.get_all_records()
        
        if not records:
            st.info(f"'{latest_version_name}' 시트에 데이터가 없습니다.")
            return pd.DataFrame(), latest_version_name
            
        df = pd.DataFrame(records)
        if '날짜' not in df.columns:
            st.info(f"'{latest_version_name}' 시트의 형식이 올바르지 않습니다.")
            return pd.DataFrame(), latest_version_name

        # ▼▼▼ [핵심 수정] 데이터프레임 자체를 필요한 열만 남도록 필터링합니다. ▼▼▼
        essential_columns = ['날짜', '요일'] + [str(i) for i in range(1, 13)] + ['오전당직(온콜)'] + [f'오후{i}' for i in range(1, 5)]
        columns_to_keep = [col for col in essential_columns if col in df.columns]
        df = df[columns_to_keep]
        # ▲▲▲ [핵심 수정] ▲▲▲
            
        df.fillna('', inplace=True)
        df['날짜_dt'] = pd.to_datetime(YEAR_STR + '년 ' + df['날짜'].astype(str), format='%Y년 %m월 %d일', errors='coerce')
        df.dropna(subset=['날짜_dt'], inplace=True)
        
        return df, latest_version_name

    except gspread.exceptions.APIError as e:
        st.warning("⚠️ 너무 많은 요청이 접속되어 딜레이되고 있습니다. 잠시 후 재시도 해주세요.")
        st.error(f"Google Sheets API 오류 (스케줄 데이터 로드): {str(e)}")
        st.stop()
    except gspread.exceptions.WorksheetNotFound:
        st.info(f"{month_str} 스케줄이 아직 배정되지 않았습니다.")
        return pd.DataFrame(), None
    except Exception as e:
        st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
        st.error(f"스케줄 데이터 로드 중 오류 발생: {str(e)}")
        st.stop()

# 근무 가능 일자 계산
@st.cache_data(show_spinner=False)
def get_user_available_dates(name, df_schedule, month_start, month_end, month_str):
    available_dates = []
    weekday_map = {0: "월", 1: "화", 2: "수", 3: "목", 4: "금", 5: "토", 6: "일"}
    
    # [수정] month_str에서 연도를 동적으로 추출
    target_year = int(month_str.split('년')[0])

    personnel_columns = [str(i) for i in range(1, 13)] + ['오전당직(온콜)'] + [f'오후{i}' for i in range(1, 6)]
    all_personnel = set(p.strip() for col in personnel_columns if col in df_schedule.columns for p in df_schedule[col].dropna().astype(str))

    if name not in all_personnel:
        st.warning(f"'{name}'님은 이번 달 근무자로 등록되어 있지 않습니다.")
        return []

    for _, row in df_schedule.iterrows():
        date_str = row['날짜']
        try:
            if "월" in date_str:
                # [수정] 하드코딩된 연도 대신 target_year 사용
                date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=target_year).date()
            else:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            continue

        if month_start <= date_obj <= month_end: # ◀ 이렇게 수정
            oncall_person = str(row.get('오전당직(온콜)', '')).strip()
            
            morning_personnel = set(str(row.get(str(i), '')).strip() for i in range(1, 13)) - {''}
            afternoon_personnel = set(str(row.get(f'오후{i}', '')).strip() for i in range(1, 6)) - {''}
            
            display_date = f"{date_obj.month}월 {date_obj.day}일 ({weekday_map[date_obj.weekday()]})"
            save_date_am = f"{date_obj.strftime('%Y-%m-%d')} (오전)"
            save_date_pm = f"{date_obj.strftime('%Y-%m-%d')} (오후)"
            
            if name in morning_personnel or name == oncall_person:
                available_dates.append((date_obj, f"{display_date} 오전", save_date_am))
            if name in afternoon_personnel:
                available_dates.append((date_obj, f"{display_date} 오후", save_date_pm))
    
    unique_dates = sorted(list(set(available_dates)), key=lambda x: x[0])
    return [(display_str, save_str) for _, display_str, save_str in unique_dates]

# df_schedule_md 생성 함수
def create_df_schedule_md(df_schedule):
    """화면 표시에 적합한 형태로 스케줄을 가공하고, 이름에서 (상태)를 제거합니다."""
    display_cols = ['날짜', '요일', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '오전당직(온콜)', '오후1', '오후2', '오후3', '오후4']
    df_schedule_md = pd.DataFrame(columns=display_cols)
    if df_schedule.empty:
        return df_schedule_md

    # 날짜, 요일, 온콜 정보는 미리 복사 (온콜 이름도 clean_name 적용)
    df_schedule_md['날짜'] = df_schedule['날짜']
    df_schedule_md['요일'] = df_schedule['요일']
    if '오전당직(온콜)' in df_schedule.columns:
        df_schedule_md['오전당직(온콜)'] = df_schedule['오전당직(온콜)'].apply(clean_name)

    # 행별로 근무자 재배치
    for idx, row in df_schedule.iterrows():
        oncall_person = clean_name(str(row.get('오전당직(온콜)', '')))
        
        # 오전 근무자 (이름 정리 및 중복 제거)
        am_original_cols = [str(i) for i in range(1, 13)]
        am_personnel_list = [
            clean_name(row[col]) for col in am_original_cols
            if col in df_schedule.columns and clean_name(row[col]) and clean_name(row[col]) != oncall_person
        ]
        am_personnel_unique = list(dict.fromkeys(am_personnel_list))
        am_display_cols = [str(i) for i in range(1, 12)]
        for i, col in enumerate(am_display_cols):
            df_schedule_md.at[idx, col] = am_personnel_unique[i] if i < len(am_personnel_unique) else ''
        
        # 오후 근무자 (이름 정리 및 중복 제거)
        pm_original_cols = [f'오후{i}' for i in range(1, 6)]
        pm_personnel_list = [
            clean_name(row[col]) for col in pm_original_cols
            if col in df_schedule.columns and clean_name(row[col]) and clean_name(row[col]) != oncall_person
        ]
        pm_personnel_unique = list(dict.fromkeys(pm_personnel_list))
        pm_display_cols = [f'오후{i}' for i in range(1, 5)]
        for i, col in enumerate(pm_display_cols):
            df_schedule_md.at[idx, col] = pm_personnel_unique[i] if i < len(pm_personnel_unique) else ''
            
    return df_schedule_md

# ✂️ 복사 & 붙여넣기용 최종 apply_schedule_swaps 함수
def apply_schedule_swaps(original_schedule_df, swap_requests_df, special_df):
    df_modified = original_schedule_df.copy()
    applied_count = 0
    total_requests = len(swap_requests_df)
    swapped_assignments = st.session_state.get("swapped_assignments", set())
    batch_change_log = []
    messages = []

    for _, request_row in swap_requests_df.iterrows():
        try:
            change_request_str = str(request_row.get('변경 요청', '')).strip()
            schedule_info_str = str(request_row.get('변경 요청한 스케줄', '')).strip()
            formatted_schedule_info = format_sheet_date_for_display(schedule_info_str)

            if '➡️' not in change_request_str: continue

            person_before, person_after = [p.strip() for p in change_request_str.split('➡️')]
            date_match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', schedule_info_str)
            if not date_match: continue

            date_part, time_period_from_request = date_match.groups()
            date_obj = datetime.strptime(date_part, '%Y-%m-%d').date()
            formatted_date_in_df = f"{date_obj.month}월 {date_obj.day}일"
            target_row_indices = df_modified[df_modified['날짜'] == formatted_date_in_df].index
            if target_row_indices.empty: continue
            target_row_idx = target_row_indices[0]

            all_cols = [str(i) for i in range(1, 18)] + ['오전당직(온콜)'] + [f'오후{i}' for i in range(1, 10)]
            available_cols = [col for col in all_cols if col in df_modified.columns]
            
            on_call_person = clean_name(df_modified.at[target_row_idx, '오전당직(온콜)'])
            is_on_call_swap = (person_before == on_call_person)

            if is_on_call_swap:
                cols_with_person_before = [c for c in available_cols if clean_name(df_modified.at[target_row_idx, c]) == person_before]
                cols_with_person_after = [c for c in available_cols if clean_name(df_modified.at[target_row_idx, c]) == person_after]

                if not cols_with_person_before:
                    error_msg = f"❌ {formatted_schedule_info} - {change_request_str} 적용 실패: {formatted_date_in_df}에 '{person_before}' 당직 근무가 배정되어 있지 않습니다."
                    messages.append(('error', error_msg))
                    continue

                for col in cols_with_person_before: df_modified.at[target_row_idx, col] = person_after
                for col in cols_with_person_after: df_modified.at[target_row_idx, col] = person_before

                # --- ▼▼▼ [ 1. 핵심 수정 ] ▼▼▼ ---
                # (날짜, 시간, 변경 전, 변경 후) 4-tuple로 저장
                swapped_assignments.add((formatted_date_in_df, '오전', person_before, person_after))
                swapped_assignments.add((formatted_date_in_df, '오후', person_before, person_after))
                swapped_assignments.add((formatted_date_in_df, '오전당직(온콜)', person_before, person_after))
                # --- ▲▲▲ [ 수정 완료 ] ▲▲▲

                batch_change_log.append({
                    '날짜': f"{formatted_date_in_df} ({'월화수목금토일'[date_obj.weekday()]}) - 오전당직 변경",
                    '변경 전 인원': person_before, '변경 후 인원': person_after,
                })
                applied_count += 1
                continue

            target_cols = [str(i) for i in range(1, 18)] if time_period_from_request == '오전' else [f'오후{i}' for i in range(1, 10)]
            available_target_cols = [col for col in target_cols if col in df_modified.columns]
            
            matched_cols = [col for col in available_target_cols if clean_name(df_modified.loc[target_row_idx, col]) == person_before]
            if not matched_cols:
                error_msg = f"❌ {formatted_schedule_info} - {change_request_str} 적용 실패: {formatted_date_in_df} '{time_period_from_request}'에 '{person_before}'이(가) 배정되어 있지 않습니다."
                messages.append(('error', error_msg))
                continue
            
            personnel_in_target_period = {clean_name(df_modified.loc[target_row_idx, col]) for col in available_target_cols}
            if person_after in personnel_in_target_period:
                warning_msg = f"🟡 {formatted_schedule_info} - {change_request_str} 적용 건너뜀: '{person_after}'님은 이미 {formatted_date_in_df} '{time_period_from_request}' 근무에 배정되어 있습니다."
                messages.append(('warning', warning_msg))
                continue
            
            for col in matched_cols:
                df_modified.at[target_row_idx, col] = person_after
            
            # --- ▼▼▼ [ 2. 핵심 수정 ] ▼▼▼ ---
            # (날짜, 시간, 변경 전, 변경 후) 4-tuple로 저장
            swapped_assignments.add((formatted_date_in_df, time_period_from_request, person_before, person_after))
            # --- ▲▲▲ [ 수정 완료 ] ▲▲▲
            
            batch_change_log.append({
                '날짜': f"{formatted_schedule_info}", '변경 전 인원': person_before, '변경 후 인원': person_after,
            })
            applied_count += 1

        except Exception as e:
            messages.append(('error', f"요청 처리 중 오류 발생: {type(e).__name__} - {str(e)}"))
            continue
    
    if applied_count > 0 or messages:
        summary = f"✅ 총 {total_requests}건 중 {applied_count}건의 스케줄 변경 요청이 성공적으로 반영되었습니다."
        messages.insert(0, ('success', summary))
    elif not messages:
        messages.append(('info', "새롭게 적용할 스케줄 변경 요청이 없습니다."))

    st.session_state["swapped_assignments_log"] = batch_change_log
    st.session_state["swapped_assignments"] = swapped_assignments

    return create_df_schedule_md(df_modified), messages

def format_sheet_date_for_display(date_string):
    match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', date_string)
    if match:
        date_part, shift_part = match.groups()
        try:
            dt_obj = datetime.strptime(date_part, '%Y-%m-%d').date()
            weekday_str = ['월', '화', '수', '목', '금', '토', '일'][dt_obj.weekday()]
            return f"{dt_obj.month}월 {dt_obj.day}일 ({weekday_str}) - {shift_part}"
        except ValueError:
            return date_string
    return date_string

def format_date_str_to_display(date_str, weekday, time_period):
    if '요일' in weekday:
        weekday = weekday.replace('요일', '')
    return f"{date_str} ({weekday}) - {time_period}"

@st.cache_data(ttl=600, show_spinner=False)
def load_special_schedules(month_str):
    """
    'YYYY년 토요/휴일 스케줄' 시트에서 데이터를 로드하는 함수입니다.
    """
    try:
        gc = get_gspread_client()
        if not gc: return pd.DataFrame()

        spreadsheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
        
        # month_str에서 연도를 동적으로 추출하여 시트 이름을 생성합니다.
        target_year = month_str.split('년')[0]
        sheet_name = f"{target_year}년 토요/휴일 스케줄"
        
        worksheet = spreadsheet.worksheet(sheet_name)
        records = worksheet.get_all_records()
        
        if not records:
            return pd.DataFrame()
        
        df = pd.DataFrame(records)
        df.fillna('', inplace=True)
        return df
        
    except gspread.exceptions.WorksheetNotFound:
        target_year = month_str.split('년')[0]
        sheet_name = f"{target_year}년 토요/휴일 스케줄"
        st.info(f"'{sheet_name}' 시트를 찾을 수 없습니다. 토요/휴일 일정 없이 진행합니다.")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"토요/휴일 데이터 로드 중 오류 발생: {str(e)}")
        return pd.DataFrame()

def generate_excel_output(df_room, stats_df, columns, special_dates, special_df, date_cache, request_cells, swapped_assignments, morning_duty_slot, month_str, change_log_map=None):
    """
    [수정됨]
    현재 방배정/통계 DataFrame을 기반으로 '방배정'이라는 단일 시트에
    스케줄과 통계 테이블을 모두 포함하는 Excel 파일을 생성(BytesIO)합니다.
    """
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    sheet.title = "방배정"

    if platform.system() == "Windows":
        font_name = "맑은 고딕"
    else:
        font_name = "Arial"

    # (스타일 정의)
    highlight_fill = PatternFill(start_color="F2DCDB", fill_type="solid")
    duty_font = Font(name=font_name, size=9, bold=True, color="FF00FF")
    default_font = Font(name=font_name, size=9, bold=False) 
    special_day_fill = PatternFill(start_color="BFBFBF", fill_type="solid")
    no_person_day_fill = PatternFill(start_color="808080", fill_type="solid")
    default_yoil_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    special_person_fill = PatternFill(start_color="DDEBF7", fill_type="solid")
    
    thin_side = Side(style='thin')
    thick_side = Side(style='medium')
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # (헤더 렌더링 - df_room 기준)
    for col_idx, header in enumerate(columns, 1):
        cell = sheet.cell(1, col_idx, header)
        cell.font = Font(bold=True, name=font_name, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        if header.startswith('8:30') or header == '온콜':
            cell.fill = PatternFill(start_color="FFE699", fill_type="solid")
        elif header.startswith('9:00'):
            cell.fill = PatternFill(start_color="F8CBAD", fill_type="solid")
        elif header.startswith('9:30'):
            cell.fill = PatternFill(start_color="B4C6E7", fill_type="solid")
        elif header.startswith('10:00'):
            cell.fill = PatternFill(start_color="C6E0B4", fill_type="solid")
        elif header.startswith('13:30'):
            cell.fill = PatternFill(start_color="CC99FF", fill_type="solid")

    # (데이터 렌더링 - df_room 기준)
    result_data = df_room.fillna('').values.tolist()
    target_year = int(month_str.split('년')[0])
    
    last_data_row = 1 
    
    for row_idx, row_data in enumerate(result_data, 2):
        current_date_str = row_data[0]
        
        is_special_day = current_date_str in special_dates
        duty_person_for_the_day = None
        
        if is_special_day:
            try:
                date_obj_lookup = datetime.strptime(current_date_str, '%m월 %d일').replace(year=target_year)
                duty_person_row = special_df[special_df['날짜_dt'].dt.date == date_obj_lookup.date()]
                if not duty_person_row.empty:
                    duty_person_raw = duty_person_row['당직'].iloc[0]
                    if pd.notna(duty_person_raw) and str(duty_person_raw).strip() and str(duty_person_raw).strip() != '당직 없음':
                        duty_person_for_the_day = str(duty_person_raw).strip()
            except Exception as e:
                pass 

        assignment_cells = row_data[2:]
        personnel_in_row = [p for p in assignment_cells if p]
        is_no_person_day = not any(personnel_in_row)
        is_small_team_day_for_bg = (0 < len(personnel_in_row) < 15) or is_special_day

        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border 
            
            if col_idx == 1: cell.fill = no_person_day_fill
            elif col_idx == 2:
                if is_no_person_day: cell.fill = no_person_day_fill
                elif is_small_team_day_for_bg: cell.fill = special_day_fill
                else: cell.fill = default_yoil_fill
            elif is_no_person_day and col_idx >= 3: cell.fill = no_person_day_fill
            elif is_special_day and col_idx > 2 and value: cell.fill = special_person_fill
            
            slot_name = columns[col_idx-1]
            current_cell_date = str(current_date_str).strip()
            current_cell_value = str(value).strip()

            # --- 1. 'swapped_assignments' (스케줄 변경) 기준 ---
            cell_shift_type = ''
            if any(time_str in str(slot_name) for time_str in ['8:30', '9:00', '9:30', '10:00']): cell_shift_type = '오전'
            # [수정] '오전당직(온콜)'도 '오후'로 간주하여 '스케줄 수정'의 `apply_schedule_swaps`와 맞춤
            elif any(time_str in str(slot_name) for time_str in ['13:30', '온콜', '오전당직(온콜)']): cell_shift_type = '오후'
            
            # --- ▼▼▼ [ 3. 핵심 수정 ] ▼▼▼ ---
            is_newly_assigned = False
            original_value_schedule = None # '변경 전' 값을 저장할 변수

            if swapped_assignments:
                # 4-tuple (날짜, 시간, 이전값, 새값) 또는 3-tuple (날짜, 시간, 새값)을 순회
                for swap_tuple in swapped_assignments:
                    try:
                        if len(swap_tuple) == 4:
                            s_date, s_shift, s_old, s_new = swap_tuple
                            # [수정] '오전당직(온콜)' 특별 처리
                            if s_shift == '오전당직(온콜)':
                                s_shift = '오후' # cell_shift_type과 맞추기
                                
                            if (s_date == current_cell_date and
                                s_shift == cell_shift_type and
                                s_new == current_cell_value):
                                is_newly_assigned = True
                                original_value_schedule = s_old # '변경 전' 값 저장
                                break 
                        elif len(swap_tuple) == 3: # [호환성]
                            s_date, s_shift, s_new = swap_tuple
                            if s_shift == '오전당직(온콜)':
                                s_shift = '오후'
                                
                            if (s_date == current_cell_date and
                                s_shift == cell_shift_type and
                                s_new == current_cell_value):
                                is_newly_assigned = True
                                break
                    except Exception as e:
                        # 튜플 형식이 잘못되었을 경우 무시
                        pass 
            # --- ▲▲▲ [ 수정 완료 ] ▲▲▲

            # --- 2. 'change_log_map' (방배정 수동 변경) 기준 ---
            cell_changed_in_room_editor = False
            original_value = None
            if change_log_map is not None:
                lookup_key_map = (current_cell_date, slot_name)
                if lookup_key_map in change_log_map:
                    cell_changed_in_room_editor = True
                    original_value, _ = change_log_map[lookup_key_map]

            # --- 3. 우선순위에 따라 적용 ---
            if cell_changed_in_room_editor:
                # 1순위: '방배정' 에디터에서 수동 변경 (색상 + 메모)
                cell.fill = highlight_fill
                original_display = original_value if original_value else '빈 값'
                cell.comment = Comment(f"변경 전: {original_display}", "Edit Tracker")

            # --- ▼▼▼ [ 4. 핵심 수정 ] ▼▼▼ ---
            elif is_newly_assigned:
                # 2순위: '스케줄' 에디터에서 변경 (색상 + 메모)
                cell.fill = highlight_fill
                # '변경 전' 값이 있으면 코멘트 추가
                if original_value_schedule is not None:
                    original_display = original_value_schedule if original_value_schedule else '빈 값'
                    cell.comment = Comment(f"변경 전: {original_display}", "Edit Tracker")
            # --- ▲▲▲ [ 수정 완료 ] ▲▲▲

            cell.font = default_font
            if value:
                if is_special_day:
                    if duty_person_for_the_day and value == duty_person_for_the_day:
                        cell.font = duty_font
                else:
                    if slot_name.endswith('_당직') or slot_name == '온콜':
                        cell.font = duty_font

            # --- ▼▼▼ [ 5. 핵심 수정 ] ▼▼▼ ---
            # '변경 전:' 코멘트가 *없을 때만* '방배정 요청' 코멘트를 추가 (덮어쓰기 방지)
            if cell.comment is None and col_idx > 2 and value and date_cache.get(current_date_str):
                formatted_date_for_comment = date_cache[current_date_str]
                if (formatted_date_for_comment, slot_name) in request_cells and value == request_cells[(formatted_date_for_comment, slot_name)]['이름']:
                    cell.comment = Comment(f"{request_cells[(formatted_date_for_comment, slot_name)]['분류']}", "System")
            # --- ▲▲▲ [ 수정 완료 ] ▲▲▲
            
        last_data_row = row_idx
    
    sheet.column_dimensions['A'].width = 11
    
    schedule_col_count = len(columns)
    stats_col_count = len(stats_df.columns)
    max_cols = max(schedule_col_count, stats_col_count)
    
    for i in range(2, max_cols + 1): 
        col_letter = openpyxl.utils.get_column_letter(i)
        sheet.column_dimensions[col_letter].width = 10

    # --- (이하 Stats 테이블 로직은 동일) ---
    stats_start_row = last_data_row + 3 
    stats_columns = stats_df.columns.tolist()  
    stats_end_col = len(stats_columns)
    stats_end_row = stats_start_row + len(stats_df)
    stats_header_fill = PatternFill(start_color="E7E6E6", fill_type="solid")

    for col_idx, header in enumerate(stats_columns, 1):
        cell = sheet.cell(stats_start_row, col_idx, header) 
        cell.font = Font(bold=True, name=font_name, size=9) 
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = stats_header_fill
        cell.border = Border(
            left=thick_side if col_idx == 1 else thin_side,
            right=thick_side if col_idx == 1 or col_idx == stats_end_col else thin_side, 
            top=thick_side,
            bottom=thick_side
        )

    item_names_series = stats_df['항목']
    separator_rows_set = {"늦은방 합계", "오전당직 누적", "오후당직 누적"}
    time_prefixes = ["8:30(", "9:00(", "9:30(", "10:00("] 
    for prefix in time_prefixes:
        matching_rows = item_names_series[item_names_series.str.startswith(prefix)]
        if not matching_rows.empty:
            separator_rows_set.add(matching_rows.iloc[-1])

    for stats_row_idx, row_data in enumerate(stats_df.fillna('').values.tolist(), stats_start_row + 1):
        item_name = str(row_data[0]) 
        is_last_row = (stats_row_idx == stats_end_row)
        is_separator_row = (item_name in separator_rows_set)
        
        summary_fill = None
        if item_name == '이른방 합계': summary_fill = PatternFill(start_color="FFE699", fill_type="solid")
        elif item_name == '늦은방 합계': summary_fill = PatternFill(start_color="C6E0B4", fill_type="solid")
        elif item_name == '오전당직 합계': summary_fill = PatternFill(start_color="B8CCE4", fill_type="solid")
        elif item_name == '오후당직 합계': summary_fill = PatternFill(start_color="B8CCE4", fill_type="solid")
        elif item_name == '오전당직 누적': summary_fill = PatternFill(start_color="FFC8CD", fill_type="solid")
        elif item_name == '오후당직 누적': summary_fill = PatternFill(start_color="FFC8CD", fill_type="solid")

        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(stats_row_idx, col_idx, value) 
            
            if col_idx == 1:
                cell.font = Font(name=font_name, size=9, bold=True)
            else:
                cell.font = default_font
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=thick_side if col_idx == 1 else thin_side, 
                right=thick_side if col_idx == 1 or col_idx == stats_end_col else thin_side,
                top=thin_side, 
                bottom=thick_side if is_last_row or is_separator_row else thin_side
            )
            
            if summary_fill:
                cell.fill = summary_fill
            else:
                if col_idx == 1:
                    cell.fill = PatternFill(start_color="D0CECE", fill_type="solid")
                else:
                    cell.fill = PatternFill(fill_type=None)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# 메인
from zoneinfo import ZoneInfo
kst = ZoneInfo("Asia/Seoul")
now = datetime.now(kst)
today = now.date()
next_month_date = today.replace(day=1) + relativedelta(months=1)
month_str = next_month_date.strftime("%Y년 %-m월")
month_str = '2025년 10월'
YEAR_STR = month_str.split('년')[0]
this_month_start = next_month_date.replace(day=1)

# 다음 달의 마지막 날 계산
if this_month_start.month == 12:
    this_month_end = date(this_month_start.year, 12, 31)
else:
    this_month_end = (date(this_month_start.year, this_month_start.month + 1, 1) - timedelta(days=1))

# 다음 달 계산 (기존 코드 유지, 필요 시 사용)
if today.month == 12:
    next_month_start = date(today.year + 1, 1, 1)
else:
    next_month_start = date(today.year, today.month + 1, 1)
next_month_end = (next_month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

# ▼▼▼ [추가] month_str을 기준으로 지난달 생성 ▼▼▼
# 1. month_str을 datetime 객체로 변환
current_target_dt = datetime.strptime(month_str, "%Y년 %m월")

# 2. 한 달을 빼서 '지난달' datetime 객체를 만듦
prev_month_dt = current_target_dt - relativedelta(months=1)

# 3. '지난달'을 month_str과 동일한 형식의 문자열로 만듦
prev_month_str = prev_month_dt.strftime("%Y년 %-m월")

# 세션 상태 초기화
initialize_session_state()

st.header("🚪 방배정", divider='rainbow')

# 새로고침 버튼
st.write("- 먼저 새로고침 버튼으로 최신 데이터를 불러온 뒤, 배정을 진행해주세요.")
if st.button("🔄 새로고침 (R)"):
    st.session_state["data_loaded"] = False
    st.cache_data.clear()

    # 모든 로그 및 메시지 초기화
    if "final_change_log" in st.session_state:
        st.session_state["final_change_log"] = []
    if "swapped_assignments_log" in st.session_state:
        st.session_state["swapped_assignments_log"] = []
    if "batch_apply_messages" in st.session_state:
        st.session_state["batch_apply_messages"] = []
    
    if "swapped_assignments" in st.session_state:
        st.session_state["swapped_assignments"] = set()

    # 수정된 스케줄 및 결과 초기화
    if "df_schedule_md_modified" in st.session_state:
        del st.session_state["df_schedule_md_modified"]
        
    # >>>>>>>>> [핵심 수정] 이 두 줄을 추가/수정하세요 <<<<<<<<<
    if "assignment_results" in st.session_state:
        del st.session_state["assignment_results"]
    st.session_state.show_assignment_results = False # 결과 보기 스위치 끄기
    
    st.rerun()

df_schedule, loaded_version = load_schedule_data(month_str)

if df_schedule.empty:
    st.stop()
else:
    # 불러온 스케줄의 버전 정보를 화면에 표시합니다.
    if loaded_version:
        # ' 스케줄 '을 기준으로 시트 이름을 분리하여 마지막 부분을 버전으로 인식합니다.
        # 예: "2025년 10월 스케줄 ver1.0" -> "ver1.0"
        # 예: "2025년 10월 스케줄 최종" -> "최종"
        version_str = loaded_version.split(' 스케줄 ')[-1]

        # version_str이 비어있다면 "2025년 10월 스케줄"과 같은 기본 시트입니다.
        if version_str:
            st.info(f"현재 표시되는 스케줄 버전은 '**{version_str}**' 입니다.")
        else:
            st.info(f"현재 표시되는 스케줄 버전은 **기본 버전**입니다.")
        
        # ▼▼▼ [사용자 요청] '최종' 버전일 때만 초기화 익스팬더 표시 ▼▼▼
        if version_str == "최종":
            # {next_month_str} 계산 (시트 삭제에 필요)
            try:
                target_month_dt = datetime.strptime(month_str, "%Y년 %m월")
                next_month_dt = target_month_dt + relativedelta(months=1)
                # (예: '2025년 1월')
                next_month_str = next_month_dt.strftime("%Y년 %-m월") 
            except Exception as e:
                st.error(f"다음 달 문자열 생성 중 오류: {e}")
                next_month_str = "[날짜 계산 오류]"

            col_delete, none = st.columns([2, 4])

            with col_delete:
                # 삭제는 위험한 작업이므로 확인 절차를 거칩니다.
                with st.expander("🗑️ 스케줄 최종 버전 초기화"):    
                    st.error(
                        "이 작업은 되돌릴 수 없습니다!\n 기존 '스케줄 최종' 버전 시트를 삭제하고 스케줄 변경을 다시 수행하시겠습니까?"
                    )
                
                    if st.button("네, 삭제합니다.", type="primary", use_container_width=True, key="delete_final_confirm"):
                        if next_month_str == "[날짜 계산 오류]":
                            st.error("날짜 계산 오류로 인해 삭제를 진행할 수 없습니다.")
                        else:
                            with st.spinner("최종 버전 시트를 삭제하는 중입니다..."):
                                try:
                                    gc = get_gspread_client()
                                    if gc is None:
                                        raise Exception("Google Sheets 클라이언트를 초기화할 수 없습니다.")
                                    
                                    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
                                    
                                    sheets_to_delete = [
                                        f"{month_str} 스케줄 최종",
                                        f"{next_month_str} 누적 최종", 
                                    ]
                                    
                                    deleted_count = 0
                                    for sheet_name in sheets_to_delete:
                                        try:
                                            worksheet = sheet.worksheet(sheet_name)
                                            sheet.del_worksheet(worksheet)
                                            st.success(f"✅ '{sheet_name}' 시트를 삭제했습니다.")
                                            deleted_count += 1
                                        except gspread.exceptions.WorksheetNotFound:
                                            st.info(f"ℹ️ '{sheet_name}' 시트가 이미 존재하지 않습니다.")
                                    
                                    if deleted_count > 0:
                                        st.info("초기화가 완료되었습니다. 페이지를 새로고침합니다.")
                                        
                                        # ▼▼▼ [사용자 요청 수정] 세션 상태 초기화 추가 ▼▼▼
                                        st.session_state["data_loaded"] = False
                                        st.cache_data.clear()
                                        
                                        # 스케줄 수정 관련 상태 초기화
                                        if "df_schedule_md_modified" in st.session_state:
                                            del st.session_state["df_schedule_md_modified"]
                                        if "final_change_log" in st.session_state:
                                            st.session_state["final_change_log"] = []
                                        if "swapped_assignments_log" in st.session_state:
                                            st.session_state["swapped_assignments_log"] = []
                                        if "batch_apply_messages" in st.session_state:
                                            st.session_state["batch_apply_messages"] = []
                                        if "changed_cells_set" in st.session_state: # 하이라이트 정보도 초기화
                                            del st.session_state["changed_cells_set"]

                                        # 방배정 결과 초기화
                                        if "assignment_results" in st.session_state:
                                                del st.session_state["assignment_results"]
                                        st.session_state.show_assignment_results = False
                                        time.sleep(2)
                                        st.rerun()
                                    else:
                                        st.warning("삭제할 '최종' 시트가 없습니다.")

                                except Exception as e:
                                    st.error(f"시트 삭제 중 오류 발생: {e}")

# 데이터 로드 (페이지 첫 로드 시에만 실행)
if not st.session_state.get("data_loaded", False):
    with st.spinner("데이터를 로드하고 있습니다..."):
        # 반환값에 loaded_version 추가
        df_schedule, df_room_request, worksheet_room_request, df_cumulative, df_swap_requests, loaded_version = load_data_page6_no_cache(month_str)

        # 로드된 데이터를 세션 상태에 저장
        st.session_state["df_schedule"] = df_schedule if df_schedule is not None else pd.DataFrame()
        st.session_state["df_schedule_original"] = st.session_state["df_schedule"].copy()
        st.session_state["df_room_request"] = df_room_request if df_room_request is not None else pd.DataFrame()
        st.session_state["worksheet_room_request"] = worksheet_room_request
        st.session_state["df_cumulative"] = df_cumulative if df_cumulative is not None else pd.DataFrame()
        st.session_state["df_swap_requests"] = df_swap_requests if df_swap_requests is not None else pd.DataFrame()
        st.session_state["df_schedule_md"] = create_df_schedule_md(st.session_state["df_schedule"])
        st.session_state["df_schedule_md_initial"] = st.session_state["df_schedule_md"].copy()
        st.session_state["loaded_version"] = loaded_version # 버전 정보 세션에 저장

        special_schedules_data = []
        special_dates_data = set()
        special_df_data = pd.DataFrame() # 기본 빈 데이터프레임

        try:
            gc = get_gspread_client()
            spreadsheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
            target_year = month_str.split('년')[0]
            special_sheet_name = f"{target_year}년 토요/휴일 스케줄"
            worksheet = spreadsheet.worksheet(special_sheet_name)
            schedule_records = worksheet.get_all_records()

            if schedule_records:
                df_yearly = pd.DataFrame(schedule_records)
                df_yearly['날짜_dt'] = pd.to_datetime(df_yearly['날짜'])
                
                target_month_dt = datetime.strptime(month_str, "%Y년 %m월")
                special_df_data = df_yearly[
                    (df_yearly['날짜_dt'].dt.year == target_month_dt.year) &
                    (df_yearly['날짜_dt'].dt.month == target_month_dt.month)
                ].copy()

                for _, row in special_df_data.iterrows():
                    date_obj = row['날짜_dt'].date()
                    formatted_date_str = f"{date_obj.month}월 {date_obj.day}일"
                    personnel_str = row.get('근무', '')
                    personnel = [p.strip() for p in personnel_str.split(',')] if personnel_str else []
                    special_schedules_data.append((date_obj, formatted_date_str, personnel))
                    special_dates_data.add(formatted_date_str)
        
        except gspread.exceptions.WorksheetNotFound:
            st.info(f"'{special_sheet_name}' 시트를 찾을 수 없어 토요/휴일 정보 없이 진행합니다.")
        except Exception as e:
            st.error(f"토요/휴일 데이터 로드 중 오류: {e}")

        st.session_state["special_schedules"] = special_schedules_data
        st.session_state["special_dates"] = special_dates_data
        st.session_state["special_df"] = special_df_data

        st.session_state["data_loaded"] = True

# 근무자 명단 수정
st.divider()
st.subheader("📋 스케줄 변경 요청 목록")
if "df_schedule" not in st.session_state or st.session_state["df_schedule"].empty:
    st.warning("⚠️ 스케줄 데이터가 로드되지 않았습니다. 새로고침 버튼을 눌러 데이터를 다시 로드해주세요.")
    st.stop()

# --- 표시할 데이터프레임 결정 ---
# data_editor에 들어갈 데이터를 먼저 결정합니다. 이것이 현재 화면의 기준이 됩니다.
df_to_display = st.session_state.get("df_schedule_md_modified", st.session_state.get("df_schedule_md_initial", pd.DataFrame()))

# --- '스케줄 변경 요청 목록' 섹션 ---
df_swaps_raw = st.session_state.get("df_swap_requests", pd.DataFrame())
if not df_swaps_raw.empty:
    cols_to_display = {'요청일시': '요청일시', '요청자': '요청자', '변경 요청': '변경 요청', '변경 요청한 스케줄': '변경 요청한 스케줄'}
    existing_cols = [col for col in cols_to_display.keys() if col in df_swaps_raw.columns]
    df_swaps_display = df_swaps_raw[existing_cols].rename(columns=cols_to_display)
    if '변경 요청한 스케줄' in df_swaps_display.columns:
        df_swaps_display['변경 요청한 스케줄'] = df_swaps_display['변경 요청한 스케줄'].apply(format_sheet_date_for_display)
    st.dataframe(df_swaps_display, use_container_width=True, hide_index=True)

    # >>>>>>>>> [핵심 수정] '일괄 적용' 전 상태일 때만 아래의 충돌 검사를 실행 <<<<<<<<<
    if "df_schedule_md_modified" not in st.session_state:
        # --- 충돌 경고 로직 (수정됨) ---
        request_sources = []
        request_destinations = []

        schedule_df_to_check = st.session_state.get("df_schedule_original", pd.DataFrame()) # 원본 스케줄로 검사
        target_year = int(month_str.split('년')[0])

        for index, row in df_swaps_raw.iterrows():
            change_request_str = str(row.get('변경 요청', '')).strip()
            schedule_info_str = str(row.get('변경 요청한 스케줄', '')).strip()
            
            if '➡️' in change_request_str and schedule_info_str:
                person_before, person_after = [p.strip() for p in change_request_str.split('➡️')]
                
                is_on_call_request = False
                date_match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', schedule_info_str)
                if date_match:
                    date_part, time_period = date_match.groups()
                    try:
                        date_obj = datetime.strptime(date_part, '%Y-%m-%d').date()
                        formatted_date_in_df = f"{date_obj.month}월 {date_obj.day}일"
                        
                        target_row = schedule_df_to_check[schedule_df_to_check['날짜'] == formatted_date_in_df]
                        
                        if not target_row.empty:
                            on_call_person_of_the_day = clean_name(target_row.iloc[0].get('오전당직(온콜)', ''))
                            if person_before == on_call_person_of_the_day:
                                is_on_call_request = True
                    except Exception:
                        pass 
                
                if not is_on_call_request:
                    request_sources.append(f"{person_before} - {schedule_info_str}")
                
                if date_match:
                    date_part, time_period = date_match.groups()
                    request_destinations.append((date_part, time_period, person_after))

        # [검사 1: 출처 충돌]
        source_counts = Counter(request_sources)
        source_conflicts = [item for item, count in source_counts.items() if count > 1]
        if source_conflicts:
            # 1. 기본 경고 메시지 생성
            warning_message = (
                "⚠️ **요청 출처 충돌**: 동일한 근무에 대한 변경 요청이 2개 이상 있습니다. "
                "목록의 가장 위에 있는 요청이 먼저 반영되며, 이후 요청은 무시될 수 있습니다."
            )
            # 2. 충돌 항목들을 리스트에 저장
            conflict_details = []
            for conflict_item in source_conflicts:
                person, schedule = conflict_item.split(' - ', 1)
                formatted_schedule = format_sheet_date_for_display(schedule)
                conflict_details.append(f"- '{person}' 님의 {formatted_schedule} 근무 요청이 중복되었습니다.")
            
            # 3. 모든 메시지를 합쳐서 st.warning으로 한 번에 출력
            warning_message += "\n" + "\n".join(conflict_details)
            st.warning(warning_message)

        # [검사 2: 도착지 중복]
        dest_counts = Counter(request_destinations)
        dest_conflicts = [item for item, count in dest_counts.items() if count > 1]
        if dest_conflicts:
            # 1. 기본 경고 메시지 생성
            warning_message = (
                "⚠️ **요청 도착지 중복**: 한 사람이 같은 날, 같은 시간대에 여러 근무를 받게 되는 요청이 있습니다. "
                "이 경우, 먼저 처리되는 요청만 반영됩니다."
            )
            # 2. 충돌 항목들을 리스트에 저장
            conflict_details = []
            for date, period, person in dest_conflicts:
                formatted_date = format_sheet_date_for_display(f"{date} ({period})")
                conflict_details.append(f"- '{person}' 님이 {formatted_date} 근무에 중복으로 배정될 가능성이 있습니다.")

            # 3. 모든 메시지를 합쳐서 st.warning으로 한 번에 출력
            warning_message += "\n" + "\n".join(conflict_details)
            st.warning(warning_message)
else:
    st.info("표시할 교환 요청 데이터가 없습니다.")

st.divider()
st.subheader("✍️ 스케줄 수정")
st.write("- 요청사항을 **일괄 적용/취소**하거나, 셀을 더블클릭하여 직접 수정한 후 **최종 저장 버튼**을 누르세요.\n- 하단에서 방배정 수행 버튼을 누르면 위 변경사항이 반영된 '**스케줄 최종**' 버전이 저장됩니다.")

# 표시할 데이터프레임 결정
df_to_display = st.session_state.get("df_schedule_md_modified", st.session_state.get("df_schedule_md_initial", pd.DataFrame()))

col1, col2 = st.columns(2)
with col1:
    if st.button("🔄 요청사항 일괄 적용"):
        df_swaps = st.session_state.get("df_swap_requests", pd.DataFrame())
        if not df_swaps.empty:
            modified_schedule, messages = apply_schedule_swaps(
                st.session_state.get("df_schedule_md_modified", st.session_state["df_schedule_original"]), # 수정된 스케줄 기반으로 계속 수정
                df_swaps,
                st.session_state.get("special_df", pd.DataFrame())
            )
            st.session_state["df_schedule_md_modified"] = modified_schedule
            
            # 💡 [로그 수정] 기존 로그에 새로운 일괄 적용 로그를 추가
            existing_log = st.session_state.get("final_change_log", [])
            new_batch_log = st.session_state.get("swapped_assignments_log", [])
            st.session_state["final_change_log"] = existing_log + new_batch_log
            
            st.session_state["batch_apply_messages"] = messages
            st.rerun()
        else:
            st.session_state["batch_apply_messages"] = [('info', "ℹ️ 처리할 교환 요청이 없습니다.")]
            st.rerun()
            
with col2:
    if st.button("⏪ 적용 취소", disabled="df_schedule_md_modified" not in st.session_state):
        if "df_schedule_md_modified" in st.session_state:
            del st.session_state["df_schedule_md_modified"]

        # ▼▼▼ [핵심 수정] 하이라이트 기록을 초기화합니다. ▼▼▼
        st.session_state["swapped_assignments"] = set()
        # ▲▲▲ [핵심 수정] ▲▲▲

        st.session_state["swapped_assignments_log"] = []
        st.session_state["final_change_log"] = []
        st.session_state["batch_apply_messages"] = [('info', "변경사항이 취소되고 원본 스케줄로 돌아갑니다.")]
        st.rerun()

# 세션에 저장된 메시지를 항상 표시하는 로직 추가 (수정됨)
if "batch_apply_messages" in st.session_state and st.session_state["batch_apply_messages"]:
    messages = st.session_state["batch_apply_messages"]
    
    # 메시지를 종류별로 분리
    summary_msg = ""
    error_details = []
    warning_details = []
    info_msgs = []

    for msg_type, msg_text in messages:
        if msg_type == 'success':
            summary_msg = msg_text
        elif msg_type == 'error':
            error_details.append(f"• {msg_text[2:]}") # 이모티콘 제거 후 리스트 아이템으로 추가
        elif msg_type == 'warning':
            warning_details.append(f"• {msg_text[2:]}")
        elif msg_type == 'info':
            info_msgs.append(msg_text)

    # 'info' 타입의 메시지는 그대로 표시
    for msg in info_msgs:
        st.info(msg)

    # 요약 및 상세 로그 표시
    if summary_msg or error_details or warning_details:
        if summary_msg:
            st.success(summary_msg)
        
        # '적용됨' 로그는 batch_change_log에서 가져옴
        success_log = st.session_state.get("swapped_assignments_log", [])
        success_details = [
            f"• {log['날짜']}: {log['변경 전 인원']} ➡️ {log['변경 후 인원']}으로 변경되었습니다."
            for log in success_log
        ]

        with st.expander("🔍 스케줄 변경 상세 로그 보기", expanded=True):
            st.write("**⛔️ 요청사항 적용 불가**")
            error_text = "\n".join(sorted(error_details)) if error_details else "해당 없음"
            st.code(error_text, language='text')
            
            st.divider()
            
            st.write("**⚠️ 요청사항 적용 건너뜀**")
            warning_text = "\n".join(sorted(warning_details)) if warning_details else "해당 없음"
            st.code(warning_text, language='text')
            
            st.divider()
            
            st.write("**✅ 요청사항 적용됨**")
            success_text = "\n".join(sorted(success_details)) if success_details else "해당 없음"
            st.code(success_text, language='text')

    # 메시지를 표시한 후 세션 상태에서 제거
    del st.session_state["batch_apply_messages"]

# 데이터 에디터 UI
edited_df_md = st.data_editor(df_to_display, use_container_width=True, key="schedule_editor", disabled=['날짜', '요일'])

# --- 실시간 변경사항 로그 ---
st.write(" ")
st.caption("📝 변경사항 미리보기")

# 1. 수동 변경사항 계산
base_df_for_manual_diff = st.session_state.get("df_schedule_md_modified", st.session_state.get("df_schedule_md_initial"))
manual_change_log = []
oncall_warning_messages = []
if not edited_df_md.equals(base_df_for_manual_diff):
    diff_indices = np.where(edited_df_md.ne(base_df_for_manual_diff))
    for row_idx, col_idx in zip(diff_indices[0], diff_indices[1]):
        date_str_raw = edited_df_md.iloc[row_idx, 0]
        col_name = edited_df_md.columns[col_idx]
        old_value = base_df_for_manual_diff.iloc[row_idx, col_idx]
        new_value = edited_df_md.iloc[row_idx, col_idx]
        try:
            original_row = st.session_state["df_schedule_original"][st.session_state["df_schedule_original"]['날짜'] == date_str_raw].iloc[0]
            weekday = original_row['요일']
        except IndexError:
            weekday = ''
        if col_name == '오전당직(온콜)':
            time_period = '오전당직'
            
            # --- ▼▼▼ [경고 메시지 생성 로직 추가] ▼▼▼ ---
            old_val_str = str(old_value).strip()
            new_val_str = str(new_value).strip()
            
            # 1. 값이 둘 다 있는 경우 (A -> B)
            if old_val_str and new_val_str:
                msg = f"• {date_str_raw}: '{old_val_str}' 님의 오전당직 누적 -1, '{new_val_str}' 님의 누적 +1"
                oncall_warning_messages.append(msg)
            # 2. 값 B가 비어있는 경우 (A -> 빈 값)
            elif old_val_str:
                msg = f"• {date_str_raw}: '{old_val_str}' 님의 오전당직 누적 -1"
                oncall_warning_messages.append(msg)
            # 3. 값 A가 비어있는 경우 (빈 값 -> B)
            elif new_val_str:
                msg = f"• {date_str_raw}: '{new_val_str}' 님의 오전당직 누적 +1"
                oncall_warning_messages.append(msg)
            # --- ▲▲▲ [추가 완료] ▲▲▲ ---

        elif col_name.startswith('오후'):
            time_period = '오후'
        else:
            time_period = '오전'
        # ▲▲▲ [수정 완료] ▲▲▲

        formatted_date_str = f"{date_str_raw} ({weekday.replace('요일', '')}) - {time_period}"
        manual_change_log.append({
            '날짜': formatted_date_str, 
            '변경 전 인원': str(old_value), 
            '변경 후 인원': str(new_value)
        })

# 2. 일괄 적용 로그와 수동 변경 로그를 합쳐서 표시
batch_log = st.session_state.get("swapped_assignments_log", [])
st.session_state["final_change_log"] = batch_log + manual_change_log

if st.session_state["final_change_log"]:
    log_df = pd.DataFrame(st.session_state["final_change_log"])
    st.dataframe(log_df, use_container_width=True, hide_index=True)
else:
    st.info("기록된 변경사항이 없습니다.")

# --- ▼▼▼ [경고 메시지 표시 로직 추가] (L1448 다음 줄) ▼▼▼ ---
if oncall_warning_messages:
    # 리스트의 중복을 제거하고 날짜순으로 정렬
    sorted_warnings = sorted(list(set(oncall_warning_messages)))
    
    # [수정] 경고 메시지에 안내 문구 추가
    warning_text = (
        "🔔 **오전당직 누적 수치 변경 알림**\n\n" +
        "\n".join(sorted_warnings) +
        "\n\n(하단 '방배정 수행' 버튼을 누르면 이 누적 수치가 최종 저장됩니다.)"
    )
    st.warning(warning_text)
# --- ▲▲▲ [추가 완료] ▲▲▲

# --- 기존 '변경사항 저장' 버튼 관련 코드를 아래 블록으로 교체하세요 ---
# --- 1. [수정] 버튼을 표시하기 전에 다음 버전 번호를 미리 계산 ---
loaded_version = st.session_state.get("loaded_version", "")
version_match = re.search(r'ver\s*(\d+)\.\d+', loaded_version)
if version_match:
    current_major_version = int(version_match.group(1))
    next_major_version = current_major_version + 1
else:
    # 기본 시트에서 작업한 경우, 다음 버전은 2.0
    next_major_version = 2

# 방 설정 UI
st.divider()
st.subheader("⚙️ 방 설정")

special_schedules = []
special_dates = set()
special_df = pd.DataFrame(columns=["날짜", "근무", "당직"])

tab_weekday, tab_weekend = st.tabs(["평일 방 설정", "토요/휴일 방 설정"])

with tab_weekday:
    room_options = [str(i) for i in range(1, 13)]

    tab830, tab900, tab930, tab1000, tab1330 = st.tabs([
        "🕘 08:30", "🕘 09:00", "🕤 09:30", "🕙 10:00", "🕜 13:30 (오후)"
    ])
    with tab830:
        col1, col2 = st.columns([1, 2.5])
        with col1:
            st.markdown("###### **방 개수**")
            num_830 = st.number_input("830_rooms_count", min_value=0, max_value=12, value=4, key="830_rooms", label_visibility="collapsed")
            st.markdown("###### **오전 당직방**")
            duty_830_options = st.session_state["room_settings"]["830_room_select"]
            try:
                duty_index_830 = duty_830_options.index(st.session_state["room_settings"].get("830_duty"))
            except ValueError:
                duty_index_830 = 0
            duty_830 = st.selectbox("830_duty_room", duty_830_options, index=duty_index_830, key="830_duty", label_visibility="collapsed", help="8:30 시간대의 당직 방을 선택합니다.")
            st.session_state["room_settings"]["830_duty"] = duty_830
        with col2:
            st.markdown("###### **방 번호 선택**")
            if len(st.session_state["room_settings"]["830_room_select"]) > num_830:
                st.session_state["room_settings"]["830_room_select"] = st.session_state["room_settings"]["830_room_select"][:num_830]
            rooms_830 = st.multiselect("830_room_select_numbers", room_options, default=st.session_state["room_settings"]["830_room_select"], max_selections=num_830, key="830_room_select", label_visibility="collapsed")
            if len(rooms_830) < num_830:
                st.warning(f"방 번호를 {num_830}개 선택해주세요.")
            st.session_state["room_settings"]["830_room_select"] = rooms_830
    with tab900:
        col1, col2 = st.columns([1, 2.5])
        with col1:
            st.markdown("###### **방 개수**")
            num_900 = st.number_input("900_rooms_count", min_value=0, max_value=12, value=3, key="900_rooms", label_visibility="collapsed")
        with col2:
            st.markdown("###### **방 번호 선택**")
            if len(st.session_state["room_settings"]["900_room_select"]) > num_900:
                st.session_state["room_settings"]["900_room_select"] = st.session_state["room_settings"]["900_room_select"][:num_900]
            rooms_900 = st.multiselect("900_room_select_numbers", room_options, default=st.session_state["room_settings"]["900_room_select"], max_selections=num_900, key="900_room_select", label_visibility="collapsed")
            if len(rooms_900) < num_900:
                st.warning(f"방 번호를 {num_900}개 선택해주세요.")
            st.session_state["room_settings"]["900_room_select"] = rooms_900
    with tab930:
        col1, col2 = st.columns([1, 2.5])
        with col1:
            st.markdown("###### **방 개수**")
            num_930 = st.number_input("930_rooms_count", min_value=0, max_value=12, value=3, key="930_rooms", label_visibility="collapsed")
        with col2:
            st.markdown("###### **방 번호 선택**")
            if len(st.session_state["room_settings"]["930_room_select"]) > num_930:
                st.session_state["room_settings"]["930_room_select"] = st.session_state["room_settings"]["930_room_select"][:num_930]
            rooms_930 = st.multiselect("930_room_select_numbers", room_options, default=st.session_state["room_settings"]["930_room_select"], max_selections=num_930, key="930_room_select", label_visibility="collapsed")
            if len(rooms_930) < num_930:
                st.warning(f"방 번호를 {num_930}개 선택해주세요.")
            st.session_state["room_settings"]["930_room_select"] = rooms_930
    with tab1000:
        col1, col2 = st.columns([1, 2.5])
        with col1:
            st.markdown("###### **방 개수**")
            num_1000 = st.number_input("1000_rooms_count", min_value=0, max_value=12, value=2, key="1000_rooms", label_visibility="collapsed")
        with col2:
            st.markdown("###### **방 번호 선택**")
            if len(st.session_state["room_settings"]["1000_room_select"]) > num_1000:
                st.session_state["room_settings"]["1000_room_select"] = st.session_state["room_settings"]["1000_room_select"][:num_1000]
            rooms_1000 = st.multiselect("1000_room_select_numbers", room_options, default=st.session_state["room_settings"]["1000_room_select"], max_selections=num_1000, key="1000_room_select", label_visibility="collapsed")
            if len(rooms_1000) < num_1000:
                st.warning(f"방 번호를 {num_1000}개 선택해주세요.")
            st.session_state["room_settings"]["1000_room_select"] = rooms_1000
    with tab1330:
        col1, col2 = st.columns([1, 2.5])
        with col1:
            st.markdown("###### **방 개수**")
            st.info("4개 고정")
            num_1330 = 4
            st.markdown("###### **오후 당직방**")
            duty_1330_options = st.session_state["room_settings"]["1330_room_select"]
            try:
                duty_index_1330 = duty_1330_options.index(st.session_state["room_settings"].get("1330_duty"))
            except ValueError:
                duty_index_1330 = 0
            duty_1330 = st.selectbox("1330_duty_room", duty_1330_options, index=duty_index_1330, key="1330_duty", label_visibility="collapsed", help="13:30 시간대의 당직 방을 선택합니다.")
            st.session_state["room_settings"]["1330_duty"] = duty_1330
        with col2:
            st.markdown("###### **방 번호 선택**")
            if len(st.session_state["room_settings"]["1330_room_select"]) > num_1330:
                st.session_state["room_settings"]["1330_room_select"] = st.session_state["room_settings"]["1330_room_select"][:num_1330]
            rooms_1330 = st.multiselect("1330_room_select_numbers", room_options, default=st.session_state["room_settings"]["1330_room_select"], max_selections=num_1330, key="1330_room_select", label_visibility="collapsed")
            if len(rooms_1330) < num_1330:
                st.warning(f"방 번호를 {num_1330}개 선택해주세요.")
            st.session_state["room_settings"]["1330_room_select"] = rooms_1330

with tab_weekend:
    # 페이지 로드 시 session_state에 저장된 데이터를 사용합니다.
    special_schedules = st.session_state.get("special_schedules", [])
    special_df = st.session_state.get("special_df", pd.DataFrame())
    
    # --- 토요/휴일 UI 렌더링 (이하 로직은 기존과 거의 동일) ---
    if special_schedules:
        for date_obj, date_str, personnel_for_day in sorted(special_schedules):
            weekday_map = {5: "토", 6: "일"}
            weekday_str = weekday_map.get(date_obj.weekday(), '휴')
            
            duty_person_for_date = ""
            if not special_df.empty:
                duty_row = special_df[special_df['날짜_dt'].dt.date == date_obj]
                if not duty_row.empty: 
                    duty_person_for_date = str(duty_row['당직'].iloc[0]).strip()

            expander_title = (f"🗓️ {date_str} ({weekday_str}) | "
                              f"근무: {len(personnel_for_day)}명 | "
                              f"당직: {duty_person_for_date or '없음'}")

            with st.expander(expander_title):
                col1, col2 = st.columns([1, 1])
                duty_room = None
                with col1:
                    st.markdown("###### **당직 방**")
                    if duty_person_for_date and duty_person_for_date != "당직 없음":
                        duty_room_options = ["선택 안 함"] + [str(i) for i in range(1, 13)]
                        default_duty_room = st.session_state.weekend_room_settings.get(date_str, {}).get("duty_room", "1")
                        duty_room = st.selectbox("당직 방 선택", duty_room_options, key=f"duty_room_{date_str}", 
                                                 index=duty_room_options.index(default_duty_room) if default_duty_room in duty_room_options else 0, label_visibility="collapsed")
                    else: 
                        st.info("당직 인원 없음")
                
                with col2:
                    st.markdown("###### **총 방 개수**")
                    default_room_count = st.session_state.weekend_room_settings.get(date_str, {}).get("total_room_count", len(personnel_for_day))
                    total_room_count = st.number_input("총 방 개수", min_value=0, max_value=12, value=default_room_count, 
                                                       key=f"total_rooms_{date_str}", label_visibility="collapsed")
                
                st.markdown("###### **방 번호 선택**")
                room_options = [str(i) for i in range(1, 13)]
                default_rooms = st.session_state.weekend_room_settings.get(date_str, {}).get("selected_rooms", room_options[:total_room_count])
                selected_rooms = st.multiselect("방 번호 선택", room_options, default=default_rooms, max_selections=total_room_count, 
                                                key=f"rooms_{date_str}", label_visibility="collapsed")

                st.session_state.weekend_room_settings[date_str] = {
                    "duty_room": duty_room if duty_room and duty_room != "선택 안 함" else None,
                    "total_room_count": total_room_count, "selected_rooms": selected_rooms
                }
    else: 
        st.info("이번 달은 토요/휴일 근무가 없습니다.")
        
all_selected_rooms = (st.session_state["room_settings"]["830_room_select"] + 
                     st.session_state["room_settings"]["900_room_select"] + 
                     st.session_state["room_settings"]["930_room_select"] + 
                     st.session_state["room_settings"]["1000_room_select"] + 
                     st.session_state["room_settings"]["1330_room_select"])

# 배정 요청 입력 UI
st.divider()
st.subheader("📋 배정 요청 관리")
st.write("- 모든 인원의 배정 요청을 추가 및 수정할 수 있습니다.")
요청분류 = ["1번방", "2번방", "3번방", "4번방", "5번방", "6번방", "7번방", "8번방", "9번방", "10번방", "11번방", "12번방", 
            "8:30", "9:00", "9:30", "10:00", "당직 아닌 이른방", "이른방 제외", "늦은방 제외", "오후 당직 제외"]

st.write(" ")
st.markdown("**🙋‍♂️ 현재 방배정 요청 목록**")
if st.session_state["df_room_request"].empty:
    st.info("☑️ 현재 방배정 요청이 없습니다.")
else:
    st.dataframe(st.session_state["df_room_request"], use_container_width=True, hide_index=True)


st.write(" ")

# 기존 save_to_gsheet 함수를 이 코드로 교체하세요.
def save_to_gsheet(name, categories, selected_save_dates, month_str, worksheet):
    try:
        if not name or not categories or not selected_save_dates:
            return None, "input_error"

        # [수정] 시트에서 직접 최신 데이터를 읽어 중복 검사
        all_requests = worksheet.get_all_records()
        df_live_requests = pd.DataFrame(all_requests)
        
        new_requests_to_append = []
        is_duplicate = False

        for category in categories:
            for date in selected_save_dates:
                date = date.strip()
                # 라이브 데이터로 중복 확인
                if not df_live_requests[(df_live_requests['이름'] == name) &
                                        (df_live_requests['날짜정보'] == date) &
                                        (df_live_requests['분류'] == category)].empty:
                    is_duplicate = True
                    continue # 중복이면 추가 목록에 넣지 않음
                
                new_requests_to_append.append([name, category, date])

        if not new_requests_to_append:
            # 추가할 요청은 없는데 중복이 발견된 경우
            return st.session_state["df_room_request"], "duplicate"

        # [수정] append_rows로 안전하게 새 요청만 추가
        worksheet.append_rows(new_requests_to_append, value_input_option='USER_ENTERED')
        
        # 성공 후 최신 데이터 다시 로드하여 반환
        updated_df = pd.DataFrame(worksheet.get_all_records())
        return updated_df, "success"

    except Exception as e:
        st.error(f"요청 추가 중 오류 발생: {type(e).__name__} - {str(e)}")
        return None, "error"

st.markdown("**🟢 방배정 요청 추가**")

# Reset flag to control form clearing
if "reset_form" not in st.session_state:
    st.session_state.reset_form = False

# Clear widget states on reset
if st.session_state.reset_form:
    for key in ["add_name", "add_categories", "add_dates", "add_time"]:
        if key in st.session_state:
            del st.session_state[key]
    st.session_state.reset_form = False  # Reset the flag after clearing

# --- UI 위젯 정의 ---
col1, col2, col3, col_button_add = st.columns([2, 2, 4, 1])

with col1:
    names = sorted([str(name).strip() for name in st.session_state["df_schedule"].iloc[:, 2:].stack().dropna().unique() if str(name).strip()])
    selected_name = st.selectbox(
        "근무자 선택",  # ✅ label을 직접 사용
        names,
        key="add_name",
        index=None,
        placeholder="근무자 선택",
    )

with col2:
    selected_categories = st.multiselect(
        "요청 분류", # ✅ label을 직접 사용
        요청분류,
        key="add_categories",
        default=[],
    )

with col3:
    processed_dates = {}
    date_to_obj_map = {}
    if st.session_state.get("add_name"):
        st.cache_data.clear()
        available_dates = get_user_available_dates(st.session_state.add_name, st.session_state["df_schedule"], this_month_start, this_month_end, month_str)
        for display_str, save_str in available_dates:
            parts = display_str.split(' ')
            date_part, time_part = ' '.join(parts[:-1]), parts[-1]
            if date_part not in processed_dates:
                processed_dates[date_part] = {}
                date_obj_str = save_str.split(' ')[0]
                date_to_obj_map[date_part] = datetime.strptime(date_obj_str, '%Y-%m-%d')
            processed_dates[date_part][time_part] = save_str
    
    date_options = sorted(processed_dates.keys(), key=lambda k: date_to_obj_map.get(k, datetime.max))
    
    sub_col_date, sub_col_time = st.columns([3, 1.5])
    with sub_col_date:
        selected_dates = st.multiselect(
            "요청 일자", # ✅ label을 직접 사용
            date_options,
            key="add_dates",
            default=[],
        )

    with sub_col_time:
        time_options = ["오전", "오후"]
        selected_time = st.selectbox(
            "시간대", # ✅ label을 직접 사용
            time_options,
            key="add_time",
            index=None,
        )

with col_button_add:
    st.markdown("<div>&nbsp;</div>", unsafe_allow_html=True)
    add_button_clicked = st.button("📅 추가")

# 메시지 출력을 버튼 아래로
st.write(" ")  # 버튼과 메시지 사이 공백

# 버튼 클릭 처리
if add_button_clicked:
    name_to_add = st.session_state.get("add_name")
    categories_to_add = st.session_state.get("add_categories", [])
    dates_to_add = st.session_state.get("add_dates", [])
    time_to_add = st.session_state.get("add_time")

    if not name_to_add or not categories_to_add or not dates_to_add:
        st.session_state.add_request_status = "input_error"
    else:
        selected_save_dates = []
        if name_to_add:
            for date_display in dates_to_add:
                if date_display in processed_dates and time_to_add in processed_dates[date_display]:
                    selected_save_dates.append(processed_dates[date_display][time_to_add])
        
        if not selected_save_dates:
            st.session_state.add_request_status = "no_slot_error"
        else:
            with st.spinner("요청을 기록중입니다..."):
                df_room_request, status = save_to_gsheet(name_to_add, categories_to_add, selected_save_dates, month_str, st.session_state["worksheet_room_request"])
                st.session_state.add_request_status = status
                if df_room_request is not None:
                    st.session_state["df_room_request"] = df_room_request

# 메시지 출력
if "add_request_status" in st.session_state:
    status = st.session_state.add_request_status
    if status == "success":
        st.success("요청이 성공적으로 추가되었습니다.")
        st.session_state.reset_form = True
        time.sleep(2)  # 메시지 표시를 위해 2초 대기
    elif status == "duplicate":
        st.warning("이미 존재하는 요청사항입니다.")
    elif status == "input_error":
        st.error("근무자, 요청 분류, 요청 일자를 모두 선택해주세요.")
    elif status == "no_slot_error":
        st.warning("선택하신 날짜에 해당하는 근무 시간대가 없습니다.")
    
    # 상태 초기화 및 성공 시 새로고침
    del st.session_state.add_request_status
    if status == "success":
        st.rerun()

st.write(" ")
st.markdown("**🔴 방배정 요청 삭제**")
if not st.session_state["df_room_request"].empty:
    col0, col1, col_button_del = st.columns([2, 6, 1])
    with col0:
        unique_names = st.session_state["df_room_request"]["이름"].unique()
        selected_employee = st.selectbox("근무자 선택", unique_names, key="delete_request_employee_select", index=None, placeholder="근무자 선택")
    with col1:
        selected_items = []
        if selected_employee:
            df_request_filtered = st.session_state["df_room_request"][st.session_state["df_room_request"]["이름"] == selected_employee]
            if not df_request_filtered.empty:
                options = [f"{row['분류']} - {row['날짜정보']}" for _, row in df_request_filtered.iterrows()]
                selected_items = st.multiselect("삭제할 항목", options, key="delete_request_select")
            else:
                st.multiselect("삭제할 항목", [], disabled=True, key="delete_request_select", help="해당 근무자의 요청이 없습니다.")
        else:
            st.multiselect("삭제할 항목", [], key="delete_request_select", disabled=True)
    with col_button_del:
        st.markdown("<div>&nbsp;</div>", unsafe_allow_html=True)
        delete_button_clicked = st.button("📅 삭제", key="request_delete_button")
    # '방배정 요청 삭제'의 if delete_button_clicked: 블록 전체를 이 코드로 교체하세요.
    if delete_button_clicked:
        if not selected_employee or not selected_items:
            st.error("삭제할 근무자와 항목을 선택해주세요.")
        else:
            with st.spinner("요청을 삭제하는 중입니다..."):
                try:
                    worksheet = st.session_state["worksheet_room_request"]
                    all_requests = worksheet.get_all_records()
                    
                    # 삭제할 항목 정보를 set으로 만들어 빠른 조회 가능
                    items_to_delete_set = set(selected_items)
                    
                    # 삭제할 행의 인덱스를 뒤에서부터 찾아서 기록 (삭제 시 인덱스 밀림 방지)
                    rows_to_delete_indices = []
                    for i, record in reversed(list(enumerate(all_requests))):
                        record_str = f"{record.get('분류')} - {record.get('날짜정보')}"
                        if record.get('이름') == selected_employee and record_str in items_to_delete_set:
                            rows_to_delete_indices.append(i + 2) # gspread는 1-based, 헤더 포함
                    
                    # 찾은 행들을 삭제
                    if rows_to_delete_indices:
                        for row_idx in rows_to_delete_indices:
                            worksheet.delete_rows(row_idx)

                    st.cache_data.clear()
                    st.success("요청사항이 삭제되었습니다.")
                    time.sleep(1.5)
                    st.rerun()

                except Exception as e:
                    st.error(f"요청 삭제 중 오류 발생: {type(e).__name__} - {e}")
else:
    st.info("📍 방배정 요청이 없습니다.")

def parse_date_info(date_info):
    try:
        date_part = date_info.split('(')[0].strip()
        date_obj = datetime.strptime(date_part, '%Y-%m-%d')
        is_morning = '오전' in date_info
        parsed_date = date_obj.strftime('%Y-%m-%d')
        return parsed_date, is_morning
    except ValueError as e:
        st.warning(f"Failed to parse date_info: {date_info}, error: {str(e)}")
        return None, False

# 🔼 기존 assign_special_date 함수를 지우고 아래 코드로 교체하세요.

def assign_special_date(personnel_for_day, date_str, formatted_date, settings, special_df_for_month, df_room_request):
    """
    [수정된 함수]
    토요/휴일의 방배정을 수행합니다.
    - 1순위: 당직자 배정
    - 2순위: '방 지정 요청'이 있는 인원 배정
    - 3순위: 나머지 인원 랜덤 배정
    """
    assignment_dict = {}
    assigned_personnel = set()
    
    duty_room = settings.get("duty_room", None)
    selected_rooms = settings.get("selected_rooms", [])
    
    # 선호도 순서에 따라 선택된 방을 정렬
    preferred_room_order = ['1', '8', '4', '7', '10', '2', '5', '6', '9', '3']
    sorted_rooms = [room for room in preferred_room_order if room in selected_rooms]
    
    duty_person = None
    if not special_df_for_month.empty:
        try:
            target_year = int(month_str.split('년')[0])
            date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=target_year).date()
            duty_person_row = special_df_for_month[special_df_for_month['날짜_dt'].dt.date == date_obj]
            if not duty_person_row.empty:
                duty_person = duty_person_row['당직'].iloc[0]
        except Exception as e:
            st.warning(f"당직자 정보 조회 중 오류: {e}")

    # 1. 당직 인원 우선 배정
    if duty_person and duty_person in personnel_for_day and duty_room and duty_room != "선택 안 함":
        if duty_person not in assigned_personnel:
            assignment_dict[f"방({duty_room})"] = duty_person
            assigned_personnel.add(duty_person)

    # 2. 방 지정 요청 처리 (새로 추가된 로직)
    if not df_room_request.empty:
        # 현재 날짜(오전)에 해당하는 요청만 필터링 (토요/휴일 근무는 모두 '오전'으로 간주)
        requests_for_day = df_room_request[
            df_room_request['날짜정보'].str.startswith(formatted_date)
        ]

        for _, req in requests_for_day.iterrows():
            person = req['이름']
            category = req['분류']  # 예: "1번방"

            # 요청자가 오늘 근무자이고 아직 배정되지 않았는지 확인
            if person in personnel_for_day and person not in assigned_personnel:
                room_match = re.match(r'(\d+)번방', category)
                if room_match:
                    req_room_num = room_match.group(1)
                    slot_key = f"방({req_room_num})"

                    # 요청한 방이 오늘 운영되는 방이고, 아직 비어있는지 확인
                    if req_room_num in selected_rooms and slot_key not in assignment_dict:
                        assignment_dict[slot_key] = person
                        assigned_personnel.add(person)

    # 3. 나머지 인원을 랜덤 배정
    remaining_personnel = [p for p in personnel_for_day if p not in assigned_personnel]
    random.shuffle(remaining_personnel)
    
    # 배정되지 않은 방 목록
    unassigned_rooms = [r for r in sorted_rooms if f"방({r})" not in assignment_dict]

    for room in unassigned_rooms:
        if remaining_personnel:
            person = remaining_personnel.pop(0)
            assignment_dict[f"방({room})"] = person
            assigned_personnel.add(person)
    
    return assignment_dict, sorted_rooms

from collections import Counter
import random
import streamlit as st

def random_assign(personnel, slots, request_assignments, time_groups, total_stats, morning_personnel, afternoon_personnel, afternoon_duty_counts):
    assignment = [None] * len(slots)
    assigned_personnel_morning = set()
    assigned_personnel_afternoon = set()
    daily_stats = {
        'early': Counter(),
        'late': Counter(),
        'morning_duty': Counter(),
        'afternoon_duty': Counter(),
        'rooms': {str(i): Counter() for i in range(1, 13)},
        'time_room_slots': {}  # 시간대-방 쌍에 대한 Counter 객체를 딕셔너리로 관리
    }

    # time_room_slots 초기화
    for slot in slots:
        daily_stats['time_room_slots'][slot] = Counter()

    # total_stats['time_room_slots'] 초기화 (외부 코드 수정 없이 함수 내에서 처리)
    if 'time_room_slots' not in total_stats:
        total_stats['time_room_slots'] = {}
    for slot in slots:
        total_stats['time_room_slots'].setdefault(slot, Counter())

    morning_slots = [s for s in slots if s.startswith(('8:30', '9:00', '9:30', '10:00')) and '_당직' not in s]
    afternoon_slots = [s for s in slots if s.startswith('13:30')]
    afternoon_duty_slot = [s for s in slots if s.startswith('13:30') and s.endswith('_당직')]

    # 요청된 배정 처리
    for slot, person in request_assignments.items():
        if person in personnel and slot in slots:
            slot_idx = slots.index(slot)
            if assignment[slot_idx] is None:
                if (slot in morning_slots and person in morning_personnel) or \
                   (slot in afternoon_slots and person in afternoon_personnel):
                    if slot in morning_slots and person in assigned_personnel_morning:
                        st.warning(f"중복 배정 방지: {person}은 이미 오전 시간대({slot})에 배정됨")
                        continue
                    if slot in afternoon_slots and person in assigned_personnel_afternoon:
                        st.warning(f"중복 배정 방지: {person}은 이미 오후 시간대({slot})에 배정됨")
                        continue

                    assignment[slot_idx] = person
                    if slot in morning_slots:
                        assigned_personnel_morning.add(person)
                    else:
                        assigned_personnel_afternoon.add(person)
                    room_num = slot.split('(')[1].split(')')[0]
                    daily_stats['rooms'][room_num][person] += 1
                    daily_stats['time_room_slots'][slot][person] += 1
                    if slot.startswith('8:30') and '_당직' not in slot:
                        daily_stats['early'][person] += 1
                    elif slot.startswith('10:00'):
                        daily_stats['late'][person] += 1
                    if slot.startswith('8:30') and slot.endswith('_당직'):
                        daily_stats['morning_duty'][person] += 1
                    elif slot.startswith('13:30') and slot.endswith('_당직'):
                        daily_stats['afternoon_duty'][person] += 1
                else:
                    st.warning(f"{date_str}({slot}): {person}님의 방배정 요청 무시됨: 해당 시간대({'오전' if slot in morning_slots else '오후'})에 근무하지 않습니다.")
            else:
                st.warning(f"배정 요청 충돌: {person}을 {date_str}({slot})에 배정할 수 없음. 이미 배정됨: {assignment[slot_idx]}")

    # 오후 당직 배정
    afternoon_duty_slot_idx = slots.index(afternoon_duty_slot[0]) if afternoon_duty_slot else None
    if afternoon_duty_slot_idx is not None and assignment[afternoon_duty_slot_idx] is None:
        
        # 1. 후보자는 배정되지 않은 모든 오후 근무자입니다.
        candidates = [p for p in afternoon_personnel if p not in assigned_personnel_afternoon]
        
        # [수정] '오후 당직 제외' 요청이 있는 인원을 후보에서 제외합니다.
        if (date_str, afternoon_duty_slot[0]) in request_assignments:
            excluded_person = request_assignments[(date_str, afternoon_duty_slot[0])]
            if excluded_person in candidates:
                candidates.remove(excluded_person)
                
        if candidates:
            # [수정] best_person, min_score 초기화 제거
            
            # 1. 모든 후보자의 '실시간 누적 점수'를 계산합니다.
            scores = {}
            for person in candidates:
                last_month_cumulative = afternoon_duty_counts.get(person, 0)
                this_month_so_far = total_stats['afternoon_duty'][person]
                scores[person] = last_month_cumulative + this_month_so_far

            # [수정] 'scores' 딕셔너리가 비어있지 않은지(True) 확인한 후에 min 값을 찾습니다.
            if scores:
                # 2. 가장 낮은 '실시간 누적 점수'를 찾습니다.
                min_score = min(scores.values())
                
                # 3. 가장 낮은 점수를 가진 모든 후보자를 리스트로 만듭니다. (동점자 처리)
                best_candidates = [person for person, score in scores.items() if score == min_score]
                
                # 4. 동점자들 중에서 무작위로 1명을 선택하여 편향을 제거합니다.
                if best_candidates:
                    best_person = random.choice(best_candidates)
                    
                    if best_person:
                        assignment[afternoon_duty_slot_idx] = best_person
                        assigned_personnel_afternoon.add(best_person)
                        room_num = afternoon_duty_slot[0].split('(')[1].split(')')[0]
                        daily_stats['rooms'][room_num][best_person] += 1
                        daily_stats['time_room_slots'][afternoon_duty_slot[0]][best_person] += 1
                        daily_stats['afternoon_duty'][best_person] += 1

    # 오전 슬롯 배정
    morning_remaining = [p for p in morning_personnel if p not in assigned_personnel_morning]
    afternoon_remaining = [p for p in afternoon_personnel if p not in assigned_personnel_afternoon]
    remaining_slots = [i for i, a in enumerate(assignment) if a is None]
    
    morning_slot_indices = [i for i in remaining_slots if slots[i] in morning_slots]
    random.shuffle(morning_remaining)
    while morning_remaining and morning_slot_indices:
        best_person = None
        best_slot_idx = None
        min_score = float('inf')
        
        for slot_idx in morning_slot_indices:
            if assignment[slot_idx] is not None:
                continue
            slot = slots[slot_idx]
            
            for person in morning_remaining:
                time_room_count = total_stats['time_room_slots'][slot][person] + \
                                  daily_stats['time_room_slots'][slot][person]
                if slot.startswith('8:30') and '_당직' not in slot:
                    early_count = total_stats['early'][person] + daily_stats['early'][person]
                    score = early_count * 100 + time_room_count
                elif slot.startswith('10:00'):
                    late_count = total_stats['late'][person] + daily_stats['late'][person]
                    score = 10000 + late_count * 100 + time_room_count
                else:
                    score = 20000 + time_room_count
                
                if score < min_score:
                    min_score = score
                    best_person = person
                    best_slot_idx = slot_idx
        
        if best_slot_idx is None or best_person is None:
            st.warning(f"오전 슬롯 배정 불가: 더 이상 배정 가능한 인원 없음")
            break
        
        slot = slots[best_slot_idx]
        assignment[best_slot_idx] = best_person
        assigned_personnel_morning.add(best_person)
        morning_remaining.remove(best_person)
        morning_slot_indices.remove(best_slot_idx)
        remaining_slots.remove(best_slot_idx)
        room_num = slot.split('(')[1].split(')')[0]
        daily_stats['rooms'][room_num][best_person] += 1
        daily_stats['time_room_slots'][slot][best_person] += 1
        if slot.startswith('8:30') and '_당직' not in slot:
            daily_stats['early'][best_person] += 1
        elif slot.startswith('10:00'):
            daily_stats['late'][best_person] += 1
        if slot.startswith('8:30') and slot.endswith('_당직'):
            daily_stats['morning_duty'][best_person] += 1

    # 오후 슬롯 배정
    afternoon_slot_indices = [i for i in remaining_slots if slots[i] in afternoon_slots]
    while afternoon_remaining and afternoon_slot_indices:
        best_person = None
        best_slot_idx = None
        min_score = float('inf')
        
        for slot_idx in afternoon_slot_indices:
            if assignment[slot_idx] is not None:
                continue
            slot = slots[slot_idx]
            
            for person in afternoon_remaining:
                time_room_count = total_stats['time_room_slots'][slot][person] + \
                                  daily_stats['time_room_slots'][slot][person]
                duty_count = total_stats['afternoon_duty'][person] + daily_stats['afternoon_duty'][person] if slot.endswith('_당직') else float('inf')
                if slot.endswith('_당직'):
                    score = duty_count * 100 + time_room_count
                else:
                    score = time_room_count
                
                if score < min_score:
                    min_score = score
                    best_person = person
                    best_slot_idx = slot_idx
        
        if best_slot_idx is None or best_person is None:
            st.warning(f"오후 슬롯 배정 불가: 더 이상 배정 가능한 인원 없음")
            break
        
        slot = slots[best_slot_idx]
        assignment[best_slot_idx] = best_person
        assigned_personnel_afternoon.add(best_person)
        afternoon_remaining.remove(best_person)
        afternoon_slot_indices.remove(best_slot_idx)
        room_num = slot.split('(')[1].split(')')[0]
        daily_stats['rooms'][room_num][best_person] += 1
        daily_stats['time_room_slots'][slot][best_person] += 1
        if slot.endswith('_당직'):
            daily_stats['afternoon_duty'][best_person] += 1

    # 남은 빈 슬롯 처리
    for slot_idx in range(len(slots)):
        if assignment[slot_idx] is None:
            slot = slots[slot_idx]
            available_personnel = morning_personnel if slot in morning_slots else afternoon_personnel
            assigned_set = assigned_personnel_morning if slot in morning_slots else assigned_personnel_afternoon
            candidates = [p for p in available_personnel if p not in assigned_set]
            
            if candidates:
                room_num = slot.split('(')[1].split(')')[0]
                best_person = None
                min_score = float('inf')
                for person in candidates:
                    early_count = total_stats['early'][person] + daily_stats['early'][person] if slot.startswith('8:30') and '_당직' not in slot else float('inf')
                    late_count = total_stats['late'][person] + daily_stats['late'][person] if slot.startswith('10:00') else float('inf')
                    morning_duty_count = total_stats['morning_duty'][person] + daily_stats['morning_duty'][person] if slot.startswith('8:30') and slot.endswith('_당직') else float('inf')
                    afternoon_duty_count = total_stats['afternoon_duty'][person] + daily_stats['afternoon_duty'][person] if slot.startswith('13:30') and slot.endswith('_당직') else float('inf')
                    time_room_count = total_stats['time_room_slots'][slot][person] + \
                                      daily_stats['time_room_slots'][slot][person]
                    
                    if slot.startswith('8:30') and '_당직' not in slot:
                        score = early_count * 100 + time_room_count
                    elif slot.startswith('10:00'):
                        score = late_count * 100 + time_room_count
                    elif slot.startswith('8:30') and slot.endswith('_당직'):
                        score = morning_duty_count * 100 + time_room_count
                    elif slot.startswith('13:30') and slot.endswith('_당직'):
                        score = afternoon_duty_count * 100 + time_room_count
                    else:
                        score = time_room_count
                    
                    if score < min_score:
                        min_score = score
                        best_person = person
                
                person = best_person
                if slot in morning_slots:
                    assigned_personnel_morning.add(person)
                else:
                    assigned_personnel_afternoon.add(person)
                st.warning(f"슬롯 {slot} 공란 방지: {person} 배정 (스코어: {min_score})")
            else:
                available_personnel = morning_personnel if slot in morning_slots else afternoon_personnel
                if available_personnel:
                    room_num = slot.split('(')[1].split(')')[0]
                    best_person = None
                    min_score = float('inf')
                    for person in available_personnel:
                        early_count = total_stats['early'][person] + daily_stats['early'][person] if slot.startswith('8:30') and '_당직' not in slot else float('inf')
                        late_count = total_stats['late'][person] + daily_stats['late'][person] if slot.startswith('10:00') else float('inf')
                        morning_duty_count = total_stats['morning_duty'][person] + daily_stats['morning_duty'][person] if slot.startswith('8:30') and slot.endswith('_당직') else float('inf')
                        afternoon_duty_count = total_stats['afternoon_duty'][person] + daily_stats['afternoon_duty'][person] if slot.startswith('13:30') and slot.endswith('_당직') else float('inf')
                        time_room_count = total_stats['time_room_slots'][slot][person] + \
                                          daily_stats['time_room_slots'][slot][person]
                        
                        if slot.startswith('8:30') and '_당직' not in slot:
                            score = early_count * 100 + time_room_count
                        elif slot.startswith('10:00'):
                            score = late_count * 100 + time_room_count
                        elif slot.startswith('8:30') and slot.endswith('_당직'):
                            score = morning_duty_count * 100 + time_room_count
                        elif slot.startswith('13:30') and slot.endswith('_당직'):
                            score = afternoon_duty_count * 100 + time_room_count
                        else:
                            score = time_room_count
                        
                        if score < min_score:
                            min_score = score
                            best_person = person
                    
                    person = best_person
                    st.warning(f"슬롯 {slot} 공란 방지: 이미 배정된 {person} 재배정 (스코어: {min_score})")
                else:
                    st.warning(f"슬롯 {slot} 공란 방지 불가: 배정 가능한 인원 없음")
                    continue
            
            assignment[slot_idx] = person
            daily_stats['rooms'][room_num][person] += 1
            daily_stats['time_room_slots'][slot][person] += 1
            if slot.startswith('8:30') and '_당직' not in slot:
                daily_stats['early'][person] += 1
            elif slot.startswith('10:00'):
                daily_stats['late'][person] += 1
            if slot.startswith('8:30') and slot.endswith('_당직'):
                daily_stats['morning_duty'][person] += 1
            elif slot.startswith('13:30') and slot.endswith('_당직'):
                daily_stats['afternoon_duty'][person] += 1

    # total_stats 업데이트
    for key in ['early', 'late', 'morning_duty', 'afternoon_duty']:
        total_stats[key].update(daily_stats[key])
    for room in daily_stats['rooms']:
        total_stats['rooms'][room].update(daily_stats['rooms'][room])
    for slot in daily_stats['time_room_slots']:
        total_stats['time_room_slots'][slot].update(daily_stats['time_room_slots'][slot])

    return assignment, daily_stats

st.divider()
st.markdown(f"**➕ 이번달 배정에 반영될 누적 테이블(오후당직)**")

# 세션 상태에 저장된 원본 df_cumulative_original 데이터를 가져옵니다.
df_cumulative_original = st.session_state.get("df_cumulative_original", pd.DataFrame())

if not df_cumulative_original.empty:
    # 1. 표시할 데이터 필터링: '오후당직' 행만 선택
    df_to_display = df_cumulative_original[df_cumulative_original['항목'] == '오후당직누적'].copy()

    # 2. data_editor를 사용하여 편집 가능한 테이블을 만듭니다.
    edited_df_display = st.data_editor(
        df_to_display,
        use_container_width=True,
        hide_index=True,
        key="cumulative_editor",
        disabled=["항목"]  # '항목' 열은 편집 불가
    )

    if st.button("💾 누적 데이터 저장", type="primary"):
        with st.spinner("수정된 누적 데이터를 Google Sheets에 저장 중입니다..."):
            try:
                # 3. 저장할 전체 데이터 준비
                # 원본 데이터의 복사본을 만듭니다.
                df_to_save = df_cumulative_original.copy()
                
                # 원본에서 '오후당직누적' 행의 인덱스를 찾습니다.
                idx_to_update = df_to_save.index[df_to_save['항목'] == '오후당직누적'][0]

                # 해당 인덱스의 행을 편집된 내용으로 교체합니다.
                df_to_save.iloc[idx_to_update] = edited_df_display.iloc[0]

                # 4. Google Sheets에 저장
                gc = get_gspread_client()
                sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
                cumulative_sheet_name = st.session_state.get("latest_cumulative_name")

                if cumulative_sheet_name:
                    worksheet_to_update = sheet.worksheet(cumulative_sheet_name)
                    
                    data_to_save_list = [df_to_save.columns.tolist()] + df_to_save.fillna('').values.tolist()
                    
                    if update_sheet_with_retry(worksheet_to_update, data_to_save_list):
                        st.success(f"🎉 '{cumulative_sheet_name}' 시트가 성공적으로 업데이트되었습니다.")
                        # 변경사항 즉시 반영을 위해 세션 데이터도 업데이트
                        st.session_state["df_cumulative_original"] = df_to_save
                        df_transposed = df_to_save.set_index('항목')
                        st.session_state["df_cumulative"] = df_transposed.transpose().reset_index().rename(columns={'index': '이름'})
                        time.sleep(1.5)
                        st.rerun()
                    else:
                        st.error("Google Sheets 업데이트에 실패했습니다.")
                else:
                    st.error("업데이트할 누적 시트의 이름을 찾을 수 없습니다. 페이지를 새로고침 해주세요.")

            except Exception as e:
                st.error(f"저장 중 오류 발생: {e}")

# afternoon_duty_counts는 기존 로직(transpose된 df_cumulative)을 그대로 사용해야 하므로 이 부분은 수정하지 않습니다.
df_cumulative = st.session_state.get("df_cumulative", pd.DataFrame())
if df_cumulative.empty:
    st.error("❌ 방배정 실패: 누적 데이터를 불러올 수 없습니다.")
    st.stop()
    
afternoon_duty_counts = {row['이름']: int(row['오후당직누적']) for _, row in df_cumulative.iterrows() if pd.notna(row.get('오후당직누적'))}

# 1. 다음 달 문자열 계산
target_month_dt = datetime.strptime(month_str, "%Y년 %m월")
next_month_dt = target_month_dt + relativedelta(months=1)
next_month_str_for_check = next_month_dt.strftime("%Y년 %-m월")

# 2. 캐시된 함수를 호출하여 3개 시트의 존재 여부 확인
final_sheets_exist = check_final_sheets_exist(month_str, next_month_str_for_check)

st.divider()

if final_sheets_exist:
    st.warning(
        "⚠️ **덮어쓰기 경고**\n\n"
        "이미 Google Sheets에 다음달의 방배정 결과 시트가 존재합니다.\n\n"
        "배정을 다시 수행하면 '이어서 작업'되지 않으며, 현재 화면의 설정을 기준으로 **처음부터 다시 계산하여 기존 시트들을 덮어쓰기**합니다."
    )

if st.button("🚀 방배정 수행", type="primary", use_container_width=True):
    # base_df_for_diff 비교 대상 결정 (현재 화면 기준)
    base_df_for_diff = st.session_state.get("df_schedule_md_modified", st.session_state.get("df_schedule_md_initial"))
    # edited_df_md는 data_editor 결과 변수

    manual_change_log_for_save = []
    swapped_set_for_save = st.session_state.get("swapped_assignments", set())

    diff_indices = np.where(edited_df_md.ne(base_df_for_diff))
    for row_idx, col_idx in zip(diff_indices[0], diff_indices[1]):
        date_str_raw = edited_df_md.iloc[row_idx, 0]
        col_name = edited_df_md.columns[col_idx]
        old_value = base_df_for_diff.iloc[row_idx, col_idx]
        new_value = edited_df_md.iloc[row_idx, col_idx]
        try:
            original_row = st.session_state["df_schedule_original"][st.session_state["df_schedule_original"]['날짜'] == date_str_raw].iloc[0]
            weekday = original_row['요일']
        except IndexError:
            weekday = ''
        time_period = '오후' if col_name.startswith('오후') else '오전'
        formatted_date_str = f"{date_str_raw} ({weekday.replace('요일', '')}) - {time_period}"

        # 로그 기록 (기존과 동일)
        manual_change_log_for_save.append({'날짜': formatted_date_str, '변경 전 인원': str(old_value), '변경 후 인원': str(new_value)})

        swapped_set_for_save.add((date_str_raw.strip(), time_period, str(old_value).strip(), str(new_value).strip()))

    # "일괄 적용" 시 생성된 세트도 여기에 합쳐줍니다.
    batch_swapped_set = st.session_state.get("swapped_assignments", set()) # 이미 위에서 가져왔으므로 중복될 수 있으나 set이 처리
    swapped_set_for_save.update(batch_swapped_set)

    # 최종 변경 기록 세트를 세션에 저장 -> 엑셀 생성 시 사용됨
    st.session_state["swapped_assignments"] = swapped_set_for_save
    st.session_state["final_change_log_at_run"] = st.session_state.get("swapped_assignments_log", []) + manual_change_log_for_save

    # 버튼을 누를 때마다 항상 새로 계산하도록 이전 결과를 삭제합니다.
    if "assignment_results" in st.session_state:
        del st.session_state["assignment_results"]

    if "assignment_results" not in st.session_state or st.session_state.assignment_results is None:
        with st.spinner("방배정 중..."):
            # --- 요청사항 처리 결과 추적을 위한 초기화 ---
            applied_messages = []
            unapplied_messages = []
            weekday_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
            save_errors = []

            # 날짜 파싱 성능을 위해 근무일 정보 미리 생성
            work_days_map = {}
            target_year = int(month_str.split('년')[0])
            df_schedule_for_check = st.session_state.get("df_schedule_md_modified", st.session_state.get("df_schedule_md_initial"))

            for _, schedule_row in df_schedule_for_check.iterrows():
                date_str = schedule_row['날짜']
                try:
                    date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=target_year)
                    formatted_date = date_obj.strftime('%Y-%m-%d')
                    
                    morning_workers = set(p.strip() for p in schedule_row.iloc[2:13].dropna() if p and p.strip())
                    afternoon_workers = set(p.strip() for p in schedule_row.iloc[13:].dropna() if p and p.strip())
                    on_call_worker = str(schedule_row.get('오전당직(온콜)', '')).strip()

                    if on_call_worker:
                        morning_workers.add(on_call_worker)
                        afternoon_workers.add(on_call_worker)
                    
                    work_days_map[formatted_date] = {
                        "morning": morning_workers,
                        "afternoon": afternoon_workers,
                        "on_call": on_call_worker
                    }
                except (ValueError, TypeError):
                    continue

            # 유효성 검사를 통과한 요청만 실제 배정에 사용
            valid_requests_indices = []
            for index, req in st.session_state["df_room_request"].iterrows():
                req_date, is_morning = parse_date_info(req['날짜정보'])
                person = req['이름']
                category = req['분류']
                
                is_valid = True

                # 날짜 포맷팅 ('MM월 DD일(요일) (오전/오후)')
                date_obj = datetime.strptime(req_date, '%Y-%m-%d')
                day_of_week = weekday_map[date_obj.weekday()]
                date_str_display = f"{date_obj.strftime('%m월 %d일')}({day_of_week})"
                time_str_display = '오전' if is_morning else '오후'
                
                # 1. 근무일이 아닌 경우 검사
                time_period_key = "morning" if is_morning else "afternoon"
                if req_date not in work_days_map or person not in work_days_map[req_date][time_period_key]:
                    msg = f"⚠️ {person}: {date_str_display} ({time_str_display})이 근무일이 아니므로 '{category}' 요청을 처리할 수 없습니다."
                    unapplied_messages.append(msg)
                    is_valid = False

                # 2. [수정] 평일 당직방 및 토요/휴일 요청 검사
                is_special_day = req_date in [d.strftime('%Y-%m-%d') for d, _, _ in st.session_state.get("special_schedules", [])]
                
                if is_valid and is_special_day:
                    # 2-1. 토요/휴일 방 요청 차단 (어떤 요청이든)
                    msg = f"⛔️ {person}: {date_str_display} ({time_str_display})의 '{category}' 요청은 토요/휴일입니다. (요청사항 반영 안 됨)"
                    unapplied_messages.append(msg)
                    is_valid = False

                elif is_valid and not is_special_day:
                    # 2-2. 평일 당직방 요청 차단
                    room_match = re.match(r'(\d+)번방', category)
                    if room_match:
                        req_room_num = room_match.group(1)
                        morning_duty_room = st.session_state["room_settings"].get("830_duty")
                        afternoon_duty_room = st.session_state["room_settings"].get("1330_duty")

                        if is_morning and req_room_num == morning_duty_room:
                            # (오전 당직방)
                            msg = f"⛔️ {person}: {date_str_display} ({time_str_display})의 '{req_room_num}번방' 요청은 '오전 당직방'입니다. (요청사항 반영 안 됨)"
                            unapplied_messages.append(msg)
                            is_valid = False
                        elif not is_morning and req_room_num == afternoon_duty_room:
                            # (오후 당직방)
                            msg = f"⛔️ {person}: {date_str_display} ({time_str_display})의 '{req_room_num}번방' 요청은 '오후 당직방'입니다. (요청사항 반영 안 됨)"
                            unapplied_messages.append(msg)
                            is_valid = False

                if is_valid:
                    valid_requests_indices.append(index)
            
            # 유효한 요청들만 필터링하여 DataFrame 생성
            valid_requests_df = st.session_state["df_room_request"].loc[valid_requests_indices].copy()
            time.sleep(1)

            try:
                target_year = int(month_str.split('년')[0])
                
                # 토요/휴일 날짜 목록 ('m월 d일' 형식)
                special_dates_str_set = {s[1] for s in st.session_state.get("special_schedules", [])}
                
                # edited_df_md에서 토요/휴일 데이터만 필터링
                final_special_df_md = edited_df_md[edited_df_md['날짜'].isin(special_dates_str_set)].copy()

                date_to_personnel_map = {}
                if not final_special_df_md.empty:
                    # 날짜 형식 변환 및 근무자 목록 생성
                    for _, row in final_special_df_md.iterrows():
                        try:
                            # 'm월 d일' -> 'YYYY-MM-DD'
                            date_obj = datetime.strptime(row['날짜'], '%m월 %d일').replace(year=target_year)
                            date_key = date_obj.strftime('%Y-%m-%d')
                            
                            # 해당 날짜의 모든 근무자 추출 (중복 제거 및 정렬)
                            personnel_cols = [str(i) for i in range(1, 12)] + ['오전당직(온콜)'] + [f'오후{i}' for i in range(1, 5)]
                            personnel_list = [str(row[col]).strip() for col in personnel_cols if col in row and pd.notna(row[col]) and str(row[col]).strip()]
                            unique_personnel = sorted(list(dict.fromkeys(personnel_list)))
                            
                            date_to_personnel_map[date_key] = ", ".join(unique_personnel)
                        except (ValueError, TypeError):
                            continue

                # Google Sheets 업데이트
                if date_to_personnel_map:
                    gc = get_gspread_client()
                    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
                    special_sheet_name = f"{target_year}년 토요/휴일 스케줄"
                    worksheet_special = sheet.worksheet(special_sheet_name)
                    df_yearly = pd.DataFrame(worksheet_special.get_all_records())
                    
                    if not df_yearly.empty:
                        # '날짜' 열을 기준으로 '근무' 열 업데이트
                        df_yearly['근무'] = df_yearly.apply(lambda r: date_to_personnel_map.get(str(r['날짜']), r['근무']), axis=1)
                        
                        if update_sheet_with_retry(worksheet_special, [df_yearly.columns.tolist()] + df_yearly.fillna('').values.tolist()):
                            # st.success(f"✅ '{special_sheet_name}' 시트의 근무 정보가 성공적으로 동기화되었습니다.")
                            pass
                        else:
                            st.error(f"❌ '{special_sheet_name}' 시트 동기화에 실패했습니다.")
            
            except gspread.exceptions.WorksheetNotFound:
                st.warning(f"'{special_sheet_name}' 시트를 찾을 수 없어 토요/휴일 스케줄을 동기화할 수 없습니다.")
            except Exception as e:
                st.error(f"토요/휴일 스케줄 동기화 중 오류 발생: {type(e).__name__} - {e}")

            # --- 최종 당직 정보 입력 검증 ---
            # 이 부분은 이전에 생성된 special_schedules와 special_df를 사용합니다.
            for date_obj, date_str, _ in special_schedules:
                settings = st.session_state.get("weekend_room_settings", {}).get(date_str, {})
                total_room_count = settings.get("total_room_count", 0)
                duty_room_selected = settings.get("duty_room")
                
                duty_person_val = ""
                if not special_df.empty:
                    duty_row = special_df[special_df['날짜_dt'].dt.date == date_obj]
                    if not duty_row.empty: 
                        duty_person_val = str(duty_row['당직'].iloc[0]).strip()

                if total_room_count > 0 and duty_person_val and duty_person_val != "당직 없음" and not duty_room_selected:
                    st.error(f"⚠️ {date_str}: 당직 인원({duty_person_val})이 지정되어 있으므로, '토요/휴일 방 설정'에서 당직 방을 선택해야 합니다.")
                    st.stop()

            # --- 평일 방 설정 검증 및 슬롯 정보 생성 (기존과 동일) ---
            time_slots, time_groups, memo_rules = {}, {}, {}
            if num_830 + num_900 + num_930 + num_1000 != 12:
                st.error(f"오전 방 개수 합계는 12개여야 합니다. (온콜 제외) 현재: {num_830 + num_900 + num_930 + num_1000}개")
                st.stop()
            elif len(rooms_830) != num_830 or len(rooms_900) != num_900 or len(rooms_930) != num_930 or len(rooms_1000) != num_1000 or len(rooms_1330) != num_1330:
                st.error("각 시간대의 방 번호 선택을 완료해주세요.")
                st.stop()
            else:
                for room in rooms_830:
                    slot = f"8:30({room})_당직" if room == duty_830 else f"8:30({room})"
                    time_slots[slot] = len(time_slots)
                    time_groups.setdefault('8:30', []).append(slot)
                for room in rooms_900:
                    slot = f"9:00({room})"
                    time_slots[slot] = len(time_slots)
                    time_groups.setdefault('9:00', []).append(slot)
                for room in rooms_930:
                    slot = f"9:30({room})"
                    time_slots[slot] = len(time_slots)
                    time_groups.setdefault('9:30', []).append(slot)
                for room in rooms_1000:
                    slot = f"10:00({room})"
                    time_slots[slot] = len(time_slots)
                    time_groups.setdefault('10:00', []).append(slot)
                for room in rooms_1330:
                    slot = f"13:30({room})_당직" if room == duty_1330 else f"13:30({room})"
                    time_slots[slot] = len(time_slots)
                    time_groups.setdefault('13:30', []).append(slot)

                memo_rules = {
                    **{f'{i}번방': [s for s in time_slots if f'({i})' in s and '_당직' not in s] for i in range(1, 13)},
                    '당직 아닌 이른방': [s for s in time_slots if s.startswith('8:30') and '_당직' not in s],
                    '이른방 제외': [s for s in time_slots if s.startswith(('9:00', '9:30', '10:00'))],
                    '늦은방 제외': [s for s in time_slots if s.startswith(('8:30', '9:00', '9:30'))],
                    '8:30': [s for s in time_slots if s.startswith('8:30') and '_당직' not in s],
                    '9:00': [s for s in time_slots if s.startswith('9:00')],
                    '9:30': [s for s in time_slots if s.startswith('9:30')],
                    '10:00': [s for s in time_slots if s.startswith('10:00')],
                    '오후 당직 제외': [s for s in time_slots if s.startswith('13:30') and '_당직' not in s]
                }

                st.session_state.update({"time_slots": time_slots, "time_groups": time_groups, "memo_rules": memo_rules})

            # 1. time_slots 딕셔너리는 L1007-L1022에서 이미 사용자 입력 순서대로 생성됨
            #    (Python 3.7+ 딕셔너리는 삽입 순서를 보장함)
            all_time_slots_ordered = list(time_slots.keys())
            
            # 2. 시간대별로 (순서가 유지된 채) 리스트를 분리
            slots_830 = [s for s in all_time_slots_ordered if s.startswith('8:30')]
            slots_900 = [s for s in all_time_slots_ordered if s.startswith('9:00')]
            slots_930 = [s for s in all_time_slots_ordered if s.startswith('9:30')]
            slots_1000 = [s for s in all_time_slots_ordered if s.startswith('10:00')]
            slots_1330 = [s for s in all_time_slots_ordered if s.startswith('13:30')]

            # 3. [수정] sorted() 함수를 제거하고, 순서가 보장된 리스트를 그대로 조합
            all_slots = slots_830 + slots_900 + slots_930 + slots_1000 + \
                        ['온콜'] + \
                        slots_1330
            
            # 4. (필수) 'morning_duty_slot' 변수는 이후 로직(L1048, L1082)에서 사용되므로, 
            #    all_slots 생성 후에도 올바른 값을 갖도록 여기서 정의합니다.
            morning_duty_slot = f"8:30({duty_830})_당직"

            columns = ['날짜', '요일'] + all_slots

            # --- 배정 로직 ---
            total_stats = {'early': Counter(), 'late': Counter(), 'morning_duty': Counter(), 'afternoon_duty': Counter(), 'rooms': {str(i): Counter() for i in range(1, 13)}, 'time_room_slots': {s: Counter() for s in time_slots}}
            df_cumulative = st.session_state["df_cumulative"]
            afternoon_duty_counts = {row['이름']: int(row['오후당직누적']) for _, row in df_cumulative.iterrows() if pd.notna(row.get('오후당직누적'))}
            
            assignments, date_cache, request_cells, result_data = {}, {}, {}, []
            assignable_slots = [s for s in st.session_state["time_slots"].keys() if not (s.startswith('8:30') and s.endswith('_당직'))]
            weekday_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}

            special_dates = [date_str for _, date_str, _ in special_schedules]

            target_year = int(month_str.split('년')[0])
            
            # [수정] for 루프 이전에 special_df 변수를 명확히 정의
            special_df_for_assignment = special_df 

            for _, row in edited_df_md.iterrows():
                date_str = row['날짜']
                try:
                    date_obj = datetime.strptime(date_str, '%m월 %d일').replace(year=target_year) if "월" in date_str else datetime.strptime(date_str, '%Y-%m-%d')
                    formatted_date = date_obj.strftime('%Y-%m-%d').strip()
                    date_cache[date_str] = formatted_date
                    day_of_week = weekday_map[date_obj.weekday()]
                except (ValueError, TypeError):
                    continue

                result_row = [date_str, day_of_week]

                # --- 토요/휴일 배정 로직 ---
                if date_str in special_dates:
                    personnel = [p for p in row.iloc[2:].dropna() if p]
                    settings = st.session_state["weekend_room_settings"].get(date_str, {})
                    
                    assignment_dict, sorted_rooms = assign_special_date(personnel, date_str, formatted_date, settings, special_df_for_assignment, valid_requests_df)

                    # (이하 로직은 기존 코드를 그대로 따르되, 하드코딩된 부분만 제거)
                    room_to_first_slot_idx = {}
                    for slot_idx, slot_name in enumerate(columns[2:]):
                        room_match = re.search(r'\((\d+)\)', str(slot_name))
                        if room_match:
                            room_num = room_match.group(1)
                            if room_num not in room_to_first_slot_idx:
                                room_to_first_slot_idx[room_num] = slot_idx
                    
                    # ▼▼▼ [새로 추가할 부분] 토요/휴일 요청사항도 request_cells에 기록하여 메모 기능 활성화 ▼▼▼
                    if not st.session_state["df_room_request"].empty:
                        requests_for_day = st.session_state["df_room_request"][
                            st.session_state["df_room_request"]['날짜정보'].str.startswith(formatted_date)
                        ]
                        for _, req in requests_for_day.iterrows():
                            person_req = req['이름']
                            category_req = req['분류'] # 예: "7번방"
                            room_match_req = re.match(r'(\d+)번방', category_req)

                            if room_match_req:
                                room_num_req = room_match_req.group(1)
                                # 이 요청이 실제로 배정에 반영되었는지 확인
                                if f"방({room_num_req})" in assignment_dict and assignment_dict[f"방({room_num_req})"] == person_req:
                                    # 해당 방 번호에 해당하는 슬롯 이름을 찾음
                                    if room_num_req in room_to_first_slot_idx:
                                        slot_idx = room_to_first_slot_idx[room_num_req]
                                        slot_name = columns[slot_idx + 2] # +2 for '날짜', '요일' columns
                                        request_cells[(formatted_date, slot_name)] = {'이름': person_req, '분류': category_req}
        
                    mapped_assignment = [None] * (len(columns) - 2)
                    
                    # sorted_rooms를 기준으로 배정하여 순서 보장
                    for room_num in sorted_rooms:
                        slot_key = f"방({room_num})"
                        if slot_key in assignment_dict:
                            person = assignment_dict[slot_key]
                            if room_num in room_to_first_slot_idx:
                                slot_idx = room_to_first_slot_idx[room_num]
                                mapped_assignment[slot_idx] = person
                            
                    result_data.append(result_row + mapped_assignment)
                    continue # 평일 로직 건너뛰기
                    
                has_person = any(val for val in row.iloc[2:-1] if pd.notna(val) and val)
                personnel_for_the_day = [p for p in row.iloc[2:].dropna() if p]
                        
                # 이 코드는 사용자의 기존 `if date_str in special_dates:` 블록을 대체합니다.
                if date_str in special_dates:
                    found_special_schedule = False
                    # 해당 날짜의 특별 근무 일정을 찾습니다.
                    for date_obj, special_date_str, personnel in special_schedules:
                        if special_date_str == date_str:
                            # Streamlit 세션 상태에서 해당 날짜의 주말/공휴일 설정을 가져옵니다.
                            settings = st.session_state["weekend_room_settings"].get(date_str, {})
                            duty_person = settings.get("duty_person", None)
                            duty_room = settings.get("duty_room", None)

                            # 설정된 인원과 방 정보를 바탕으로 배정 계획을 생성합니다.
                            # assignment_dict는 {"방번호": "담당자"} 형태의 딕셔너리입니다.
                            assignment_dict, sorted_rooms = assign_special_date(personnel, date_str, formatted_date, settings, special_df_for_assignment, valid_requests_df)
                            
                            # 배정된 인원 수가 방 수보다 적을 경우 경고 메시지를 표시합니다.
                            if len(assignment_dict) < len(sorted_rooms):
                                st.warning(f"{date_str}: 인원 수({len(personnel)}) 부족으로 {len(sorted_rooms) - len(assignment_dict)}개 방배정 안 됨.")
                            
                            # 1. 각 방 번호와 매칭되는 첫 번째 오전 슬롯의 인덱스를 찾습니다.
                            room_to_first_slot_idx = {}
                            # DataFrame의 최종 컬럼을 기준으로 슬롯을 순회하여 길이 불일치 문제를 해결합니다.
                            for slot_idx, slot in enumerate(columns[2:]):
                                # 오후(13:30) 슬롯이나 '온콜' 등 배정 대상이 아닌 슬롯은 건너뜁니다.
                                slot_str = str(slot)
                                if '13:30' in slot_str or '온콜' in slot_str:
                                    continue
                                
                                # 정규식을 사용해 슬롯 이름에서 방 번호를 추출합니다. 예: "8:30(1)_당직" -> "1"
                                room_match = re.search(r'\((\d+)\)', slot_str)
                                if room_match:
                                    room_num = room_match.group(1)
                                    # 아직 맵에 없는 방 번호일 경우에만 추가하여, 각 방의 '첫 번째' 슬롯만 매핑되도록 합니다.
                                    if room_num not in room_to_first_slot_idx:
                                        room_to_first_slot_idx[room_num] = slot_idx
                            
                            # 2. 배정 결과를 최종 슬롯 리스트에 매핑합니다.
                            # 최종 결과(엑셀의 한 행)를 담을 리스트를 'columns' 길이에 맞춰 초기화합니다.
                            mapped_assignment = [None] * (len(columns) - 2)
                            # 중복 배정을 방지하기 위해 이미 배정된 인원을 기록하는 세트입니다.
                            assigned_personnel = set()
                            
                            # `assignment_dict`의 모든 항목(방-사람)을 순회하며 배정합니다.
                            for room_num, person_with_room in assignment_dict.items():
                                # 담당자 이름만 추출합니다. (예: "강승주[3]" -> "강승주")
                                person = person_with_room.split('[')[0].strip()

                                # 해당 방 번호가 배정 대상인 오전 슬롯에 포함되어 있는지 확인합니다.
                                if room_num in room_to_first_slot_idx:
                                    # 이미 다른 방에 배정된 인원인지 확인하여 중복을 방지합니다.
                                    if person in assigned_personnel:
                                        st.warning(f"{date_str}: {person}님이 중복 배정되었습니다. 확인이 필요합니다.")
                                        continue

                                    # 배정할 슬롯의 인덱스를 가져옵니다.
                                    slot_idx = room_to_first_slot_idx[room_num]
                                    
                                    # 최종 배정 리스트의 해당 위치에 담당자 이름을 할당합니다.
                                    mapped_assignment[slot_idx] = person
                                    # 이 담당자를 '배정 완료' 세트에 추가합니다.
                                    assigned_personnel.add(person)
                            
                            # 완성된 배정 결과를 전체 결과 데이터에 추가합니다.
                            full_row = result_row + mapped_assignment
                            result_data.append(full_row)
                            found_special_schedule = True
                            break  # 해당 날짜의 처리가 끝났으므로 내부 루프를 종료합니다.

                    # 특별 근무 일정이 없는 경우 (예: 공휴일이지만 근무자가 없는 날) 빈 행을 추가합니다.
                    if not found_special_schedule:
                        result_data.append(result_row + [None] * (len(columns) - 2))

                    # special_date 처리가 끝났으므로, 평일 배정 로직을 건너뛰고 다음 날짜로 넘어갑니다.
                    continue

                # 기존 평일 처리
                # 2. '소수 인원 근무'로 판단할 기준 인원수를 설정합니다.
                SMALL_TEAM_THRESHOLD = 15

                # 3. 근무 인원수가 설정된 기준보다 적으면, 방배정 없이 순서대로 나열합니다.
                if len(personnel_for_the_day) < SMALL_TEAM_THRESHOLD and has_person:
                    result_row.append(None)
                    result_row.extend(personnel_for_the_day)
                    num_slots_to_fill = len(all_slots)
                    slots_filled_count = len(personnel_for_the_day) + 1  # 근무자 수 + 비워둔 1칸
                    padding_needed = num_slots_to_fill - slots_filled_count
                    if padding_needed > 0:
                        result_row.extend([None] * padding_needed)
                    result_data.append(result_row)
                    continue
                
                morning_personnel = [row[str(i)] for i in range(1, 12) if pd.notna(row[str(i)]) and row[str(i)]]
                afternoon_personnel = [row[f'오후{i}'] for i in range(1, 5) if pd.notna(row[f'오후{i}']) and row[f'오후{i}']]
                
                if not (morning_personnel or afternoon_personnel):
                    result_row.extend([None] * len(all_slots))
                    result_data.append(result_row)
                    continue
                
                # ✨ --- 여기가 수정된 요청사항 처리 로직입니다 --- ✨
                request_assignments = {}
                # 그날에 해당하는 유효한 요청만 필터링
                requests_for_day = valid_requests_df[valid_requests_df['날짜정보'].str.startswith(formatted_date)]
                
                if not requests_for_day.empty:
                    # 1단계: '특정 방' 요청 먼저 처리 (충돌 가능성 높음)
                    room_reqs = requests_for_day[requests_for_day['분류'].str.contains('번방')].sort_index()
                    for _, req in room_reqs.iterrows():
                        person, category = req['이름'], req['분류']
                        # 이 사람/시간대에 대한 요청이 이미 처리되었는지 확인
                        if any(p == person for p in request_assignments.values()): continue

                        slots_for_category = st.session_state["memo_rules"].get(category, [])
                        if slots_for_category:
                            # '1번방' 요청은 슬롯이 하나뿐이므로, 그 슬롯이 비어있으면 배정
                            target_slot = slots_for_category[0]
                            if target_slot not in request_assignments:
                                request_assignments[target_slot] = person
                                request_cells[(formatted_date, target_slot)] = {'이름': person, '분류': category}

                    # 2단계: '특정 시간대' 및 기타 요청 처리
                    other_reqs = requests_for_day[~requests_for_day['분류'].str.contains('번방')].sort_index()
                    for _, req in other_reqs.iterrows():
                        person, category, date_info = req['이름'], req['분류'], req['날짜정보']
                        is_morning = '(오전)' in date_info
                        if any(p == person for p in request_assignments.values()): continue

                        # 요청을 만족하는 '아직 비어있는' 슬롯 찾기
                        possible_slots = [s for s in st.session_state["memo_rules"].get(category, []) if s not in request_assignments]
                        if possible_slots:
                            selected_slot = random.choice(possible_slots)
                            request_assignments[selected_slot] = person
                            request_cells[(formatted_date, selected_slot)] = {'이름': person, '분류': category}

                # `random_assign` 호출은 기존과 동일합니다.
                assignment, _ = random_assign(list(set(morning_personnel)|set(afternoon_personnel)), assignable_slots, request_assignments, st.session_state["time_groups"], total_stats, list(morning_personnel), list(afternoon_personnel), afternoon_duty_counts)

                for slot in all_slots:
                    person = row['오전당직(온콜)'] if slot == morning_duty_slot or slot == '온콜' else (assignment[assignable_slots.index(slot)] if slot in assignable_slots and assignment else None)
                    result_row.append(person if has_person else None)
                
                # [추가] 중복 배정 검증 로직
                assignments_for_day = dict(zip(all_slots, result_row[2:]))
                morning_slots_check = [s for s in all_slots if s.startswith(('8:30', '9:00', '9:30', '10:00'))]
                afternoon_slots_check = [s for s in all_slots if s.startswith('13:30')]

                morning_counts = Counter(p for s, p in assignments_for_day.items() if s in morning_slots_check and p)
                for person, count in morning_counts.items():
                    if count > 1:
                        duplicated_slots = [s for s, p in assignments_for_day.items() if p == person and s in morning_slots_check]
                        st.error(f"⚠️ {date_str}: '{person}'님이 오전에 중복 배정되었습니다 (슬롯: {', '.join(duplicated_slots)}).")
                
                afternoon_counts = Counter(p for s, p in assignments_for_day.items() if s in afternoon_slots_check and p)
                for person, count in afternoon_counts.items():
                    if count > 1:
                        duplicated_slots = [s for s, p in assignments_for_day.items() if p == person and s in afternoon_slots_check]
                        st.error(f"⚠️ {date_str}: '{person}'님이 오후/온콜에 중복 배정되었습니다 (슬롯: {', '.join(duplicated_slots)}).")

                result_data.append(result_row)
            
            df_room = pd.DataFrame(result_data, columns=columns)

            # 1. 'df_room' (최종 방배정 결과)를 기반으로 'total_stats'를 (재)계산합니다.
            total_stats = {
                'early': Counter(), 
                'late': Counter(), 
                'morning_duty': Counter(), 
                'afternoon_duty': Counter(), 
                'rooms': {str(i): Counter() for i in range(1, 13)}, 
                'time_room_slots': {s: Counter() for s in st.session_state["time_slots"].keys()}
            }
            
            for _, row in df_room.iterrows():
                current_date_str = row['날짜']
                if current_date_str in special_dates:
                    continue # 토요/휴일은 통계에 포함 안 함

                # 'columns' 리스트와 'row' (Series)를 매핑
                assignment_for_day = row[columns[2:]] # '날짜', '요일' 제외

                for slot_name, person in assignment_for_day.items():
                    if not person:
                        continue
                    
                    # 1. 오전 당직
                    if slot_name == morning_duty_slot:
                        total_stats['morning_duty'][person] += 1
                    
                    # 2. 오후 당직 (13:30 당직)
                    elif slot_name.startswith('13:30') and slot_name.endswith('_당직'):
                        total_stats['afternoon_duty'][person] += 1
                    
                    # 3. 이른방 (8:30, 당직 제외)
                    elif slot_name.startswith('8:30') and '_당직' not in slot_name:
                        total_stats['early'][person] += 1
                    
                    # 4. 늦은방 (10:00)
                    elif slot_name.startswith('10:00'):
                        total_stats['late'][person] += 1
                    
                    # 5. 시간대별/방별 통계 (for stats_df)
                    if slot_name in total_stats['time_room_slots']:
                        total_stats['time_room_slots'][slot_name][person] += 1

            # --- [수정 2] 통계 DataFrame 생성 (목표치 열 추가) ---
            stats_data, all_personnel_stats = [], set(p for _, r in st.session_state["df_schedule_md"].iterrows() for p in r[2:].dropna() if p)
            
            df_cumulative_stats = st.session_state.get("df_cumulative", pd.DataFrame())
            
            # ▼▼▼ [수정 시작] ▼▼▼
            duty_target_map = {}
            # [수정] "진짜 베이스 누적"을 계산하기 위해 4개의 맵을 모두 로드합니다.
            old_pm_cumulative_map = {}
            old_pm_source_map = {}
            old_am_cumulative_map = {} 
            old_am_source_map = {}
            
            if not df_cumulative_stats.empty and '이름' in df_cumulative_stats.columns:
                if '오후당직누적' in df_cumulative_stats.columns:
                    old_pm_cumulative_map = df_cumulative_stats.set_index('이름')['오후당직누적'].to_dict()
                if '오후당직' in df_cumulative_stats.columns:
                    old_pm_source_map = df_cumulative_stats.set_index('이름')['오후당직'].to_dict()
                if '오전당직누적' in df_cumulative_stats.columns:
                    old_am_cumulative_map = df_cumulative_stats.set_index('이름')['오전당직누적'].to_dict()
                if '오전당직' in df_cumulative_stats.columns:
                    old_am_source_map = df_cumulative_stats.set_index('이름')['오전당직'].to_dict()

            for person in sorted(all_personnel_stats):
                
                # --- [수정] 오후 당직: (Old 누적 - Old 합계) + New 합계 ---
                pm_assigned_this_month = total_stats['afternoon_duty'][person] # New 합계
                pm_old_cumulative = 0
                pm_old_source = 0
                try:
                    pm_old_cumulative = int(old_pm_cumulative_map.get(person, 0)) # Old 누적
                    pm_old_source = int(old_pm_source_map.get(person, 0))     # Old 합계
                except (ValueError, TypeError):
                    pass
                true_pm_base = pm_old_cumulative - pm_old_source # 진짜 베이스
                pm_final_cumulative = true_pm_base + pm_assigned_this_month # 최종 누적
                
                # --- [수정] 오전 당직: (Old 누적 - Old 합계) + New 합계 ---
                am_assigned_this_month = total_stats['morning_duty'][person] # New 합계
                am_old_cumulative = 0
                am_old_source = 0
                try:
                    am_old_cumulative = int(old_am_cumulative_map.get(person, 0)) # Old 누적
                    am_old_source = int(old_am_source_map.get(person, 0))     # Old 합계
                except (ValueError, TypeError):
                    pass
                true_am_base = am_old_cumulative - am_old_source # 진짜 베이스
                am_final_cumulative = true_am_base + am_assigned_this_month # 최종 누적
                
                stats_entry = {
                    '인원': person,
                    '이른방 합계': total_stats['early'][person],
                    '늦은방 합계': total_stats['late'][person],
                    '오전당직 합계': am_assigned_this_month,
                    '오전당직 누적': am_final_cumulative,
                    '오후당직 합계': pm_assigned_this_month,
                    '오후당직 누적': pm_final_cumulative
                }
                for slot in st.session_state["time_slots"].keys():
                    if not slot.endswith('_당직'):
                        stats_entry[f'{slot} 합계'] = total_stats['time_room_slots'].get(slot, Counter())[person]
                stats_data.append(stats_entry)

            time_order = ['8:30', '9:00', '9:30', '10:00', '13:30']

            # [수정] 컬럼 목록에 '오전당직 누적' 추가
            sorted_columns = ['인원', '이른방 합계', '늦은방 합계', '오전당직 합계', '오전당직 누적', '오후당직 합계', '오후당직 누적']
            
            time_slots_sorted = sorted(
                [slot for slot in st.session_state["time_slots"].keys() if not slot.endswith('_당직')],
                key=lambda x: (time_order.index(x.split('(')[0]), int(x.split('(')[1].split(')')[0]))
            )
            sorted_columns.extend([f'{slot} 합계' for slot in time_slots_sorted])
            stats_df_names_as_rows = pd.DataFrame(stats_data)[sorted_columns]
            # [신규] (항목-행) 기준으로 Transpose 하여 최종 stats_df 생성
            stats_df = stats_df_names_as_rows.set_index('인원').transpose().reset_index().rename(columns={'index': '항목'})

            # --- [수정 3] 배정 완료 후, 모든 로그 생성 ---
            
            # 3-1. 방배정 요청 로그 생성
            applied_request_keys = set((key[0], value['이름'], value['분류']) for key, value in request_cells.items())
            
            # ▼▼▼ [수정] 이 블록 전체를 교체하세요. (1594~1608 라인) ▼▼▼
            
            # [FIX] 휴일 날짜 문자열 세트를 미리 생성 (예: '10월 18일')
            special_dates_str_set = {s[1] for s in st.session_state.get("special_schedules", [])}

            for _, req in valid_requests_df.iterrows():
                req_date, is_morning = parse_date_info(req['날짜정보'])
                if not req_date: continue
                person, category = req['이름'], req['분류']
                
                date_obj = datetime.strptime(req_date, '%Y-%m-%d')
                # [FIX] 날짜 형식을 '10월 18일' (m월 d일)로 변경 (휴일 세트와 비교하기 위함)
                date_str_display_for_check = f"{date_obj.month}월 {date_obj.day}일"
                # 화면 표시용 날짜 (요일 포함)
                date_str_display = f"{date_str_display_for_check}({weekday_map[date_obj.weekday()]})"
                time_str_display = '오전' if is_morning else '오후'
                
                if (req_date, person, category) in applied_request_keys:
                    msg = f"✅ {person}: {date_str_display} ({time_str_display})의 '{category}' 요청이 적용되었습니다."
                    applied_messages.append(msg)
                else:
                    # [FIX] 이 요청이 실패한 요청인지 확인
                    
                    # 1. 이 날짜가 휴일인지 확인
                    is_special_day_log = date_str_display_for_check in special_dates_str_set
                    
                    if is_special_day_log:
                        # [FIX] 날짜가 휴일(토요일 포함)이기만 하면, 실패 시 무조건 수기 수정 메시지
                        msg = f"⛔️ {person}: {date_str_display} ({time_str_display})의 '{category}' 요청은 토요 방입니다. 수기로 수정해 주십시오."
                    else:
                        # 평일 요청이 실패한 경우 (배정 균형)
                        msg = f"⚠️ {person}: {date_str_display} ({time_str_display})의 '{category}' 요청이 배정 균형을 위해 반영되지 않았습니다."
                    
                    unapplied_messages.append(msg)
                # ▲▲▲ [수정] 교체 완료 ▲▲▲

            # 3-2. 오후당직 배정 로그 생성
            oncall_logs = []
            actual_duty_counts = total_stats.get('afternoon_duty', Counter())
            
            # [수정] 오후당직 배정 로그를 횟수별로 그룹화합니다.
            # 1. 횟수별로 인원을 저장할 딕셔너리 초기화
            counts_to_workers = {}

            # 2. 모든 인원(all_personnel_stats)을 순회하며 횟수별 딕셔너리에 추가
            # (actual_duty_counts에 없는 0회 배정자도 포함하기 위함)
            for person in all_personnel_stats: # L1879에서 정의된 변수
                count = actual_duty_counts.get(person, 0) # 배정 못 받았으면 0
                if count not in counts_to_workers:
                    counts_to_workers[count] = []
                counts_to_workers[count].append(person)
                
            # 3. oncall_logs 리스트 생성 (횟수(key) 기준으로 정렬)
            for count, workers in sorted(counts_to_workers.items()):
                if workers: # 해당 횟수에 배정된 사람이 있을 경우에만 로그 생성
                    sorted_workers = sorted(workers) # 이름순으로 정렬
                    log_message = f"- {count}회 배정: {', '.join(sorted_workers)}"
                    oncall_logs.append(log_message)

            # --- [수정] Google Sheets 연결 및 시트 저장 로직 (누락된 부분 복원) ---
            try:
                gc = get_gspread_client()
                sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])

                # --- 1. [신규 삽입] "스케줄 최종" 저장 로직 (L1448-L1542 코드) ---
                try:
                    df_schedule_to_save = st.session_state["df_schedule_original"].copy()
                    
                    # 'edited_df_md' (L1377에서 정의됨)는 현재 data_editor의 상태입니다.
                    edited_df_md_for_save = edited_df_md 

                    # L1450-L1455: robust_parse_date 함수 정의
                    target_year_for_save = int(month_str.split('년')[0]) 
                    def robust_parse_date(date_str, year=target_year_for_save):
                        try:
                            if "월" in str(date_str): return datetime.strptime(str(date_str), '%m월 %d일').replace(year=year).date()
                            else: return pd.to_datetime(date_str).date()
                        except: return None
                    
                    df_schedule_to_save['parsed_date'] = df_schedule_to_save['날짜'].apply(robust_parse_date)
                    
                    # L1456-L1522: edited_df_md_for_save를 순회하며 df_schedule_to_save 업데이트
                    for _, edited_row in edited_df_md_for_save.iterrows():
                        edited_date_obj = robust_parse_date(edited_row['날짜'])
                        if edited_date_obj is None: continue
                        target_indices = df_schedule_to_save[df_schedule_to_save['parsed_date'] == edited_date_obj].index
                        if target_indices.empty: continue
                        original_row_idx = target_indices[0]

                        # (L1464) 수정된 내용 가져오기
                        oncall_person = str(edited_row.get('오전당직(온콜)', '')).strip()
                        am_editor_cols = [str(i) for i in range(1, 12)]
                        am_personnel = [str(edited_row[col]).strip() for col in am_editor_cols if col in edited_row and pd.notna(edited_row[col]) and str(edited_row[col]).strip()]
                        pm_editor_cols = [f'오후{i}' for i in range(1, 5)]
                        pm_personnel = [str(edited_row[col]).strip() for col in pm_editor_cols if col in edited_row and pd.notna(edited_row[col]) and str(edited_row[col]).strip()]

                        # (L1477) 원본 스케줄의 숨겨진 근무자 파싱
                        original_schedule_row = st.session_state["df_schedule_original"].loc[original_row_idx]
                        pm_hidden_cols = [f'오후{i}' for i in range(5, 10)]
                        pm_hidden_personnel = [
                            clean_name(original_schedule_row[col]) for col in pm_hidden_cols
                            if col in original_schedule_row and pd.notna(original_schedule_row[col]) and clean_name(original_schedule_row[col])
                        ]
                        am_hidden_cols = [str(i) for i in range(12, 18)]
                        am_hidden_personnel = [
                            clean_name(original_schedule_row[col]) for col in am_hidden_cols
                            if col in original_schedule_row and pd.notna(original_schedule_row[col]) and clean_name(original_schedule_row[col])
                        ]

                        # (L1501) 저장할 DataFrame의 모든 관련 열 초기화
                        cols_to_clear_am = [str(i) for i in range(1, 18)]
                        for col in cols_to_clear_am:
                            if col in df_schedule_to_save.columns: df_schedule_to_save.at[original_row_idx, col] = ''
                        cols_to_clear_pm = [f'오후{i}' for i in range(1, 10)]
                        for col in cols_to_clear_pm:
                            if col in df_schedule_to_save.columns: df_schedule_to_save.at[original_row_idx, col] = ''

                        # (L1510) 수정된 내용으로 다시 채워넣기
                        df_schedule_to_save.at[original_row_idx, '오전당직(온콜)'] = oncall_person

                        am_save_list = list(dict.fromkeys(am_personnel + ([oncall_person] if oncall_person else []) + am_hidden_personnel))
                        for i, person in enumerate(am_save_list, 1):
                            col_name_am = str(i)
                            if col_name_am in df_schedule_to_save.columns:
                                df_schedule_to_save.at[original_row_idx, col_name_am] = person

                        pm_save_list = list(dict.fromkeys(pm_personnel + pm_hidden_personnel))
                        for i, person in enumerate(pm_save_list, 1):
                            col_name_pm = f'오후{i}'
                            if col_name_pm in df_schedule_to_save.columns:
                                df_schedule_to_save.at[original_row_idx, col_name_pm] = person
                    
                    df_schedule_to_save.drop(columns=['parsed_date'], inplace=True, errors='ignore')

                    # (L1524) "스케줄 최종" 시트에 저장
                    schedule_sheet_name = f"{month_str} 스케줄 최종"
                    try:
                        worksheet_schedule = sheet.worksheet(schedule_sheet_name)
                    except gspread.exceptions.WorksheetNotFound:
                        worksheet_schedule = sheet.add_worksheet(title=schedule_sheet_name, rows=100, cols=30) 
                    
                    # [수정] 원본 df_schedule_original의 컬럼을 기준으로 저장할 컬럼을 정함
                    columns_to_save_requested = st.session_state["df_schedule_original"].columns.tolist()
                    
                    # 불필요한 열 제거
                    if 'parsed_date' in columns_to_save_requested:
                         columns_to_save_requested.remove('parsed_date')
                    if '날짜_dt' in columns_to_save_requested:
                         columns_to_save_requested.remove('날짜_dt')

                    columns_to_save = [col for col in columns_to_save_requested if col in df_schedule_to_save.columns]
                    schedule_data = [columns_to_save] + df_schedule_to_save[columns_to_save].fillna('').values.tolist()

                    if update_sheet_with_retry(worksheet_schedule, schedule_data):
                        st.success(f"✅ '{schedule_sheet_name}' 시트 저장을 완료하였습니다.")
                    else:
                        save_errors.append(f"'{schedule_sheet_name}' 시트 저장에 실패하였습니다.")

                except Exception as e:
                    st.error(f"Google Sheets '스케줄 최종' 시트 저장 중 오류 발생: {type(e).__name__} - {e}")
                    save_errors.append(f"'스케줄 최종' 시트 저장 실패: {e}")
                # --- ▲▲▲ [신규 삽입] "스케줄 최종" 저장 로직 종료 ▲▲▲ ---


                # --- 2. '방배정' 시트 저장 (기존 로직) ---
                try:
                    try:
                        worksheet_result = sheet.worksheet(f"{month_str} 방배정")
                    except gspread.exceptions.WorksheetNotFound:
                        worksheet_result = sheet.add_worksheet(f"{month_str} 방배정", rows=100, cols=len(df_room.columns))
                    
                    if update_sheet_with_retry(worksheet_result, [df_room.columns.tolist()] + df_room.fillna('').values.tolist()):
                        st.success(f"✅ {month_str} 방배정 테이블이 Google Sheets에 저장되었습니다.")
                        time.sleep(2)
                    else:
                        save_errors.append(f"'{month_str} 방배정' 시트 저장에 실패하였습니다.")
                        
                except Exception as e:
                    st.error(f"Google Sheets '방배정' 시트 저장 중 오류 발생: {type(e).__name__} - {e}")
                    save_errors.append(f"'방배정' 시트 저장 실패: {e}")

                # --- 3. '다음달 누적 최종' 시트 업데이트 (기존 로직) ---
                try:
                    target_month_dt = datetime.strptime(month_str, "%Y년 %m월")
                    next_month_dt = target_month_dt + relativedelta(months=1)
                    next_month_str = next_month_dt.strftime("%Y년 %-m월")
                    
                    latest_cumulative_name_next = st.session_state.get("latest_cumulative_name") 

                    if not latest_cumulative_name_next:
                        save_errors.append(f"'{next_month_str} 누적' 시트를 찾을 수 없어 업데이트 불가.")
                    else:
                        worksheet_cumulative_next = sheet.worksheet(latest_cumulative_name_next)
                        all_data = worksheet_cumulative_next.get_all_values() 
                        
                        if not all_data or len(all_data) < 2:
                            save_errors.append(f"'{latest_cumulative_name_next}' 시트가 비어있거나 형식이 잘못됨.")
                        else:
                            headers, rows = all_data[0], all_data[1:]
                            
                            for r_idx, row_data in enumerate(rows):
                                for c_idx, cell_value in enumerate(row_data):
                                    if c_idx == 0: continue
                                    try:
                                        if cell_value != '':
                                            rows[r_idx][c_idx] = int(cell_value)
                                    except (ValueError, TypeError):
                                        pass

                            # [1단계] 오전/오후 인덱스 변수 초기화
                            pm_target_row_index = -1 # '오후당직누적'
                            pm_source_row_index = -1 # '오후당직'
                            am_target_row_index = -1 # '오전당직누적'
                            am_source_row_index = -1 # '오전당직'

                            # [2단계] 4개 행의 인덱스 찾기
                            for i, row_data in enumerate(rows):
                                if row_data[0] == '오후당직누적':
                                    pm_target_row_index = i
                                if row_data[0] == '오후당직':
                                    pm_source_row_index = i
                                if row_data[0] == '오전당직누적':
                                    am_target_row_index = i
                                if row_data[0] == '오전당직':
                                    am_source_row_index = i
                            
                            # [3단계] 4개 행을 모두 찾았는지 확인
                            if all(idx != -1 for idx in [pm_target_row_index, pm_source_row_index, am_target_row_index, am_source_row_index]):
                                
                                # [4단계] 오전/오후 누적값 맵(Map) 불러오기
                                df_cumulative_updated = st.session_state.get("df_cumulative", pd.DataFrame())
                                old_pm_cumulative_map = {}
                                old_am_cumulative_map = {}
                                
                                # --- ▼▼▼ [수정] 누락된 '합계(Source)' 맵 로드 ▼▼▼ ---
                                old_pm_source_map = {}
                                old_am_source_map = {}
                                # --- ▲▲▲ [수정] 완료 ▲▲▲ ---

                                if not df_cumulative_updated.empty and '이름' in df_cumulative_updated.columns:
                                    if '오후당직누적' in df_cumulative_updated.columns:
                                        old_pm_cumulative_map = df_cumulative_updated.set_index('이름')['오후당직누적'].to_dict()
                                    if '오전당직누적' in df_cumulative_updated.columns:
                                        old_am_cumulative_map = df_cumulative_updated.set_index('이름')['오전당직누적'].to_dict()
                                    
                                    # --- ▼▼▼ [수정] '합계' 맵을 로드하는 코드 추가 ▼▼▼ ---
                                    if '오후당직' in df_cumulative_updated.columns:
                                        old_pm_source_map = df_cumulative_updated.set_index('이름')['오후당직'].to_dict()
                                    if '오전당직' in df_cumulative_updated.columns:
                                        old_am_source_map = df_cumulative_updated.set_index('이름')['오전당직'].to_dict()
                                    # --- ▲▲▲ [수정] 완료 ▲▲▲ ---
                                else:
                                    save_errors.append("세션에서 'df_cumulative' (수정된 누적 데이터)를 읽어오는 데 실패했습니다.")

                                # [5단계] '오전당직' 횟수 계산 및 저장 로직 추가
                                for col_idx, name in enumerate(headers):
                                    if col_idx == 0: continue
                                    name_strip = name.strip()
                                    
                                    # --- 1. 오후 당직 (계산식 수정) ---
                                    pm_assigned_this_month = total_stats['afternoon_duty'].get(name_strip, 0) # New 합계
                                    pm_old_cumulative_val = 0
                                    pm_old_source_val = 0 # [추가]
                                    try:
                                        pm_old_cumulative_val = int(old_pm_cumulative_map.get(name_strip, 0)) # Old 누적
                                        pm_old_source_val = int(old_pm_source_map.get(name_strip, 0)) # Old 합계 [추가]
                                    except (ValueError, TypeError):
                                        pm_old_cumulative_val = 0
                                        pm_old_source_val = 0 # [추가]
                                    
                                    true_pm_base = pm_old_cumulative_val - pm_old_source_val # [수정] 진짜 베이스
                                    pm_final_cumulative_count = true_pm_base + pm_assigned_this_month # [수정] 최종 누적
                                    
                                    rows[pm_source_row_index][col_idx] = pm_assigned_this_month
                                    rows[pm_target_row_index][col_idx] = pm_final_cumulative_count
                                    
                                    # --- 2. 오전 당직 (계산식 수정) ---
                                    am_assigned_this_month = total_stats['morning_duty'].get(name_strip, 0) # New 합계
                                    am_old_cumulative_val = 0
                                    am_old_source_val = 0 # [추가]
                                    try:
                                        am_old_cumulative_val = int(old_am_cumulative_map.get(name_strip, 0)) # Old 누적
                                        am_old_source_val = int(old_am_source_map.get(name_strip, 0)) # Old 합계 [추가]
                                    except (ValueError, TypeError):
                                        am_old_cumulative_val = 0
                                        am_old_source_val = 0 # [추가]
                                    
                                    true_am_base = am_old_cumulative_val - am_old_source_val # [수정] 진짜 베이스
                                    am_final_cumulative_count = true_am_base + am_assigned_this_month # [수정] 최종 누적
                                    
                                    rows[am_source_row_index][col_idx] = am_assigned_this_month
                                    rows[am_target_row_index][col_idx] = am_final_cumulative_count

                                final_cumulative_sheet_name = f"{next_month_str} 누적 최종"
                                try:
                                    worksheet_final_cumulative = sheet.worksheet(final_cumulative_sheet_name)
                                except gspread.exceptions.WorksheetNotFound:
                                    worksheet_final_cumulative = sheet.add_worksheet(title=final_cumulative_sheet_name,
                                                                                    rows=len(all_data) + 5,
                                                                                    cols=len(headers) + 5)

                                if update_sheet_with_retry(worksheet_final_cumulative, [headers] + rows):
                                    st.success(f"✅ '{final_cumulative_sheet_name}' 시트 업데이트가 완료되었습니다.")
                                else:
                                    save_errors.append(f"'{final_cumulative_sheet_name}' 시트 업데이트에 실패하였습니다.")
                            
                            # [6단계] 오류 메시지 수정
                            elif pm_target_row_index == -1: 
                                save_errors.append(f"'{latest_cumulative_name_next}' 시트에서 '오후당직누적' 항목을 찾을 수 없습니다.")
                            elif pm_source_row_index == -1: 
                                save_errors.append(f"'{latest_cumulative_name_next}' 시트에서 '오후당직' 항목을 찾을 수 없습니다.")
                            elif am_target_row_index == -1: 
                                save_errors.append(f"'{latest_cumulative_name_next}' 시트에서 '오전당직누적' 항목을 찾을 수 없습니다.")
                            elif am_source_row_index == -1: 
                                save_errors.append(f"'{latest_cumulative_name_next}' 시트에서 '오전당직' 항목을 찾을 수 없습니다.")

                except Exception as e:
                    st.error(f"Google Sheets '누적 최종' 시트 저장 중 오류 발생: {type(e).__name__} - {e}")
                    save_errors.append(f"'누적 최종' 시트 저장 실패: {e}")
                    time.sleep(5)

            except Exception as e:
                # gc = get_gspread_client() 자체에서 오류가 난 경우
                st.error(f"Google Sheets 연결 중 오류 발생: {type(e).__name__} - {e}")
                save_errors.append(f"Google Sheets 연결 실패: {e}")

            # --- [수정] Excel 생성 함수 호출 ---
            output = generate_excel_output(
                df_room=df_room,
                stats_df=stats_df,
                columns=columns,
                special_dates=special_dates,
                special_df=special_df,
                date_cache=date_cache,
                request_cells=request_cells,
                swapped_assignments=st.session_state.get("swapped_assignments", set()), # 이 시점엔 비어있음
                morning_duty_slot=morning_duty_slot,
                month_str=month_str,
            )
            # --- [수정 완료] ---

            st.session_state["assignment_results"] = {
                "df_room": df_room,
                "stats_df": stats_df,
                "excel_output": output, # <-- 방금 생성한 output
                "applied_messages": applied_messages,
                "unapplied_messages": unapplied_messages,
                "oncall_logs": oncall_logs,
                "save_errors": save_errors,
                "actual_duty_counts": actual_duty_counts, # <-- [핵심 수정] 이 줄을 추가하세요.

                # ▼▼▼ [수정] 다운로드 재성성을 위해 이 변수들을 추가합니다 ▼▼▼
                "columns": columns,
                "special_dates": special_dates,
                "date_cache": date_cache,
                "request_cells": request_cells,
                "morning_duty_slot": morning_duty_slot,
                "all_personnel_stats": all_personnel_stats,
                "time_slots": st.session_state["time_slots"],
                "time_order": time_order,
                "all_slots": all_slots,
            }
            
            save_errors = [] # 저장 오류 기록용

            # 저장 오류가 있으면 화면에 표시
            if save_errors:
                for error_msg in save_errors:
                    st.error(f"❌ {error_msg}")

            # --- [수정] assignment_results에 통계 재계산을 위한 '설정값'도 함께 저장 ---
            st.session_state["assignment_results"] = {
                "df_room": df_room,
                "stats_df": stats_df,
                "excel_output": output,
                "applied_messages": applied_messages,
                "unapplied_messages": unapplied_messages,
                "oncall_logs": oncall_logs,
                "save_errors": save_errors,
                "actual_duty_counts": actual_duty_counts, # <-- [핵심 수정] 이 줄을 추가하세요.
                
                # ▼▼▼ 통계 재계산을 위해 이 7줄을 추가하세요 ▼▼▼
                "columns": columns,
                "date_cache": date_cache,
                "all_slots": all_slots,
                "request_cells": request_cells, # <-- 이 줄이 누락되었습니다.
                "morning_duty_slot": morning_duty_slot,
                "special_dates": special_dates,
                "time_slots": st.session_state["time_slots"],
                "time_order": time_order,
                "all_personnel_stats": all_personnel_stats 
            }
            # --- [수정 완료] ---

            st.rerun()
        
        # 계산이 끝나면 스크립트를 재실행하여 아래의 '결과 표시' 로직을 타게 합니다.
        st.rerun()

def get_sort_key_from_log(log_message):
    """로그 메시지에서 (월, 일)을 추출하여 정렬 키로 반환합니다."""
    # '10월 10일' 형식의 날짜를 찾습니다.
    match = re.search(r'(\d+)월 (\d+)일', log_message)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        # (월, 일) 튜플을 반환하여 날짜순으로 정렬되도록 함
        return (month, day)
    else:
        # 날짜를 찾지 못한 경우, 정렬 순서의 맨 뒤로 보냅니다.
        return (99, 99)
    
# [L2114 부터 L2178까지의 '결과 표시' 로직 전체를 이 코드로 교체하세요]

if "assignment_results" in st.session_state and st.session_state["assignment_results"] is not None:
    results = st.session_state["assignment_results"]
    df_room = results["df_room"]
    stats_df = results["stats_df"]
    output = results["excel_output"]
    applied_messages = results["applied_messages"]
    unapplied_messages = results["unapplied_messages"]
    
    # 1. 생성된 메시지를 심각도에 따라 세 그룹으로 분류합니다.
    critical_unapplied = [msg for msg in unapplied_messages if msg.strip().startswith('⛔️')]
    warning_unapplied = [msg for msg in unapplied_messages if not msg.strip().startswith('⛔️')]
    sorted_applied = sorted(applied_messages)
    
    # ▼▼▼ [수정] 로그 표시 로직 (기존과 동일) ▼▼▼
    st.write("---")
    with st.expander("🔍 방배정 상세 로그 보기", expanded=True):
        st.write(" ")

        st.write("**📞 오후당직 배정 로그**")
        oncall_logs = results.get("oncall_logs", [])
        oncall_log_text = "\n".join(oncall_logs) if oncall_logs else "모든 오후당직이 목표치에 맞게 정상 배정되었습니다."
        st.code(oncall_log_text, language='text')
        
        st.divider()
        st.write("**✅ 방배정 요청사항 적용됨**")
        applied_log_text = "\n".join(f"• {msg[2:]}" for msg in sorted(applied_messages, key=get_sort_key_from_log)) if applied_messages else "해당 없음"
        st.code(applied_log_text, language='text')

        st.divider()
        st.write("**⚠️ 방배정 요청사항 적용 안 됨**")
        warning_log_text = "\n".join(f"• {msg[2:]}" for msg in sorted(warning_unapplied, key=get_sort_key_from_log)) if warning_unapplied else "해당 없음"
        st.code(warning_log_text, language='text')

        st.divider()
        st.write("**⛔️ 방배정 요청사항 적용 안 됨 (수기 수정 필요)**")
        critical_log_text = "\n".join(f"• {msg[2:]}" for msg in sorted(critical_unapplied, key=get_sort_key_from_log)) if critical_unapplied else "해당 없음"
        st.code(critical_log_text, language='text')

    # --- ▲▲▲ 로그 표시 로직 끝 ---

    st.divider()
    st.markdown("**✅ 방배정 스케줄 (수정 가능)**") 
    edited_df_room = st.data_editor(
        df_room,
        hide_index=True,
        key="room_editor",
        disabled=['날짜', '요일'], # 날짜/요일은 수정 불가
        # on_change=set_editor_changed_flag <-- 이 줄을 삭제하거나 주석 처리하세요.
    )

    # --- ▼▼▼ [신규] 통계 자동 재계산 로직 ▼▼▼ ---
    # 1. 세션에서 설정값과 원본 데이터를 불러옵니다.
    df_cumulative = st.session_state["df_cumulative"]
    all_personnel_stats = results["all_personnel_stats"] # <-- 이 줄을 추가하세요.
    columns = results["columns"]
    all_slots = results["all_slots"]
    morning_duty_slot = results["morning_duty_slot"]
    special_dates = results["special_dates"]
    time_slots = results["time_slots"]
    time_order = results["time_order"]
    
    # --- [수정] 통계(Stats) 계산 로직 (L2208의 올바른 로직을 여기로 가져옴) ---
    
    # 1. 'df_room' (최종 방배정 결과)를 기반으로 'total_stats'를 (재)계산합니다.
    # (이것이 '오전당직(온콜)'이 포함된 가장 정확한 통계입니다)
    total_stats = {
        'early': Counter(), 
        'late': Counter(), 
        'morning_duty': Counter(), 
        'afternoon_duty': Counter(), 
        'rooms': {str(i): Counter() for i in range(1, 13)}, 
        'time_room_slots': {s: Counter() for s in st.session_state["time_slots"].keys()}
    }
    
    for _, row in edited_df_room.iterrows():
        current_date_str = row['날짜']
        if current_date_str in special_dates:
            continue # 토요/휴일은 통계에 포함 안 함

        # 'columns' 리스트와 'row' (Series)를 매핑
        assignment_for_day = row[columns[2:]] # '날짜', '요일' 제외

        for slot_name, person in assignment_for_day.items():
            if not person:
                continue
            
            # 1. 오전 당직
            if slot_name == morning_duty_slot:
                total_stats['morning_duty'][person] += 1
            
            # 2. 오후 당직 (13:30 당직)
            elif slot_name.startswith('13:30') and slot_name.endswith('_당직'):
                total_stats['afternoon_duty'][person] += 1
            
            # 3. [삭제] 온콜 (오후 당직으로 합산) - 이 로직은 화면 재계산 로직과 동일하게 삭제
            
            # 4. 이른방 (8:30, 당직 제외)
            elif slot_name.startswith('8:30') and '_당직' not in slot_name:
                total_stats['early'][person] += 1
            
            # 5. 늦은방 (10:00)
            elif slot_name.startswith('10:00'):
                total_stats['late'][person] += 1
            
            # 6. 시간대별/방별 통계 (for stats_df)
            if slot_name in total_stats['time_room_slots']:
                total_stats['time_room_slots'][slot_name][person] += 1

    time_order = ['8:30', '9:00', '9:30', '10:00', '13:30']

    # 2. 통계 DataFrame을 생성합니다. (L2200 로직과 동일)
    stats_data, all_personnel_stats = [], set(p for _, r in st.session_state["df_schedule_md"].iterrows() for p in r[2:].dropna() if p)
    df_cumulative_stats = st.session_state.get("df_cumulative", pd.DataFrame())
    
# [수정] "진짜 베이스 누적"을 계산하기 위해 4개의 맵을 모두 로드합니다.
    old_pm_cumulative_map = {}
    old_pm_source_map = {}
    old_am_cumulative_map = {} 
    old_am_source_map = {}
    
    if not df_cumulative.empty and '이름' in df_cumulative.columns:
        if '오후당직누적' in df_cumulative.columns:
            old_pm_cumulative_map = df_cumulative.set_index('이름')['오후당직누적'].to_dict()
        if '오후당직' in df_cumulative.columns:
            old_pm_source_map = df_cumulative.set_index('이름')['오후당직'].to_dict()
        if '오전당직누적' in df_cumulative.columns:
            old_am_cumulative_map = df_cumulative.set_index('이름')['오전당직누적'].to_dict()
        if '오전당직' in df_cumulative.columns:
            old_am_source_map = df_cumulative.set_index('이름')['오전당직'].to_dict()

    for person in sorted(all_personnel_stats):
        
        # --- [수정] 오후 당직: (Old 누적 - Old 합계) + New 합계 ---
        pm_assigned_this_month = total_stats['afternoon_duty'][person] # New 합계
        pm_old_cumulative = 0
        pm_old_source = 0
        try:
            pm_old_cumulative = int(old_pm_cumulative_map.get(person, 0)) # Old 누적
            pm_old_source = int(old_pm_source_map.get(person, 0))     # Old 합계
        except (ValueError, TypeError):
            pass
        true_pm_base = pm_old_cumulative - pm_old_source # 진짜 베이스
        pm_final_cumulative = true_pm_base + pm_assigned_this_month # 최종 누적
        
        # --- [수정] 오전 당직: (Old 누적 - Old 합계) + New 합계 ---
        am_assigned_this_month = total_stats['morning_duty'][person] # New 합계
        am_old_cumulative = 0
        am_old_source = 0
        try:
            am_old_cumulative = int(old_am_cumulative_map.get(person, 0)) # Old 누적
            am_old_source = int(old_am_source_map.get(person, 0))     # Old 합계
        except (ValueError, TypeError):
            pass
        true_am_base = am_old_cumulative - am_old_source # 진짜 베이스
        am_final_cumulative = true_am_base + am_assigned_this_month # 최종 누적
        
        stats_entry = {
            '인원': person,
            '이른방 합계': total_stats['early'][person],
            '늦은방 합계': total_stats['late'][person],
            '오전당직 합계': am_assigned_this_month,
            '오전당직 누적': am_final_cumulative,
            '오후당직 합계': pm_assigned_this_month,
            '오후당직 누적': pm_final_cumulative
        }
        for slot in time_slots.keys():
            if not slot.endswith('_당직'):
                stats_entry[f'{slot} 합계'] = total_stats['time_room_slots'].get(slot, Counter())[person]
        stats_data.append(stats_entry)

    # [수정] '오전당직 누적'을 포함하도록 컬럼 목록 수정
    sorted_columns = ['인원', '이른방 합계', '늦은방 합계', '오전당직 합계', '오전당직 누적', '오후당직 합계', '오후당직 누적']
    
    # 시간대별 합계 컬럼 추가
    time_slots_sorted = sorted(
        [slot for slot in time_slots.keys() if not slot.endswith('_당직')],
        key=lambda x: (time_order.index(x.split('(')[0]), int(x.split('(')[1].split(')')[0]))
    )
    sorted_columns.extend([f'{slot} 합계' for slot in time_slots_sorted])
    
    # 4. 'recalculated_stats_df' 라는 새 변수에 자동 재계산된 통계를 저장
    recalculated_stats_df_names_as_rows = pd.DataFrame(stats_data)[sorted_columns]
    # [신규] (항목-행) 기준으로 Transpose 하여 최종 df 생성
    recalculated_stats_df = recalculated_stats_df_names_as_rows.set_index('인원').transpose().reset_index().rename(columns={'index': '항목'})

    # --- ▼▼▼ [수정] 방배정 로그 로직 (기존과 동일) ▼▼▼ ---
    st.markdown("📝 **방배정 스케줄 수정사항**")
    room_change_log = []
    original_room_df = results["df_room"]
    if not edited_df_room.equals(original_room_df):
        try:
            diff_indices = np.where(edited_df_room.astype(str).ne(original_room_df.astype(str)))
            changed_cells = set(zip(diff_indices[0], diff_indices[1]))
            for row_idx, col_idx in changed_cells:
                date_str = edited_df_room.iloc[row_idx, 0]
                slot_name = edited_df_room.columns[col_idx]
                old_value = original_room_df.iloc[row_idx, col_idx]
                new_value = edited_df_room.iloc[row_idx, col_idx]
                log_msg = f"{date_str} '{slot_name}' 변경: '{old_value}' → '{new_value}'"
                room_change_log.append(log_msg)
        except Exception as e:
            room_change_log.append(f"[로그 오류] 방배정 변경사항을 비교하는 중 오류: {e}")
    if room_change_log:
        st.code("\n".join(f"• {msg}" for msg in sorted(room_change_log)), language='text')
    else:
        st.info("수정된 사항이 없습니다.")
    # --- ▲▲▲ 방배정 로그 끝 ---
    

    st.divider()
    st.markdown("**☑️ 통계 테이블 (수정 가능)**")
    st.write("- 통계 테이블은 '방배정 스케줄' 편집기에 반영된 내용을 바탕으로 자동 재계산됩니다.")
    
    # --- ▼▼▼ [수정] 통계 편집기에 'recalculated_stats_df'를 전달 ---
    edited_stats_df = st.data_editor(
        recalculated_stats_df, # [수정] stats_df -> recalculated_stats_df
        hide_index=True,
        key="stats_editor",
        disabled=['항목'],
        on_change=set_editor_changed_flag
    )
    # --- ▲▲▲ [수정] 완료 ---

    # --- ▼▼▼ [수정] 통계 로그 로직 (기존과 동일) ▼▼▼ ---
# --- [수정] 통계 로그 로직 (L2331 ~ L2368 교체) ---

    st.markdown("📝 **통계 테이블 수정사항**")
    stats_change_log = [] # 리스트 초기화

    # --- ▼▼▼ [신규] 항목 순서 정렬을 위한 맵(Map) 정의 ▼▼▼ ---
    # 사용자가 요청한 정확한 순서
    desired_order = [
        "이른방 합계", "늦은방 합계", 
        "오전당직 합계", "오전당직 누적", 
        "오후당직 합계", "오후당직 누적"
    ]
    # 항목 이름을 정렬 순서(숫자)로 매핑
    # (예: '이른방 합계': 0, '오전당직 누적': 3)
    order_map = {item_name: index for index, item_name in enumerate(desired_order)}
    # --- ▲▲▲ [신규] 완료 ---

    # 2. 인원별 통계 비교
    # (L2334) 비교 대상을 '최초 원본'으로 설정 (이전 단계에서 수정 완료됨)
    original_stats_df = results["stats_df"] 
    
    # '최종 편집본'(edited_stats_df)과 '최초 원본'(original_stats_df)을 비교
    if not edited_stats_df.equals(original_stats_df): 
        try:
            stats_orig_str = original_stats_df.astype(str)
            stats_edit_str = edited_stats_df.astype(str)
            
            diff_indices = np.where(stats_edit_str.ne(stats_orig_str))
            changed_cells_stats = set(zip(diff_indices[0], diff_indices[1])) 

            for row_idx, col_idx in changed_cells_stats:
                stat_name = edited_stats_df.iloc[row_idx, 0] # '항목' (예: "오전당직 합계")
                person_name = edited_stats_df.columns[col_idx] # '인원' (예: "강승주")
                
                old_value = original_stats_df.iloc[row_idx, col_idx]
                new_value = edited_stats_df.iloc[row_idx, col_idx]

                log_msg = f"{person_name} '{stat_name}' 변경: {old_value} → {new_value}"
                
                # --- ▼▼▼ [수정] 로그를 튜플로 저장 (정렬 기준 포함) ▼▼▼ ---
                
                # 1. 맵에서 정렬 순서(숫자)를 조회.
                #    '이른방' 등은 0-5, 맵에 없는 '8:30(4) 합계' 등은 99로 설정
                item_sort_key = order_map.get(stat_name, 99) 
                
                # 2. (사람이름, 항목순서, 실제로그메시지) 튜플로 저장
                stats_change_log.append((person_name, item_sort_key, log_msg))
                
                # --- ▲▲▲ [수정] (기존 log_msg.append(log_msg) 줄은 삭제) ---
                
        except Exception as e:
            # (예외 발생 시 튜플 대신 문자열로 추가)
            stats_change_log.append(("[로그 오류]", 99, f"[로그 오류] 통계 변경사항을 비교하는 중 오류: {e}"))
            
    if stats_change_log:
        # --- ▼▼▼ [수정] 튜플을 기준으로 정렬 (사람이름 > 항목순서) ▼▼▼ ---
        
        # 1. 튜플의 첫 번째 요소 (x[0] = person_name)로 먼저 정렬
        # 2. 튜플의 두 번째 요소 (x[1] = item_sort_key)로 두 번째 정렬
        stats_change_log.sort(key=lambda x: (x[0], x[1]))
        
        # 3. 정렬된 튜플 리스트에서 실제 로그 메시지(x[2])만 추출
        log_text_stats = "\n".join(f"• {msg_tuple[2]}" for msg_tuple in stats_change_log)
        st.code(log_text_stats, language='text')
        # --- ▲▲▲ [수정] 완료 ---
    else:
        st.info("수정된 사항이 없습니다.")
    
    # --- [수정 완료] ---
    st.divider()

    # --- ▼▼▼ [수정] 저장/다운로드 버튼 영역 수정 ▼▼▼ ---
    
    # 1. 변경사항 여부를 col1, col2를 정의하기 *전에* 계산합니다.
    has_unsaved_changes = bool(room_change_log or stats_change_log)
    
    col1, col2 = st.columns(2)
    with col1:
        # 2. '저장' 버튼에 disabled 파라미터를 적용합니다.
        if st.button("💾 수정사항 Google Sheet에 저장", 
                    type="primary", 
                    use_container_width=True,
                    disabled=not has_unsaved_changes): # <-- 이 부분이 추가/수정되었습니다.
            
            room_change_map = {}
            # 'results'에서 'df_room' (배정 직후 원본)을 가져옵니다.
            original_room_df_for_map = st.session_state.assignment_results["df_room"]
            
            # 'edited_df_room'은 L2551에서 data_editor의 결과로 이미 정의됨
            if not edited_df_room.equals(original_room_df_for_map):
                diff_indices = np.where(edited_df_room.astype(str).ne(original_room_df_for_map.astype(str)))
                changed_cells = set(zip(diff_indices[0], diff_indices[1]))
                
                for row_idx, col_idx in changed_cells:
                    date_key = edited_df_room.iloc[row_idx, 0] # '날짜' (e.g., '10월 1일')
                    slot_key = edited_df_room.columns[col_idx] # '슬롯' (e.g., '8:30(1)')
                    
                    old_value = original_room_df_for_map.iloc[row_idx, col_idx]
                    new_value = edited_df_room.iloc[row_idx, col_idx]
                    
                    # 키: (날짜, 슬롯이름), 값: (이전 값, 새 값)
                    room_change_map[(date_key, slot_key)] = (str(old_value).strip(), str(new_value).strip())

            # (이하 저장 로직은 동일)
            if not edited_df_room.empty:
                with st.spinner("수정된 '방배정 결과' 저장 중..."):
                    try:
                        gc = get_gspread_client()
                        sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
                        schedule_sheet_name = f"{month_str} 방배정" 

                        try: 
                            ws_sched = sheet.worksheet(schedule_sheet_name)
                        except gspread.exceptions.WorksheetNotFound: 
                            ws_sched = sheet.add_worksheet(title=schedule_sheet_name, rows=100, cols=len(edited_df_room.columns)+5)
                        
                        # [수정] edited_df_room과 edited_stats_df를 저장해야 함
                        success_sched = update_sheet_with_retry(ws_sched, [edited_df_room.columns.tolist()] + edited_df_room.astype(str).fillna('').values.tolist())

                        # [필요시] 통계 테이블(edited_stats_df) 저장 로직 추가 (현재는 '방배정' 시트만 저장)
                        if success_sched:
                            st.success(f"✅ '{schedule_sheet_name}' 시트에 수정된 내용이 저장되었습니다.")

                            # --- ▼▼▼ [수정] Excel 파일(output)을 여기서 갱신합니다 ▼▼▼ ---
                            
                            # 1. 갱신에 필요한 변수들을 results에서 다시 로드
                            results = st.session_state["assignment_results"]
                            swapped_assignments = st.session_state.get("swapped_assignments", set()) # 수동/일괄 적용 포함
                            special_df = st.session_state.get("special_df", pd.DataFrame())

                            # change_log_map_for_save = results.get("change_log_map", {})

                            # 2. 갱신된 df(edited_df_room, edited_stats_df)로 Excel 재생성
                            new_output = generate_excel_output(
                                df_room=edited_df_room,
                                stats_df=edited_stats_df,
                                columns=results["columns"],
                                special_dates=results["special_dates"],
                                special_df=special_df,
                                date_cache=results["date_cache"],
                                request_cells=results["request_cells"],
                                swapped_assignments=swapped_assignments,
                                morning_duty_slot=results["morning_duty_slot"],
                                month_str=month_str,
                                change_log_map=room_change_map
                            )
                            
                            # 3. 세션의 원본 데이터 및 'excel_output'을 갱신
                            st.session_state.assignment_results["df_room"] = edited_df_room.copy()
                            st.session_state.assignment_results["stats_df"] = edited_stats_df.copy()
                            st.session_state.assignment_results["excel_output"] = new_output # <-- 핵심
                            # --- ▲▲▲ [수정] 완료 ---
                            
                            
                            st.session_state.editor_has_changes = False
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error("Google Sheets 업데이트가 완료되지 않았습니다.")
                    except Exception as e:
                        st.error(f"Google Sheets 저장 중 오류 발생: {e}")
                        st.session_state.editor_has_changes = True
            else:
                st.error("편집된 데이터가 없습니다.")

# [L2331 부근 - 수정 필요 없음, 참고용]
    with col2:
        # [수정] 다운로드 버튼의 오류 감지 로직 (기존과 동일)
        has_unsaved_changes = bool(room_change_log or stats_change_log)

        if has_unsaved_changes:
            st.error("⚠️ 수정사항이 감지되었습니다. 먼저 '수정사항 Google Sheet에 저장' 버튼을 눌러주세요.")
        else:
            st.download_button(
                label="📥 방배정 다운로드",
                data=output, # <-- 이 output이 L2308에서 갱신된 최신 파일임
                file_name=f"{month_str} 방배정.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )