import re
import streamlit as st
import pandas as pd
import numpy as np
import gspread
from collections import Counter
from google.oauth2.service_account import Credentials
import time
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
import menu
import os
from dateutil.relativedelta import relativedelta
from openpyxl.utils import get_column_letter

# --- 페이지 기본 설정 ---
st.set_page_config(page_title="방배정 변경", page_icon="🔄", layout="wide")
st.session_state.current_page = os.path.basename(__file__)
menu.menu()

# --- 로그인 확인 ---
if not st.session_state.get("login_success", False):
    st.warning("⚠️ Home 페이지에서 먼저 로그인해주세요.")
    st.error("1초 후 Home 페이지로 돌아갑니다...")
    time.sleep(1)
    st.switch_page("Home.py")
    st.stop()

# --- 세션 상태 초기화 ---
if "change_data_loaded" not in st.session_state:
    st.session_state["change_data_loaded"] = False
if "saved_changes_log" not in st.session_state:
    st.session_state["saved_changes_log"] = []
if "df_final_assignment" not in st.session_state:
    st.session_state["df_final_assignment"] = pd.DataFrame()
if "df_change_requests" not in st.session_state:
    st.session_state["df_change_requests"] = pd.DataFrame()
if "changed_cells_log" not in st.session_state:
    st.session_state["changed_cells_log"] = []
if "df_before_apply" not in st.session_state:
    st.session_state["df_before_apply"] = pd.DataFrame()
if "has_changes_to_revert" not in st.session_state:
    st.session_state["has_changes_to_revert"] = False
if 'download_file' not in st.session_state:
    st.session_state.download_file = None
if 'download_filename' not in st.session_state:
    st.session_state.download_filename = None
if 'page7_messages' not in st.session_state:
    st.session_state['page7_messages'] = []
if "editor_key" not in st.session_state:
    st.session_state["editor_key"] = 0
    
# --- Google Sheets 연동 함수 ---
def get_gspread_client():
    scope = ["https://www.googleapis.com/auth/spreadsheets"]
    try:
        service_account_info = dict(st.secrets["gspread"])
        service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
        credentials = Credentials.from_service_account_info(service_account_info, scopes=scope)
        return gspread.authorize(credentials)
    except gspread.exceptions.APIError as e:
        st.warning("⚠️ 너무 많은 요청이 접속되어 딜레이되고 있습니다. 잠시 후 재시도 해주세요.")
        st.error(f"Google Sheets API 오류 (클라이언트 초기화): {e.response.status_code} - {e.response.text}")
        st.stop()
    except NameError as e:
        st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
        st.error(f"Google Sheets 인증 정보 로드 중 오류: {type(e).__name__} - {e}")
        st.stop()
    except Exception as e:
        st.warning("⚠️ 새로고침 버튼을 눌러 데이터를 다시 로드해주십시오.")
        st.error(f"Google Sheets 클라이언트 초기화 또는 인증 실패: {type(e).__name__} - {e}")
        st.stop()

def update_sheet_with_retry(worksheet, data, retries=5, delay=10):
    for attempt in range(retries):
        try:
            worksheet.clear()
            worksheet.update('A1', data, value_input_option='RAW')
            return
        except Exception as e:
            if "Quota exceeded" in str(e):
                st.warning(f"API 쿼터 초과, {delay}초 후 재시도 ({attempt+1}/{retries})")
                time.sleep(delay)
            else:
                st.error(f"업데이트 실패, {delay}초 후 재시도 ({attempt+1}/{retries}): {str(e)}")
                time.sleep(delay)
    st.error("Google Sheets 업데이트 실패: 재시도 횟수 초과")

@st.cache_data(ttl=600, show_spinner=False)
def load_data_for_change_page(month_str):
    try:
        gc = get_gspread_client()
        sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
    except Exception as e:
        st.error(f"스프레드시트 열기 실패: {e}")
        return "STOP", None, None, None # 반환값 개수 4개로 변경

    # 1. 방배정 시트 (우선순위: 최종 > ver1.0)
    final_name = f"{month_str} 방배정 최종"
    ver1_name = f"{month_str} 방배정 ver1.0"
    load_status = "" # 'Final' 또는 'ver1.0'

    try:
        # 모든 시트 이름 가져오기
        all_ws_titles = [ws.title for ws in sheet.worksheets()]

        if final_name in all_ws_titles:
            worksheet_final = sheet.worksheet(final_name)
            load_status = "Final"
        elif ver1_name in all_ws_titles:
            worksheet_final = sheet.worksheet(ver1_name)
            load_status = "ver1.0"
        else:
            st.info(f"{month_str} 방배정이 아직 수행되지 않았습니다.")
            return "STOP", None, None, None

        df_final = pd.DataFrame(worksheet_final.get_all_records()).fillna('')
        
    except Exception as e:
        st.error(f"방배정 데이터 로드 중 오류: {e}")
        return "STOP", None, None, None

    # 2. 변경요청 시트
    try:
        worksheet_req = sheet.worksheet(f"{month_str} 방배정 변경요청")
        df_req = pd.DataFrame(worksheet_req.get_all_records())
    except:
        df_req = pd.DataFrame(columns=['RequestID', '요청일시', '요청자', '변경 요청', '변경 요청한 방배정'])

    # 3. 누적 데이터 시트 (기존 로직 유지)
    df_cumulative = pd.DataFrame()
    try:
        target_dt = datetime.strptime(month_str, "%Y년 %m월")
        next_dt = target_dt + relativedelta(months=1)
        next_month_str = next_dt.strftime("%Y년 %-m월")

        cum_name = f"{next_month_str} 누적 최종"
        all_titles = [ws.title for ws in sheet.worksheets()]
        if cum_name not in all_titles:
            cum_name = f"{next_month_str} 누적"

        if cum_name in all_titles:
            ws = sheet.worksheet(cum_name)
            vals = ws.get_all_values()
            if len(vals) > 1:
                headers = vals[0]
                data = vals[1:]
                df_cumulative = pd.DataFrame(data, columns=headers)
                if '항목' in df_cumulative.columns:
                    df_cumulative.set_index('항목', inplace=True)
                df_cumulative = df_cumulative.apply(pd.to_numeric, errors='coerce').fillna(0).astype(int)
                
    except Exception as e:
        print(f"누적 로드 실패: {e}")

    # [수정] 4개의 값을 반환합니다.
    return df_final, df_req, df_cumulative, load_status

@st.cache_data(ttl=600, show_spinner=False)
def load_special_schedules(month_str):
    """
    'YYYY년 토요/휴일 스케줄' 시트에서 특정 월의 데이터를 로드합니다.
    연도는 month_str에서 동적으로 추출합니다.
    """
    try:
        gc = get_gspread_client()
        if not gc: return pd.DataFrame()
        
        spreadsheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
        
        # 1. month_str에서 연도를 동적으로 추출하여 시트 이름을 생성합니다.
        target_year = month_str.split('년')[0]
        sheet_name = f"{target_year}년 토요/휴일 스케줄"
        
        worksheet = spreadsheet.worksheet(sheet_name)
        records = worksheet.get_all_records()
        
        if not records:
            return pd.DataFrame()
        
        df = pd.DataFrame(records)

        # 2. '날짜'와 '근무' 열이 있는지 확인합니다.
        if '날짜' not in df.columns or '근무' not in df.columns:
            st.error(f"'{sheet_name}' 시트에 '날짜' 또는 '근무' 열이 없습니다.")
            return pd.DataFrame()

        df.fillna('', inplace=True)
        df['날짜_dt'] = pd.to_datetime(df['날짜'], format='%Y-%m-%d', errors='coerce')
        df.dropna(subset=['날짜_dt'], inplace=True)

        # 3. 'month_str'에 해당하는 월의 데이터만 필터링합니다.
        target_month_dt = datetime.strptime(month_str, "%Y년 %m월")
        df_filtered = df[
            (df['날짜_dt'].dt.year == target_month_dt.year) &
            (df['날짜_dt'].dt.month == target_month_dt.month)
        ].copy()

        return df_filtered
        
    except gspread.exceptions.WorksheetNotFound:
        target_year = month_str.split('년')[0]
        sheet_name = f"{target_year}년 토요/휴일 스케줄"
        st.info(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"토요/휴일 데이터 로드 중 오류 발생: {str(e)}")
        return pd.DataFrame()

def apply_assignment_swaps(df_assignment, df_requests, df_special):
    df_modified = df_assignment.copy()
    df_special_modified = df_special.copy() if df_special is not None else pd.DataFrame()
    changed_log = []
    applied_count = 0
    # [수정] 메시지를 담을 리스트 생성
    messages = []

    for _, req in df_requests.iterrows():
        try:
            swap_request_str = str(req.get('변경 요청', '')).strip()
            raw_slot_info = str(req.get('변경 요청한 방배정', '')).strip()

            if '➡️' not in swap_request_str: continue
            old_person, new_person = [p.strip() for p in swap_request_str.split('➡️')]
            
            slot_match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', raw_slot_info)
            if not slot_match: continue
            
            date_str, target_slot = slot_match.groups()
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            target_date_str = f"{date_obj.month}월 {date_obj.day}일"
            
            row_indices = df_modified.index[df_modified['날짜'] == target_date_str].tolist()
            if not row_indices:
                # [수정] 메시지 리스트에 추가
                messages.append(('warning', f"⚠️ 요청 처리 불가: 방배정표에서 날짜 '{target_date_str}'를 찾을 수 없습니다."))
                continue
            target_row_idx = row_indices[0]

            target_col_found = None
            for col in df_modified.columns[2:]: 
                person_in_cell = str(df_modified.at[target_row_idx, col]).strip()
                if person_in_cell == old_person and col == target_slot:
                    target_col_found = col
                    break
            
            if target_col_found:
                df_modified.at[target_row_idx, target_col_found] = new_person
                applied_count += 1
                
                is_special_date = False
                if df_special is not None and not df_special.empty and '날짜_dt' in df_special.columns:
                    is_special_date = not df_special[df_special['날짜_dt'].dt.date == date_obj.date()].empty
                
                if is_special_date and not df_special_modified.empty:
                    duty_row = df_special_modified[df_special_modified['날짜_dt'].dt.date == date_obj.date()]
                    if not duty_row.empty:
                        current_duty_person = str(duty_row['당직'].iloc[0]).strip()
                        if current_duty_person == old_person:
                            df_special_modified.loc[duty_row.index, '당직'] = new_person
                            # [수정] 메시지 리스트에 추가
                            messages.append(('info', f"ℹ️ {target_date_str}의 토요/휴일 당직자가 '{new_person}' (으)로 함께 변경됩니다."))

                changed_log.append({
                    '날짜': f"{target_date_str} ({'월화수목금토일'[date_obj.weekday()]})",
                    '방배정': target_slot,
                    '변경 전 인원': old_person,
                    '변경 후 인원': new_person,
                })
            else:
                # [수정] 메시지 리스트에 추가
                messages.append(('error', f"❌ 적용 실패: {target_date_str}의 '{target_slot}'에 '{old_person}'님이 배정되어 있지 않습니다."))
                
        except Exception as e:
            # [수정] 메시지 리스트에 추가
            messages.append(('error', f"⚠️ 요청 처리 중 시스템 오류 발생: {e}"))

    if applied_count > 0:
        # [수정] 메시지 리스트에 추가 (가장 위로)
        messages.insert(0, ('success', f"🎉 총 {applied_count}건의 변경 요청이 반영되었습니다."))
    elif not df_requests.empty and not messages:
        messages.append(('info', "ℹ️ 새롭게 반영할 유효한 변경 요청이 없습니다."))

    # [수정] df_modified, 로그, 그리고 '메시지 리스트'를 함께 반환
    return df_modified, changed_log, df_special_modified, messages
    
# --- 시간대 순서 정의 ---
time_order = ['8:30', '9:00', '9:30', '10:00', '13:30']

def calculate_statistics(result_df: pd.DataFrame, df_special: pd.DataFrame, df_cumulative: pd.DataFrame) -> pd.DataFrame:
    # 1. 화면의 스케줄표 카운팅 (이름별 카운트)
    # (행렬 연산을 위해 이름을 키로 하는 딕셔너리 생성)
    total_stats = {
        'early': Counter(), 'late': Counter(),
        'morning_duty': Counter(), 'afternoon_duty': Counter(),
        'time_room_slots': {} 
    }
    
    # 날짜/인원 처리
    special_dates = []
    if df_special is not None and not df_special.empty and '날짜_dt' in df_special.columns:
        special_dates = df_special['날짜_dt'].dt.strftime('%#m월 %#d일').tolist() if os.name != 'nt' else df_special['날짜_dt'].dt.strftime('%m월 %d일').apply(lambda x: x.lstrip("0").replace(" 0", " "))
    
    all_personnel_raw = pd.unique(result_df.iloc[:, 2:].values.ravel('K'))
    all_personnel = sorted(list({re.sub(r'\[\d+\]', '', str(p)).strip() for p in all_personnel_raw if pd.notna(p) and str(p).strip()}))
    SMALL_TEAM_THRESHOLD = 13
    
    # 슬롯 초기화
    for col in result_df.columns[2:]:
        if col != '온콜': total_stats['time_room_slots'].setdefault(col, Counter())

    # 카운팅
    for _, row in result_df.iterrows():
        if str(row['날짜']).strip() in special_dates: continue
        personnel = [p for p in row.iloc[2:].dropna() if p]
        if 0 < len(personnel) < 13: continue

        for col in result_df.columns[2:]:
            person = row.get(col)
            if not person: continue
            p = re.sub(r'\[\d+\]', '', str(person)).strip()
            
            if col != '온콜':
                total_stats['time_room_slots'][col][p] += 1
            if col.startswith('8:30') and '_당직' not in col:
                total_stats['early'][p] += 1
            elif col.startswith('10:00'):
                total_stats['late'][p] += 1
            
            # 당직 카운팅 (화면 기준 실시간)
            if col == '온콜' or (col.startswith('8:30') and '_당직' in col):
                total_stats['morning_duty'][p] += 1
            elif col.startswith('13:30') and '_당직' in col:
                total_stats['afternoon_duty'][p] += 1

    # 2. 결과 데이터프레임 생성 (Index=항목, Columns=이름)
    # 시트 형식을 따름
    rows_list = [
        '이른방 합계', '늦은방 합계', 
        '오전당직', '오전당직 누적', 
        '오후당직', '오후당직 누적'
    ]
    
    # 시간대별 합계 행 추가
    time_order = ['8:30', '9:00', '9:30', '10:00', '13:30']
    sorted_slots = sorted([s for s in total_stats['time_room_slots'].keys() if '_당직' not in s],
                          key=lambda x: (time_order.index(x.split('(')[0]), x))
    for s in sorted_slots:
        rows_list.append(f"{s} 합계")
    
    # 빈 DataFrame 생성 (시트와 같은 모양)
    stats_df = pd.DataFrame(index=rows_list, columns=all_personnel)
    stats_df = stats_df.fillna(0) # 기본값 0

    # 3. 데이터 채우기
    # df_cumulative는 이미 (Index=항목, Columns=이름) 상태임.
    
    for p in all_personnel:
        # (1) 기본 카운트 채우기
        stats_df.at['이른방 합계', p] = total_stats['early'][p]
        stats_df.at['늦은방 합계', p] = total_stats['late'][p]
        stats_df.at['오전당직', p] = total_stats['morning_duty'][p] # 화면 값
        stats_df.at['오후당직', p] = total_stats['afternoon_duty'][p] # 화면 값

        for s in sorted_slots:
             stats_df.at[f"{s} 합계", p] = total_stats['time_room_slots'][s][p]

        # (2) 누적 계산 (시트 값 참조)
        # df_cumulative에서 해당 사람(p)의 데이터를 가져옴
        old_am_cum = 0
        old_am_sum = 0
        old_pm_cum = 0
        old_pm_sum = 0

        if not df_cumulative.empty and p in df_cumulative.columns:
            try:
                # df_cumulative는 Index가 '항목'임
                if '오전당직누적' in df_cumulative.index: old_am_cum = int(df_cumulative.at['오전당직누적', p])
                if '오전당직' in df_cumulative.index: old_am_sum = int(df_cumulative.at['오전당직', p])
                
                if '오후당직누적' in df_cumulative.index: old_pm_cum = int(df_cumulative.at['오후당직누적', p])
                if '오후당직' in df_cumulative.index: old_pm_sum = int(df_cumulative.at['오후당직', p])
            except: pass

        # 계산: (시트누적 - 시트합계) + 화면합계
        stats_df.at['오전당직 누적', p] = (old_am_cum - old_am_sum) + total_stats['morning_duty'][p]
        stats_df.at['오후당직 누적', p] = (old_pm_cum - old_pm_sum) + total_stats['afternoon_duty'][p]

    # 최종: '항목'을 컬럼으로 꺼내서 반환 (Streamlit 표시용)
    return stats_df.reset_index().rename(columns={'index': '항목'})

@st.cache_data(ttl=300, show_spinner=False)
def check_final_sheets_exist(month_str, next_month_str):
    """
    지정된 월의 '방배정 최종' 시트와 다음 달의 '누적 최종' 시트가 
    이미 존재하는지 확인하여 True/False를 반환합니다.
    """
    try:
        # 1. 구글 시트 연결
        gc = get_gspread_client()
        if not gc:
            return False
            
        sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
        
        # 2. 현재 존재하는 모든 시트 이름 가져오기
        all_titles = [ws.title for ws in sheet.worksheets()]
        
        # 3. 확인할 시트 이름 정의
        # (1) 현재 달의 최종 방배정 결과
        schedule_sheet_name = f"{month_str} 방배정 최종"
        # (2) 다음 달의 최종 누적 데이터
        cumulative_sheet_name = f"{next_month_str} 누적 최종"
        
        # 4. 존재 여부 확인 (둘 중 하나라도 있으면 True 반환)
        if schedule_sheet_name in all_titles or cumulative_sheet_name in all_titles:
            return True
            
        return False

    except Exception as e:
        # 에러 발생 시 (연결 실패 등) False 반환하여 진행 막지 않음
        # 필요 시 st.error(f"확인 중 오류: {e}") 추가 가능
        return False


# =============================================================================
# ▼▼▼ [통계 재계산 로직] 함수 정의 및 실행 ▼▼▼
# =============================================================================
    
# --- [수정됨] 통계 재계산 함수 ---
def calculate_stats_from_schedule(schedule_df):
    """스케줄 DataFrame을 입력받아 통계 DataFrame을 반환하는 함수"""
    if schedule_df is None or schedule_df.empty:
        return pd.DataFrame(columns=['항목'])

    # 1. 집계 카운터 초기화
    temp = {
        'early': Counter(), 'late': Counter(), 
        'morning_duty': Counter(), 'afternoon_duty': Counter(),
        'time_slots': Counter()
    }
    
    # 휴일 날짜 처리
    special_dates_s = set()
    if "df_special_schedules" in st.session_state and not st.session_state.df_special_schedules.empty:
        try: special_dates_s = set(st.session_state.df_special_schedules['날짜'].astype(str).tolist())
        except: pass

    # 2. 스케줄 순회 및 카운트
    for _, row in schedule_df.iterrows():
        if str(row.iloc[0]) in special_dates_s: continue # 휴일 제외
        
        for col_name, val in row.items():
            if col_name in ['날짜', '요일'] or not val: continue
            person = str(val).replace(u'\xa0', ' ').strip()
            if not person: continue
            
            is_duty_slot = '_당직' in col_name or col_name == '오전당직(온콜)'
            
            # --- ▼▼▼ [핵심 수정] 방 번호 포함하여 카운트 ---
            # 기존: time_prefix = col_name.split('(')[0]  -> "8:30"만 추출됨
            # 수정: col_name 전체를 사용하여 "8:30(4)" 형식 유지 (단, 당직/온콜 제외)
            
            if not is_duty_slot and '온콜' not in col_name:
                # 슬롯 이름 그대로를 키로 사용 (예: "8:30(4)")
                # 단, 괄호가 없는 경우(예: 9:00)도 고려해야 함.
                # 하지만 방배정 페이지 특성상 모든 일반 방은 괄호를 포함하므로 col_name 사용이 안전함.
                
                # (추가) 만약 컬럼명에 '_당직' 같은 접미사가 없으면 일반 방으로 간주
                temp['time_slots'][(col_name, person)] += 1
            # --- ▲▲▲ [수정 완료] ---
            
            if col_name == '오전당직(온콜)' or (col_name.startswith('8:30') and '_당직' in col_name):
                temp['morning_duty'][person] += 1
            elif col_name.startswith('13:30') and '_당직' in col_name:
                temp['afternoon_duty'][person] += 1
            elif col_name.startswith('8:30') and not is_duty_slot:
                temp['early'][person] += 1
            elif col_name.startswith('10:00'):
                temp['late'][person] += 1

    # 3. 누적 데이터와 결합
    df_cum_base = st.session_state.get("df_cumulative", pd.DataFrame())
    # 누적값 로드
    map_am_cum = df_cum_base.set_index('이름')['오전당직누적'].to_dict() if not df_cum_base.empty and '오전당직누적' in df_cum_base.columns else {}
    map_am_src = df_cum_base.set_index('이름')['오전당직'].to_dict() if not df_cum_base.empty and '오전당직' in df_cum_base.columns else {}
    map_pm_cum = df_cum_base.set_index('이름')['오후당직누적'].to_dict() if not df_cum_base.empty and '오후당직누적' in df_cum_base.columns else {}
    map_pm_src = df_cum_base.set_index('이름')['오후당직'].to_dict() if not df_cum_base.empty and '오후당직' in df_cum_base.columns else {}
    
    # 인원 목록 추출
    active_p = set(temp['morning_duty'].keys()) | set(temp['afternoon_duty'].keys()) | \
            set(temp['early'].keys()) | set(temp['late'].keys()) | {p for t, p in temp['time_slots'].keys()}
    all_p = sorted(list(active_p | set(map_am_cum.keys())))
    
    rows_list = []
    
    # --- ▼▼▼ [핵심 수정] 슬롯 헤더 동적 생성 및 정렬 ---
    # 1. 카운트된 모든 슬롯 이름(Unique)을 수집
    collected_slots = sorted(list({t for t, p in temp['time_slots'].keys()}))
    
    # 2. 시간순 -> 방 번호순 정렬 함수
    def sort_key(slot_name):
        time_order = ['8:30', '9:00', '9:30', '10:00', '13:30']
        # 시간 추출 (예: "8:30")
        time_part = slot_name.split('(')[0]
        time_idx = time_order.index(time_part) if time_part in time_order else 99
        
        # 방 번호 추출 (예: "4")
        room_num = 0
        match = re.search(r'\((\d+)\)', slot_name)
        if match:
            room_num = int(match.group(1))
            
        return (time_idx, room_num)

    t_headers = sorted(collected_slots, key=sort_key)
    # --- ▲▲▲ [수정 완료] ---
    
    for p in all_p:
        # 누적 재계산
        am_fin = (int(map_am_cum.get(p, 0)) - int(map_am_src.get(p, 0))) + temp['morning_duty'][p]
        pm_fin = (int(map_pm_cum.get(p, 0)) - int(map_pm_src.get(p, 0))) + temp['afternoon_duty'][p]
        
        r = {
            '인원': p,
            '이른방 합계': temp['early'][p], '늦은방 합계': temp['late'][p],
            '오전당직': temp['morning_duty'][p], '오전당직 누적': am_fin,
            '오후당직': temp['afternoon_duty'][p], '오후당직 누적': pm_fin
        }
        # [수정] 동적으로 생성된 헤더를 사용하여 값 매핑
        for t in t_headers: 
            r[f'{t} 합계'] = temp['time_slots'][(t, p)]
            
        rows_list.append(r)
        
    if not rows_list: return pd.DataFrame(columns=['항목'])
    
    # [수정] 고정 컬럼 + 동적 슬롯 컬럼 합치기
    fixed_cols = ['인원', '이른방 합계', '늦은방 합계', '오전당직', '오전당직 누적', '오후당직', '오후당직 누적'] + [f'{t} 합계' for t in t_headers]
    
    res_df = pd.DataFrame(rows_list)
    for c in fixed_cols: 
        if c not in res_df.columns: res_df[c] = 0
        
    return res_df[fixed_cols].set_index('인원').transpose().reset_index().rename(columns={'index': '항목'})

# --- 엑셀 생성 함수 (중복 방지 및 서식 포함) ---
def create_formatted_excel(df_sched, df_stats):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "방배정 최종"

    # 스타일 정의
    import platform
    font_name = "맑은 고딕" if platform.system() == "Windows" else "Arial"
    
    font_bold = Font(name=font_name, size=9, bold=True)
    font_default = Font(name=font_name, size=9)
    font_duty = Font(name=font_name, size=9, bold=True, color="FF00FF")

    thin_side = Side(style='thin')
    thick_side = Side(style='medium')
    border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 색상
    fill_header_830 = PatternFill(start_color="FFE699", fill_type="solid")
    fill_header_900 = PatternFill(start_color="F8CBAD", fill_type="solid")
    fill_header_930 = PatternFill(start_color="B4C6E7", fill_type="solid")
    fill_header_1000 = PatternFill(start_color="C6E0B4", fill_type="solid")
    fill_header_1330 = PatternFill(start_color="CC99FF", fill_type="solid")
    fill_gray = PatternFill(start_color="808080", fill_type="solid")
    fill_light_gray = PatternFill(start_color="BFBFBF", fill_type="solid")
    fill_yoil = PatternFill(start_color="FFF2CC", fill_type="solid")
    fill_holiday = PatternFill(start_color="DDEBF7", fill_type="solid")
    fill_change = PatternFill(start_color="F2DCDB", fill_type="solid")
    
    fill_stats_header = PatternFill(start_color="E7E6E6", fill_type="solid")
    fill_stats_label = PatternFill(start_color="D0CECE", fill_type="solid")
    fill_row_early = PatternFill(start_color="FFE699", fill_type="solid")
    fill_row_late = PatternFill(start_color="C6E0B4", fill_type="solid")
    fill_row_am = PatternFill(start_color="B8CCE4", fill_type="solid")
    fill_row_cum = PatternFill(start_color="FFC8CD", fill_type="solid")

    # 1. 스케줄 테이블 작성
    cols = df_sched.columns.tolist()
    
    # 헤더
    for i, col in enumerate(cols, 1):
        cell = sheet.cell(1, i, col)
        cell.font = font_bold
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
        
        if '8:30' in col or '온콜' in col: cell.fill = fill_header_830
        elif '9:00' in col: cell.fill = fill_header_900
        elif '9:30' in col: cell.fill = fill_header_930
        elif '10:00' in col: cell.fill = fill_header_1000
        elif '13:30' in col: cell.fill = fill_header_1330
        # else: cell.fill = fill_gray

    # 데이터
    special_dates = []
    if st.session_state.df_special_schedules is not None:
        try: special_dates = [d.strftime('%-m월 %-d일').lstrip('0').replace(' 0', ' ') for d in st.session_state.df_special_schedules['날짜_dt']]
        except: pass

    # 비교용 원본 (없으면 현재 데이터)
    df_compare_base = st.session_state.get("df_before_apply", pd.DataFrame())

    last_row = 1
    for r, row in enumerate(df_sched.itertuples(index=False), 2):
        date_str = row[0]
        is_special = date_str in special_dates
        
        duty_name = None
        if is_special:
            try:
                # (주의: month_str은 외부 변수 참조)
                dt = datetime.strptime(date_str, '%m월 %d일').replace(year=int(month_str[:4]))
                d_str = dt.strftime('%Y-%m-%d')
                res = st.session_state.df_special_schedules[st.session_state.df_special_schedules['날짜']==d_str]
                if not res.empty: duty_name = str(res.iloc[0]['당직']).strip()
            except: pass

        personnel = [x for x in row[2:] if x]
        is_no_person = not any(personnel)
        is_small = 0 < len(personnel) < 15

        for c, val in enumerate(row, 1):
            cell = sheet.cell(r, c, val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
            cell.font = font_default

            # 배경색
            if c == 1: cell.fill = fill_gray
            elif c == 2:
                if is_no_person: cell.fill = fill_gray
                elif is_small or is_special: cell.fill = fill_light_gray
                else: cell.fill = fill_yoil
            elif is_no_person and c > 2: cell.fill = fill_gray
            
            if is_special and val and c > 2: cell.fill = fill_holiday

            # 변경사항 코멘트 (원본과 비교)
            val_str = str(val).strip() if pd.notna(val) else ""
            old_str = ""
            try:
                if not df_compare_base.empty and r-2 < len(df_compare_base):
                    old_str = str(df_compare_base.iat[r-2, c-1]).strip()
            except: pass
            
            if val_str != old_str:
                cell.fill = fill_change
                cell.comment = Comment(f"변경 전: {old_str if old_str else '빈 값'}", "Edit Tracker")

            # 폰트 (당직)
            if val:
                head = cols[c-1]
                if is_special:
                    if duty_name and val == duty_name: cell.font = font_duty
                else:
                    if '_당직' in head or '온콜' in head: cell.font = font_duty
        
        last_row = r

    # 2. 통계 테이블
    stats_start = last_row + 2
    stats_cols = df_stats.columns.tolist()
    
    # 헤더
    for i, col in enumerate(stats_cols, 1):
        cell = sheet.cell(stats_start, i, col)
        cell.font = font_bold
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = fill_stats_header
        
        cell.border = Border(
            left=thick_side if i==1 else thin_side,
            right=thick_side if i==len(stats_cols) else thin_side,
            top=thick_side,
            bottom=thick_side
        )

    # 데이터
    sep_items = ["늦은방 합계", "오전당직 누적", "오후당직 누적"]
    item_list = df_stats['항목'].tolist()
    prefixes = ["8:30(", "9:00(", "9:30(", "10:00("]
    for pf in prefixes:
        matches = [x for x in item_list if str(x).startswith(pf)]
        if matches: sep_items.append(matches[-1])

    for r, row in enumerate(df_stats.itertuples(index=False), stats_start + 1):
        item_name = str(row[0])
        is_last = (r == stats_start + len(df_stats))
        is_sep = (item_name in sep_items)

        row_fill = None
        if '이른방' in item_name: row_fill = fill_row_early
        elif '늦은방' in item_name: row_fill = fill_row_late
        elif item_name in ['오전당직', '오후당직']: row_fill = fill_row_am
        elif '누적' in item_name: row_fill = fill_row_cum

        for c, val in enumerate(row, 1):
            cell = sheet.cell(r, c, val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            cell.border = Border(
                left=thick_side if c==1 else thin_side,
                right=thick_side if c==len(stats_cols) else thin_side,
                top=thin_side,
                bottom=thick_side if is_last or is_sep else thin_side
            )

            if c == 1:
                cell.font = font_bold
                cell.fill = fill_stats_label
            else:
                cell.font = font_default
                if row_fill: cell.fill = row_fill

    # 열 너비
    sheet.column_dimensions['A'].width = 11
    for i in range(2, 50):
        sheet.column_dimensions[get_column_letter(i)].width = 10

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
    
# --- UI 및 데이터 핸들링 ---
from zoneinfo import ZoneInfo
kst = ZoneInfo("Asia/Seoul")
now = datetime.now(kst)
today = now.date()
next_month_date = today.replace(day=1) + relativedelta(months=1)
month_str = next_month_date.strftime("%Y년 %-m월")
st.header(f"🔄 {month_str} 방배정 변경", divider='rainbow')

def load_and_initialize_data():
    with st.spinner("데이터를 로드하고 있습니다..."):
        # [수정] 4개 반환값 언패킹
        df_final, df_req, df_cumulative, load_status = load_data_for_change_page(month_str)
    
    if isinstance(df_final, str) and df_final == "STOP":
        st.stop()
        
    df_special = load_special_schedules(month_str)
    
    st.session_state.df_final_assignment = df_final
    st.session_state.df_change_requests = df_req
    st.session_state.df_cumulative_stats = df_cumulative
    st.session_state.df_special_schedules = df_special
    
    # [추가] 로드된 버전 정보를 세션에 저장
    st.session_state.loaded_version_status = load_status 
    
    st.session_state.changed_cells_log = []
    st.session_state.df_before_apply = df_final.copy()
    st.session_state.has_changes_to_revert = False
    st.session_state.change_data_loaded = True

# 새로고침 버튼
col_text, col_btn = st.columns([3, 1], vertical_alignment="center")

with col_text:
    st.caption("ℹ️ 먼저 새로고침 버튼으로 최신 데이터를 불러온 뒤 진행해주세요.")

with col_btn:
    if st.button("🔄 새로고침 (R)", use_container_width=True):
        st.cache_data.clear()
        
        # [핵심 수정] 새로고침 시 '누적 최종' 시트도 강제 재로드
        try:
            gc = get_gspread_client()
            sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
            
            # 다음 달 계산 (현재 10월 -> 다음달 11월)
            curr_dt = datetime.strptime(month_str, "%Y년 %m월")
            next_dt = curr_dt + relativedelta(months=1)
            next_month_str = next_dt.strftime("%Y년 %-m월")
            
            # [수정] 여기서 변수를 정의해야 에러가 나지 않습니다.
            cum_sheet_name = f"{next_month_str} 누적 최종"
            
            # (디버깅용 출력은 주석 처리하거나 필요시 사용)
            # st.write(cum_sheet_name) 

            df_cumulative_reloaded = pd.DataFrame()
            
            try:
                # 1. '누적 최종' 시도
                ws_cum = sheet.worksheet(cum_sheet_name)
            except gspread.exceptions.WorksheetNotFound:
                # 2. 없으면 '누적' 시도 (혹시 모를 호환성)
                try:
                    cum_sheet_name = f"{next_month_str} 누적"
                    ws_cum = sheet.worksheet(cum_sheet_name)
                except gspread.exceptions.WorksheetNotFound:
                    ws_cum = None

            if ws_cum:
                d_cum = ws_cum.get_all_values()
                if len(d_cum) > 1:
                    headers = d_cum[0]
                    data = d_cum[1:]
                    df_cumulative_reloaded = pd.DataFrame(data, columns=headers)
                    
                    # 숫자 변환 및 인덱스 설정 (기존 로직과 동일하게)
                    if '항목' in df_cumulative_reloaded.columns:
                        df_cumulative_reloaded.set_index('항목', inplace=True)
                    
                    df_cumulative_reloaded = df_cumulative_reloaded.apply(pd.to_numeric, errors='coerce').fillna(0).astype(int)
            
            # 세션에 강제 주입 (이제 데이터가 살아있음)
            st.session_state.df_cumulative_stats = df_cumulative_reloaded
            # [추가] 통계 계산 함수가 참조하는 키에도 넣어줌
            st.session_state.df_cumulative = df_cumulative_reloaded
            
        except Exception as e:
            st.error(f"누적 데이터 재로드 실패: {e}")

        # 나머지 초기화 로직
        st.session_state.change_data_loaded = False
        st.session_state['page7_messages'] = []
        if 'show_final_results' in st.session_state:
            st.session_state['show_final_results'] = False
            
        st.rerun()

curr_dt = datetime.strptime(month_str, "%Y년 %m월")
next_dt = curr_dt + relativedelta(months=1)
next_month_str = next_dt.strftime("%Y년 %-m월")
final_sheets_exist = check_final_sheets_exist(month_str, next_month_str)

if check_final_sheets_exist(month_str, next_month_str):
    st.warning("방배정이 변경 완료되어, 현재 '방배정 최종' 버전이 이미 존재합니다.")

    c_dl, c_reset = st.columns([1, 1])

# --- A. 방배정 최종 다운로드 버튼 (V8: 캐시 강제 삭제 및 자료형 완전 통일) ---
    with c_dl:
        # [핵심 원인 해결] 기존에 잘못 생성된 파일이 세션에 남아있으면 무조건 지웁니다.
        # 이걸 안 하면 코드를 아무리 고쳐도 옛날 파일을 계속 다운로드하게 됩니다.
        if "final_download_ready" in st.session_state:
            del st.session_state["final_download_ready"]

        try:
            gc_tmp = get_gspread_client()
            if gc_tmp:
                sh_tmp = gc_tmp.open_by_url(st.secrets["google_sheet"]["url"])
                
                # 1. '방배정 최종' 스케줄 로드
                schedule_sheet_name = f"{month_str} 방배정 최종"
                try:
                    ws_final = sh_tmp.worksheet(schedule_sheet_name)
                    d_final = ws_final.get_all_values()
                    if len(d_final) < 2:
                        raise ValueError("스케줄 데이터가 비어있습니다.")
                    df_final_exist = pd.DataFrame(d_final[1:], columns=d_final[0])
                except gspread.exceptions.WorksheetNotFound:
                    st.warning("아직 '방배정 최종' 시트가 없습니다.")
                    st.stop()

                # -----------------------------------------------------------
                # [핵심 1] 배경색 해결: 자료형(Type) 강제 통일
                # -----------------------------------------------------------
                # NaN과 ""(빈문자열)은 다릅니다. 이걸 안 맞추면 엑셀 함수는 다르다고 판단해 분홍색을 칠합니다.
                # 불러온 데이터를 무조건 문자열로 변환하고 빈 값을 통일합니다.
                df_final_exist = df_final_exist.fillna("").astype(str)
                
                # 이제 "정제된 데이터"를 비교 기준(변경 전)으로 설정합니다.
                # 이러면 원본 vs 원본 비교가 되어 배경색이 칠해지지 않습니다.
                st.session_state["df_before_apply"] = df_final_exist.copy()

                # -----------------------------------------------------------
                # [핵심 2] 통계 계산 준비 (누적 데이터 로드)
                # -----------------------------------------------------------
                if "df_cumulative" not in st.session_state or st.session_state.df_cumulative is None:
                    _, _, df_base_cum = load_data_for_change_page(month_str)
                    # 누적 데이터도 안전하게 문자열로 변환해 둡니다.
                    st.session_state["df_cumulative"] = df_base_cum.fillna("").astype(str)
                    st.session_state["df_cumulative_stats"] = st.session_state["df_cumulative"]

                # 휴일 데이터 로드
                if "df_special_schedules" not in st.session_state or st.session_state.df_special_schedules is None:
                    st.session_state.df_special_schedules = load_special_schedules(month_str)

                # -----------------------------------------------------------
                # 3. 통계표 재계산 (화면에 찍힌 그 정상 데이터프레임 생성)
                # -----------------------------------------------------------
                df_stats_calculated = calculate_stats_from_schedule(df_final_exist)
                
                # (디버깅용: 사용자님이 확인하신 것과 동일한지 확인)
                # st.dataframe(df_final_exist)      # 스케줄 확인
                # st.dataframe(df_stats_calculated) # 통계 확인

                # -----------------------------------------------------------
                # 4. 엑셀 생성 (이제 진짜 새 데이터로 만듭니다)
                # -----------------------------------------------------------
                excel_bytes = create_formatted_excel(df_final_exist, df_stats_calculated)
                
                # 생성된 새 파일을 세션에 저장
                st.session_state.final_download_ready = excel_bytes
                st.session_state.load_error = None 

        except Exception as e:
            st.session_state.final_download_ready = None
            st.session_state.load_error = str(e)
            st.error(f"엑셀 생성 중 오류: {e}")

        # 버튼 표시
        if st.session_state.get("final_download_ready"):
            st.download_button(
                label="📥 방배정 최종 다운로드",
                data=st.session_state.final_download_ready,
                file_name=f"{month_str} 방배정_최종.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                key="download_btn_top_fixed_final_v8" # 키 변경으로 강제 리렌더링 유도
            )
        else:
            if st.session_state.get("load_error"):
                st.error(f"데이터 로드 실패: {st.session_state.load_error}")
                
    # --- B. 방배정 최종 버전 초기화 ---
    with c_reset:
        with st.expander("🗑️ 방배정 최종 버전 초기화"):
            st.error(
                    "이 작업은 되돌릴 수 없습니다!\n 기존 '방배정 최종' 버전 시트를 삭제하고 스케줄 변경을 다시 수행하시겠습니까?"
                )
                    
            if st.button("네, 삭제합니다.", type="primary", use_container_width=True, key="delete_final_confirm"):
                with st.spinner("최종 버전 시트를 삭제하는 중입니다..."):
                    try:
                        gc_del = get_gspread_client()
                        sh_del = gc_del.open_by_url(st.secrets["google_sheet"]["url"])
                        
                        sheets_to_del = [f"{month_str} 방배정 최종"]
                        deleted_cnt = 0
                        
                        for s_name in sheets_to_del:
                            try:
                                ws_del = sh_del.worksheet(s_name)
                                sh_del.del_worksheet(ws_del)
                                deleted_cnt += 1
                            except gspread.exceptions.WorksheetNotFound:
                                pass
                        
                        if deleted_cnt > 0:
                            st.success("✅ 초기화 완료. 페이지를 새로고침합니다.")
                            keys_to_clear = ["final_download_ready", "show_final_results", "change_data_loaded", "load_error"]
                            for k in keys_to_clear:
                                if k in st.session_state: del st.session_state[k]
                            time.sleep(1.5)
                            st.rerun()
                        else:
                            st.warning("삭제할 시트가 존재하지 않습니다.")
                            
                    except Exception as e:
                        st.error(f"초기화 실패: {e}")

# 초기 데이터 로드
if not st.session_state.change_data_loaded:
    load_and_initialize_data()

st.divider()

st.subheader("📋 방배정 변경 요청 목록")
# --- st.subheader("📋 방배정 변경 요청 목록") 섹션 내부 ---

if not st.session_state.df_change_requests.empty:
    df_display = st.session_state.df_change_requests.copy()
    
    # 날짜 포맷을 보기 좋게 변경하는 함수
    def convert_date_format(x):
        x = str(x).strip()
        match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', x)
        if match:
            date_str, slot = match.groups()
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                weekday_str = '월화수목금토일'[date_obj.weekday()]
                return f"{date_obj.month}월 {date_obj.day}일 ({weekday_str}) - {slot}"
            except ValueError:
                return x
        return x

    df_display['변경 요청한 방배정'] = df_display['변경 요청한 방배정'].apply(convert_date_format)
    if 'RequestID' in df_display.columns:
        df_display = df_display.drop(columns=['RequestID'])
    if '요청자 사번' in df_display.columns:
        df_display = df_display.drop(columns=['요청자 사번'])
    
    st.dataframe(df_display, use_container_width=True, hide_index=True)

    # --- 💡 [추가] 충돌 감지 경고 메시지 로직 ---
    request_sources = []
    request_destinations = []

    for index, row in st.session_state.df_change_requests.iterrows():
        change_request_str = str(row.get('변경 요청', '')).strip()
        slot_info_str = str(row.get('변경 요청한 방배정', '')).strip()
        
        if '➡️' in change_request_str and slot_info_str:
            person_before, person_after = [p.strip() for p in change_request_str.split('➡️')]
            
            # 1. 출처 충돌 검사 리스트 추가
            # 동일한 슬롯에 대한 요청이 여러 개 있는지 확인
            request_sources.append(slot_info_str)
            
            # 2. 도착지 중복 검사 리스트 추가
            date_match = re.match(r'(\d{4}-\d{2}-\d{2}) \((.+)\)', slot_info_str)
            if date_match:
                date_part, slot_name = date_match.groups()
                # 시간대만 추출 (예: "8:30(1)_당직" -> "8:30")
                time_part_match = re.match(r'(\d{1,2}:\d{2})', slot_name)
                if time_part_match:
                    time_part = time_part_match.group(1)
                    # (날짜, 시간대, 변경 후 사람)을 기준으로 중복 확인
                    request_destinations.append((date_part, time_part, person_after))

    # [검사 1: 출처 충돌]
    source_counts = Counter(request_sources)
    source_conflicts = [item for item, count in source_counts.items() if count > 1]
    if source_conflicts:
        st.warning(
            "⚠️ **요청 출처 충돌**: 동일한 방(시간대)에 대한 변경 요청이 2개 이상 있습니다. "
            "목록의 가장 위에 있는 요청이 먼저 반영되며, 이후 요청은 무시될 수 있습니다."
        )
        for conflict_item in source_conflicts:
            formatted_slot = convert_date_format(conflict_item)
            st.info(f"- **{formatted_slot}** 에 대한 요청이 중복되었습니다.")

    # [검사 2: 도착지 중복]
    dest_counts = Counter(request_destinations)
    dest_conflicts = [item for item, count in dest_counts.items() if count > 1]
    if dest_conflicts:
        st.warning(
            "⚠️ **요청 도착지 중복**: 한 사람이 같은 날, 같은 시간대에 여러 방에 배정될 가능성이 있는 요청이 있습니다. "
            "이 경우, 먼저 처리되는 요청만 반영됩니다."
        )
        for date, period, person in dest_conflicts:
            # 날짜 포맷팅을 위해 임시 문자열 생성
            temp_slot_info = f"{date} ({period})"
            formatted_date = convert_date_format(temp_slot_info)
            # 시간대만 표시하도록 재조정 (예: "10월 23일 (목) - 8:30")
            display_text = formatted_date.split(' - ')[0] + f" - {period} 시간대"
            st.info(f"- **'{person}'** 님이 **{display_text}** 에 중복으로 배정될 가능성이 있습니다.")

else:
    st.info("접수된 변경 요청이 없습니다.")
st.divider()

# --- UI 및 데이터 핸들링 (수정된 부분) ---
st.subheader("✍️ 방배정 최종 수정")

# [추가] 현재 로드된 버전에 따른 안내 메시지 표시
version_status = st.session_state.get("loaded_version_status", "")

if version_status == "Final":
    st.success("✅ 현재 표시되는 방배정 버전은 '**최종**'입니다. 방배정 변경이 이미 완료되었습니다.")
elif version_status == "ver1.0":
    st.success("ℹ️ 현재 표시되는 방배정 버전은 '**ver1.0**'입니다.")

st.write("- 요청사항을 **일괄 적용/취소**하거나, 셀을 더블클릭하여 직접 수정한 후 **최종 저장 버튼**을 누르세요.\n- 하단에서 방배정 수행 버튼을 누르면 위 변경사항이 반영된 '**방배정 최종**' 버전이 저장됩니다.")
col1, col2 = st.columns(2)
# [추가] 세션에 저장된 메시지를 항상 표시하는 로직
if "page7_messages" in st.session_state and st.session_state["page7_messages"]:
    for msg_type, msg_text in st.session_state["page7_messages"]:
        if msg_type == 'success':
            st.success(msg_text)
        elif msg_type == 'warning':
            st.warning(msg_text)
        elif msg_type == 'error':
            st.error(msg_text)
        elif msg_type == 'info':
            st.info(msg_text)

with col1:
    if st.button("🔄 요청사항 일괄 적용"):
        # 메시지 리스트를 먼저 비워줍니다.
        st.session_state['page7_messages'] = []
        if not st.session_state.df_change_requests.empty:
            current_df = st.session_state.df_final_assignment
            requests_df = st.session_state.df_change_requests
            special_df = st.session_state.df_special_schedules
            st.session_state.df_before_apply = current_df.copy()
            
            # [수정] 4개의 반환값을 모두 받음
            modified_df, new_changes, modified_special_df, messages = apply_assignment_swaps(current_df, requests_df, special_df)
            
            # [수정] 반환된 메시지를 세션에 저장
            st.session_state['page7_messages'] = messages
            
            st.session_state.df_final_assignment = modified_df
            st.session_state.df_special_schedules = modified_special_df
            if not isinstance(st.session_state.changed_cells_log, list):
                st.session_state.changed_cells_log = []
            existing_keys = {(log['날짜'], log['방배정']) for log in st.session_state.changed_cells_log}
            for change in new_changes:
                if (change['날짜'], change['방배정']) not in existing_keys:
                    st.session_state.changed_cells_log.append(change)
                    existing_keys.add((change['날짜'], change['방배정']))
            st.session_state.has_changes_to_revert = True
            st.rerun()
        else:
            # [수정] 직접 메시지를 표시하는 대신 세션에 저장
            st.session_state['page7_messages'] = [('info', "ℹ️ 처리할 변경 요청이 없습니다.")]
            st.rerun()
with col2:
    if st.button("⏪ 적용 취소", disabled=not st.session_state.has_changes_to_revert):
        st.session_state.df_final_assignment = st.session_state.df_before_apply.copy()
        st.session_state.changed_cells_log = []
        st.session_state.has_changes_to_revert = False
        # [수정] 직접 메시지를 표시하는 대신 세션에 저장
        st.session_state['page7_messages'] = [('info', "변경사항이 취소되고 원본 스케줄로 돌아갑니다.")]
        st.rerun()

# 실시간 차이 비교 및 로그 생성 준비
batch_log = st.session_state.get("changed_cells_log", [])
manual_change_log = []
oncall_warnings = []

base_df = st.session_state.df_final_assignment 

edited_df = st.data_editor(
    st.session_state.df_final_assignment,
    use_container_width=True,
    # [수정] 키를 변수로 설정하여 버튼 누를 때마다 강제 리셋
    key=f"assignment_editor_top_{st.session_state['editor_key']}", 
    disabled=['날짜', '요일'],
    hide_index=True
)

# 변경 사항 감지 및 로그 생성 (통합 로직)
if not edited_df.equals(base_df):
    diff_mask = (edited_df != base_df) & (edited_df.notna() | base_df.notna())
    
    for col in diff_mask.columns:
        if diff_mask[col].any():
            for idx in diff_mask.index[diff_mask[col]]:
                date_val = edited_df.at[idx, '날짜']
                day_val = edited_df.at[idx, '요일']
                
                new_val = str(edited_df.at[idx, col]).strip() if pd.notna(edited_df.at[idx, col]) else ""
                old_val = str(base_df.at[idx, col]).strip() if pd.notna(base_df.at[idx, col]) else ""

                if new_val != old_val:
                    # 일반 로그 추가
                    manual_change_log.append({
                        '날짜': f"{date_val} ({day_val})",
                        '방배정': col,
                        '변경 전 인원': old_val,
                        '변경 후 인원': new_val
                    })
                    
                    # [수정 2] 당직/온콜 경고 메시지 통합
                    if '온콜' in col or '당직' in col:
                        # A -> B
                        if old_val and new_val:
                             oncall_warnings.append(f"• {date_val}: '{old_val}' 오전당직 누적 -1, '{new_val}' 누적 +1")
                        # A -> 빈 값
                        elif old_val:
                             oncall_warnings.append(f"• {date_val}: '{old_val}' 오전당직 누적 -1")
                        # 빈 값 -> B
                        elif new_val:
                             oncall_warnings.append(f"• {date_val}: '{new_val}' 오전당직 누적 +1")

# 로그 표시
final_log_to_display = batch_log + manual_change_log

st.write(" ")
st.caption("📝 변경사항 미리보기")

# 2. 일괄 적용 로그와 수동 변경 로그를 합쳐서 표시
batch_log = st.session_state.get("swapped_assignments_log", [])
st.session_state["final_change_log"] = batch_log + manual_change_log

if st.session_state["final_change_log"]:
    log_df = pd.DataFrame(st.session_state["final_change_log"])
    st.dataframe(log_df, use_container_width=True, hide_index=True)
else:
    st.info("기록된 변경사항이 없습니다.")

# --- ▼▼▼ [경고 메시지 표시 로직 추가] (L1448 다음 줄) ▼▼▼ ---
if oncall_warnings:
    # 리스트의 중복을 제거하고 날짜순으로 정렬
    sorted_warnings = sorted(list(set(oncall_warnings)))
    
    # [수정] 경고 메시지에 안내 문구 추가
    warning_text = (
        "🔔 **오전당직 누적 수치 변경 알림**\n\n" +
        "\n".join(sorted_warnings) +
        "\n\n(하단 '방배정 수행' 버튼을 누르면 이 누적 수치가 최종 저장됩니다.)"
    )
    st.warning(warning_text)
# --- ▲▲▲ [추가 완료] ▲▲▲

st.divider()

# --- 2. 방배정 수행 버튼 (저장 및 결과 보기) ---
# [핵심 변경] '변경사항 저장' 버튼 삭제하고, '수행' 버튼 하나로 통합

# 2. 캐시된 함수를 호출하여 3개 시트의 존재 여부 확인
curr_dt = datetime.strptime(month_str, "%Y년 %m월")
next_dt = curr_dt + relativedelta(months=1)
next_month_str = next_dt.strftime("%Y년 %-m월")
final_sheets_exist = check_final_sheets_exist(month_str, next_month_str)

if final_sheets_exist:
    st.warning(
        "⚠️ **덮어쓰기 경고**\n\n"
        "이미 Google Sheets에 다음달의 방배정 최종 결과 시트가 존재합니다.\n\n"
        "배정을 다시 수행하면 현재 화면의 설정을 기준으로 **기존 시트들을 덮어쓰기**합니다."
    )

if st.button("🚀 최종 방배정 수행", type="primary", use_container_width=True):
    with st.spinner("수기 수정사항을 초기화하고, 원본 상태로 '방배정 최종' 시트에 저장합니다..."):
        try:
            gc = get_gspread_client()
            sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
            
            final_sheet_name = f"{month_str} 방배정 최종"

            try:
                worksheet_final = sheet.worksheet(final_sheet_name)
            except gspread.exceptions.WorksheetNotFound:
                worksheet_final = sheet.add_worksheet(title=final_sheet_name, rows=100, cols=30)
            
            original_df = st.session_state.df_final_assignment
            
            # 1. 시트 저장
            final_data_list = [original_df.columns.tolist()] + original_df.fillna('').values.tolist()
            update_sheet_with_retry(worksheet_final, final_data_list)
            
            # 2. 화면 상태 업데이트
            st.session_state['show_final_results'] = True
            st.session_state["editor_key"] += 1 

            st.session_state.changed_cells_log = []
            st.session_state.has_changes_to_revert = False
            
            # 3. 기준점 재설정 (스케줄)
            st.session_state.df_final_assignment_base = original_df.copy()
            
            # 4. [핵심 추가] 기준점 재설정 (통계)
            # 이제 저장된 기준 통계도 방금 확정된 스케줄을 바탕으로 재계산된 값을 가집니다.
            # 그래야 하단 에디터와 비교할 때 '차이 없음'이 나옵니다.
            new_base_stats = calculate_stats_from_schedule(original_df)
            st.session_state.df_cumulative_stats = new_base_stats.copy()

            st.success(f"✅ '{final_sheet_name}' 시트가 원본 상태로 저장/초기화 되었습니다.")
            time.sleep(1)
            st.rerun()

        except Exception as e:
            st.error(f"저장 및 수행 중 오류 발생: {e}")

# ---------------------------------------------------------------------------
# [하단 섹션] 방배정 결과 검토 및 수정 (덮어쓰기 모드)
# ---------------------------------------------------------------------------
if st.session_state.get('show_final_results', False):
    st.divider()
    
    # 1. 기준 데이터 로드
    if 'df_final_assignment_base' not in st.session_state:
        st.session_state.df_final_assignment_base = st.session_state.df_final_assignment.copy()
    
    # 2. 방배정 스케줄 에디터
    st.markdown("**✅ 방배정 스케줄 (수정 가능)**") 
    edited_final_schedule = st.data_editor(
        st.session_state.df_final_assignment_base, 
        use_container_width=True,
        hide_index=True,
        disabled=['날짜', '요일'],
        key=f"final_schedule_editor_{st.session_state['editor_key']}"
    )
    
    # 3. 통계 자동 재계산
    with st.spinner("통계 재계산 중..."):
        # (A) 현재 화면 데이터로 계산
        recalculated_stats = calculate_stats_from_schedule(edited_final_schedule)
        # (B) 저장된 원본 데이터로 계산 (비교 기준)
        original_stats_df = calculate_stats_from_schedule(st.session_state.df_final_assignment_base)

    # 4. 스케줄 변경 로그
    st.markdown("📝 **방배정 스케줄 수정사항**")
    schedule_logs = []
    if not edited_final_schedule.equals(st.session_state.df_final_assignment_base):
        try:
            diff_indices = np.where(edited_final_schedule.astype(str).ne(st.session_state.df_final_assignment_base.astype(str)))
            changed_cells = set(zip(diff_indices[0], diff_indices[1]))
            for row_idx, col_idx in changed_cells:
                date_str = edited_final_schedule.iloc[row_idx, 0]
                slot_name = edited_final_schedule.columns[col_idx]
                old_value = st.session_state.df_final_assignment_base.iloc[row_idx, col_idx]
                new_value = edited_final_schedule.iloc[row_idx, col_idx]
                schedule_logs.append(f"{date_str} '{slot_name}': '{old_value}' → '{new_value}'")
        except: pass
            
    if schedule_logs:
        st.code("\n".join(sorted(schedule_logs)), language='text')
    else:
        st.info("수정된 사항이 없습니다.")

    st.divider()

    # 5. 통계 테이블 에디터
    st.markdown("**☑️ 통계 테이블 (수정 가능)**")
    edited_final_stats = st.data_editor(
        recalculated_stats,
        use_container_width=True,
        hide_index=True,
        disabled=['항목'],
        key="final_stats_editor_unique"
    )

    # 6. 통계 변경 로그
    st.markdown("📝 **통계 테이블 수정사항**")
    stats_change_log = []
    if not edited_final_stats.equals(original_stats_df):
        try:
            diffs = np.where(edited_final_stats.astype(str).ne(original_stats_df.astype(str)))
            changed_indices = set(zip(diffs[0], diffs[1]))
            for r_idx, c_idx in changed_indices:
                stat_name = edited_final_stats.iloc[r_idx, 0]
                person_name = edited_final_stats.columns[c_idx]
                old_val = original_stats_df.iloc[r_idx, c_idx]
                new_val = edited_final_stats.iloc[r_idx, c_idx]
                if str(old_val) != str(new_val):
                    stats_change_log.append(f"{person_name} '{stat_name}': {old_val} → {new_val}")
        except: pass

    if stats_change_log:
        st.code("\n".join(sorted(stats_change_log)), language='text')
    else:
        st.info("수정된 사항이 없습니다.")

    st.divider()

    # ---------------------------------------------------------------------------
    # [변경 감지 및 버튼 제어]
    # ---------------------------------------------------------------------------
    def check_diff(df1, df2):
        if df1 is None: df1 = pd.DataFrame()
        if df2 is None: df2 = pd.DataFrame()
        
        d1 = df1.reset_index(drop=True)
        d2 = df2.reset_index(drop=True)
        
        # 모양 맞추기
        cols = sorted(list(set(d1.columns) | set(d2.columns)))
        d1 = d1.reindex(columns=cols).fillna("").astype(str)
        d2 = d2.reindex(columns=cols).fillna("").astype(str)
        
        max_len = max(len(d1), len(d2))
        d1 = d1.reindex(range(max_len)).fillna("")
        d2 = d2.reindex(range(max_len)).fillna("")
        
        # 공백/특수문자/숫자형식 제거 후 비교
        def clean(x): 
            return x.strip().lower().replace(u'\xa0', ' ').replace('nan', '').replace('.0', '')
        return not d1.map(clean).equals(d2.map(clean))

    is_modified = check_diff(edited_final_schedule, st.session_state.df_final_assignment_base) or \
                  check_diff(edited_final_stats, original_stats_df)

    # [자동 파일 생성] 변경사항이 없고 파일이 없으면 -> 자동 생성 (다운로드 버튼 활성화를 위해)
    if not is_modified and st.session_state.get('download_file') is None:
        try:
            excel_data = create_formatted_excel(edited_final_schedule, edited_final_stats)
            st.session_state.download_file = excel_data
            st.session_state.download_filename = f"{month_str} 방배정_최종.xlsx"
        except: pass

    # ---------------------------------------------------------------------------
    # [버튼 UI]
    # ---------------------------------------------------------------------------
    c1, c2 = st.columns(2)
    
    with c1:
        # 수정사항이 있을 때만 저장 가능
        if st.button("💾 수정사항 Google Sheet에 저장", type="primary", use_container_width=True, disabled=not is_modified):
            with st.spinner("저장 중..."):
                try:
                    gc = get_gspread_client()
                    sheet = gc.open_by_url(st.secrets["google_sheet"]["url"])
                    
                    # A. 방배정 저장
                    ws_name = f"{month_str} 방배정 최종"
                    try: ws = sheet.worksheet(ws_name)
                    except: ws = sheet.add_worksheet(ws_name, 100, 30)
                    ws.clear()
                    ws.update('A1', [edited_final_schedule.columns.tolist()] + edited_final_schedule.fillna('').values.tolist())
                    
                    # B. 통계 저장
                    next_m = (datetime.strptime(month_str, "%Y년 %m월") + relativedelta(months=1)).strftime("%Y년 %-m월")
                    cum_name = f"{next_m} 누적 최종"
                    try: ws_cum = sheet.worksheet(cum_name)
                    except: ws_cum = sheet.add_worksheet(cum_name, 100, 30)
                    ws_cum.clear()
                    ws_cum.update('A1', [edited_final_stats.columns.tolist()] + edited_final_stats.fillna('').values.tolist())
                    
                    # [중요] 엑셀 파일도 최신 데이터로 갱신
                    new_excel = create_formatted_excel(edited_final_schedule, edited_final_stats)
                    st.session_state.download_file = new_excel
                    st.session_state.download_filename = f"{month_str} 방배정_최종.xlsx"

                    # 기준점 업데이트 (수정사항 없음 상태로 전환)
                    st.session_state.df_final_assignment_base = edited_final_schedule.copy()
                    st.session_state.df_final_assignment = edited_final_schedule.copy()
                    
                    st.success(f"✅ '{ws_name}' 시트에 수정된 내용이 저장되었습니다.")
                    time.sleep(1)
                    st.rerun()
                except Exception as e:
                    st.error(f"저장 실패: {e}")

    with c2:
        if is_modified:
            st.error("⚠️ 수정사항이 감지되었습니다. 먼저 '수정사항 Google Sheet에 저장' 버튼을 눌러주세요.")
            st.button("📥 방배정 최종 다운로드", disabled=True, use_container_width=True)
        else:
            if st.session_state.get('download_file'):
                st.download_button(
                    label="📥 방배정 최종 다운로드",
                    data=st.session_state.download_file,
                    file_name=st.session_state.download_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
            else:
                pass